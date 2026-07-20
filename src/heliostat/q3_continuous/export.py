"""独立第三问连续 Campo 参数、逐镜结果和提交表输出。"""

from __future__ import annotations

import csv
import json
from copy import copy
from dataclasses import asdict
from pathlib import Path
from typing import Any, Iterable, Sequence

import numpy as np
from openpyxl import load_workbook

from .evaluate import EvaluationProfile, HeterogeneousEvaluation
from .model import CampoMotherField, ContinuousDesign, SUPPORTED_ZONES
from .search import CampoDiagnostics


TARGET_ANNUAL_POWER_MW = 42.0


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        raise ValueError(f"没有可写入 {path.name} 的结果。")
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def write_result3_workbook(
    *,
    template_path: str | Path,
    output_path: str | Path,
    evaluation: HeterogeneousEvaluation,
    tower_x: float,
    tower_y: float,
) -> Path:
    """按题目模板写出塔坐标和每面镜子的异构规格、位置。"""

    template = Path(template_path)
    if not template.exists():
        raise FileNotFoundError(f"找不到 result3.xlsx 模板：{template}")
    destination = Path(output_path)
    destination.parent.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(template)
    sheet = workbook.active
    if sheet.max_column < 8:
        workbook.close()
        raise ValueError("result3.xlsx 模板列数不足 8 列。")

    style_row = 2 if sheet.max_row >= 2 else 1
    styles = [
        copy(sheet.cell(style_row, column)._style)
        for column in range(1, 9)
    ]
    number_formats = [
        sheet.cell(style_row, column).number_format
        for column in range(1, 9)
    ]
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)

    for index in range(evaluation.mirror_count):
        row_index = index + 2
        values = (
            tower_x,
            tower_y,
            index + 1,
            float(evaluation.widths[index]),
            float(evaluation.heights[index]),
            float(evaluation.coordinates[index, 0]),
            float(evaluation.coordinates[index, 1]),
            float(evaluation.installation_heights[index]),
        )
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column, value)
            cell._style = copy(styles[column - 1])
            cell.number_format = number_formats[column - 1]

    workbook.save(destination)
    workbook.close()
    return destination


def _zone_rows(
    evaluation: HeterogeneousEvaluation,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for zone in SUPPORTED_ZONES:
        active = evaluation.zone_indices == zone
        rows.append(
            {
                "zone": zone,
                "ring_count": int(
                    np.unique(evaluation.ring_indices[active]).size
                ),
                "mirror_count": int(np.count_nonzero(active)),
                "minimum_width_m": float(np.min(evaluation.widths[active])),
                "maximum_width_m": float(np.max(evaluation.widths[active])),
                "minimum_height_m": float(
                    np.min(evaluation.heights[active])
                ),
                "maximum_height_m": float(
                    np.max(evaluation.heights[active])
                ),
                "minimum_installation_height_m": float(
                    np.min(evaluation.installation_heights[active])
                ),
                "maximum_installation_height_m": float(
                    np.max(evaluation.installation_heights[active])
                ),
                "total_area_m2": float(
                    np.sum(
                        evaluation.widths[active]
                        * evaluation.heights[active]
                    )
                ),
            }
        )
    return rows


def _stage_rows(
    stages: Iterable[tuple[str, HeterogeneousEvaluation]],
) -> list[dict[str, Any]]:
    return [
        {
            "stage": name,
            "profile": evaluation.profile_name,
            "mirror_count": evaluation.mirror_count,
            "total_area_m2": evaluation.total_area_m2,
            "annual_power_mw": evaluation.annual_power_mw,
            "unit_area_power_kw_m2": (
                evaluation.unit_area_power_kw_m2
            ),
        }
        for name, evaluation in stages
    ]


def _parameter_rows(
    design: ContinuousDesign,
) -> tuple[tuple[str, str, float], ...]:
    return (
        ("size_zone1_slope", "区域 1 尺寸行号趋势", design.size_zone1_slope),
        ("size_zone2_slope", "区域 2 尺寸行号趋势", design.size_zone2_slope),
        ("size_zone2_offset", "进入区域 2 的尺寸修正", design.size_zone2_offset),
        ("size_azimuth", "同环南北尺寸修正", design.size_azimuth),
        ("height_offset", "全场安装高度平移", design.height_offset),
        (
            "height_zone1_slope",
            "区域 1 高度行号趋势",
            design.height_zone1_slope,
        ),
        (
            "height_zone2_slope",
            "区域 2 高度行号趋势",
            design.height_zone2_slope,
        ),
        (
            "height_zone2_offset",
            "进入区域 2 的高度修正",
            design.height_zone2_offset,
        ),
        ("height_azimuth", "同环南北高度修正", design.height_azimuth),
        ("area_ratio", "相对起始镜场总面积比例", design.area_ratio),
    )


def write_question3_results(
    *,
    output_dir: str | Path,
    mother: CampoMotherField,
    design: ContinuousDesign,
    evaluation: HeterogeneousEvaluation,
    result3_template: str | Path,
    stages: Iterable[tuple[str, HeterogeneousEvaluation]] = (),
    diagnostics: CampoDiagnostics | None = None,
    model_name: str = "campo-monotone-radial",
    legacy_comparison: dict[str, Any] | None = None,
) -> dict[str, Path]:
    destination = Path(output_dir)
    destination.mkdir(parents=True, exist_ok=True)

    coordinate_rows = [
        {
            "mirror_id": index + 1,
            "original_campo_mirror_id": (
                int(evaluation.original_indices[index]) + 1
            ),
            "ring_index": int(evaluation.ring_indices[index]),
            "campo_zone": int(evaluation.zone_indices[index]),
            "zone_row_index": int(evaluation.zone_row_indices[index]),
            "normalized_zone_row": float(
                evaluation.normalized_rows[index]
            ),
            "azimuth_rad": float(evaluation.azimuth_angles[index]),
            "centered_azimuth_feature": float(
                evaluation.azimuth_features[index]
            ),
            "nominal_ring_count": int(
                evaluation.nominal_ring_counts[index]
            ),
            "actual_ring_count": int(
                evaluation.actual_ring_counts[index]
            ),
            "mirror_width_m": float(evaluation.widths[index]),
            "mirror_height_m": float(evaluation.heights[index]),
            "x_m": float(evaluation.coordinates[index, 0]),
            "y_m": float(evaluation.coordinates[index, 1]),
            "z_m": float(evaluation.installation_heights[index]),
        }
        for index in range(evaluation.mirror_count)
    ]
    monthly_rows = [
        asdict(record)
        for record in evaluation.solution.monthly_results
    ]
    mirror_rows = [
        {
            **asdict(record),
            "original_campo_mirror_id": (
                int(evaluation.original_indices[index]) + 1
            ),
            "ring_index": int(evaluation.ring_indices[index]),
            "campo_zone": int(evaluation.zone_indices[index]),
            "zone_row_index": int(
                evaluation.zone_row_indices[index]
            ),
            "normalized_zone_row": float(
                evaluation.normalized_rows[index]
            ),
            "centered_azimuth_feature": float(
                evaluation.azimuth_features[index]
            ),
            "mirror_width_m": float(evaluation.widths[index]),
            "mirror_height_m": float(evaluation.heights[index]),
            "installation_height_m": float(
                evaluation.installation_heights[index]
            ),
            "mirror_area_m2": float(
                evaluation.widths[index] * evaluation.heights[index]
            ),
        }
        for index, record in enumerate(
            evaluation.solution.mirror_annual_results
        )
    ]
    annual = asdict(evaluation.solution.annual_result)
    zones = _zone_rows(evaluation)
    stage_data = _stage_rows(stages)

    stages_path = destination / "02_分阶段方案比较.json"
    coordinates_path = destination / "03_最终逐镜参数与坐标.csv"
    monthly_path = destination / "04_月平均计算结果.csv"
    annual_path = destination / "05_年平均计算结果.json"
    mirror_path = destination / "06_单镜年平均结果.csv"
    summary_path = destination / "07_最终方案摘要.json"
    table_path = destination / "08_论文结果与验证表.md"
    workbook_path = destination / "10_第三问提交结果.xlsx"

    stages_path.write_text(
        json.dumps(stage_data, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    _write_csv(coordinates_path, coordinate_rows)
    _write_csv(monthly_path, monthly_rows)
    _write_csv(mirror_path, mirror_rows)
    annual_path.write_text(
        json.dumps(annual, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    summary = {
        "layout": "q2-campo-continuous-heterogeneous",
        "model": model_name,
        "annual_power_constraint_mw": TARGET_ANNUAL_POWER_MW,
        "annual_power_margin_mw": (
            evaluation.annual_power_mw - TARGET_ANNUAL_POWER_MW
        ),
        "constraint_satisfied": evaluation.is_feasible(
            TARGET_ANNUAL_POWER_MW
        ),
        "tower": {
            "x_m": mother.parameters.tower_x,
            "y_m": mother.parameters.tower_y,
        },
        "mirror_count": evaluation.mirror_count,
        "ring_count": int(np.unique(evaluation.ring_indices).size),
        "total_area_m2": evaluation.total_area_m2,
        "continuous_design": asdict(design),
        "zone_summaries": zones,
        "diagnostics": (
            asdict(diagnostics) if diagnostics is not None else None
        ),
        "geometry": asdict(evaluation.geometry),
        "annual": annual,
    }
    summary_path.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    lines = [
        "# 第三问结果与验证表",
        "",
        "## 表 1 功率约束与优化目标",
        "",
        "| 年平均功率下限 (MW) | 年平均功率 (MW) | 功率余量 (MW) | 总镜面面积 (m²) | 单位面积年平均输出 (kW/m²) | 是否满足约束 |",
        "| ---: | ---: | ---: | ---: | ---: | :---: |",
        (
            f"| {TARGET_ANNUAL_POWER_MW:.6f} "
            f"| {evaluation.annual_power_mw:.6f} "
            f"| {evaluation.annual_power_mw - TARGET_ANNUAL_POWER_MW:.6f} "
            f"| {evaluation.total_area_m2:.3f} "
            f"| {evaluation.unit_area_power_kw_m2:.6f} "
            f"| {'是' if evaluation.is_feasible() else '否'} |"
        ),
        "",
        "## 表 2 连续规格参数",
        "",
        "| 参数 | 含义 | 结果 |",
        "| --- | --- | ---: |",
    ]
    for name, meaning, value in _parameter_rows(design):
        lines.append(f"| `{name}` | {meaning} | {value:.9f} |")

    lines.extend(
        [
            "",
            "## 表 3 Campo 区域规格范围",
            "",
            "| 区域 | 圆环数 | 镜子数 | 宽度范围 (m) | 高度范围 (m) | 安装高度范围 (m) | 区域总面积 (m²) |",
            "| ---: | ---: | ---: | ---: | ---: | ---: | ---: |",
        ]
    )
    for row in zones:
        lines.append(
            f"| {row['zone']} "
            f"| {row['ring_count']} "
            f"| {row['mirror_count']} "
            f"| {row['minimum_width_m']:.6f}–{row['maximum_width_m']:.6f} "
            f"| {row['minimum_height_m']:.6f}–{row['maximum_height_m']:.6f} "
            f"| {row['minimum_installation_height_m']:.6f}–"
            f"{row['maximum_installation_height_m']:.6f} "
            f"| {row['total_area_m2']:.3f} |"
        )

    lines.extend(
        [
            "",
            "## 表 4 分阶段与模型消融",
            "",
            "| 阶段 | 评价精度 | 镜子数 | 总面积 (m²) | 年平均功率 (MW) | 单位面积输出 (kW/m²) |",
            "| --- | --- | ---: | ---: | ---: | ---: |",
        ]
    )
    for row in stage_data:
        lines.append(
            f"| {row['stage']} "
            f"| {row['profile']} "
            f"| {row['mirror_count']} "
            f"| {row['total_area_m2']:.3f} "
            f"| {row['annual_power_mw']:.6f} "
            f"| {row['unit_area_power_kw_m2']:.6f} |"
        )

    geometry = evaluation.geometry
    lines.extend(
        [
            "",
            "## 表 5 异构几何约束复核",
            "",
            "| 检查项 | 实际值 | 约束 | 结果 |",
            "| --- | ---: | ---: | :---: |",
            (
                "| 最小镜心距离 (m) "
                f"| {geometry.minimum_center_distance_m:.9f} "
                "| - "
                f"| {'通过' if geometry.valid else '未通过'} |"
            ),
            (
                "| 最小异构宽度安全余量 (m) "
                f"| {geometry.minimum_width_clearance_m:.9f} "
                "| ≥ 0.010000000 "
                f"| {'通过' if geometry.minimum_width_clearance_m >= 0.01 - 1e-9 else '未通过'} |"
            ),
            (
                "| 最大场地半径 (m) "
                f"| {geometry.maximum_field_radius_m:.6f} "
                "| ≤ 350 "
                f"| {'通过' if geometry.maximum_field_radius_m <= 350.0 + 1e-9 else '未通过'} |"
            ),
            (
                "| 最小塔距 (m) "
                f"| {geometry.minimum_tower_distance_m:.6f} "
                "| ≥ 100 "
                f"| {'通过' if geometry.minimum_tower_distance_m >= 100.0 - 1e-9 else '未通过'} |"
            ),
            (
                "| 最小不触地余量 (m) "
                f"| {geometry.minimum_ground_clearance_m:.6f} "
                "| ≥ 0 "
                f"| {'通过' if geometry.minimum_ground_clearance_m >= -1e-9 else '未通过'} |"
            ),
            "",
            "## 表 6 每月 21 日平均光学效率及输出功率",
            "",
            "| 月份 | 光学效率 | 余弦效率 | 阴影遮挡效率 | 截断效率 | 输出热功率 (MW) | 单位面积输出 (kW/m²) |",
            "| ---: | ---: | ---: | ---: | ---: | ---: | ---: |",
        ]
    )
    for record in evaluation.solution.monthly_results:
        lines.append(
            f"| {record.month} "
            f"| {record.average_optical_efficiency:.6f} "
            f"| {record.average_cosine_efficiency:.6f} "
            f"| {record.average_shadow_blocking_efficiency:.6f} "
            f"| {record.average_truncation_efficiency:.6f} "
            f"| {record.field_output_mw:.6f} "
            f"| {record.unit_area_output_kw_m2:.6f} |"
        )
    table_path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    write_result3_workbook(
        template_path=result3_template,
        output_path=workbook_path,
        evaluation=evaluation,
        tower_x=mother.parameters.tower_x,
        tower_y=mother.parameters.tower_y,
    )
    written = {
        "stages": stages_path,
        "coordinates": coordinates_path,
        "monthly": monthly_path,
        "annual": annual_path,
        "mirror_annual": mirror_path,
        "summary": summary_path,
        "paper_table": table_path,
        "result3": workbook_path,
    }
    if legacy_comparison is not None:
        legacy_path = destination / "11_原六组对照结果.json"
        legacy_path.write_text(
            json.dumps(
                legacy_comparison,
                ensure_ascii=False,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )
        written["legacy_six_group_comparison"] = legacy_path
    return written


def write_dense_validation(
    *,
    output_dir: str | Path,
    evaluation: HeterogeneousEvaluation,
    profile: EvaluationProfile,
    sensitivity_evaluations: Sequence[
        tuple[EvaluationProfile, HeterogeneousEvaluation]
    ] = (),
) -> Path:
    destination = Path(output_dir)
    path = destination / "09_高精度加密验证.json"
    evaluations = ((profile, evaluation), *sensitivity_evaluations)

    def validation_record(
        item_profile: EvaluationProfile,
        item_evaluation: HeterogeneousEvaluation,
    ) -> dict[str, Any]:
        return {
            "profile": {
                "months": len(item_profile.months),
                "solar_times_per_month": len(item_profile.solar_times),
                "shadow_grid_size": item_profile.solver.shadow_grid_size,
                "truncation_rays": item_profile.solver.truncation_rays,
                "neighbor_radius_m": (
                    item_profile.solver.neighbor_radius_m
                ),
            },
            "annual_power_mw": item_evaluation.annual_power_mw,
            "annual_power_margin_mw": (
                item_evaluation.annual_power_mw
                - TARGET_ANNUAL_POWER_MW
            ),
            "unit_area_power_kw_m2": (
                item_evaluation.unit_area_power_kw_m2
            ),
            "constraint_satisfied": item_evaluation.is_feasible(),
        }

    payload = {
        "profile": {
            "months": len(profile.months),
            "solar_times_per_month": len(profile.solar_times),
            "shadow_grid_size": profile.solver.shadow_grid_size,
            "truncation_rays": profile.solver.truncation_rays,
            "neighbor_radius_m": profile.solver.neighbor_radius_m,
        },
        "mirror_count": evaluation.mirror_count,
        "total_area_m2": evaluation.total_area_m2,
        "annual_power_constraint_mw": TARGET_ANNUAL_POWER_MW,
        "annual_power_mw": evaluation.annual_power_mw,
        "annual_power_margin_mw": (
            evaluation.annual_power_mw - TARGET_ANNUAL_POWER_MW
        ),
        "unit_area_power_kw_m2": evaluation.unit_area_power_kw_m2,
        "constraint_satisfied": evaluation.is_feasible(),
        "neighbor_radius_sensitivity": [
            validation_record(item_profile, item_evaluation)
            for item_profile, item_evaluation in evaluations
        ],
    }
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    table_path = destination / "08_论文结果与验证表.md"
    lines = [
        "",
        "## 表 7 高精度加密与邻域敏感性验证",
        "",
        "| 阴影网格 | 截断光线 | 邻镜半径 (m) | 年平均功率 (MW) | 功率余量 (MW) | 单位面积输出 (kW/m²) | 是否满足约束 |",
        "| ---: | ---: | ---: | ---: | ---: | ---: | :---: |",
    ]
    for item_profile, item_evaluation in evaluations:
        lines.append(
            f"| {item_profile.solver.shadow_grid_size}×"
            f"{item_profile.solver.shadow_grid_size} "
            f"| {item_profile.solver.truncation_rays} "
            f"| {item_profile.solver.neighbor_radius_m:.0f} "
            f"| {item_evaluation.annual_power_mw:.6f} "
            f"| {item_evaluation.annual_power_mw - TARGET_ANNUAL_POWER_MW:.6f} "
            f"| {item_evaluation.unit_area_power_kw_m2:.6f} "
            f"| {'是' if item_evaluation.is_feasible() else '否'} |"
        )
    table_content = table_path.read_text(encoding="utf-8")
    marker = "\n## 表 7 "
    if marker in table_content:
        table_content = table_content.split(marker, maxsplit=1)[0].rstrip()
    table_path.write_text(
        table_content + "\n" + "\n".join(lines) + "\n",
        encoding="utf-8",
    )
    return path
