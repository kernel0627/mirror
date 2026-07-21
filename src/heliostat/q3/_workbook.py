"""第三问 result3.xlsx 提交表输出。"""

from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import load_workbook

from ._optics import HeterogeneousEvaluation


def write_result3_workbook(
    *,
    template_path: str | Path,
    output_path: str | Path,
    evaluation: HeterogeneousEvaluation,
    tower_x: float,
    tower_y: float,
) -> Path:
    """按题目模板写出塔坐标和每面镜子的异构规格、位置。"""

    template = Path(template_path)
    if not template.exists():
        raise FileNotFoundError(f"找不到 result3.xlsx 模板：{template}")
    destination = Path(output_path)
    destination.parent.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(template)
    sheet = workbook.active
    if sheet.max_column < 8:
        workbook.close()
        raise ValueError("result3.xlsx 模板列数不足 8 列。")

    style_row = 2 if sheet.max_row >= 2 else 1
    styles = [
        copy(sheet.cell(style_row, column)._style)
        for column in range(1, 9)
    ]
    number_formats = [
        sheet.cell(style_row, column).number_format
        for column in range(1, 9)
    ]
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)

    for index in range(evaluation.mirror_count):
        row_index = index + 2
        values = (
            tower_x,
            tower_y,
            index + 1,
            float(evaluation.widths[index]),
            float(evaluation.heights[index]),
            float(evaluation.coordinates[index, 0]),
            float(evaluation.coordinates[index, 1]),
            float(evaluation.installation_heights[index]),
        )
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column, value)
            cell._style = copy(styles[column - 1])
            cell.number_format = number_formats[column - 1]

    workbook.save(destination)
    workbook.close()
    return destination
