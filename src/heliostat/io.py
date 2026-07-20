"""坐标输入与结果文件输出。"""

from __future__ import annotations

import csv
import json
from dataclasses import asdict
from pathlib import Path
from typing import Any, Iterable, Sequence

import numpy as np
from numpy.typing import NDArray
from openpyxl import load_workbook

from .config import FieldConfig, SolverConfig


FloatArray = NDArray[np.float64]


def load_mirror_xy(path: str | Path, expected_count: int | None = 1745) -> FloatArray:
    """从题目附件读取定日镜 x、y 坐标。"""

    source = Path(path)
    if not source.exists():
        raise FileNotFoundError(f"找不到定日镜坐标文件：{source}")

    if source.suffix.lower() == ".xlsx":
        workbook = load_workbook(source, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        workbook.close()
        values = [(row[0], row[1]) for row in rows if row[0] is not None]
    elif source.suffix.lower() == ".csv":
        with source.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            next(reader, None)
            values = [(row[0], row[1]) for row in reader if row]
    else:
        raise ValueError("坐标文件只支持 .xlsx 或 .csv。")

    try:
        mirror_xy = np.asarray(values, dtype=float)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"坐标文件包含非数值数据：{source}") from exc

    if mirror_xy.ndim != 2 or mirror_xy.shape[1] != 2:
        raise ValueError(f"坐标数据应为 N×2，实际形状为 {mirror_xy.shape}。")
    if not np.all(np.isfinite(mirror_xy)):
        raise ValueError("坐标数据包含 NaN 或无穷值。")
    if expected_count is not None and mirror_xy.shape[0] != expected_count:
        raise ValueError(
            f"应读取 {expected_count} 面定日镜，实际读取 {mirror_xy.shape[0]} 面。"
        )
    return mirror_xy


def _write_csv(path: Path, rows: Sequence[dict[str, Any]]) -> None:
    if not rows:
        raise ValueError(f"没有可写入 {path.name} 的结果。")
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def write_question1_results(
    output_dir: str | Path,
    time_records: Iterable[Any],
    monthly_records: Iterable[Any],
    annual_record: Any,
    field_config: FieldConfig,
    solver_config: SolverConfig,
    source_path: str | Path,
    mirror_count: int,
) -> dict[str, Path]:
    """将 60 时刻、月均、年均和运行参数写入可审计文件。"""

    destination = Path(output_dir)
    destination.mkdir(parents=True, exist_ok=True)

    time_rows = [asdict(record) for record in time_records]
    monthly_rows = [asdict(record) for record in monthly_records]
    annual_row = asdict(annual_record)
    months = sorted({row["month"] for row in time_rows})
    solar_times = sorted({row["solar_time"] for row in time_rows})

    time_path = destination / "time_results.csv"
    monthly_path = destination / "monthly_results.csv"
    annual_path = destination / "annual_results.json"
    run_path = destination / "run_config.json"

    _write_csv(time_path, time_rows)
    _write_csv(monthly_path, monthly_rows)
    annual_path.write_text(
        json.dumps(annual_row, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    run_path.write_text(
        json.dumps(
            {
                "source": str(Path(source_path).resolve()),
                "field": field_config.to_dict(),
                "solver": solver_config.to_dict(),
                "run": {
                    "mirror_count": mirror_count,
                    "months": months,
                    "solar_times": solar_times,
                    "time_state_count": len(time_rows),
                },
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )

    return {
        "time": time_path,
        "monthly": monthly_path,
        "annual": annual_path,
        "config": run_path,
    }
