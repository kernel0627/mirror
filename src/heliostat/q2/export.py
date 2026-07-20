"""第二问结果文件、论文表格和 result2.xlsx 输出。"""

from __future__ import annotations

import csv
import json
from copy import copy
from dataclasses import asdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .evaluate import EvaluationProfile, FieldEvaluation, LayoutParameters
from .layout import validate_layout


TARGET_ANNUAL_POWER_MW = 42.0


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        raise ValueError(f"没有可写入 {path.name} 的结果。")
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def write_result2_workbook(
    *,
    template_path: str | Path,
    output_path: str | Path,
    evaluation: FieldEvaluation,
    parameters: LayoutParameters,
) -> Path:
    """按题目模板写出塔坐标、统一尺寸、高度和全部镜位。"""

    template = Path(template_path)
    if not template.exists():
        raise FileNotFoundError(f"找不到 result2.xlsx 模板：{template}")
    destination = Path(output_path)
    destination.parent.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(template)
    sheet = workbook.active
    if sheet.max_column < 8:
        workbook.close()
        raise ValueError("result2.xlsx 模板列数不足 8 列。")

    style_source = [copy(sheet.cell(2, column)._style) for column in range(1, 9)]
    number_formats = [sheet.cell(2, column).number_format for column in range(1, 9)]
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)

    for row_index, (x_m, y_m) in enumerate(
        evaluation.coordinates,
        start=2,
    ):
        values = (
            parameters.tower_x,
            parameters.tower_y,
            row_index - 1,
            parameters.mirror_width,
            parameters.mirror_height,
            float(x_m),
            float(y_m),
            parameters.installation_height,
        )
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column, value)
            cell._style = copy(style_source[column - 1])
            cell.number_format = number_formats[column - 1]

    workbook.save(destination)
    workbook.close()
    return destination


def write_question2_results(
    *,
    output_dir: str | Path,
    layout_name: str,
    parameters: LayoutParameters,
    evaluation: FieldEvaluation,
    result2_template: str | Path,
    comparison: dict[str, Any] | None = None,
) -> dict[str, Path]:
    """写出坐标、月年平均结果、配置摘要、论文表和 result2.xlsx。"""

    destination = Path(output_dir)
    destination.mkdir(parents=True, exist_ok=True)

    coordinate_rows = [
        {
            "mirror_id": index + 1,
            "mirror_width_m": parameters.mirror_width,
            "mirror_height_m": parameters.mirror_height,
            "x_m": float(x_m),
            "y_m": float(y_m),
            "z_m": parameters.installation_height,
        }
        for index, (x_m, y_m) in enumerate(evaluation.coordinates)
    ]
    monthly_rows = [asdict(record) for record in evaluation.solution.monthly_results]
    mirror_rows = [
        asdict(record) for record in evaluation.solution.mirror_annual_results
    ]
    annual = asdict(evaluation.solution.annual_result)

    coordinates_path = destination / "03_最终镜位坐标.csv"
    monthly_path = destination / "04_月平均计算结果.csv"
    annual_path = destination / "05_年平均计算结果.json"
    mirror_path = destination / "06_单镜年平均结果.csv"
    summary_path = destination / "07_最终方案摘要.json"
    table_path = destination / "08_论文结果与验证表.md"
    workbook_path = destination / "10_第二问提交结果.xlsx"

    _write_csv(coordinates_path, coordinate_rows)
    _write_csv(monthly_path, monthly_rows)
    _write_csv(mirror_path, mirror_rows)
    annual_path.write_text(
        json.dumps(annual, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    summary = {
        "layout": layout_name,
        "annual_power_constraint_mw": TARGET_ANNUAL_POWER_MW,
        "annual_power_margin_mw": (evaluation.annual_power_mw - TARGET_ANNUAL_POWER_MW),
        "constraint_satisfied": (evaluation.annual_power_mw >= TARGET_ANNUAL_POWER_MW),
        "parameters": asdict(parameters),
        "ring_count": evaluation.ring_count,
        "mirror_count": evaluation.mirror_count,
        "mirror_area_m2": evaluation.mirror_area_m2,
        "total_area_m2": evaluation.total_area_m2,
        "annual": annual,
    }
    summary_path.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    lines = [
        "# 第二问结果与验证表",
        "",
        "## 表 1 功率约束与优化目标",
        "",
        "| 年平均输出热功率下限 (MW) | 最终年平均输出热功率 (MW) | 功率余量 (MW) | 是否满足约束 | 单位镜面面积年平均输出热功率 (kW/m²) |",
        "| ---: | ---: | ---: | :---: | ---: |",
        (
            f"| {TARGET_ANNUAL_POWER_MW:.6f} "
            f"| {evaluation.annual_power_mw:.6f} "
            f"| {evaluation.annual_power_mw - TARGET_ANNUAL_POWER_MW:.6f} "
            f"| {'是' if evaluation.annual_power_mw >= TARGET_ANNUAL_POWER_MW else '否'} "
            f"| {evaluation.unit_area_power_kw_m2:.6f} |"
        ),
        "",
        "> 本题中的 42 MW 是年平均输出热功率下限；优化目标是在满足该下限后最大化单位镜面面积年平均输出热功率。",
        "",
        "## 表 2 最终设计参数",
        "",
        "| 布局 | 塔坐标 | 镜面尺寸 | 安装高度 | 镜子数 | 总镜面面积 (m²) | 年平均功率 (MW) | 单位面积功率 (kW/m²) |",
        "| --- | --- | ---: | ---: | ---: | ---: | ---: | ---: |",
        (
            f"| {layout_name} "
            f"| ({parameters.tower_x:.3f}, {parameters.tower_y:.3f}) "
            f"| {parameters.mirror_width:.3f}×{parameters.mirror_height:.3f} "
            f"| {parameters.installation_height:.3f} "
            f"| {evaluation.mirror_count} "
            f"| {evaluation.total_area_m2:.3f} "
            f"| {evaluation.annual_power_mw:.6f} "
            f"| {evaluation.unit_area_power_kw_m2:.6f} |"
        ),
    ]
    if comparison is not None and {
        "partitioned",
        "campo",
    }.issubset(comparison):
        lines.extend(
            [
                "",
                "## 表 3 两种候选布局的正式精度对比",
                "",
                "| 布局 | 安全余量 (m) | 镜子数 | 总镜面面积 (m²) | 年平均功率 (MW) | 单位面积功率 (kW/m²) |",
                "| --- | ---: | ---: | ---: | ---: | ---: |",
            ]
        )
        for kind, label in (
            ("partitioned", "分区交错同心圆"),
            ("campo", "改进 Campo"),
        ):
            record = comparison[kind]
            lines.append(
                f"| {label} "
                f"| {record['parameters']['safety_epsilon']:.6f} "
                f"| {record['mirror_count']} "
                f"| {record['total_area_m2']:.3f} "
                f"| {record['annual_power_mw']:.6f} "
                f"| {record['unit_area_power_kw_m2']:.6f} |"
            )

    geometry = validate_layout(evaluation.coordinates, parameters)
    lines.extend(
        [
            "",
            "## 表 4 几何约束复核",
            "",
            "| 检查项 | 实际值 | 约束 | 结果 |",
            "| --- | ---: | ---: | :---: |",
            (
                "| 最小镜心距离 (m) "
                f"| {geometry.minimum_center_distance:.9f} "
                f"| > {parameters.mirror_width + 5.0:.9f} "
                f"| {'通过' if geometry.valid else '未通过'} |"
            ),
            (
                "| 镜心距离安全余量 (m) "
                f"| {geometry.minimum_center_distance - parameters.mirror_width - 5.0:.9f} "
                "| > 0 | "
                f"{'通过' if geometry.minimum_center_distance > parameters.mirror_width + 5.0 else '未通过'} |"
            ),
            (
                "| 最大场地半径 (m) "
                f"| {geometry.maximum_field_radius:.6f} "
                f"| ≤ {parameters.field_radius:.3f} "
                f"| {'通过' if geometry.maximum_field_radius <= parameters.field_radius + 1e-9 else '未通过'} |"
            ),
            (
                "| 最小塔距 (m) "
                f"| {geometry.minimum_tower_distance:.6f} "
                f"| ≥ {parameters.exclusion_radius:.3f} "
                f"| {'通过' if geometry.minimum_tower_distance >= parameters.exclusion_radius - 1e-9 else '未通过'} |"
            ),
            (
                "| 不触地高度余量 (m) "
                f"| {parameters.installation_height - parameters.mirror_height / 2.0:.6f} "
                "| ≥ 0 "
                f"| {'通过' if parameters.installation_height >= parameters.mirror_height / 2.0 else '未通过'} |"
            ),
            "",
            "## 表 5 月平均光学效率及输出热功率",
            "",
            "| 月份 | 光学效率 | 余弦效率 | 阴影遮挡效率 | 截断效率 | 输出热功率 (MW) | 单位面积功率 (kW/m²) |",
            "| ---: | ---: | ---: | ---: | ---: | ---: | ---: |",
        ]
    )
    for record in evaluation.solution.monthly_results:
        lines.append(
            f"| {record.month} "
            f"| {record.average_optical_efficiency:.6f} "
            f"| {record.average_cosine_efficiency:.6f} "
            f"| {record.average_shadow_blocking_efficiency:.6f} "
            f"| {record.average_truncation_efficiency:.6f} "
            f"| {record.field_output_mw:.6f} "
            f"| {record.unit_area_output_kw_m2:.6f} |"
        )
    table_path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    write_result2_workbook(
        template_path=result2_template,
        output_path=workbook_path,
        evaluation=evaluation,
        parameters=parameters,
    )
    return {
        "coordinates": coordinates_path,
        "monthly": monthly_path,
        "annual": annual_path,
        "mirror_annual": mirror_path,
        "summary": summary_path,
        "paper_table": table_path,
        "result2": workbook_path,
    }


def write_high_precision_validation(
    *,
    output_dir: str | Path,
    evaluation: FieldEvaluation,
    profile: EvaluationProfile,
) -> Path:
    """写出并追加 20×20、512 条光线的高精度可行性复核。"""

    destination = Path(output_dir)
    validation_path = destination / "09_高精度加密验证.json"
    annual = evaluation.solution.annual_result
    payload = {
        "profile": {
            "months": len(profile.months),
            "solar_times_per_month": len(profile.solar_times),
            "shadow_grid_size": profile.solver.shadow_grid_size,
            "truncation_rays": profile.solver.truncation_rays,
            "neighbor_radius_m": profile.solver.neighbor_radius_m,
        },
        "mirror_count": evaluation.mirror_count,
        "annual_power_constraint_mw": TARGET_ANNUAL_POWER_MW,
        "annual_power_mw": evaluation.annual_power_mw,
        "annual_power_margin_mw": (evaluation.annual_power_mw - TARGET_ANNUAL_POWER_MW),
        "unit_area_power_kw_m2": evaluation.unit_area_power_kw_m2,
        "average_optical_efficiency": annual.average_optical_efficiency,
        "average_shadow_blocking_efficiency": (
            annual.average_shadow_blocking_efficiency
        ),
        "average_truncation_efficiency": (annual.average_truncation_efficiency),
        "constraint_satisfied": evaluation.is_feasible(),
    }
    validation_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    table_path = destination / "08_论文结果与验证表.md"
    lines = [
        "",
        "## 表 6 高精度加密验证",
        "",
        "| 阴影网格 | 截断光线 | 邻镜半径 (m) | 年平均功率 (MW) | 功率余量 (MW) | 单位面积功率 (kW/m²) | 是否满足约束 |",
        "| ---: | ---: | ---: | ---: | ---: | ---: | :---: |",
        (
            f"| {profile.solver.shadow_grid_size}×{profile.solver.shadow_grid_size} "
            f"| {profile.solver.truncation_rays} "
            f"| {profile.solver.neighbor_radius_m:.0f} "
            f"| {evaluation.annual_power_mw:.6f} "
            f"| {evaluation.annual_power_mw - TARGET_ANNUAL_POWER_MW:.6f} "
            f"| {evaluation.unit_area_power_kw_m2:.6f} "
            f"| {'是' if evaluation.is_feasible() else '否'} |"
        ),
    ]
    with table_path.open("a", encoding="utf-8") as handle:
        handle.write("\n".join(lines) + "\n")
    return validation_path
