"""Campo2D 搜索、正式结果、结构统计和提交表导出。"""

from __future__ import annotations

import csv
import json
from copy import copy
from dataclasses import asdict
from pathlib import Path
from typing import Any, Sequence

import numpy as np
from openpyxl import load_workbook

from ..q2.evaluate import EvaluationProfile
from .evaluate import Campo2DEvaluation
from .model import Campo2DBase, Campo2DDesign
from .search import InitialScreenRecord, MultiStartOutcome, SearchTraceRecord


TARGET_POWER_MW = 42.0


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        raise ValueError(f"没有可写入 {path.name} 的结果。")
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def _design_record(design: Campo2DDesign) -> dict[str, Any]:
    record: dict[str, Any] = {
        "tower_x_m": design.tower_x,
        "tower_y_m": design.tower_y,
        "initial_spacing_m": design.initial_spacing,
        "spacing_growth_m_per_ring": design.spacing_growth,
        "ring_count": design.ring_count,
        "area_scale_lambda": design.area_scale,
    }
    for index, value in enumerate(design.size_nodes, start=1):
        record[f"alpha_{index}"] = value
    for index, value in enumerate(design.height_nodes, start=1):
        record[f"beta_{index}_m"] = value
    for index, value in enumerate(design.size_angles, start=1):
        record[f"size_angle_{index}"] = value
    for index, value in enumerate(design.height_angles, start=1):
        record[f"height_angle_{index}_m"] = value
    return record


def _initial_row(record: InitialScreenRecord) -> dict[str, Any]:
    return {
        "initial_index": record.index,
        "source": record.source,
        **_design_record(record.design),
        "geometry_valid": record.geometry_valid,
        "mirror_count": record.mirror_count,
        "total_area_m2": record.total_area_m2,
        "coarse_power_mw": record.coarse_power_mw,
        "coarse_q_kw_m2": record.coarse_q_kw_m2,
        "retained": record.retained,
        "reason": record.reason,
    }


def _trace_row(record: SearchTraceRecord) -> dict[str, Any]:
    previous = {
        f"previous_{name}": value
        for name, value in _design_record(record.previous_design).items()
    }
    candidate = {
        f"candidate_{name}": value
        for name, value in _design_record(record.design).items()
    }
    return {
        "sequence": record.sequence,
        "start_name": record.start_name,
        "joint_cycle": record.joint_cycle,
        "phase": record.phase,
        "block": record.block,
        "step": record.step,
        "round_index": record.round_index,
        "candidate_count": record.candidate_count,
        "legal_candidate_count": record.legal_candidate_count,
        "medium_candidate_count": record.medium_candidate_count,
        "action": record.action,
        "accepted": record.accepted,
        "evaluation_profile": record.evaluation_profile,
        "previous_annual_power_mw": record.previous_annual_power_mw,
        "previous_total_area_m2": record.previous_total_area_m2,
        "previous_unit_area_power_kw_m2": record.previous_unit_area_power_kw_m2,
        "annual_power_mw": record.annual_power_mw,
        "total_area_m2": record.total_area_m2,
        "unit_area_power_kw_m2": record.unit_area_power_kw_m2,
        "power_margin_mw": record.power_margin_mw,
        **previous,
        **candidate,
    }


def _evaluation_record(evaluation: Campo2DEvaluation) -> dict[str, Any]:
    return {
        "profile": evaluation.profile_name,
        "mirror_count": evaluation.mirror_count,
        "ring_count": evaluation.field.ring_count,
        "total_area_m2": evaluation.total_area_m2,
        "annual_power_mw": evaluation.annual_power_mw,
        "annual_power_margin_mw": evaluation.annual_power_mw - TARGET_POWER_MW,
        "unit_area_power_kw_m2": evaluation.unit_area_power_kw_m2,
        "constraint_satisfied": evaluation.is_feasible(TARGET_POWER_MW),
        "annual_efficiencies": asdict(evaluation.solution.annual_result),
        "geometry": asdict(evaluation.geometry),
    }


def _mirror_rows(evaluation: Campo2DEvaluation) -> list[dict[str, Any]]:
    field = evaluation.field
    specifications = evaluation.specifications
    rows: list[dict[str, Any]] = []
    for index in range(evaluation.mirror_count):
        rows.append(
            {
                "mirror_id": index + 1,
                "ring_id": int(field.ring_indices[index]),
                "campo_zone": int(field.zone_indices[index]),
                "ring_member_id": int(field.ring_member_indices[index]),
                "x_m": float(field.coordinates[index, 0]),
                "y_m": float(field.coordinates[index, 1]),
                "radius_to_tower_m": float(field.ring_radii[index]),
                "theta_rad": float(field.azimuth_angles[index]),
                "normalized_radius": float(field.normalized_radii[index]),
                "mirror_width_m": float(specifications.widths[index]),
                "mirror_height_m": float(specifications.heights[index]),
                "installation_height_m": float(specifications.installation_heights[index]),
                "mirror_area_m2": float(specifications.areas[index]),
                "radial_size_component": float(specifications.radial_size_component[index]),
                "angular_size_component": float(specifications.angular_size_component[index]),
                "radial_height_component_m": float(
                    specifications.radial_height_component[index]
                ),
                "angular_height_component_m": float(
                    specifications.angular_height_component[index]
                ),
            }
        )
    return rows


def _mirror_annual_rows(evaluation: Campo2DEvaluation) -> list[dict[str, Any]]:
    specifications = evaluation.specifications
    rows: list[dict[str, Any]] = []
    for index, record in enumerate(evaluation.solution.mirror_annual_results):
        area = float(specifications.areas[index])
        rows.append(
            {
                **asdict(record),
                "ring_id": int(evaluation.field.ring_indices[index]),
                "theta_rad": float(evaluation.field.azimuth_angles[index]),
                "mirror_area_m2": area,
                "unit_area_contribution_kw_m2": record.average_output_power_kw / area,
            }
        )
    return rows


def _ring_rows(evaluation: Campo2DEvaluation) -> list[dict[str, Any]]:
    field = evaluation.field
    specs = evaluation.specifications
    mirror_results = evaluation.solution.mirror_annual_results
    optical = np.asarray([row.average_optical_efficiency for row in mirror_results])
    power = np.asarray([row.average_output_power_kw for row in mirror_results])
    rows: list[dict[str, Any]] = []
    for ring in range(1, field.ring_count + 1):
        active = field.ring_indices == ring
        first = int(np.flatnonzero(active)[0])
        rows.append(
            {
                "ring_id": ring,
                "radius_m": float(field.ring_radii[first]),
                "nominal_count": int(field.nominal_ring_counts[first]),
                "actual_count": int(np.count_nonzero(active)),
                "retention_ratio": float(
                    np.count_nonzero(active) / field.nominal_ring_counts[first]
                ),
                "average_width_m": float(np.mean(specs.widths[active])),
                "average_height_m": float(np.mean(specs.heights[active])),
                "average_installation_height_m": float(
                    np.mean(specs.installation_heights[active])
                ),
                "average_mirror_area_m2": float(np.mean(specs.areas[active])),
                "average_optical_efficiency": float(np.mean(optical[active])),
                "average_mirror_output_kw": float(np.mean(power[active])),
                "average_unit_area_contribution_kw_m2": float(
                    np.mean(power[active] / specs.areas[active])
                ),
            }
        )
    return rows


def _angle_bin_rows(
    evaluation: Campo2DEvaluation,
    *,
    bin_count: int = 12,
) -> list[dict[str, Any]]:
    field = evaluation.field
    specs = evaluation.specifications
    mirror_results = evaluation.solution.mirror_annual_results
    optical = np.asarray([row.average_optical_efficiency for row in mirror_results])
    power = np.asarray([row.average_output_power_kw for row in mirror_results])
    representative = tuple(dict.fromkeys(field.control_ring_indices))
    edges = np.linspace(-np.pi, np.pi, bin_count + 1)
    rows: list[dict[str, Any]] = []
    for ring in representative:
        ring_active = field.ring_indices == ring
        bins = np.digitize(field.azimuth_angles, edges[1:-1], right=False)
        for bin_index in range(bin_count):
            active = ring_active & (bins == bin_index)
            if not np.any(active):
                continue
            rows.append(
                {
                    "ring_id": ring,
                    "theta_left_rad": float(edges[bin_index]),
                    "theta_right_rad": float(edges[bin_index + 1]),
                    "mirror_count": int(np.count_nonzero(active)),
                    "average_scale": float(np.mean(specs.scales[active])),
                    "average_installation_height_m": float(
                        np.mean(specs.installation_heights[active])
                    ),
                    "average_optical_efficiency": float(np.mean(optical[active])),
                    "average_unit_area_contribution_kw_m2": float(
                        np.mean(power[active] / specs.areas[active])
                    ),
                }
            )
    return rows


def write_result3_workbook(
    *,
    template_path: str | Path,
    output_path: str | Path,
    evaluation: Campo2DEvaluation,
) -> Path:
    template = Path(template_path)
    if not template.exists():
        raise FileNotFoundError(f"找不到 result3.xlsx 模板：{template}")
    destination = Path(output_path)
    workbook = load_workbook(template)
    sheet = workbook.active
    if sheet.max_column < 8:
        workbook.close()
        raise ValueError("result3.xlsx 模板列数不足 8 列。")
    style_row = 2 if sheet.max_row >= 2 else 1
    styles = [copy(sheet.cell(style_row, column)._style) for column in range(1, 9)]
    formats = [sheet.cell(style_row, column).number_format for column in range(1, 9)]
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)
    specs = evaluation.specifications
    for index in range(evaluation.mirror_count):
        values = (
            evaluation.field.parameters.tower_x,
            evaluation.field.parameters.tower_y,
            index + 1,
            float(specs.widths[index]),
            float(specs.heights[index]),
            float(evaluation.field.coordinates[index, 0]),
            float(evaluation.field.coordinates[index, 1]),
            float(specs.installation_heights[index]),
        )
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(index + 2, column, value)
            cell._style = copy(styles[column - 1])
            cell.number_format = formats[column - 1]
    workbook.save(destination)
    workbook.close()
    return destination


def write_primary_results(
    *,
    output_dir: str | Path,
    base: Campo2DBase,
    design: Campo2DDesign,
    evaluation: Campo2DEvaluation,
    search: MultiStartOutcome,
    formal_candidates: Sequence[tuple[str, Campo2DEvaluation]],
    q2_baseline: dict[str, Any],
    six_group_baseline: dict[str, Any],
    result3_template: str | Path,
    validation_evaluations: Sequence[
        tuple[EvaluationProfile, Campo2DEvaluation]
    ] = (),
) -> dict[str, Path]:
    destination = Path(output_dir)
    destination.mkdir(parents=True, exist_ok=True)
    paths = {
        "initials": destination / "02_Sobol初值与筛选结果.csv",
        "trace": destination / "03_搜索轨迹.csv",
        "mirrors": destination / "04_最终逐镜参数与坐标.csv",
        "time": destination / "05_逐时刻计算结果.csv",
        "monthly": destination / "06_月平均计算结果.csv",
        "annual": destination / "07_年平均计算结果.json",
        "mirror_annual": destination / "08_单镜年平均结果.csv",
        "rings": destination / "09_逐环统计.csv",
        "angles": destination / "10_角度分箱统计.csv",
        "baseline": destination / "11_baseline比较.json",
        "summary": destination / "12_最终方案摘要.json",
        "geometry": destination / "13_几何验证.json",
        "workbook": destination / "15_第三问提交结果.xlsx",
        "paper": destination / "16_论文结果与验证表.md",
    }
    _write_csv(paths["initials"], [_initial_row(record) for record in search.initial_screen])
    _write_csv(
        paths["trace"],
        [_trace_row(record) for outcome in search.starts for record in outcome.trace],
    )
    _write_csv(paths["mirrors"], _mirror_rows(evaluation))
    _write_csv(paths["time"], [asdict(record) for record in evaluation.solution.time_results])
    _write_csv(
        paths["monthly"],
        [asdict(record) for record in evaluation.solution.monthly_results],
    )
    annual_payload = {
        "mirror_count": evaluation.mirror_count,
        "total_area_m2": evaluation.total_area_m2,
        **asdict(evaluation.solution.annual_result),
    }
    paths["annual"].write_text(
        json.dumps(annual_payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    _write_csv(paths["mirror_annual"], _mirror_annual_rows(evaluation))
    _write_csv(paths["rings"], _ring_rows(evaluation))
    _write_csv(paths["angles"], _angle_bin_rows(evaluation))

    is_smoke = "smoke" in evaluation.profile_name.lower()
    six_q = float(six_group_baseline["annual"]["unit_area_output_kw_m2"])
    comparison = {
        "result_status": "smoke_nonformal" if is_smoke else "formal_dense_accepted",
        "q2_uniform": q2_baseline,
        "six_group": six_group_baseline,
        "q3_campo2d": _evaluation_record(evaluation),
        "q3_minus_six_q_kw_m2": (
            None if is_smoke else evaluation.unit_area_power_kw_m2 - six_q
        ),
        "recommended_final_model": (
            None
            if is_smoke
            else (
                "q3_campo2d"
                if evaluation.unit_area_power_kw_m2 >= six_q
                else "six_group"
            )
        ),
    }
    paths["baseline"].write_text(
        json.dumps(comparison, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    convergence = [
        {
            "start_name": outcome.start_name,
            "joint_cycles": outcome.joint_cycles,
            "stable_joint_cycles": outcome.stable_joint_cycles,
            "stopped_by": outcome.stopped_by,
            "medium_result": _evaluation_record(outcome.best_evaluation),
            "final_design": _design_record(outcome.best_design),
        }
        for outcome in search.starts
    ]
    summary = {
        "result_status": "smoke_nonformal" if is_smoke else "formal_dense_accepted",
        "layout": "q3-campo2d-radial-angular-continuous",
        "base_ring_count": base.ring_count,
        "base_structural_exclusions": [list(item) for item in base.excluded_ring_angles],
        "selected_start": next(
            name for name, candidate in formal_candidates if candidate is evaluation
        ),
        "design": _design_record(design),
        "control_ring_indices": list(evaluation.field.control_ring_indices),
        "control_radii_m": list(evaluation.field.control_radii),
        "formal_result": _evaluation_record(evaluation),
        "formal_candidates": [
            {"start_name": name, **_evaluation_record(candidate)}
            for name, candidate in formal_candidates
        ],
        "six_group_q_kw_m2": six_q,
        "q3_minus_six_q_kw_m2": (
            None if is_smoke else evaluation.unit_area_power_kw_m2 - six_q
        ),
        "recommended_final_model": comparison["recommended_final_model"],
        "convergence": convergence,
    }
    paths["summary"].write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    geometry_payload = {
        **asdict(evaluation.geometry),
        "minimum_width_m": float(np.min(evaluation.specifications.widths)),
        "maximum_width_m": float(np.max(evaluation.specifications.widths)),
        "minimum_height_m": float(np.min(evaluation.specifications.heights)),
        "maximum_height_m": float(np.max(evaluation.specifications.heights)),
        "minimum_installation_height_m": float(
            np.min(evaluation.specifications.installation_heights)
        ),
        "maximum_installation_height_m": float(
            np.max(evaluation.specifications.installation_heights)
        ),
    }
    paths["geometry"].write_text(
        json.dumps(geometry_payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    write_result3_workbook(
        template_path=result3_template,
        output_path=paths["workbook"],
        evaluation=evaluation,
    )
    if is_smoke:
        lines = [
            "# 第三问 Campo2D smoke 链路验证（非正式）",
            "",
            "| 模型 | 镜子数 | 总面积 (m²) | 单时刻功率 (MW) | 单位面积输出 (kW/m²) |",
            "| --- | ---: | ---: | ---: | ---: |",
            (
                f"| 新 Campo2D smoke | {evaluation.mirror_count} "
                f"| {evaluation.total_area_m2:.3f} "
                f"| {evaluation.annual_power_mw:.6f} "
                f"| {evaluation.unit_area_power_kw_m2:.6f} |"
            ),
            "",
            "> 本文件只证明搜索与导出链路可以运行；当前评价仅含六月正午一个状态，不能作为正式年平均结果。",
        ]
    else:
        q2_annual = q2_baseline["annual"]
        six_annual = six_group_baseline["annual"]
        lines = [
            "# 第三问结果与验证表",
            "",
            "## 表 3-2 正式结果与 baseline 比较",
            "",
            "| 模型 | 镜子数 | 总面积 (m²) | 年平均功率 (MW) | 单位面积输出 (kW/m²) |",
            "| --- | ---: | ---: | ---: | ---: |",
            (
                f"| 问题二统一规格 | {q2_baseline['mirror_count']} "
                f"| {q2_baseline['total_area_m2']:.3f} "
                f"| {q2_annual['field_output_mw']:.6f} "
                f"| {q2_annual['unit_area_output_kw_m2']:.6f} |"
            ),
            (
                f"| 六组 baseline | {six_group_baseline['mirror_count']} "
                f"| {six_group_baseline['total_area_m2']:.3f} "
                f"| {six_annual['field_output_mw']:.6f} "
                f"| {six_annual['unit_area_output_kw_m2']:.6f} |"
            ),
            (
                f"| 新 Campo2D | {evaluation.mirror_count} "
                f"| {evaluation.total_area_m2:.3f} "
                f"| {evaluation.annual_power_mw:.6f} "
                f"| {evaluation.unit_area_power_kw_m2:.6f} |"
            ),
            "",
            "## 表 3-3 几何与加密验证",
            "",
            "| 检查项 | 结果 |",
            "| --- | ---: |",
            f"| 最大场地半径 / m | {evaluation.geometry.maximum_field_radius_m:.6f} |",
            f"| 最小塔距 / m | {evaluation.geometry.minimum_tower_distance_m:.6f} |",
            f"| 最小异构镜间距余量 / m | {evaluation.geometry.minimum_width_clearance_m:.6f} |",
            f"| 最小不触地余量 / m | {evaluation.geometry.minimum_ground_clearance_m:.6f} |",
        ]
        for profile, candidate in validation_evaluations:
            lines.append(
                f"| {profile.name} 功率 / MW | {candidate.annual_power_mw:.6f} |"
            )
        lines.extend(
            (
                "",
                f"正式与加密精度功率约束：$\\overline P\\ge {TARGET_POWER_MW:g}\\ \\mathrm{{MW}}$。",
            )
        )
    paths["paper"].write_text("\n".join(lines) + "\n", encoding="utf-8")
    return paths


def write_dense_validation(
    *,
    output_dir: str | Path,
    evaluations: Sequence[tuple[EvaluationProfile, Campo2DEvaluation]],
) -> Path:
    path = Path(output_dir) / "14_高精度加密验证.json"
    payload = {
        "evaluations": [
            {
                "profile": {
                    "state_count": len(profile.months) * len(profile.solar_times),
                    "shadow_grid_size": profile.solver.shadow_grid_size,
                    "truncation_rays": profile.solver.truncation_rays,
                    "neighbor_radius_m": profile.solver.neighbor_radius_m,
                },
                **_evaluation_record(evaluation),
            }
            for profile, evaluation in evaluations
        ]
    }
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return path
