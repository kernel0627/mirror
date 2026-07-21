from __future__ import annotations

import csv
import json
import unittest
from dataclasses import replace
from pathlib import Path

import numpy as np
from openpyxl import load_workbook

from heliostat.q3.closure import close_formal_neighborhood
from heliostat.q3.evaluate import (
    EvaluationCache,
    evaluate_design,
    prepare_candidate,
    smoke_profile,
)
from heliostat.q3.model import load_baseline
from heliostat.q3.sensitivity import specification_perturbations
from heliostat.q3.search import coordinate_search
from heliostat.q3.tower_modes import build_refine_field


class SixGroupRefineModelTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.baseline = load_baseline(
            q2_summary_path="outputs/q2/07_最终方案摘要.json",
            six_group_summary_path="src/heliostat/q3/six_group_baseline.json",
        )

    def test_zero_increment_reproduces_mother_field_and_specs(self) -> None:
        field, specifications, check = prepare_candidate(
            baseline=self.baseline,
            design=self.baseline.design,
        )
        self.assertTrue(check.valid, check.reason)
        self.assertEqual(field.mirror_count, 1471)
        self.assertEqual(field.group_counts, (72, 269, 283, 224, 357, 266))
        np.testing.assert_allclose(
            field.coordinates,
            self.baseline.mother.coordinates,
            rtol=0.0,
            atol=1e-12,
        )
        np.testing.assert_array_equal(
            specifications.widths,
            np.asarray(self.baseline.design.widths)[field.group_indices],
        )
        np.testing.assert_array_equal(
            specifications.heights,
            np.asarray(self.baseline.design.mirror_heights)[field.group_indices],
        )

    def test_mode_b_moves_only_tower(self) -> None:
        design = replace(
            self.baseline.design,
            tower_mode="B",
            tower_y=self.baseline.design.tower_y - 0.5,
        )
        field = build_refine_field(self.baseline, design)
        baseline_field = build_refine_field(self.baseline, self.baseline.design)
        np.testing.assert_array_equal(field.coordinates, baseline_field.coordinates)
        self.assertEqual(field.mirror_set_hash, baseline_field.mirror_set_hash)
        self.assertEqual(field.geometry_center_y, self.baseline.design.tower_y)

    def test_mode_a_rebuilds_around_candidate_tower(self) -> None:
        design = replace(
            self.baseline.design,
            tower_mode="A",
            tower_y=self.baseline.design.tower_y - 0.5,
        )
        field = build_refine_field(self.baseline, design)
        self.assertEqual(field.geometry_center_y, design.tower_y)
        first_ring = field.ring_indices == 1
        baseline_first = self.baseline.mother.ring_indices == 1
        np.testing.assert_allclose(
            field.coordinates[first_ring, 1],
            self.baseline.mother.coordinates[baseline_first, 1] - 0.5,
            rtol=0.0,
            atol=1e-12,
        )

    def test_sensitivity_has_36_single_parameter_candidates(self) -> None:
        perturbations = specification_perturbations(self.baseline.design)
        self.assertEqual(len(perturbations), 36)
        self.assertEqual(len({item.parameter for item in perturbations}), 18)
        candidate = next(item for item in perturbations if item.parameter == "w3" and item.direction == "+")
        self.assertAlmostEqual(
            candidate.design.widths[2],
            self.baseline.design.widths[2] + 0.1,
        )
        np.testing.assert_array_equal(
            candidate.design.mirror_heights,
            self.baseline.design.mirror_heights,
        )

    def test_smoke_evaluation_accepts_zero_increment(self) -> None:
        evaluation = evaluate_design(
            baseline=self.baseline,
            design=self.baseline.design,
            profile=smoke_profile(),
            cache=EvaluationCache(),
        )
        self.assertEqual(evaluation.mirror_count, 1471)
        self.assertTrue(evaluation.geometry.valid)
        self.assertGreater(evaluation.annual_power_mw, 0.0)

    def test_search_trace_uses_explicit_six_group_reference(self) -> None:
        evaluation = evaluate_design(
            baseline=self.baseline,
            design=self.baseline.design,
            profile=smoke_profile(),
            cache=EvaluationCache(),
        )
        outcome = coordinate_search(
            initial_design=self.baseline.design,
            initial_evaluation=evaluation,
            active_variables=(),
            evaluator=lambda design: evaluation,
            baseline_q_kw_m2=0.5,
            maximum_sweeps=0,
        )
        self.assertEqual(outcome.trace, ())

    def test_closure_smoke_checks_both_sides(self) -> None:
        cache = EvaluationCache()
        evaluation = evaluate_design(
            baseline=self.baseline,
            design=self.baseline.design,
            profile=smoke_profile(),
            cache=cache,
        )
        outcome = close_formal_neighborhood(
            initial_design=self.baseline.design,
            initial_evaluation=evaluation,
            evaluator=lambda design: evaluate_design(
                baseline=self.baseline,
                design=design,
                profile=smoke_profile(),
                cache=cache,
            ),
            coarse_step_limit=1,
            fine_radius_steps=1,
            maximum_local_sweeps=0,
        )
        fine_rows = [
            row for row in outcome.trace if row["phase"] == "tower_fine"
        ]
        self.assertEqual(len(fine_rows), 3)
        self.assertEqual(
            {round(float(row["step"]), 1) for row in fine_rows},
            {-0.1, 0.0, 0.1},
        )


class Question3DeliverableTests(unittest.TestCase):
    def test_result3_matches_template_and_every_csv_row(self) -> None:
        output = Path("outputs/q3/result3.xlsx")
        self.assertTrue(output.is_file())
        self.assertFalse(Path("outputs/q3/14_第三问最终提交结果.xlsx").exists())

        template = load_workbook(
            "task/A/result3.xlsx",
            read_only=True,
            data_only=True,
        )
        workbook = load_workbook(output, read_only=True, data_only=True)
        self.assertEqual(workbook.sheetnames, template.sheetnames)
        sheet = workbook.active
        template_sheet = template.active
        self.assertEqual(sheet.title, template_sheet.title)
        self.assertEqual(sheet.max_column, 8)
        self.assertEqual(
            [sheet.cell(1, column).value for column in range(1, 9)],
            [template_sheet.cell(1, column).value for column in range(1, 9)],
        )

        with Path("outputs/q3/10_最终逐镜参数与坐标.csv").open(
            encoding="utf-8-sig",
            newline="",
        ) as handle:
            rows = list(csv.DictReader(handle))
        comparison = json.loads(
            Path("outputs/q3/11_正式结果比较.json").read_text(
                encoding="utf-8"
            )
        )
        design = comparison["selected_design"]
        self.assertEqual(len(rows), 1471)
        self.assertEqual(sheet.max_row, 1472)
        workbook_rows = list(
            sheet.iter_rows(min_row=2, max_col=8, values_only=True)
        )
        self.assertEqual(len(workbook_rows), len(rows))
        for index, (row, workbook_row) in enumerate(
            zip(rows, workbook_rows),
            start=1,
        ):
            self.assertEqual(workbook_row[0], design["tower_x_m"])
            self.assertEqual(workbook_row[1], design["tower_y_m"])
            self.assertEqual(workbook_row[2], index)
            for column_index, key in (
                (3, "mirror_width_m"),
                (4, "mirror_height_m"),
                (5, "x_m"),
                (6, "y_m"),
                (7, "z_m"),
            ):
                self.assertAlmostEqual(
                    workbook_row[column_index],
                    float(row[key]),
                )
        workbook.close()
        template.close()

    def test_final_metrics_are_consistent_across_deliverables(self) -> None:
        formal = json.loads(
            Path("outputs/q3/11_正式结果比较.json").read_text(
                encoding="utf-8"
            )
        )
        dense = json.loads(
            Path("outputs/q3/12_加密验收比较.json").read_text(
                encoding="utf-8"
            )
        )
        selected = formal["selected"]
        texts = [
            Path("README.md").read_text(encoding="utf-8"),
            Path("docs/questions/第三问.md").read_text(encoding="utf-8"),
            Path("outputs/q3/15_论文结果与验证表.md").read_text(
                encoding="utf-8"
            ),
        ]
        expected = (
            f"{selected['annual_power_mw']:.9f}",
            f"{selected['total_area_m2']:.6f}",
            f"{selected['unit_area_power_kw_m2']:.9f}",
            f"{dense['candidate']['80']['annual_power_mw']:.9f}",
            f"{dense['candidate']['80']['unit_area_power_kw_m2']:.9f}",
        )
        for text in texts:
            for value in expected:
                self.assertIn(value, text)
        for radius in ("80", "100"):
            self.assertTrue(dense["candidate"][radius]["power_margin_mw"] >= 0.0)
        for path in (
            "outputs/q3/16_参数敏感性图.png",
            "outputs/q3/17_六区宽高与安装高度图.png",
            "outputs/q3/18_六组与优化方案指标比较图.png",
            "outputs/q3/19_最终六区镜场与塔位平面图.png",
        ):
            self.assertGreater(Path(path).stat().st_size, 0)


if __name__ == "__main__":
    unittest.main()
