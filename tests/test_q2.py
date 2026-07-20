from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from heliostat.config import SolverConfig
from heliostat.q2.evaluate import (
    EvaluationProfile,
    scan_layout_extents,
)
from heliostat.q2.export import write_result2_workbook
from heliostat.q2.layout import (
    CampoParameters,
    LayoutError,
    PartitionedRingParameters,
    generate_campo_layout,
    generate_partitioned_layout,
    validate_layout,
)
from heliostat.q2.search import (
    sample_campo_parameters,
    sample_partitioned_parameters,
)


class Question2LayoutTests(unittest.TestCase):
    def test_partitioned_layout_is_legal_and_clipped_to_field(self) -> None:
        parameters = PartitionedRingParameters(
            tower_y=-100.0,
            mirror_width=6.2,
            mirror_height=6.2,
            installation_height=4.5,
            split_radius=220.0,
            near_spacing=12.0,
            far_spacing=15.0,
        )
        layout = generate_partitioned_layout(parameters)
        check = validate_layout(layout.coordinates, parameters)

        self.assertTrue(check.valid)
        self.assertGreater(layout.mirror_count, 1500)
        self.assertLessEqual(check.maximum_field_radius, 350.0 + 1e-9)
        self.assertGreater(check.minimum_center_distance, 11.2)
        self.assertTrue(
            any(ring.mirror_count < ring.nominal_count for ring in layout.rings)
        )

    def test_partitioned_layout_rejects_cross_ring_conflict(self) -> None:
        parameters = PartitionedRingParameters(
            tower_y=-100.0,
            mirror_width=6.2,
            mirror_height=6.2,
            installation_height=4.5,
            split_radius=220.0,
            near_spacing=2.0,
            far_spacing=2.0,
        )
        with self.assertRaises(LayoutError):
            generate_partitioned_layout(parameters)

    def test_campo_uses_zone_multipliers_and_growing_spacing(self) -> None:
        parameters = CampoParameters(
            tower_y=-100.0,
            mirror_width=6.2,
            mirror_height=6.2,
            installation_height=4.5,
            first_ring_count=56,
            initial_spacing=12.0,
            spacing_growth=0.2,
        )
        layout = generate_campo_layout(parameters)
        check = validate_layout(layout.coordinates, parameters)

        self.assertTrue(check.valid)
        self.assertEqual(layout.rings[0].zone, 1)
        self.assertEqual(layout.rings[0].nominal_count, 56)
        self.assertTrue(
            any(ring.zone == 2 and ring.nominal_count == 112 for ring in layout.rings)
        )
        self.assertTrue(
            any(ring.zone == 3 and ring.nominal_count == 224 for ring in layout.rings)
        )
        first_gap = layout.rings[1].radius - layout.rings[0].radius
        second_gap = layout.rings[2].radius - layout.rings[1].radius
        self.assertAlmostEqual(first_gap, 12.0)
        self.assertAlmostEqual(second_gap, 12.2)
        self.assertAlmostEqual(parameters.safety_epsilon, 0.01)
        self.assertGreaterEqual(
            check.minimum_center_distance - parameters.mirror_width - 5.0,
            0.01 - 1e-9,
        )

    def test_sobol_samples_respect_coupled_constraints(self) -> None:
        partitioned = sample_partitioned_parameters(9, seed=7)
        campo = sample_campo_parameters(9, seed=8)

        self.assertEqual(len(partitioned), 9)
        self.assertEqual(len(campo), 9)
        self.assertTrue(
            all(
                value.mirror_height <= value.mirror_width
                and value.installation_height >= value.mirror_height / 2.0
                and value.far_spacing >= value.near_spacing
                for value in partitioned
            )
        )
        self.assertTrue(
            all(
                value.mirror_height <= value.mirror_width
                and value.installation_height >= value.mirror_height / 2.0
                and isinstance(value.first_ring_count, int)
                for value in campo
            )
        )


class Question2EvaluationTests(unittest.TestCase):
    def setUp(self) -> None:
        self.parameters = PartitionedRingParameters(
            tower_y=-50.0,
            mirror_width=6.2,
            mirror_height=6.2,
            installation_height=4.5,
            split_radius=220.0,
            near_spacing=14.0,
            far_spacing=18.0,
        )
        self.layout = generate_partitioned_layout(self.parameters)
        self.profile = EvaluationProfile(
            name="unit-test",
            solver=SolverConfig(
                shadow_grid_size=2,
                truncation_rays=4,
                calculate_shadow=False,
                calculate_truncation=False,
            ),
            months=(6,),
            solar_times=(12.0,),
        )

    def test_extent_scan_reuses_question1_model(self) -> None:
        result = scan_layout_extents(
            self.layout,
            self.parameters,
            self.profile,
            target_power_mw=0.01,
            coarse_stride=4,
            window=1,
        )

        self.assertTrue(result.best.is_feasible(0.01))
        self.assertEqual(result.first_feasible_ring_count, 1)
        self.assertGreater(result.best.mirror_count, 0)
        self.assertGreater(result.best.unit_area_power_kw_m2, 0.0)
        self.assertEqual(
            result.best.solution.time_results[0].month,
            6,
        )

    def test_result2_export_contains_every_mirror(self) -> None:
        result = scan_layout_extents(
            self.layout,
            self.parameters,
            self.profile,
            target_power_mw=0.01,
            coarse_stride=4,
            window=0,
        )

        with tempfile.TemporaryDirectory() as temporary:
            output = Path(temporary) / "result2.xlsx"
            write_result2_workbook(
                template_path="task/A/result2.xlsx",
                output_path=output,
                evaluation=result.best,
                parameters=self.parameters,
            )
            workbook = load_workbook(output, read_only=True, data_only=True)
            sheet = workbook.active
            self.assertEqual(sheet.max_row, result.best.mirror_count + 1)
            self.assertEqual(sheet.cell(2, 1).value, self.parameters.tower_x)
            self.assertEqual(sheet.cell(2, 2).value, self.parameters.tower_y)
            self.assertEqual(sheet.cell(2, 3).value, 1)
            self.assertAlmostEqual(
                sheet.cell(2, 6).value,
                float(result.best.coordinates[0, 0]),
            )
            workbook.close()


if __name__ == "__main__":
    unittest.main()
