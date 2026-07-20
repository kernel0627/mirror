from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import numpy as np
from openpyxl import load_workbook

from heliostat.q3_continuous.evaluate import (
    EvaluationCache,
    evaluate_design,
    smoke_profile,
)
from heliostat.q3_continuous.export import write_result3_workbook
from heliostat.q3_continuous.model import (
    ContinuousDesign,
    build_campo_mother_field,
    expand_continuous_design,
    validate_heterogeneous_field,
)
from heliostat.q3_continuous.search import diagnose_campo_structure


class ContinuousCampoModelTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.mother = build_campo_mother_field(
            "outputs/q2/07_最终方案摘要.json",
            selected_coordinates_path="outputs/q2/03_最终镜位坐标.csv",
        )

    def test_q2_field_keeps_campo_structure_labels(self) -> None:
        self.assertEqual(self.mother.mirror_count, 1469)
        self.assertEqual(self.mother.zone_counts, (624, 845))
        self.assertEqual(self.mother.zone_ring_counts, (11, 17))
        self.assertEqual(
            int(np.unique(self.mother.ring_indices).size),
            28,
        )
        for ring in np.unique(self.mother.ring_indices):
            active = self.mother.ring_indices == ring
            self.assertAlmostEqual(
                float(np.mean(self.mother.azimuth_features[active])),
                0.0,
                places=12,
            )
            self.assertTrue(
                np.all(
                    self.mother.actual_ring_counts[active]
                    == np.count_nonzero(active)
                )
            )

    def test_uniform_design_reproduces_q2_specifications(self) -> None:
        specifications = expand_continuous_design(
            self.mother,
            ContinuousDesign.uniform(),
        )
        np.testing.assert_allclose(
            specifications.widths,
            self.mother.base_width,
        )
        np.testing.assert_allclose(
            specifications.heights,
            self.mother.base_height,
        )
        np.testing.assert_allclose(
            specifications.installation_heights,
            self.mother.base_installation_height,
        )
        self.assertAlmostEqual(
            specifications.total_area_m2,
            self.mother.base_total_area_m2,
            places=8,
        )

    def test_shape_parameters_keep_target_area_exact(self) -> None:
        design = ContinuousDesign(
            size_zone1_slope=-0.08,
            size_zone2_slope=-0.04,
            size_zone2_offset=0.03,
            size_azimuth=0.02,
            area_ratio=0.985,
        )
        specifications = expand_continuous_design(self.mother, design)
        self.assertAlmostEqual(
            specifications.total_area_m2,
            self.mother.base_total_area_m2 * design.area_ratio,
            places=8,
        )

    def test_nonuniform_design_passes_geometry_when_conservative(self) -> None:
        design = ContinuousDesign(
            size_zone1_slope=-0.02,
            size_zone2_slope=-0.02,
            height_offset=0.2,
            height_zone1_slope=0.1,
            height_zone2_slope=0.1,
            area_ratio=0.96,
        )
        specifications = expand_continuous_design(self.mother, design)
        check = validate_heterogeneous_field(
            coordinates=self.mother.coordinates,
            widths=specifications.widths,
            heights=specifications.heights,
            installation_heights=(
                specifications.installation_heights
            ),
            tower_x=self.mother.parameters.tower_x,
            tower_y=self.mother.parameters.tower_y,
            safety_epsilon=self.mother.parameters.safety_epsilon,
        )
        self.assertTrue(check.valid, check.reason)


class ContinuousCampoEvaluationTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.mother = build_campo_mother_field(
            "outputs/q2/07_最终方案摘要.json",
            selected_coordinates_path="outputs/q2/03_最终镜位坐标.csv",
        )
        cls.evaluation = evaluate_design(
            mother=cls.mother,
            design=ContinuousDesign.uniform(),
            profile=smoke_profile(),
            cache=EvaluationCache(),
        )

    def test_smoke_evaluation_and_diagnostics(self) -> None:
        self.assertEqual(self.evaluation.mirror_count, 1469)
        self.assertGreater(self.evaluation.annual_power_mw, 0.0)
        diagnostics = diagnose_campo_structure(
            self.mother,
            self.evaluation,
        )
        self.assertGreaterEqual(
            diagnostics.relative_rmse_reduction,
            0.0,
        )
        self.assertLessEqual(
            diagnostics.relative_rmse_reduction,
            1.0,
        )

    def test_result3_export_contains_individual_specs(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            output = Path(temporary) / "result3.xlsx"
            write_result3_workbook(
                template_path="task/A/result3.xlsx",
                output_path=output,
                evaluation=self.evaluation,
                tower_x=self.mother.parameters.tower_x,
                tower_y=self.mother.parameters.tower_y,
            )
            workbook = load_workbook(
                output,
                read_only=True,
                data_only=True,
            )
            sheet = workbook.active
            self.assertEqual(
                sheet.max_row,
                self.evaluation.mirror_count + 1,
            )
            self.assertAlmostEqual(
                sheet.cell(2, 4).value,
                float(self.evaluation.widths[0]),
            )
            workbook.close()


if __name__ == "__main__":
    unittest.main()
