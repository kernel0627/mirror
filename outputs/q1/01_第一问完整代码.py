"""第一问完整代码展示稿。

本文件把正式工程中的共享光学核心和第一问流程合并为单文件，便于成果浏览与代码审阅。
"""

from __future__ import annotations


# ========================================================================
# 参数配置
# ========================================================================

from dataclasses import asdict, dataclass


@dataclass(frozen=True)
class FieldConfig:
    """第一问给定的镜场和环境参数。"""

    latitude_deg: float = 39.4
    altitude_km: float = 3.0
    field_radius: float = 350.0
    exclusion_radius: float = 100.0
    tower_x: float = 0.0
    tower_y: float = 0.0
    receiver_center_z: float = 86.0
    receiver_radius: float = 4.0
    receiver_height: float = 8.0
    mirror_width: float = 6.2
    mirror_height: float = 6.2
    mirror_center_z: float = 4.5
    reflectivity: float = 0.92
    solar_angular_radius_rad: float = 4.65e-3

    @property
    def receiver_z_min(self) -> float:
        return self.receiver_center_z - self.receiver_height / 2.0

    @property
    def receiver_z_max(self) -> float:
        return self.receiver_center_z + self.receiver_height / 2.0

    @property
    def mirror_area(self) -> float:
        return self.mirror_width * self.mirror_height

    def to_dict(self) -> dict[str, float]:
        return asdict(self)


@dataclass(frozen=True)
class SolverConfig:
    """可调的采样精度和加速参数。"""

    shadow_grid_size: int = 15
    truncation_rays: int = 256
    neighbor_radius_m: float = 60.0
    candidate_margin: float = 1.05
    ray_epsilon: float = 1e-7
    truncation_chunk_size: int = 128
    sobol_seed: int = 2023
    calculate_shadow: bool = True
    calculate_truncation: bool = True

    def __post_init__(self) -> None:
        if self.shadow_grid_size < 1:
            raise ValueError("shadow_grid_size 必须大于等于 1。")
        if self.truncation_rays < 1:
            raise ValueError("truncation_rays 必须大于等于 1。")
        if self.neighbor_radius_m <= 0.0:
            raise ValueError("neighbor_radius_m 必须大于 0。")
        if self.candidate_margin < 1.0:
            raise ValueError("candidate_margin 不能小于 1。")
        if self.ray_epsilon <= 0.0:
            raise ValueError("ray_epsilon 必须大于 0。")
        if self.truncation_chunk_size < 1:
            raise ValueError("truncation_chunk_size 必须大于等于 1。")

    def to_dict(self) -> dict[str, int | float | bool]:
        return asdict(self)


# ========================================================================
# 坐标读取
# ========================================================================

import csv
from pathlib import Path

import numpy as np
from numpy.typing import NDArray
from openpyxl import load_workbook


FloatArray = NDArray[np.float64]


def load_mirror_xy(path: str | Path, expected_count: int | None = 1745) -> FloatArray:
    """从题目附件读取定日镜 x、y 坐标。"""

    source = Path(path)
    if not source.exists():
        raise FileNotFoundError(f"找不到定日镜坐标文件：{source}")

    if source.suffix.lower() == ".xlsx":
        workbook = load_workbook(source, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        workbook.close()
        values = [(row[0], row[1]) for row in rows if row[0] is not None]
    elif source.suffix.lower() == ".csv":
        with source.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            next(reader, None)
            values = [(row[0], row[1]) for row in reader if row]
    else:
        raise ValueError("坐标文件只支持 .xlsx 或 .csv。")

    try:
        mirror_xy = np.asarray(values, dtype=float)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"坐标文件包含非数值数据：{source}") from exc

    if mirror_xy.ndim != 2 or mirror_xy.shape[1] != 2:
        raise ValueError(f"坐标数据应为 N×2，实际形状为 {mirror_xy.shape}。")
    if not np.all(np.isfinite(mirror_xy)):
        raise ValueError("坐标数据包含 NaN 或无穷值。")
    if expected_count is not None and mirror_xy.shape[0] != expected_count:
        raise ValueError(
            f"应读取 {expected_count} 面定日镜，实际读取 {mirror_xy.shape[0]} 面。"
        )
    return mirror_xy


# ========================================================================
# 太阳位置与 DNI
# ========================================================================

import math
from dataclasses import dataclass

import numpy as np
from numpy.typing import NDArray


FloatArray = NDArray[np.float64]


@dataclass(frozen=True)
class SolarState:
    month: int
    solar_time: float
    direction: FloatArray
    altitude_rad: float
    azimuth_rad: float
    declination_rad: float
    dni_kw_m2: float


MONTH_DAYS = (31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31)


def day_from_spring_equinox(month: int, day: int = 21) -> int:
    """以 3 月 21 日为第 0 天，返回题面赤纬公式所需的 D。"""

    if not 1 <= month <= 12:
        raise ValueError("month 必须位于 1 到 12。")
    if not 1 <= day <= MONTH_DAYS[month - 1]:
        raise ValueError("day 不在指定月份的有效范围内。")
    day_of_year = sum(MONTH_DAYS[: month - 1]) + day
    return day_of_year - 80


def calculate_solar_state(
    month: int,
    solar_time: float,
    latitude_deg: float,
    altitude_km: float,
) -> SolarState:
    """按题面附录计算东-北-天坐标下的太阳单位方向和 DNI。"""

    if not 0.0 <= solar_time <= 24.0:
        raise ValueError("solar_time 必须位于 0 到 24 小时。")

    d = day_from_spring_equinox(month)
    declination = math.asin(
        math.sin(2.0 * math.pi * d / 365.0)
        * math.sin(math.radians(23.45))
    )
    latitude = math.radians(latitude_deg)
    hour_angle = math.pi / 12.0 * (solar_time - 12.0)

    direction = np.array(
        [
            -math.cos(declination) * math.sin(hour_angle),
            math.cos(latitude) * math.sin(declination)
            - math.sin(latitude) * math.cos(declination) * math.cos(hour_angle),
            math.sin(latitude) * math.sin(declination)
            + math.cos(latitude) * math.cos(declination) * math.cos(hour_angle),
        ],
        dtype=float,
    )
    direction /= np.linalg.norm(direction)

    altitude = math.asin(float(np.clip(direction[2], -1.0, 1.0)))
    azimuth = math.atan2(float(direction[0]), float(direction[1])) % (
        2.0 * math.pi
    )

    h = altitude_km
    a = 0.4237 - 0.00821 * (6.0 - h) ** 2
    b = 0.5055 + 0.00595 * (6.5 - h) ** 2
    c = 0.2711 + 0.01858 * (2.5 - h) ** 2
    if altitude <= 0.0:
        dni = 0.0
    else:
        dni = 1.366 * (a + b * math.exp(-c / math.sin(altitude)))

    return SolarState(
        month=month,
        solar_time=solar_time,
        direction=direction,
        altitude_rad=altitude,
        azimuth_rad=azimuth,
        declination_rad=declination,
        dni_kw_m2=dni,
    )


# ========================================================================
# 镜场几何与镜面姿态
# ========================================================================

from dataclasses import dataclass

import numpy as np
from numpy.typing import NDArray



FloatArray = NDArray[np.float64]


def normalize_rows(vectors: FloatArray) -> FloatArray:
    norms = np.linalg.norm(vectors, axis=1, keepdims=True)
    if np.any(norms <= 1e-15):
        raise ValueError("不能归一化长度为零的向量。")
    return vectors / norms


def reflect(directions: FloatArray, normals: FloatArray) -> FloatArray:
    """根据 d - 2(d·n)n 计算反射方向。"""

    dot = np.sum(directions * normals, axis=-1, keepdims=True)
    return directions - 2.0 * dot * normals


@dataclass(frozen=True)
class PreparedField:
    config: FieldConfig
    centers: FloatArray
    receiver_center: FloatArray
    receiver_directions: FloatArray
    receiver_distances: FloatArray
    atmospheric_efficiency: FloatArray

    @property
    def mirror_count(self) -> int:
        return int(self.centers.shape[0])

    @property
    def total_mirror_area(self) -> float:
        return self.mirror_count * self.config.mirror_area


@dataclass(frozen=True)
class MirrorOrientation:
    normals: FloatArray
    width_axes: FloatArray
    height_axes: FloatArray
    cosine_efficiency: FloatArray


def prepare_field(mirror_xy: FloatArray, config: FieldConfig) -> PreparedField:
    mirror_count = int(mirror_xy.shape[0])
    centers = np.column_stack(
        (
            mirror_xy,
            np.full(mirror_count, config.mirror_center_z, dtype=float),
        )
    )
    receiver_center = np.array(
        [config.tower_x, config.tower_y, config.receiver_center_z],
        dtype=float,
    )
    receiver_vectors = receiver_center[None, :] - centers
    distances = np.linalg.norm(receiver_vectors, axis=1)
    receiver_directions = receiver_vectors / distances[:, None]
    atmospheric = (
        0.99321 - 0.0001176 * distances + 1.97e-8 * distances**2
    )
    atmospheric = np.clip(atmospheric, 0.0, 1.0)
    return PreparedField(
        config=config,
        centers=centers,
        receiver_center=receiver_center,
        receiver_directions=receiver_directions,
        receiver_distances=distances,
        atmospheric_efficiency=atmospheric,
    )


def calculate_orientation(
    prepared: PreparedField,
    sun_direction: FloatArray,
) -> MirrorOrientation:
    sun_rows = np.broadcast_to(sun_direction, prepared.receiver_directions.shape)
    normals = normalize_rows(sun_rows + prepared.receiver_directions)

    upward = np.broadcast_to(np.array([0.0, 0.0, 1.0]), normals.shape)
    width_axes = np.cross(upward, normals)
    weak = np.linalg.norm(width_axes, axis=1) < 1e-10
    width_axes[weak] = np.array([1.0, 0.0, 0.0])
    width_axes = normalize_rows(width_axes)
    height_axes = normalize_rows(np.cross(normals, width_axes))

    cosine = np.clip(normals @ sun_direction, 0.0, 1.0)
    return MirrorOrientation(
        normals=normals,
        width_axes=width_axes,
        height_axes=height_axes,
        cosine_efficiency=cosine,
    )


def maximum_reflection_error(
    prepared: PreparedField,
    orientation: MirrorOrientation,
    sun_direction: FloatArray,
) -> float:
    incoming = np.broadcast_to(-sun_direction, orientation.normals.shape)
    reflected = reflect(incoming, orientation.normals)
    errors = np.linalg.norm(reflected - prepared.receiver_directions, axis=1)
    return float(np.max(errors))


# ========================================================================
# 阴影遮挡效率
# ========================================================================

import numpy as np
from numpy.typing import NDArray
from scipy.spatial import cKDTree



IntArray = NDArray[np.int64]


def mirror_grid_offsets(
    grid_size: int,
    mirror_width: float,
    mirror_height: float,
) -> FloatArray:
    """返回位于等面积小格中心的局部二维坐标。"""

    width_step = mirror_width / grid_size
    height_step = mirror_height / grid_size
    width_values = np.linspace(
        -mirror_width / 2.0 + width_step / 2.0,
        mirror_width / 2.0 - width_step / 2.0,
        grid_size,
    )
    height_values = np.linspace(
        -mirror_height / 2.0 + height_step / 2.0,
        mirror_height / 2.0 - height_step / 2.0,
        grid_size,
    )
    width_grid, height_grid = np.meshgrid(
        width_values,
        height_values,
        indexing="xy",
    )
    return np.column_stack((width_grid.ravel(), height_grid.ravel()))


def _direction_candidates(
    target_index: int,
    neighbors: IntArray,
    centers: FloatArray,
    direction: FloatArray,
    reach: float,
    maximum_distance: float | None = None,
) -> IntArray:
    if neighbors.size == 0:
        return neighbors

    delta = centers[neighbors] - centers[target_index]
    projection = delta @ direction
    perpendicular = delta - projection[:, None] * direction[None, :]
    perpendicular_distance = np.linalg.norm(perpendicular, axis=1)

    keep = (projection > -reach) & (perpendicular_distance <= reach)
    if maximum_distance is not None:
        keep &= projection < maximum_distance + reach
    return neighbors[keep]


def ray_rectangle_hits(
    origins: FloatArray,
    direction: FloatArray,
    rectangle_center: FloatArray,
    rectangle_normal: FloatArray,
    rectangle_width_axis: FloatArray,
    rectangle_height_axis: FloatArray,
    half_width: float,
    half_height: float,
    epsilon: float,
    maximum_distance: float | None = None,
) -> NDArray[np.bool_]:
    """判断一组同向射线是否与一面有限矩形相交。"""

    denominator = float(direction @ rectangle_normal)
    if abs(denominator) <= epsilon:
        return np.zeros(origins.shape[0], dtype=bool)

    distance = (rectangle_center - origins) @ rectangle_normal / denominator
    active = distance > epsilon
    if maximum_distance is not None:
        active &= distance < maximum_distance - epsilon
    if not np.any(active):
        return active

    intersections = origins + distance[:, None] * direction[None, :]
    relative = intersections - rectangle_center[None, :]
    local_width = relative @ rectangle_width_axis
    local_height = relative @ rectangle_height_axis
    active &= np.abs(local_width) <= half_width + epsilon
    active &= np.abs(local_height) <= half_height + epsilon
    return active


def _blocked_by_candidates(
    origins: FloatArray,
    direction: FloatArray,
    candidates: IntArray,
    prepared: PreparedField,
    orientation: MirrorOrientation,
    solver: SolverConfig,
    maximum_distance: float | None = None,
) -> NDArray[np.bool_]:
    blocked = np.zeros(origins.shape[0], dtype=bool)
    config = prepared.config

    for candidate in candidates:
        active_indices = np.flatnonzero(~blocked)
        if active_indices.size == 0:
            break
        hits = ray_rectangle_hits(
            origins=origins[active_indices],
            direction=direction,
            rectangle_center=prepared.centers[candidate],
            rectangle_normal=orientation.normals[candidate],
            rectangle_width_axis=orientation.width_axes[candidate],
            rectangle_height_axis=orientation.height_axes[candidate],
            half_width=config.mirror_width / 2.0,
            half_height=config.mirror_height / 2.0,
            epsilon=solver.ray_epsilon,
            maximum_distance=maximum_distance,
        )
        blocked[active_indices[hits]] = True

    return blocked


def calculate_shadow_blocking_efficiency(
    prepared: PreparedField,
    orientation: MirrorOrientation,
    sun_direction: FloatArray,
    solver: SolverConfig,
) -> FloatArray:
    """逐镜计算入射阴影和反射遮挡损失的采样点并集。"""

    mirror_count = prepared.mirror_count
    if mirror_count == 1:
        return np.ones(1, dtype=float)

    config = prepared.config
    offsets = mirror_grid_offsets(
        solver.shadow_grid_size,
        config.mirror_width,
        config.mirror_height,
    )
    sample_count = offsets.shape[0]
    tree = cKDTree(prepared.centers[:, :2])
    bounding_radius = 0.5 * np.hypot(
        config.mirror_width,
        config.mirror_height,
    )
    reach = 2.0 * bounding_radius * solver.candidate_margin
    efficiencies = np.empty(mirror_count, dtype=float)

    for index in range(mirror_count):
        points = (
            prepared.centers[index][None, :]
            + offsets[:, :1] * orientation.width_axes[index][None, :]
            + offsets[:, 1:] * orientation.height_axes[index][None, :]
        )
        neighbors = np.asarray(
            tree.query_ball_point(
                prepared.centers[index, :2],
                solver.neighbor_radius_m,
            ),
            dtype=np.int64,
        )
        neighbors = neighbors[neighbors != index]

        incoming_candidates = _direction_candidates(
            target_index=index,
            neighbors=neighbors,
            centers=prepared.centers,
            direction=sun_direction,
            reach=reach,
        )
        incoming_blocked = _blocked_by_candidates(
            origins=points,
            direction=sun_direction,
            candidates=incoming_candidates,
            prepared=prepared,
            orientation=orientation,
            solver=solver,
        )

        reflected_direction = prepared.receiver_directions[index]
        reflected_candidates = _direction_candidates(
            target_index=index,
            neighbors=neighbors,
            centers=prepared.centers,
            direction=reflected_direction,
            reach=reach,
            maximum_distance=prepared.receiver_distances[index],
        )
        reflected_blocked = _blocked_by_candidates(
            origins=points,
            direction=reflected_direction,
            candidates=reflected_candidates,
            prepared=prepared,
            orientation=orientation,
            solver=solver,
            maximum_distance=prepared.receiver_distances[index],
        )

        blocked = incoming_blocked | reflected_blocked
        efficiencies[index] = 1.0 - np.count_nonzero(blocked) / sample_count

    return np.clip(efficiencies, 0.0, 1.0)


# ========================================================================
# 截断效率
# ========================================================================


import numpy as np
from numpy.typing import NDArray
from scipy.stats import qmc



def build_sobol_samples(sample_count: int, seed: int) -> FloatArray:
    """生成固定、可复现的四维 Sobol 样本。"""

    exponent = int(math.ceil(math.log2(sample_count)))
    sampler = qmc.Sobol(d=4, scramble=True, seed=seed)
    return sampler.random_base2(exponent)[:sample_count]


def _sun_disk_directions(
    sun_direction: FloatArray,
    samples: FloatArray,
    angular_radius: float,
) -> FloatArray:
    reference = (
        np.array([1.0, 0.0, 0.0])
        if abs(float(sun_direction[2])) > 0.9
        else np.array([0.0, 0.0, 1.0])
    )
    tangent_one = np.cross(reference, sun_direction)
    tangent_one /= np.linalg.norm(tangent_one)
    tangent_two = np.cross(sun_direction, tangent_one)

    radial_angle = angular_radius * np.sqrt(samples[:, 2])
    polar_angle = 2.0 * math.pi * samples[:, 3]
    tangent = (
        np.cos(polar_angle)[:, None] * tangent_one[None, :]
        + np.sin(polar_angle)[:, None] * tangent_two[None, :]
    )
    directions = (
        np.cos(radial_angle)[:, None] * sun_direction[None, :]
        + np.sin(radial_angle)[:, None] * tangent
    )
    directions /= np.linalg.norm(directions, axis=1, keepdims=True)
    return directions


def ray_cylinder_side_hits(
    origins: FloatArray,
    directions: FloatArray,
    tower_x: float,
    tower_y: float,
    radius: float,
    z_min: float,
    z_max: float,
    epsilon: float,
) -> NDArray[np.bool_]:
    """判断任意形状批量射线的两个正根中是否有有限圆柱侧面交点。"""

    origin_x = origins[..., 0] - tower_x
    origin_y = origins[..., 1] - tower_y
    direction_x = directions[..., 0]
    direction_y = directions[..., 1]

    a = direction_x**2 + direction_y**2
    b = 2.0 * (origin_x * direction_x + origin_y * direction_y)
    c = origin_x**2 + origin_y**2 - radius**2
    discriminant = b**2 - 4.0 * a * c

    valid = (a > epsilon) & (discriminant >= 0.0)
    square_root = np.sqrt(np.maximum(discriminant, 0.0))
    denominator = np.where(valid, 2.0 * a, 1.0)
    near = (-b - square_root) / denominator
    far = (-b + square_root) / denominator

    near_z = origins[..., 2] + near * directions[..., 2]
    far_z = origins[..., 2] + far * directions[..., 2]
    near_hit = (
        (near > epsilon)
        & (near_z >= z_min - epsilon)
        & (near_z <= z_max + epsilon)
    )
    far_hit = (
        (far > epsilon)
        & (far_z >= z_min - epsilon)
        & (far_z <= z_max + epsilon)
    )
    return valid & (near_hit | far_hit)


def calculate_truncation_efficiency(
    prepared: PreparedField,
    orientation: MirrorOrientation,
    sun_direction: FloatArray,
    solver: SolverConfig,
) -> FloatArray:
    """联合采样镜面位置和太阳圆盘方向，计算集热器截断效率。"""

    config = prepared.config
    samples = build_sobol_samples(solver.truncation_rays, solver.sobol_seed)
    local_width = (samples[:, 0] - 0.5) * config.mirror_width
    local_height = (samples[:, 1] - 0.5) * config.mirror_height
    sampled_sun = _sun_disk_directions(
        sun_direction,
        samples,
        config.solar_angular_radius_rad,
    )
    incoming = -sampled_sun

    efficiencies = np.empty(prepared.mirror_count, dtype=float)
    chunk_size = solver.truncation_chunk_size
    for start in range(0, prepared.mirror_count, chunk_size):
        stop = min(start + chunk_size, prepared.mirror_count)
        centers = prepared.centers[start:stop]
        normals = orientation.normals[start:stop]
        width_axes = orientation.width_axes[start:stop]
        height_axes = orientation.height_axes[start:stop]

        origins = (
            centers[:, None, :]
            + local_width[None, :, None] * width_axes[:, None, :]
            + local_height[None, :, None] * height_axes[:, None, :]
        )
        incoming_chunk = np.broadcast_to(
            incoming[None, :, :],
            origins.shape,
        )
        dot = np.einsum("csj,cj->cs", incoming_chunk, normals)
        reflected = (
            incoming_chunk - 2.0 * dot[:, :, None] * normals[:, None, :]
        )
        hits = ray_cylinder_side_hits(
            origins=origins,
            directions=reflected,
            tower_x=config.tower_x,
            tower_y=config.tower_y,
            radius=config.receiver_radius,
            z_min=config.receiver_z_min,
            z_max=config.receiver_z_max,
            epsilon=solver.ray_epsilon,
        )
        efficiencies[start:stop] = np.mean(hits, axis=1)

    return np.clip(efficiencies, 0.0, 1.0)


# ========================================================================
# 月平均、年平均与单镜年平均汇总
# ========================================================================

from dataclasses import dataclass
from typing import Any, Sequence

import numpy as np


@dataclass(frozen=True)
class MonthlyResult:
    month: int
    average_optical_efficiency: float
    average_cosine_efficiency: float
    average_shadow_blocking_efficiency: float
    average_atmospheric_efficiency: float
    average_truncation_efficiency: float
    field_output_mw: float
    unit_area_output_kw_m2: float


@dataclass(frozen=True)
class AnnualResult:
    average_optical_efficiency: float
    average_cosine_efficiency: float
    average_shadow_blocking_efficiency: float
    average_atmospheric_efficiency: float
    average_truncation_efficiency: float
    field_output_mw: float
    unit_area_output_kw_m2: float


@dataclass(frozen=True)
class MirrorAnnualResult:
    mirror_id: int
    x_m: float
    y_m: float
    radius_to_tower_m: float
    average_optical_efficiency: float
    average_cosine_efficiency: float
    average_shadow_blocking_efficiency: float
    average_atmospheric_efficiency: float
    average_truncation_efficiency: float
    average_output_power_kw: float


@dataclass(frozen=True)
class Question1Solution:
    time_results: tuple[Any, ...]
    monthly_results: tuple[MonthlyResult, ...]
    annual_result: AnnualResult
    mirror_annual_results: tuple[MirrorAnnualResult, ...]


_MEAN_FIELDS = (
    "average_optical_efficiency",
    "average_cosine_efficiency",
    "average_shadow_blocking_efficiency",
    "average_atmospheric_efficiency",
    "average_truncation_efficiency",
    "field_output_mw",
    "unit_area_output_kw_m2",
)


def _means(records: Sequence[Any]) -> tuple[float, ...]:
    if not records:
        raise ValueError("汇总记录不能为空。")
    return tuple(
        float(np.mean([getattr(record, field) for record in records]))
        for field in _MEAN_FIELDS
    )


def summarize_monthly(records: Sequence[Any]) -> tuple[MonthlyResult, ...]:
    """对每月规定的五个时刻等权平均。"""

    results: list[MonthlyResult] = []
    for month in sorted({record.month for record in records}):
        monthly = [record for record in records if record.month == month]
        results.append(MonthlyResult(month, *_means(monthly)))
    return tuple(results)


def summarize_annual(records: Sequence[Any]) -> AnnualResult:
    """对题目规定的全部时刻等权平均。"""

    return AnnualResult(*_means(records))


def summarize_mirror_annual(
    mirror_xy: np.ndarray,
    tower_x: float,
    tower_y: float,
    state_count: int,
    optical_efficiency_sum: np.ndarray,
    cosine_efficiency_sum: np.ndarray,
    shadow_blocking_efficiency_sum: np.ndarray,
    atmospheric_efficiency_sum: np.ndarray,
    truncation_efficiency_sum: np.ndarray,
    output_power_kw_sum: np.ndarray,
) -> tuple[MirrorAnnualResult, ...]:
    """由逐时刻运行和生成单镜年平均结果，不保留单镜逐时刻明细。"""

    if state_count < 1:
        raise ValueError("state_count 必须大于等于 1。")
    radius = np.hypot(mirror_xy[:, 0] - tower_x, mirror_xy[:, 1] - tower_y)
    means = {
        "optical": optical_efficiency_sum / state_count,
        "cosine": cosine_efficiency_sum / state_count,
        "shadow": shadow_blocking_efficiency_sum / state_count,
        "atmospheric": atmospheric_efficiency_sum / state_count,
        "truncation": truncation_efficiency_sum / state_count,
        "power": output_power_kw_sum / state_count,
    }
    return tuple(
        MirrorAnnualResult(
            mirror_id=index + 1,
            x_m=float(mirror_xy[index, 0]),
            y_m=float(mirror_xy[index, 1]),
            radius_to_tower_m=float(radius[index]),
            average_optical_efficiency=float(means["optical"][index]),
            average_cosine_efficiency=float(means["cosine"][index]),
            average_shadow_blocking_efficiency=float(means["shadow"][index]),
            average_atmospheric_efficiency=float(
                means["atmospheric"][index]
            ),
            average_truncation_efficiency=float(means["truncation"][index]),
            average_output_power_kw=float(means["power"][index]),
        )
        for index in range(mirror_xy.shape[0])
    )


# ========================================================================
# 结果与表格输出
# ========================================================================

import json
from typing import Iterable



def _write_csv(path: Path, rows: Sequence[dict[str, Any]]) -> None:
    if not rows:
        raise ValueError(f"没有可写入 {path.name} 的结果。")
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def _display_source_path(source_path: str | Path) -> str:
    path = Path(source_path)
    if not path.is_absolute():
        return path.as_posix()
    if "task" in path.parts:
        task_index = path.parts.index("task")
        return Path(*path.parts[task_index:]).as_posix()
    return path.name


def write_question1_results(
    output_dir: str | Path,
    time_records: Iterable[Any],
    monthly_records: Iterable[Any],
    annual_record: Any,
    mirror_annual_records: Iterable[Any],
    field_config: FieldConfig,
    solver_config: SolverConfig,
    source_path: str | Path,
    mirror_count: int,
) -> dict[str, Path]:
    """保持原有四类结果文件名和字段口径不变。"""

    destination = Path(output_dir)
    destination.mkdir(parents=True, exist_ok=True)

    time_rows = [asdict(record) for record in time_records]
    monthly_rows = [asdict(record) for record in monthly_records]
    annual_row = asdict(annual_record)
    mirror_annual_rows = [
        asdict(record) for record in mirror_annual_records
    ]
    months = sorted({row["month"] for row in time_rows})
    solar_times = sorted({row["solar_time"] for row in time_rows})

    time_path = destination / "02_逐时刻计算结果.csv"
    monthly_path = destination / "03_月平均计算结果.csv"
    annual_path = destination / "04_年平均计算结果.json"
    mirror_annual_path = destination / "05_单镜年平均结果.csv"
    run_path = destination / "06_运行配置.json"

    _write_csv(time_path, time_rows)
    _write_csv(monthly_path, monthly_rows)
    _write_csv(mirror_annual_path, mirror_annual_rows)
    annual_path.write_text(
        json.dumps(annual_row, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    run_path.write_text(
        json.dumps(
            {
                "source": _display_source_path(source_path),
                "field": field_config.to_dict(),
                "solver": solver_config.to_dict(),
                "run": {
                    "mirror_count": mirror_count,
                    "months": months,
                    "solar_times": solar_times,
                    "time_state_count": len(time_rows),
                },
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )

    return {
        "time": time_path,
        "monthly": monthly_path,
        "annual": annual_path,
        "mirror_annual": mirror_annual_path,
        "config": run_path,
    }


def write_paper_tables(
    output_dir: str | Path,
    monthly_records: Iterable[Any],
    annual_record: Any,
) -> dict[str, Path]:
    """将月平均、年平均和验证表集中到一个展示文件。"""

    destination = Path(output_dir)
    destination.mkdir(parents=True, exist_ok=True)
    table_path = destination / "07_论文结果与验证表.md"

    monthly_lines = [
        "# 第一问结果与验证表",
        "",
        "本文档汇总第一问的月平均、年平均和数值收敛结果。",
        "",
        "## 表 1 每月 21 日平均光学效率及输出功率",
        "",
        "| 日期 | 平均光学效率 | 平均余弦效率 | 平均阴影遮挡效率 | 平均截断效率 | 单位面积镜面平均输出热功率 (kW/m²) |",
        "| --- | ---: | ---: | ---: | ---: | ---: |",
    ]
    for record in monthly_records:
        monthly_lines.append(
            f"| {record.month} 月 21 日 "
            f"| {record.average_optical_efficiency:.6f} "
            f"| {record.average_cosine_efficiency:.6f} "
            f"| {record.average_shadow_blocking_efficiency:.6f} "
            f"| {record.average_truncation_efficiency:.6f} "
            f"| {record.unit_area_output_kw_m2:.6f} |"
        )
    annual_lines = [
        "",
        "## 表 2 年平均光学效率及输出功率",
        "",
        "| 年平均光学效率 | 年平均余弦效率 | 年平均阴影遮挡效率 | 年平均截断效率 | 年平均输出热功率 (MW) | 单位面积镜面年平均输出热功率 (kW/m²) |",
        "| ---: | ---: | ---: | ---: | ---: | ---: |",
        (
            f"| {annual_record.average_optical_efficiency:.6f} "
            f"| {annual_record.average_cosine_efficiency:.6f} "
            f"| {annual_record.average_shadow_blocking_efficiency:.6f} "
            f"| {annual_record.average_truncation_efficiency:.6f} "
            f"| {annual_record.field_output_mw:.6f} "
            f"| {annual_record.unit_area_output_kw_m2:.6f} |"
        ),
    ]
    table_path.write_text(
        "\n".join(monthly_lines + annual_lines) + "\n",
        encoding="utf-8",
    )
    return {"paper_tables": table_path}


def write_validation_table(
    output_dir: str | Path,
    validation_records: Iterable[Any],
) -> dict[str, Path]:
    """把三组收敛实验追加为一张验证表。"""

    destination = Path(output_dir)
    rows = [asdict(record) for record in validation_records]
    table_path = destination / "07_论文结果与验证表.md"

    lines = [
        "",
        "## 表 3 数值收敛验证",
        "",
        "| 验证项目 | 参数 | 观测指标 | 数值 | 相对正式配置差异 | 运行时间 (s) |",
        "| --- | ---: | --- | ---: | ---: | ---: |",
    ]
    for row in rows:
        lines.append(
            f"| {row['category']} "
            f"| {row['parameter']} "
            f"| {row['metric']} "
            f"| {row['value']:.6f} "
            f"| {row['relative_difference_percent']:.4f}% "
            f"| {row['runtime_seconds']:.3f} |"
        )
    with table_path.open("a", encoding="utf-8") as handle:
        handle.write("\n".join(lines) + "\n")
    return {"validation_table": table_path}


# ========================================================================
# 两张正式结果图
# ========================================================================

# ruff: noqa: E402


import os
import tempfile

_mpl_config = Path(tempfile.gettempdir()) / "cowork-matplotlib"
_mpl_config.mkdir(parents=True, exist_ok=True)
os.environ.setdefault("MPLCONFIGDIR", str(_mpl_config))

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
from matplotlib.colors import LinearSegmentedColormap


BLUE = "#2F5D7C"
BLUE_DARK = "#173B54"
BLUE_LIGHT = "#DCE9F1"
ORANGE = "#D97706"
DARK = "#24323D"
GREY = "#76838F"
LIGHT_GREY = "#D7DEE3"


def _configure_style() -> None:
    plt.rcParams.update(
        {
            "font.family": "sans-serif",
            "font.sans-serif": [
                "Hiragino Sans GB",
                "Arial Unicode MS",
                "PingFang SC",
                "DejaVu Sans",
            ],
            "axes.unicode_minus": False,
            "figure.facecolor": "white",
            "axes.facecolor": "white",
            "axes.edgecolor": DARK,
            "axes.labelcolor": DARK,
            "axes.titlecolor": DARK,
            "xtick.color": DARK,
            "ytick.color": DARK,
            "text.color": DARK,
            "grid.color": LIGHT_GREY,
            "grid.linewidth": 0.7,
            "grid.alpha": 0.65,
        }
    )


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def plot_monthly_performance(output_dir: str | Path) -> Path:
    """绘制月平均综合光学效率和单位面积输出热功率。"""

    _configure_style()
    destination = Path(output_dir)
    rows = _read_csv(destination / "03_月平均计算结果.csv")
    months = np.array([int(row["month"]) for row in rows])
    optical = np.array(
        [float(row["average_optical_efficiency"]) for row in rows]
    )
    unit_power = np.array(
        [float(row["unit_area_output_kw_m2"]) for row in rows]
    )
    output_path = destination / "08_月平均光学性能与输出热功率.png"

    fig, (ax_efficiency, ax_power) = plt.subplots(
        2,
        1,
        figsize=(8.0, 6.3),
        sharex=True,
        gridspec_kw={"height_ratios": (1.0, 1.15), "hspace": 0.12},
    )

    ax_efficiency.plot(
        months,
        optical,
        color=BLUE,
        linewidth=2.2,
        marker="o",
        markersize=5.0,
        markerfacecolor="white",
        markeredgewidth=1.5,
    )
    efficiency_padding = 0.02
    ax_efficiency.set_ylim(
        max(0.0, float(np.min(optical)) - efficiency_padding),
        min(1.0, float(np.max(optical)) + efficiency_padding),
    )
    ax_efficiency.set_ylabel("综合光学效率")
    ax_efficiency.grid(axis="y")
    ax_efficiency.spines[["top", "right"]].set_visible(False)

    ax_power.bar(
        months,
        unit_power,
        width=0.62,
        color=ORANGE,
        edgecolor="white",
        linewidth=0.8,
    )
    ax_power.set_ylim(0.0, float(np.max(unit_power)) * 1.14)
    ax_power.set_ylabel(
        r"单位面积输出热功率 ($\mathrm{kW\,m^{-2}}$)"
    )
    ax_power.set_xlabel("月份")
    ax_power.set_xticks(months)
    ax_power.grid(axis="y")
    ax_power.set_axisbelow(True)
    ax_power.spines[["top", "right"]].set_visible(False)

    fig.suptitle(
        "月平均光学性能与输出热功率",
        fontsize=15,
        y=0.98,
    )
    fig.subplots_adjust(left=0.13, right=0.97, top=0.91, bottom=0.10)
    fig.savefig(output_path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    return output_path


def plot_mirror_annual_efficiency_map(output_dir: str | Path) -> Path:
    """绘制 1745 面定日镜的年平均综合光学效率空间分布。"""

    _configure_style()
    destination = Path(output_dir)
    rows = _read_csv(destination / "05_单镜年平均结果.csv")
    x = np.array([float(row["x_m"]) for row in rows])
    y = np.array([float(row["y_m"]) for row in rows])
    optical = np.array(
        [float(row["average_optical_efficiency"]) for row in rows]
    )
    output_path = destination / "09_单镜年平均综合光学效率空间分布.png"

    efficiency_cmap = LinearSegmentedColormap.from_list(
        "heliostat_efficiency",
        (BLUE_LIGHT, "#8EB7CF", BLUE, BLUE_DARK),
    )
    fig, ax = plt.subplots(figsize=(7.4, 6.6))
    points = ax.scatter(
        x,
        y,
        c=optical,
        cmap=efficiency_cmap,
        vmin=float(np.min(optical)),
        vmax=float(np.max(optical)),
        s=18,
        linewidths=0,
    )
    ax.scatter(
        [0.0],
        [0.0],
        marker="*",
        s=190,
        color=ORANGE,
        edgecolor=DARK,
        linewidth=0.8,
        label="吸收塔",
        zorder=4,
    )
    limit = max(float(np.max(np.abs(x))), float(np.max(np.abs(y)))) + 20.0
    ax.set_xlim(-limit, limit)
    ax.set_ylim(-limit, limit)
    ax.set_aspect("equal", adjustable="box")
    ax.set_xlabel("x 坐标 (m)")
    ax.set_ylabel("y 坐标 (m)")
    ax.set_title("单镜年平均综合光学效率空间分布", fontsize=15, pad=12)
    ax.grid(color=LIGHT_GREY, linewidth=0.6, alpha=0.55)
    ax.set_axisbelow(True)
    ax.legend(loc="upper right", frameon=False)

    colorbar = fig.colorbar(points, ax=ax, pad=0.025, fraction=0.047)
    colorbar.set_label("年平均综合光学效率")
    colorbar.outline.set_edgecolor(GREY)
    colorbar.outline.set_linewidth(0.7)

    fig.subplots_adjust(left=0.11, right=0.91, top=0.91, bottom=0.10)
    fig.savefig(output_path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    return output_path


def build_paper_figures(output_dir: str | Path) -> dict[str, Path]:
    """生成第一问最终采用的两张结果图。"""

    return {
        "monthly_performance": plot_monthly_performance(output_dir),
        "mirror_efficiency_map": plot_mirror_annual_efficiency_map(
            output_dir
        ),
    }


# ========================================================================
# 逐时刻求解、验证与入口
# ========================================================================

import argparse
import time
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Callable

import numpy as np



SOLAR_TIMES = (9.0, 10.5, 12.0, 13.5, 15.0)
PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_INPUT = PROJECT_ROOT / "task" / "A" / "fj.xlsx"
DEFAULT_OUTPUT = PROJECT_ROOT / "outputs" / "q1"


@dataclass(frozen=True)
class TimeResult:
    month: int
    solar_time: float
    dni_kw_m2: float
    average_optical_efficiency: float
    average_cosine_efficiency: float
    average_shadow_blocking_efficiency: float
    average_atmospheric_efficiency: float
    average_truncation_efficiency: float
    field_output_mw: float
    unit_area_output_kw_m2: float
    maximum_reflection_error: float


@dataclass(frozen=True)
class ValidationResult:
    category: str
    parameter: str
    metric: str
    value: float
    relative_difference_percent: float
    runtime_seconds: float


ProgressCallback = Callable[[int, int, TimeResult], None]


def _check_efficiency(name: str, values: np.ndarray) -> None:
    tolerance = 1e-12
    if np.any(values < -tolerance) or np.any(values > 1.0 + tolerance):
        minimum = float(np.min(values))
        maximum = float(np.max(values))
        raise RuntimeError(
            f"{name} 超出 [0, 1]：min={minimum:.6g}, max={maximum:.6g}"
        )


def evaluate_time(
    prepared: PreparedField,
    month: int,
    solar_time: float,
    solver: SolverConfig,
    mirror_sums: dict[str, np.ndarray] | None = None,
) -> TimeResult:
    """计算一个月份、一个规定时刻的全场平均结果。"""

    solar = calculate_solar_state(
        month=month,
        solar_time=solar_time,
        latitude_deg=prepared.config.latitude_deg,
        altitude_km=prepared.config.altitude_km,
    )
    orientation = calculate_orientation(prepared, solar.direction)
    reflection_error = maximum_reflection_error(
        prepared,
        orientation,
        solar.direction,
    )
    if reflection_error >= 1e-8:
        raise RuntimeError(f"中心光线反射误差过大：{reflection_error:.3e}")

    if solver.calculate_shadow:
        shadow = calculate_shadow_blocking_efficiency(
            prepared,
            orientation,
            solar.direction,
            solver,
        )
    else:
        shadow = np.ones(prepared.mirror_count, dtype=float)

    if solver.calculate_truncation:
        truncation = calculate_truncation_efficiency(
            prepared,
            orientation,
            solar.direction,
            solver,
        )
    else:
        truncation = np.ones(prepared.mirror_count, dtype=float)

    cosine = orientation.cosine_efficiency
    atmospheric = prepared.atmospheric_efficiency
    optical = (
        cosine
        * shadow
        * atmospheric
        * truncation
        * prepared.config.reflectivity
    )
    for name, values in (
        ("余弦效率", cosine),
        ("阴影遮挡效率", shadow),
        ("大气透射率", atmospheric),
        ("截断效率", truncation),
        ("光学效率", optical),
    ):
        _check_efficiency(name, values)

    mirror_power_kw = solar.dni_kw_m2 * prepared.config.mirror_area * optical
    if mirror_sums is not None:
        mirror_sums["optical_efficiency_sum"] += optical
        mirror_sums["cosine_efficiency_sum"] += cosine
        mirror_sums["shadow_blocking_efficiency_sum"] += shadow
        mirror_sums["atmospheric_efficiency_sum"] += atmospheric
        mirror_sums["truncation_efficiency_sum"] += truncation
        mirror_sums["output_power_kw_sum"] += mirror_power_kw

    field_power_kw = float(np.sum(mirror_power_kw))
    return TimeResult(
        month=month,
        solar_time=solar_time,
        dni_kw_m2=solar.dni_kw_m2,
        average_optical_efficiency=float(np.mean(optical)),
        average_cosine_efficiency=float(np.mean(cosine)),
        average_shadow_blocking_efficiency=float(np.mean(shadow)),
        average_atmospheric_efficiency=float(np.mean(atmospheric)),
        average_truncation_efficiency=float(np.mean(truncation)),
        field_output_mw=field_power_kw / 1000.0,
        unit_area_output_kw_m2=field_power_kw / prepared.total_mirror_area,
        maximum_reflection_error=reflection_error,
    )


def solve_question1(
    prepared: PreparedField,
    solver: SolverConfig,
    months: Sequence[int] = tuple(range(1, 13)),
    solar_times: Sequence[float] = SOLAR_TIMES,
    progress: ProgressCallback | None = None,
) -> Question1Solution:
    """执行所选月份和时刻；默认即题目规定的 60 个状态。"""

    if not months or not solar_times:
        raise ValueError("months 和 solar_times 不能为空。")
    if any(month < 1 or month > 12 for month in months):
        raise ValueError("months 必须位于 1 到 12。")

    records: list[TimeResult] = []
    mirror_sums = {
        name: np.zeros(prepared.mirror_count, dtype=float)
        for name in (
            "optical_efficiency_sum",
            "cosine_efficiency_sum",
            "shadow_blocking_efficiency_sum",
            "atmospheric_efficiency_sum",
            "truncation_efficiency_sum",
            "output_power_kw_sum",
        )
    }
    total = len(months) * len(solar_times)
    for month in months:
        for solar_time in solar_times:
            record = evaluate_time(
                prepared,
                month,
                solar_time,
                solver,
                mirror_sums=mirror_sums,
            )
            records.append(record)
            if progress is not None:
                progress(len(records), total, record)

    time_results = tuple(records)
    return Question1Solution(
        time_results=time_results,
        monthly_results=summarize_monthly(time_results),
        annual_result=summarize_annual(time_results),
        mirror_annual_results=summarize_mirror_annual(
            mirror_xy=prepared.centers[:, :2],
            tower_x=prepared.config.tower_x,
            tower_y=prepared.config.tower_y,
            state_count=len(time_results),
            **mirror_sums,
        ),
    )


def run_validation_suite(
    prepared: PreparedField,
    base_solver: SolverConfig,
) -> tuple[ValidationResult, ...]:
    """运行三组隔离后的收敛实验，供一张验证表使用。"""

    specifications = [
        (
            "阴影网格",
            "10×10",
            replace(
                base_solver,
                shadow_grid_size=10,
                calculate_shadow=True,
                calculate_truncation=False,
            ),
            "average_shadow_blocking_efficiency",
            "年平均阴影遮挡效率",
            False,
        ),
        (
            "阴影网格",
            "15×15",
            replace(
                base_solver,
                shadow_grid_size=15,
                calculate_shadow=True,
                calculate_truncation=False,
            ),
            "average_shadow_blocking_efficiency",
            "年平均阴影遮挡效率",
            True,
        ),
        (
            "阴影网格",
            "20×20",
            replace(
                base_solver,
                shadow_grid_size=20,
                calculate_shadow=True,
                calculate_truncation=False,
            ),
            "average_shadow_blocking_efficiency",
            "年平均阴影遮挡效率",
            False,
        ),
        (
            "邻镜半径",
            "40 m",
            replace(
                base_solver,
                neighbor_radius_m=40.0,
                calculate_shadow=True,
                calculate_truncation=False,
            ),
            "average_shadow_blocking_efficiency",
            "年平均阴影遮挡效率",
            False,
        ),
        (
            "邻镜半径",
            "60 m",
            replace(
                base_solver,
                neighbor_radius_m=60.0,
                calculate_shadow=True,
                calculate_truncation=False,
            ),
            "average_shadow_blocking_efficiency",
            "年平均阴影遮挡效率",
            True,
        ),
        (
            "邻镜半径",
            "80 m",
            replace(
                base_solver,
                neighbor_radius_m=80.0,
                calculate_shadow=True,
                calculate_truncation=False,
            ),
            "average_shadow_blocking_efficiency",
            "年平均阴影遮挡效率",
            False,
        ),
        (
            "截断光线",
            "128",
            replace(
                base_solver,
                truncation_rays=128,
                calculate_shadow=False,
                calculate_truncation=True,
            ),
            "average_truncation_efficiency",
            "年平均截断效率",
            False,
        ),
        (
            "截断光线",
            "256",
            replace(
                base_solver,
                truncation_rays=256,
                calculate_shadow=False,
                calculate_truncation=True,
            ),
            "average_truncation_efficiency",
            "年平均截断效率",
            True,
        ),
        (
            "截断光线",
            "512",
            replace(
                base_solver,
                truncation_rays=512,
                calculate_shadow=False,
                calculate_truncation=True,
            ),
            "average_truncation_efficiency",
            "年平均截断效率",
            False,
        ),
    ]

    raw: list[tuple[str, str, str, float, float, bool]] = []
    for category, parameter, solver, field, metric, reference in specifications:
        started = time.perf_counter()
        solution = solve_question1(prepared, solver)
        elapsed = time.perf_counter() - started
        value = float(getattr(solution.annual_result, field))
        raw.append((category, parameter, metric, value, elapsed, reference))

    baselines = {
        category: value
        for category, _, _, value, _, reference in raw
        if reference
    }
    return tuple(
        ValidationResult(
            category=category,
            parameter=parameter,
            metric=metric,
            value=value,
            relative_difference_percent=(
                abs(value - baselines[category]) / abs(baselines[category]) * 100.0
            ),
            runtime_seconds=elapsed,
        )
        for category, parameter, metric, value, elapsed, _ in raw
    )


def _comma_ints(value: str) -> tuple[int, ...]:
    try:
        result = tuple(int(item.strip()) for item in value.split(",") if item.strip())
    except ValueError as exc:
        raise argparse.ArgumentTypeError("月份应使用逗号分隔的整数。") from exc
    if not result:
        raise argparse.ArgumentTypeError("月份列表不能为空。")
    return result


def _comma_floats(value: str) -> tuple[float, ...]:
    try:
        result = tuple(
            float(item.strip()) for item in value.split(",") if item.strip()
        )
    except ValueError as exc:
        raise argparse.ArgumentTypeError("时刻应使用逗号分隔的数字。") from exc
    if not result:
        raise argparse.ArgumentTypeError("时刻列表不能为空。")
    return result


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="计算 CUMCM 2023 A 题第一问的镜场光学效率和输出热功率"
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--shadow-grid", type=int, default=15)
    parser.add_argument("--truncation-rays", type=int, default=256)
    parser.add_argument("--neighbor-radius", type=float, default=60.0)
    parser.add_argument("--truncation-chunk-size", type=int, default=128)
    parser.add_argument("--sobol-seed", type=int, default=2023)
    parser.add_argument(
        "--months",
        type=_comma_ints,
        default=tuple(range(1, 13)),
        help="逗号分隔；默认 1 到 12 月",
    )
    parser.add_argument(
        "--times",
        type=_comma_floats,
        default=SOLAR_TIMES,
        help="逗号分隔的当地太阳时",
    )
    parser.add_argument(
        "--limit-mirrors",
        type=int,
        default=None,
        help="仅用于调试；只计算附件中的前 N 面镜子",
    )
    parser.add_argument("--skip-shadow", action="store_true")
    parser.add_argument("--skip-truncation", action="store_true")
    parser.add_argument("--skip-figures", action="store_true")
    parser.add_argument(
        "--run-validation",
        action="store_true",
        help="额外运行三组收敛实验并生成一张验证表",
    )
    parser.add_argument("--quiet", action="store_true")
    return parser


def _progress(current: int, total: int, record: TimeResult) -> None:
    hour = int(record.solar_time)
    minute = int(round((record.solar_time - hour) * 60.0))
    print(
        f"[{current:02d}/{total:02d}] "
        f"{record.month:02d}月21日 {hour:02d}:{minute:02d} "
        f"光学效率={record.average_optical_efficiency:.4f} "
        f"输出={record.field_output_mw:.3f} MW"
    )


def run(argv: Sequence[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    mirror_xy = load_mirror_xy(args.input)
    if args.limit_mirrors is not None:
        if args.limit_mirrors < 1:
            raise SystemExit("--limit-mirrors 必须大于等于 1。")
        mirror_xy = mirror_xy[: args.limit_mirrors]

    field_config = FieldConfig()
    solver_config = SolverConfig(
        shadow_grid_size=args.shadow_grid,
        truncation_rays=args.truncation_rays,
        neighbor_radius_m=args.neighbor_radius,
        truncation_chunk_size=args.truncation_chunk_size,
        sobol_seed=args.sobol_seed,
        calculate_shadow=not args.skip_shadow,
        calculate_truncation=not args.skip_truncation,
    )
    prepared = prepare_field(mirror_xy, field_config)
    solution = solve_question1(
        prepared=prepared,
        solver=solver_config,
        months=args.months,
        solar_times=args.times,
        progress=None if args.quiet else _progress,
    )
    written = write_question1_results(
        output_dir=args.output,
        time_records=solution.time_results,
        monthly_records=solution.monthly_results,
        annual_record=solution.annual_result,
        mirror_annual_records=solution.mirror_annual_results,
        field_config=field_config,
        solver_config=solver_config,
        source_path=args.input,
        mirror_count=prepared.mirror_count,
    )
    written.update(
        write_paper_tables(
            args.output,
            solution.monthly_results,
            solution.annual_result,
        )
    )

    if args.run_validation:
        validation = run_validation_suite(prepared, solver_config)
        written.update(write_validation_table(args.output, validation))

    if not args.skip_figures:
        written.update(
            build_paper_figures(
                output_dir=args.output,
            )
        )

    annual = solution.annual_result
    print("\n汇总结果")
    print(f"平均光学效率：{annual.average_optical_efficiency:.6f}")
    print(f"平均余弦效率：{annual.average_cosine_efficiency:.6f}")
    print(
        "平均阴影遮挡效率："
        f"{annual.average_shadow_blocking_efficiency:.6f}"
    )
    print(f"平均截断效率：{annual.average_truncation_efficiency:.6f}")
    print(f"平均输出热功率：{annual.field_output_mw:.6f} MW")
    print(
        "单位镜面面积平均输出热功率："
        f"{annual.unit_area_output_kw_m2:.6f} kW/m²"
    )
    print(f"结果目录：{args.output.resolve()}")
    for name, path in written.items():
        print(f"  {name}: {path.relative_to(args.output)}")
    return 0


def main() -> None:
    raise SystemExit(run())


if __name__ == "__main__":
    main()
