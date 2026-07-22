"""三问共用算法：太阳位置、镜面几何、阴影遮挡、截断效率和年度评价。"""
from __future__ import annotations
# ruff: noqa

# ---- config.py ----

from dataclasses import asdict, dataclass

@dataclass(frozen=True)
class FieldConfig:
    latitude_deg: float = 39.4
    altitude_km: float = 3.0
    field_radius: float = 350.0
    exclusion_radius: float = 100.0
    tower_x: float = 0.0
    tower_y: float = 0.0
    receiver_center_z: float = 86.0
    receiver_radius: float = 4.0
    receiver_height: float = 8.0
    mirror_width: float = 6.2
    mirror_height: float = 6.2
    mirror_center_z: float = 4.5
    reflectivity: float = 0.92
    solar_angular_radius_rad: float = 0.00465

    @property
    def receiver_z_min(self) -> float:
        return self.receiver_center_z - self.receiver_height / 2.0

    @property
    def receiver_z_max(self) -> float:
        return self.receiver_center_z + self.receiver_height / 2.0

    @property
    def mirror_area(self) -> float:
        return self.mirror_width * self.mirror_height

    def to_dict(self) -> dict[str, float]:
        return asdict(self)

@dataclass(frozen=True)
class SolverConfig:
    shadow_grid_size: int = 15
    truncation_rays: int = 256
    neighbor_radius_m: float = 60.0
    candidate_margin: float = 1.05
    ray_epsilon: float = 1e-07
    truncation_chunk_size: int = 128
    sobol_seed: int = 2023
    calculate_shadow: bool = True
    calculate_truncation: bool = True

    def __post_init__(self) -> None:
        if self.shadow_grid_size < 1:
            raise ValueError('shadow_grid_size 必须大于等于 1。')
        if self.truncation_rays < 1:
            raise ValueError('truncation_rays 必须大于等于 1。')
        if self.neighbor_radius_m <= 0.0:
            raise ValueError('neighbor_radius_m 必须大于 0。')
        if self.candidate_margin < 1.0:
            raise ValueError('candidate_margin 不能小于 1。')
        if self.ray_epsilon <= 0.0:
            raise ValueError('ray_epsilon 必须大于 0。')
        if self.truncation_chunk_size < 1:
            raise ValueError('truncation_chunk_size 必须大于等于 1。')

    def to_dict(self) -> dict[str, int | float | bool]:
        return asdict(self)

# ---- solar.py ----

import math
from dataclasses import dataclass
import numpy as np
from numpy.typing import NDArray
FloatArray = NDArray[np.float64]

@dataclass(frozen=True)
class SolarState:
    month: int
    solar_time: float
    direction: FloatArray
    altitude_rad: float
    azimuth_rad: float
    declination_rad: float
    dni_kw_m2: float
MONTH_DAYS = (31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31)

def day_from_spring_equinox(month: int, day: int=21) -> int:
    if not 1 <= month <= 12:
        raise ValueError('month 必须位于 1 到 12。')
    if not 1 <= day <= MONTH_DAYS[month - 1]:
        raise ValueError('day 不在指定月份的有效范围内。')
    day_of_year = sum(MONTH_DAYS[:month - 1]) + day
    return day_of_year - 80

def calculate_solar_state(month: int, solar_time: float, latitude_deg: float, altitude_km: float) -> SolarState:
    if not 0.0 <= solar_time <= 24.0:
        raise ValueError('solar_time 必须位于 0 到 24 小时。')
    d = day_from_spring_equinox(month)
    declination = math.asin(math.sin(2.0 * math.pi * d / 365.0) * math.sin(math.radians(23.45)))
    latitude = math.radians(latitude_deg)
    hour_angle = math.pi / 12.0 * (solar_time - 12.0)
    direction = np.array([-math.cos(declination) * math.sin(hour_angle), math.cos(latitude) * math.sin(declination) - math.sin(latitude) * math.cos(declination) * math.cos(hour_angle), math.sin(latitude) * math.sin(declination) + math.cos(latitude) * math.cos(declination) * math.cos(hour_angle)], dtype=float)
    direction /= np.linalg.norm(direction)
    altitude = math.asin(float(np.clip(direction[2], -1.0, 1.0)))
    azimuth = math.atan2(float(direction[0]), float(direction[1])) % (2.0 * math.pi)
    h = altitude_km
    a = 0.4237 - 0.00821 * (6.0 - h) ** 2
    b = 0.5055 + 0.00595 * (6.5 - h) ** 2
    c = 0.2711 + 0.01858 * (2.5 - h) ** 2
    if altitude <= 0.0:
        dni = 0.0
    else:
        dni = 1.366 * (a + b * math.exp(-c / math.sin(altitude)))
    return SolarState(month=month, solar_time=solar_time, direction=direction, altitude_rad=altitude, azimuth_rad=azimuth, declination_rad=declination, dni_kw_m2=dni)

# ---- geometry.py ----

from dataclasses import dataclass
import numpy as np
from numpy.typing import NDArray
FloatArray = NDArray[np.float64]

def normalize_rows(vectors: FloatArray) -> FloatArray:
    norms = np.linalg.norm(vectors, axis=1, keepdims=True)
    if np.any(norms <= 1e-15):
        raise ValueError('不能归一化长度为零的向量。')
    return vectors / norms

def reflect(directions: FloatArray, normals: FloatArray) -> FloatArray:
    dot = np.sum(directions * normals, axis=-1, keepdims=True)
    return directions - 2.0 * dot * normals

@dataclass(frozen=True)
class PreparedField:
    config: FieldConfig
    centers: FloatArray
    mirror_widths: FloatArray
    mirror_heights: FloatArray
    mirror_areas: FloatArray
    receiver_center: FloatArray
    receiver_directions: FloatArray
    receiver_distances: FloatArray
    atmospheric_efficiency: FloatArray

    @property
    def mirror_count(self) -> int:
        return int(self.centers.shape[0])

    @property
    def total_mirror_area(self) -> float:
        return float(np.sum(self.mirror_areas))

@dataclass(frozen=True)
class MirrorOrientation:
    normals: FloatArray
    width_axes: FloatArray
    height_axes: FloatArray
    cosine_efficiency: FloatArray

def _per_mirror_values(values: FloatArray | None, *, fallback: float, mirror_count: int, name: str) -> FloatArray:
    if values is None:
        result = np.full(mirror_count, fallback, dtype=float)
    else:
        result = np.asarray(values, dtype=float)
        if result.ndim != 1 or result.shape[0] != mirror_count:
            raise ValueError(f'{name} 必须是一维且长度等于镜子数 {mirror_count}，实际形状为 {result.shape}。')
        result = result.copy()
    if not np.all(np.isfinite(result)):
        raise ValueError(f'{name} 包含 NaN 或无穷值。')
    if np.any(result <= 0.0):
        raise ValueError(f'{name} 必须全部大于 0。')
    return result

def prepare_field(mirror_xy: FloatArray, config: FieldConfig, *, mirror_widths: FloatArray | None=None, mirror_heights: FloatArray | None=None, mirror_center_zs: FloatArray | None=None) -> PreparedField:
    xy = np.asarray(mirror_xy, dtype=float)
    if xy.ndim != 2 or xy.shape[1] != 2 or xy.shape[0] == 0:
        raise ValueError('镜位坐标必须为非空 N×2 数组。')
    if not np.all(np.isfinite(xy)):
        raise ValueError('镜位坐标包含 NaN 或无穷值。')
    mirror_count = int(xy.shape[0])
    widths = _per_mirror_values(mirror_widths, fallback=config.mirror_width, mirror_count=mirror_count, name='mirror_widths')
    heights = _per_mirror_values(mirror_heights, fallback=config.mirror_height, mirror_count=mirror_count, name='mirror_heights')
    center_zs = _per_mirror_values(mirror_center_zs, fallback=config.mirror_center_z, mirror_count=mirror_count, name='mirror_center_zs')
    areas = widths * heights
    centers = np.column_stack((xy, center_zs))
    receiver_center = np.array([config.tower_x, config.tower_y, config.receiver_center_z], dtype=float)
    receiver_vectors = receiver_center[None, :] - centers
    distances = np.linalg.norm(receiver_vectors, axis=1)
    receiver_directions = receiver_vectors / distances[:, None]
    atmospheric = 0.99321 - 0.0001176 * distances + 1.97e-08 * distances ** 2
    atmospheric = np.clip(atmospheric, 0.0, 1.0)
    return PreparedField(config=config, centers=centers, mirror_widths=widths, mirror_heights=heights, mirror_areas=areas, receiver_center=receiver_center, receiver_directions=receiver_directions, receiver_distances=distances, atmospheric_efficiency=atmospheric)

def calculate_orientation(prepared: PreparedField, sun_direction: FloatArray) -> MirrorOrientation:
    sun_rows = np.broadcast_to(sun_direction, prepared.receiver_directions.shape)
    normals = normalize_rows(sun_rows + prepared.receiver_directions)
    upward = np.broadcast_to(np.array([0.0, 0.0, 1.0]), normals.shape)
    width_axes = np.cross(upward, normals)
    weak = np.linalg.norm(width_axes, axis=1) < 1e-10
    width_axes[weak] = np.array([1.0, 0.0, 0.0])
    width_axes = normalize_rows(width_axes)
    height_axes = normalize_rows(np.cross(normals, width_axes))
    cosine = np.clip(normals @ sun_direction, 0.0, 1.0)
    return MirrorOrientation(normals=normals, width_axes=width_axes, height_axes=height_axes, cosine_efficiency=cosine)

def maximum_reflection_error(prepared: PreparedField, orientation: MirrorOrientation, sun_direction: FloatArray) -> float:
    incoming = np.broadcast_to(-sun_direction, orientation.normals.shape)
    reflected = reflect(incoming, orientation.normals)
    errors = np.linalg.norm(reflected - prepared.receiver_directions, axis=1)
    return float(np.max(errors))

# ---- shadow.py ----

import numpy as np
from numpy.typing import NDArray
from scipy.spatial import cKDTree
IntArray = NDArray[np.int64]

def mirror_grid_offsets(grid_size: int, mirror_width: float, mirror_height: float) -> FloatArray:
    width_step = mirror_width / grid_size
    height_step = mirror_height / grid_size
    width_values = np.linspace(-mirror_width / 2.0 + width_step / 2.0, mirror_width / 2.0 - width_step / 2.0, grid_size)
    height_values = np.linspace(-mirror_height / 2.0 + height_step / 2.0, mirror_height / 2.0 - height_step / 2.0, grid_size)
    (width_grid, height_grid) = np.meshgrid(width_values, height_values, indexing='xy')
    return np.column_stack((width_grid.ravel(), height_grid.ravel()))

def _direction_candidates(target_index: int, neighbors: IntArray, centers: FloatArray, direction: FloatArray, reach: float, maximum_distance: float | None=None) -> IntArray:
    if neighbors.size == 0:
        return neighbors
    delta = centers[neighbors] - centers[target_index]
    projection = delta @ direction
    perpendicular = delta - projection[:, None] * direction[None, :]
    perpendicular_distance = np.linalg.norm(perpendicular, axis=1)
    keep = (projection > -reach) & (perpendicular_distance <= reach)
    if maximum_distance is not None:
        keep &= projection < maximum_distance + reach
    return neighbors[keep]

def ray_rectangle_hits(origins: FloatArray, direction: FloatArray, rectangle_center: FloatArray, rectangle_normal: FloatArray, rectangle_width_axis: FloatArray, rectangle_height_axis: FloatArray, half_width: float, half_height: float, epsilon: float, maximum_distance: float | None=None) -> NDArray[np.bool_]:
    denominator = float(direction @ rectangle_normal)
    if abs(denominator) <= epsilon:
        return np.zeros(origins.shape[0], dtype=bool)
    distance = (rectangle_center - origins) @ rectangle_normal / denominator
    active = distance > epsilon
    if maximum_distance is not None:
        active &= distance < maximum_distance - epsilon
    if not np.any(active):
        return active
    intersections = origins + distance[:, None] * direction[None, :]
    relative = intersections - rectangle_center[None, :]
    local_width = relative @ rectangle_width_axis
    local_height = relative @ rectangle_height_axis
    active &= np.abs(local_width) <= half_width + epsilon
    active &= np.abs(local_height) <= half_height + epsilon
    return active

def _blocked_by_candidates(origins: FloatArray, direction: FloatArray, candidates: IntArray, prepared: PreparedField, orientation: MirrorOrientation, solver: SolverConfig, maximum_distance: float | None=None) -> NDArray[np.bool_]:
    blocked = np.zeros(origins.shape[0], dtype=bool)
    for candidate in candidates:
        active_indices = np.flatnonzero(~blocked)
        if active_indices.size == 0:
            break
        hits = ray_rectangle_hits(origins=origins[active_indices], direction=direction, rectangle_center=prepared.centers[candidate], rectangle_normal=orientation.normals[candidate], rectangle_width_axis=orientation.width_axes[candidate], rectangle_height_axis=orientation.height_axes[candidate], half_width=prepared.mirror_widths[candidate] / 2.0, half_height=prepared.mirror_heights[candidate] / 2.0, epsilon=solver.ray_epsilon, maximum_distance=maximum_distance)
        blocked[active_indices[hits]] = True
    return blocked

def calculate_shadow_blocking_efficiency(prepared: PreparedField, orientation: MirrorOrientation, sun_direction: FloatArray, solver: SolverConfig) -> FloatArray:
    mirror_count = prepared.mirror_count
    if mirror_count == 1:
        return np.ones(1, dtype=float)
    tree = cKDTree(prepared.centers[:, :2])
    bounding_radii = 0.5 * np.hypot(prepared.mirror_widths, prepared.mirror_heights)
    maximum_bounding_radius = float(np.max(bounding_radii))
    efficiencies = np.empty(mirror_count, dtype=float)
    offset_cache: dict[tuple[float, float], FloatArray] = {}
    for index in range(mirror_count):
        size_key = (float(prepared.mirror_widths[index]), float(prepared.mirror_heights[index]))
        offsets = offset_cache.get(size_key)
        if offsets is None:
            offsets = mirror_grid_offsets(solver.shadow_grid_size, size_key[0], size_key[1])
            offset_cache[size_key] = offsets
        sample_count = offsets.shape[0]
        reach = (float(bounding_radii[index]) + maximum_bounding_radius) * solver.candidate_margin
        points = prepared.centers[index][None, :] + offsets[:, :1] * orientation.width_axes[index][None, :] + offsets[:, 1:] * orientation.height_axes[index][None, :]
        neighbors = np.asarray(tree.query_ball_point(prepared.centers[index, :2], solver.neighbor_radius_m), dtype=np.int64)
        neighbors = neighbors[neighbors != index]
        incoming_candidates = _direction_candidates(target_index=index, neighbors=neighbors, centers=prepared.centers, direction=sun_direction, reach=reach)
        incoming_blocked = _blocked_by_candidates(origins=points, direction=sun_direction, candidates=incoming_candidates, prepared=prepared, orientation=orientation, solver=solver)
        reflected_direction = prepared.receiver_directions[index]
        reflected_candidates = _direction_candidates(target_index=index, neighbors=neighbors, centers=prepared.centers, direction=reflected_direction, reach=reach, maximum_distance=prepared.receiver_distances[index])
        reflected_blocked = _blocked_by_candidates(origins=points, direction=reflected_direction, candidates=reflected_candidates, prepared=prepared, orientation=orientation, solver=solver, maximum_distance=prepared.receiver_distances[index])
        blocked = incoming_blocked | reflected_blocked
        efficiencies[index] = 1.0 - np.count_nonzero(blocked) / sample_count
    return np.clip(efficiencies, 0.0, 1.0)

# ---- truncation.py ----

import math
import numpy as np
from numpy.typing import NDArray
from scipy.stats import qmc

def build_sobol_samples(sample_count: int, seed: int) -> FloatArray:
    exponent = int(math.ceil(math.log2(sample_count)))
    sampler = qmc.Sobol(d=4, scramble=True, seed=seed)
    return sampler.random_base2(exponent)[:sample_count]

def _sun_disk_directions(sun_direction: FloatArray, samples: FloatArray, angular_radius: float) -> FloatArray:
    reference = np.array([1.0, 0.0, 0.0]) if abs(float(sun_direction[2])) > 0.9 else np.array([0.0, 0.0, 1.0])
    tangent_one = np.cross(reference, sun_direction)
    tangent_one /= np.linalg.norm(tangent_one)
    tangent_two = np.cross(sun_direction, tangent_one)
    radial_angle = angular_radius * np.sqrt(samples[:, 2])
    polar_angle = 2.0 * math.pi * samples[:, 3]
    tangent = np.cos(polar_angle)[:, None] * tangent_one[None, :] + np.sin(polar_angle)[:, None] * tangent_two[None, :]
    directions = np.cos(radial_angle)[:, None] * sun_direction[None, :] + np.sin(radial_angle)[:, None] * tangent
    directions /= np.linalg.norm(directions, axis=1, keepdims=True)
    return directions

def ray_cylinder_side_hits(origins: FloatArray, directions: FloatArray, tower_x: float, tower_y: float, radius: float, z_min: float, z_max: float, epsilon: float) -> NDArray[np.bool_]:
    origin_x = origins[..., 0] - tower_x
    origin_y = origins[..., 1] - tower_y
    direction_x = directions[..., 0]
    direction_y = directions[..., 1]
    a = direction_x ** 2 + direction_y ** 2
    b = 2.0 * (origin_x * direction_x + origin_y * direction_y)
    c = origin_x ** 2 + origin_y ** 2 - radius ** 2
    discriminant = b ** 2 - 4.0 * a * c
    valid = (a > epsilon) & (discriminant >= 0.0)
    square_root = np.sqrt(np.maximum(discriminant, 0.0))
    denominator = np.where(valid, 2.0 * a, 1.0)
    near = (-b - square_root) / denominator
    far = (-b + square_root) / denominator
    near_z = origins[..., 2] + near * directions[..., 2]
    far_z = origins[..., 2] + far * directions[..., 2]
    near_hit = (near > epsilon) & (near_z >= z_min - epsilon) & (near_z <= z_max + epsilon)
    far_hit = (far > epsilon) & (far_z >= z_min - epsilon) & (far_z <= z_max + epsilon)
    return valid & (near_hit | far_hit)

def calculate_truncation_efficiency(prepared: PreparedField, orientation: MirrorOrientation, sun_direction: FloatArray, solver: SolverConfig) -> FloatArray:
    config = prepared.config
    samples = build_sobol_samples(solver.truncation_rays, solver.sobol_seed)
    unit_width = samples[:, 0] - 0.5
    unit_height = samples[:, 1] - 0.5
    sampled_sun = _sun_disk_directions(sun_direction, samples, config.solar_angular_radius_rad)
    incoming = -sampled_sun
    efficiencies = np.empty(prepared.mirror_count, dtype=float)
    chunk_size = solver.truncation_chunk_size
    for start in range(0, prepared.mirror_count, chunk_size):
        stop = min(start + chunk_size, prepared.mirror_count)
        centers = prepared.centers[start:stop]
        normals = orientation.normals[start:stop]
        width_axes = orientation.width_axes[start:stop]
        height_axes = orientation.height_axes[start:stop]
        local_width = prepared.mirror_widths[start:stop, None] * unit_width[None, :]
        local_height = prepared.mirror_heights[start:stop, None] * unit_height[None, :]
        origins = centers[:, None, :] + local_width[:, :, None] * width_axes[:, None, :] + local_height[:, :, None] * height_axes[:, None, :]
        incoming_chunk = np.broadcast_to(incoming[None, :, :], origins.shape)
        dot = np.einsum('csj,cj->cs', incoming_chunk, normals)
        reflected = incoming_chunk - 2.0 * dot[:, :, None] * normals[:, None, :]
        hits = ray_cylinder_side_hits(origins=origins, directions=reflected, tower_x=config.tower_x, tower_y=config.tower_y, radius=config.receiver_radius, z_min=config.receiver_z_min, z_max=config.receiver_z_max, epsilon=solver.ray_epsilon)
        efficiencies[start:stop] = np.mean(hits, axis=1)
    return np.clip(efficiencies, 0.0, 1.0)

# ---- io.py ----

import csv
from pathlib import Path
import numpy as np
from numpy.typing import NDArray
from openpyxl import load_workbook
FloatArray = NDArray[np.float64]

def load_mirror_xy(path: str | Path, expected_count: int | None=1745) -> FloatArray:
    source = Path(path)
    if not source.exists():
        raise FileNotFoundError(f'找不到定日镜坐标文件：{source}')
    if source.suffix.lower() == '.xlsx':
        workbook = load_workbook(source, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        workbook.close()
        values = [(row[0], row[1]) for row in rows if row[0] is not None]
    elif source.suffix.lower() == '.csv':
        with source.open('r', encoding='utf-8-sig', newline='') as handle:
            reader = csv.reader(handle)
            next(reader, None)
            values = [(row[0], row[1]) for row in reader if row]
    else:
        raise ValueError('坐标文件只支持 .xlsx 或 .csv。')
    try:
        mirror_xy = np.asarray(values, dtype=float)
    except (TypeError, ValueError) as exc:
        raise ValueError(f'坐标文件包含非数值数据：{source}') from exc
    if mirror_xy.ndim != 2 or mirror_xy.shape[1] != 2:
        raise ValueError(f'坐标数据应为 N×2，实际形状为 {mirror_xy.shape}。')
    if not np.all(np.isfinite(mirror_xy)):
        raise ValueError('坐标数据包含 NaN 或无穷值。')
    if expected_count is not None and mirror_xy.shape[0] != expected_count:
        raise ValueError(f'应读取 {expected_count} 面定日镜，实际读取 {mirror_xy.shape[0]} 面。')
    return mirror_xy

# ---- aggregate.py ----

from dataclasses import dataclass
from typing import Any, Sequence
import numpy as np

@dataclass(frozen=True)
class MonthlyResult:
    month: int
    average_optical_efficiency: float
    average_cosine_efficiency: float
    average_shadow_blocking_efficiency: float
    average_atmospheric_efficiency: float
    average_truncation_efficiency: float
    field_output_mw: float
    unit_area_output_kw_m2: float

@dataclass(frozen=True)
class AnnualResult:
    average_optical_efficiency: float
    average_cosine_efficiency: float
    average_shadow_blocking_efficiency: float
    average_atmospheric_efficiency: float
    average_truncation_efficiency: float
    field_output_mw: float
    unit_area_output_kw_m2: float

@dataclass(frozen=True)
class MirrorAnnualResult:
    mirror_id: int
    x_m: float
    y_m: float
    radius_to_tower_m: float
    average_optical_efficiency: float
    average_cosine_efficiency: float
    average_shadow_blocking_efficiency: float
    average_atmospheric_efficiency: float
    average_truncation_efficiency: float
    average_output_power_kw: float

@dataclass(frozen=True)
class Question1Solution:
    time_results: tuple[Any, ...]
    monthly_results: tuple[MonthlyResult, ...]
    annual_result: AnnualResult
    mirror_annual_results: tuple[MirrorAnnualResult, ...]
_MEAN_FIELDS = ('average_optical_efficiency', 'average_cosine_efficiency', 'average_shadow_blocking_efficiency', 'average_atmospheric_efficiency', 'average_truncation_efficiency', 'field_output_mw', 'unit_area_output_kw_m2')

def _means(records: Sequence[Any]) -> tuple[float, ...]:
    if not records:
        raise ValueError('汇总记录不能为空。')
    return tuple((float(np.mean([getattr(record, field) for record in records])) for field in _MEAN_FIELDS))

def summarize_monthly(records: Sequence[Any]) -> tuple[MonthlyResult, ...]:
    results: list[MonthlyResult] = []
    for month in sorted({record.month for record in records}):
        monthly = [record for record in records if record.month == month]
        results.append(MonthlyResult(month, *_means(monthly)))
    return tuple(results)

def summarize_annual(records: Sequence[Any]) -> AnnualResult:
    return AnnualResult(*_means(records))

def summarize_mirror_annual(mirror_xy: np.ndarray, tower_x: float, tower_y: float, state_count: int, optical_efficiency_sum: np.ndarray, cosine_efficiency_sum: np.ndarray, shadow_blocking_efficiency_sum: np.ndarray, atmospheric_efficiency_sum: np.ndarray, truncation_efficiency_sum: np.ndarray, output_power_kw_sum: np.ndarray) -> tuple[MirrorAnnualResult, ...]:
    if state_count < 1:
        raise ValueError('state_count 必须大于等于 1。')
    radius = np.hypot(mirror_xy[:, 0] - tower_x, mirror_xy[:, 1] - tower_y)
    means = {'optical': optical_efficiency_sum / state_count, 'cosine': cosine_efficiency_sum / state_count, 'shadow': shadow_blocking_efficiency_sum / state_count, 'atmospheric': atmospheric_efficiency_sum / state_count, 'truncation': truncation_efficiency_sum / state_count, 'power': output_power_kw_sum / state_count}
    return tuple((MirrorAnnualResult(mirror_id=index + 1, x_m=float(mirror_xy[index, 0]), y_m=float(mirror_xy[index, 1]), radius_to_tower_m=float(radius[index]), average_optical_efficiency=float(means['optical'][index]), average_cosine_efficiency=float(means['cosine'][index]), average_shadow_blocking_efficiency=float(means['shadow'][index]), average_atmospheric_efficiency=float(means['atmospheric'][index]), average_truncation_efficiency=float(means['truncation'][index]), average_output_power_kw=float(means['power'][index])) for index in range(mirror_xy.shape[0])))

# ---- q1/solve.py 中的公共评价部分 ----

import argparse
import time
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Callable, Sequence
import numpy as np
SOLAR_TIMES = (9.0, 10.5, 12.0, 13.5, 15.0)

@dataclass(frozen=True)
class TimeResult:
    month: int
    solar_time: float
    dni_kw_m2: float
    average_optical_efficiency: float
    average_cosine_efficiency: float
    average_shadow_blocking_efficiency: float
    average_atmospheric_efficiency: float
    average_truncation_efficiency: float
    field_output_mw: float
    unit_area_output_kw_m2: float
    maximum_reflection_error: float

@dataclass(frozen=True)
class ValidationResult:
    category: str
    parameter: str
    metric: str
    value: float
    relative_difference_percent: float
    runtime_seconds: float
ProgressCallback = Callable[[int, int, TimeResult], None]

def _check_efficiency(name: str, values: np.ndarray) -> None:
    tolerance = 1e-12
    if np.any(values < -tolerance) or np.any(values > 1.0 + tolerance):
        minimum = float(np.min(values))
        maximum = float(np.max(values))
        raise RuntimeError(f'{name} 超出 [0, 1]：min={minimum:.6g}, max={maximum:.6g}')

def evaluate_time(prepared: PreparedField, month: int, solar_time: float, solver: SolverConfig, mirror_sums: dict[str, np.ndarray] | None=None) -> TimeResult:
    solar = calculate_solar_state(month=month, solar_time=solar_time, latitude_deg=prepared.config.latitude_deg, altitude_km=prepared.config.altitude_km)
    orientation = calculate_orientation(prepared, solar.direction)
    reflection_error = maximum_reflection_error(prepared, orientation, solar.direction)
    if reflection_error >= 1e-08:
        raise RuntimeError(f'中心光线反射误差过大：{reflection_error:.3e}')
    if solver.calculate_shadow:
        shadow = calculate_shadow_blocking_efficiency(prepared, orientation, solar.direction, solver)
    else:
        shadow = np.ones(prepared.mirror_count, dtype=float)
    if solver.calculate_truncation:
        truncation = calculate_truncation_efficiency(prepared, orientation, solar.direction, solver)
    else:
        truncation = np.ones(prepared.mirror_count, dtype=float)
    cosine = orientation.cosine_efficiency
    atmospheric = prepared.atmospheric_efficiency
    optical = cosine * shadow * atmospheric * truncation * prepared.config.reflectivity
    for (name, values) in (('余弦效率', cosine), ('阴影遮挡效率', shadow), ('大气透射率', atmospheric), ('截断效率', truncation), ('光学效率', optical)):
        _check_efficiency(name, values)
    mirror_power_kw = solar.dni_kw_m2 * prepared.mirror_areas * optical
    if mirror_sums is not None:
        mirror_sums['optical_efficiency_sum'] += optical
        mirror_sums['cosine_efficiency_sum'] += cosine
        mirror_sums['shadow_blocking_efficiency_sum'] += shadow
        mirror_sums['atmospheric_efficiency_sum'] += atmospheric
        mirror_sums['truncation_efficiency_sum'] += truncation
        mirror_sums['output_power_kw_sum'] += mirror_power_kw
    field_power_kw = float(np.sum(mirror_power_kw))
    area_weights = prepared.mirror_areas
    return TimeResult(month=month, solar_time=solar_time, dni_kw_m2=solar.dni_kw_m2, average_optical_efficiency=float(np.average(optical, weights=area_weights)), average_cosine_efficiency=float(np.average(cosine, weights=area_weights)), average_shadow_blocking_efficiency=float(np.average(shadow, weights=area_weights)), average_atmospheric_efficiency=float(np.average(atmospheric, weights=area_weights)), average_truncation_efficiency=float(np.average(truncation, weights=area_weights)), field_output_mw=field_power_kw / 1000.0, unit_area_output_kw_m2=field_power_kw / prepared.total_mirror_area, maximum_reflection_error=reflection_error)

def solve_question1(prepared: PreparedField, solver: SolverConfig, months: Sequence[int]=tuple(range(1, 13)), solar_times: Sequence[float]=SOLAR_TIMES, progress: ProgressCallback | None=None) -> Question1Solution:
    if not months or not solar_times:
        raise ValueError('months 和 solar_times 不能为空。')
    if any((month < 1 or month > 12 for month in months)):
        raise ValueError('months 必须位于 1 到 12。')
    records: list[TimeResult] = []
    mirror_sums = {name: np.zeros(prepared.mirror_count, dtype=float) for name in ('optical_efficiency_sum', 'cosine_efficiency_sum', 'shadow_blocking_efficiency_sum', 'atmospheric_efficiency_sum', 'truncation_efficiency_sum', 'output_power_kw_sum')}
    total = len(months) * len(solar_times)
    for month in months:
        for solar_time in solar_times:
            record = evaluate_time(prepared, month, solar_time, solver, mirror_sums=mirror_sums)
            records.append(record)
            if progress is not None:
                progress(len(records), total, record)
    time_results = tuple(records)
    return Question1Solution(time_results=time_results, monthly_results=summarize_monthly(time_results), annual_result=summarize_annual(time_results), mirror_annual_results=summarize_mirror_annual(mirror_xy=prepared.centers[:, :2], tower_x=prepared.config.tower_x, tower_y=prepared.config.tower_y, state_count=len(time_results), **mirror_sums))

def run_validation_suite(prepared: PreparedField, base_solver: SolverConfig) -> tuple[ValidationResult, ...]:
    specifications = [('阴影网格', '10×10', replace(base_solver, shadow_grid_size=10, calculate_shadow=True, calculate_truncation=False), 'average_shadow_blocking_efficiency', '年平均阴影遮挡效率', False), ('阴影网格', '15×15', replace(base_solver, shadow_grid_size=15, calculate_shadow=True, calculate_truncation=False), 'average_shadow_blocking_efficiency', '年平均阴影遮挡效率', True), ('阴影网格', '20×20', replace(base_solver, shadow_grid_size=20, calculate_shadow=True, calculate_truncation=False), 'average_shadow_blocking_efficiency', '年平均阴影遮挡效率', False), ('邻镜半径', '40 m', replace(base_solver, neighbor_radius_m=40.0, calculate_shadow=True, calculate_truncation=False), 'average_shadow_blocking_efficiency', '年平均阴影遮挡效率', False), ('邻镜半径', '60 m', replace(base_solver, neighbor_radius_m=60.0, calculate_shadow=True, calculate_truncation=False), 'average_shadow_blocking_efficiency', '年平均阴影遮挡效率', True), ('邻镜半径', '80 m', replace(base_solver, neighbor_radius_m=80.0, calculate_shadow=True, calculate_truncation=False), 'average_shadow_blocking_efficiency', '年平均阴影遮挡效率', False), ('截断光线', '128', replace(base_solver, truncation_rays=128, calculate_shadow=False, calculate_truncation=True), 'average_truncation_efficiency', '年平均截断效率', False), ('截断光线', '256', replace(base_solver, truncation_rays=256, calculate_shadow=False, calculate_truncation=True), 'average_truncation_efficiency', '年平均截断效率', True), ('截断光线', '512', replace(base_solver, truncation_rays=512, calculate_shadow=False, calculate_truncation=True), 'average_truncation_efficiency', '年平均截断效率', False)]
    raw: list[tuple[str, str, str, float, float, bool]] = []
    for (category, parameter, solver, field, metric, reference) in specifications:
        started = time.perf_counter()
        solution = solve_question1(prepared, solver)
        elapsed = time.perf_counter() - started
        value = float(getattr(solution.annual_result, field))
        raw.append((category, parameter, metric, value, elapsed, reference))
    baselines = {category: value for (category, _, _, value, _, reference) in raw if reference}
    return tuple((ValidationResult(category=category, parameter=parameter, metric=metric, value=value, relative_difference_percent=abs(value - baselines[category]) / abs(baselines[category]) * 100.0, runtime_seconds=elapsed) for (category, parameter, metric, value, elapsed, _) in raw))
