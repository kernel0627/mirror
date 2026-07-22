"""第二问：双布局生成、搜索、统一复算、修剪和验收。"""
from __future__ import annotations
# ruff: noqa
import sys
from Public import *

# ---- layout.py ----

import math
from dataclasses import dataclass
from typing import Protocol
import numpy as np
from numpy.typing import NDArray
from scipy.spatial import cKDTree
FloatArray = NDArray[np.float64]

class LayoutError(ValueError):
    pass

class CommonParameters(Protocol):
    tower_x: float
    tower_y: float
    mirror_width: float
    mirror_height: float
    installation_height: float
    field_radius: float
    exclusion_radius: float
    safety_epsilon: float

@dataclass(frozen=True)
class PartitionedRingParameters:
    tower_y: float
    mirror_width: float
    mirror_height: float
    installation_height: float
    split_radius: float
    near_spacing: float
    far_spacing: float
    tower_x: float = 0.0
    field_radius: float = 350.0
    exclusion_radius: float = 100.0
    safety_epsilon: float = 0.01

    @property
    def safe_distance(self) -> float:
        return self.mirror_width + 5.0 + self.safety_epsilon

@dataclass(frozen=True)
class CampoParameters:
    tower_y: float
    mirror_width: float
    mirror_height: float
    installation_height: float
    first_ring_count: int
    initial_spacing: float
    spacing_growth: float
    tower_x: float = 0.0
    field_radius: float = 350.0
    exclusion_radius: float = 100.0
    safety_epsilon: float = 0.01

    @property
    def safe_distance(self) -> float:
        return self.mirror_width + 5.0 + self.safety_epsilon

@dataclass(frozen=True)
class LayoutRing:
    index: int
    radius: float
    zone: int
    nominal_count: int
    coordinates: FloatArray

    @property
    def mirror_count(self) -> int:
        return int(self.coordinates.shape[0])

@dataclass(frozen=True)
class GeneratedLayout:
    kind: str
    rings: tuple[LayoutRing, ...]

    @property
    def mirror_count(self) -> int:
        return sum((ring.mirror_count for ring in self.rings))

    @property
    def coordinates(self) -> FloatArray:
        if not self.rings:
            return np.empty((0, 2), dtype=float)
        return np.concatenate([ring.coordinates for ring in self.rings], axis=0)

    def prefix(self, ring_count: int) -> FloatArray:
        if ring_count < 1 or ring_count > len(self.rings):
            raise ValueError(f'ring_count 应位于 1 到 {len(self.rings)}，实际为 {ring_count}。')
        return np.concatenate([ring.coordinates for ring in self.rings[:ring_count]], axis=0)

    def prefix_mirror_count(self, ring_count: int) -> int:
        if ring_count < 1 or ring_count > len(self.rings):
            raise ValueError(f'ring_count 应位于 1 到 {len(self.rings)}，实际为 {ring_count}。')
        return sum((ring.mirror_count for ring in self.rings[:ring_count]))

@dataclass(frozen=True)
class GeometryCheck:
    valid: bool
    reason: str | None
    mirror_count: int
    minimum_center_distance: float
    maximum_field_radius: float
    minimum_tower_distance: float

def _validate_common_parameters(parameters: CommonParameters) -> None:
    values = (parameters.tower_x, parameters.tower_y, parameters.mirror_width, parameters.mirror_height, parameters.installation_height, parameters.field_radius, parameters.exclusion_radius, parameters.safety_epsilon)
    if not all((math.isfinite(value) for value in values)):
        raise LayoutError('布局参数必须全部为有限数。')
    if not 2.0 <= parameters.mirror_height <= parameters.mirror_width <= 8.0:
        raise LayoutError('镜面尺寸必须满足 2 ≤ h ≤ w ≤ 8。')
    if not 2.0 <= parameters.installation_height <= 6.0:
        raise LayoutError('安装高度必须位于 2 m 到 6 m。')
    if parameters.installation_height < parameters.mirror_height / 2.0:
        raise LayoutError('安装高度不足，镜面转动时可能触地。')
    if parameters.field_radius <= 0.0:
        raise LayoutError('场地半径必须大于 0。')
    if parameters.exclusion_radius <= 0.0:
        raise LayoutError('塔周禁区半径必须大于 0。')
    if parameters.safety_epsilon <= 0.0:
        raise LayoutError('安全距离余量必须大于 0。')

def _maximum_tower_centered_radius(parameters: CommonParameters) -> float:
    return parameters.field_radius + math.hypot(parameters.tower_x, parameters.tower_y)

def _ring_coordinates(parameters: CommonParameters, radius: float, mirror_count: int, phase: float) -> FloatArray:
    if mirror_count < 2:
        raise LayoutError('单圈镜子数必须大于等于 2。')
    angles = 2.0 * math.pi * np.arange(mirror_count, dtype=float) / mirror_count + phase
    coordinates = np.column_stack((parameters.tower_x + radius * np.sin(angles), parameters.tower_y + radius * np.cos(angles)))
    field_radius = np.hypot(coordinates[:, 0], coordinates[:, 1])
    keep = field_radius <= parameters.field_radius + 1e-09
    clipped = np.asarray(coordinates[keep], dtype=float)
    clipped.setflags(write=False)
    return clipped

def _within_ring_count(radius: float, safe_distance: float) -> int:
    ratio = safe_distance / (2.0 * radius)
    if ratio >= 1.0:
        raise LayoutError('圆环半径过小，无法放置满足安全距离的镜子。')
    return int(math.floor(math.pi / math.asin(ratio)))

def generate_partitioned_layout(parameters: PartitionedRingParameters, *, maximum_rings: int=256) -> GeneratedLayout:
    _validate_common_parameters(parameters)
    if not math.isfinite(parameters.split_radius):
        raise LayoutError('分区半径必须为有限数。')
    if parameters.split_radius <= parameters.exclusion_radius:
        raise LayoutError('分区半径必须位于塔周禁区之外。')
    if parameters.near_spacing <= 0.0:
        raise LayoutError('近区行距必须大于 0。')
    if parameters.far_spacing < parameters.near_spacing:
        raise LayoutError('远区行距必须大于等于近区行距。')
    rings: list[LayoutRing] = []
    radius = parameters.exclusion_radius
    maximum_radius = _maximum_tower_centered_radius(parameters)
    for ring_index in range(maximum_rings):
        if radius > maximum_radius + 1e-09:
            break
        count = _within_ring_count(radius, parameters.safe_distance)
        phase = 0.0 if ring_index % 2 == 0 else math.pi / count
        coordinates = _ring_coordinates(parameters, radius, count, phase)
        if coordinates.size:
            rings.append(LayoutRing(index=ring_index, radius=radius, zone=1 if radius < parameters.split_radius else 2, nominal_count=count, coordinates=coordinates))
        spacing = parameters.near_spacing if radius < parameters.split_radius else parameters.far_spacing
        radius += spacing
    else:
        raise LayoutError('达到 maximum_rings，圆环生成未正常终止。')
    layout = GeneratedLayout('partitioned', tuple(rings))
    check = validate_layout(layout.coordinates, parameters)
    if not check.valid:
        raise LayoutError(check.reason or '分区圆环布局不满足几何约束。')
    return layout

def _campo_zone(radius: float, first_radius: float) -> tuple[int, int]:
    if radius < 2.0 * first_radius:
        return (1, 1)
    if radius < 4.0 * first_radius:
        return (2, 2)
    return (3, 4)

def generate_campo_layout(parameters: CampoParameters, *, maximum_rings: int=256) -> GeneratedLayout:
    _validate_common_parameters(parameters)
    if parameters.first_ring_count < 2:
        raise LayoutError('Campo 首环镜子数必须大于等于 2。')
    if parameters.initial_spacing <= 0.0:
        raise LayoutError('Campo 初始行距必须大于 0。')
    if parameters.spacing_growth < 0.0:
        raise LayoutError('Campo 行距增长量不能小于 0。')
    first_radius = max(parameters.exclusion_radius, parameters.safe_distance / (2.0 * math.sin(math.pi / parameters.first_ring_count)))
    maximum_radius = _maximum_tower_centered_radius(parameters)
    rings: list[LayoutRing] = []
    zone_rows = {1: 0, 2: 0, 3: 0}
    radius = first_radius
    for ring_index in range(maximum_rings):
        if radius > maximum_radius + 1e-09:
            break
        (zone, multiplier) = _campo_zone(radius, first_radius)
        count = parameters.first_ring_count * multiplier
        zone_index = zone_rows[zone]
        phase = 0.0 if zone_index % 2 == 0 else math.pi / count
        coordinates = _ring_coordinates(parameters, radius, count, phase)
        if coordinates.size:
            rings.append(LayoutRing(index=ring_index, radius=radius, zone=zone, nominal_count=count, coordinates=coordinates))
        zone_rows[zone] += 1
        radius += parameters.initial_spacing + parameters.spacing_growth * ring_index
    else:
        raise LayoutError('达到 maximum_rings，Campo 圆环生成未正常终止。')
    layout = GeneratedLayout('campo', tuple(rings))
    check = validate_layout(layout.coordinates, parameters)
    if not check.valid:
        raise LayoutError(check.reason or 'Campo 布局不满足几何约束。')
    return layout

def validate_layout(coordinates: FloatArray, parameters: CommonParameters) -> GeometryCheck:
    try:
        _validate_common_parameters(parameters)
    except LayoutError as exc:
        return GeometryCheck(False, str(exc), 0, math.inf, math.inf, math.inf)
    xy = np.asarray(coordinates, dtype=float)
    if xy.ndim != 2 or xy.shape[1] != 2:
        return GeometryCheck(False, f'镜位坐标应为 N×2，实际形状为 {xy.shape}。', 0, math.inf, math.inf, math.inf)
    if xy.shape[0] == 0:
        return GeometryCheck(False, '镜场中没有定日镜。', 0, math.inf, 0.0, 0.0)
    if not np.all(np.isfinite(xy)):
        return GeometryCheck(False, '镜位坐标包含 NaN 或无穷值。', int(xy.shape[0]), math.inf, math.inf, math.inf)
    field_radii = np.hypot(xy[:, 0], xy[:, 1])
    tower_distances = np.hypot(xy[:, 0] - parameters.tower_x, xy[:, 1] - parameters.tower_y)
    maximum_field_radius = float(np.max(field_radii))
    minimum_tower_distance = float(np.min(tower_distances))
    if maximum_field_radius > parameters.field_radius + 1e-09:
        return GeometryCheck(False, '存在越过 350 m 场地边界的镜位。', int(xy.shape[0]), math.inf, maximum_field_radius, minimum_tower_distance)
    if minimum_tower_distance < parameters.exclusion_radius - 1e-09:
        return GeometryCheck(False, '存在进入塔周 100 m 禁区的镜位。', int(xy.shape[0]), math.inf, maximum_field_radius, minimum_tower_distance)
    if xy.shape[0] == 1:
        minimum_distance = math.inf
    else:
        (distances, _) = cKDTree(xy).query(xy, k=2)
        minimum_distance = float(np.min(distances[:, 1]))
    if minimum_distance <= parameters.mirror_width + 5.0:
        return GeometryCheck(False, f'最小中心距离不满足严格约束：{minimum_distance:.9f} m ≤ {parameters.mirror_width + 5.0:.9f} m。', int(xy.shape[0]), minimum_distance, maximum_field_radius, minimum_tower_distance)
    return GeometryCheck(True, None, int(xy.shape[0]), minimum_distance, maximum_field_radius, minimum_tower_distance)

# ---- evaluate.py ----

import hashlib
from dataclasses import dataclass, replace
from typing import Sequence
import numpy as np
LayoutParameters = PartitionedRingParameters | CampoParameters

@dataclass(frozen=True)
class EvaluationProfile:
    name: str
    solver: SolverConfig
    months: tuple[int, ...] = tuple(range(1, 13))
    solar_times: tuple[float, ...] = SOLAR_TIMES

@dataclass(frozen=True)
class FieldEvaluation:
    layout_kind: str
    ring_count: int
    mirror_count: int
    mirror_area_m2: float
    total_area_m2: float
    coordinates: np.ndarray
    solution: Question1Solution

    @property
    def annual_power_mw(self) -> float:
        return self.solution.annual_result.field_output_mw

    @property
    def unit_area_power_kw_m2(self) -> float:
        return self.solution.annual_result.unit_area_output_kw_m2

    def is_feasible(self, target_power_mw: float=42.0) -> bool:
        return self.annual_power_mw >= target_power_mw

@dataclass(frozen=True)
class ExtentScanResult:
    best: FieldEvaluation
    evaluations: tuple[FieldEvaluation, ...]
    first_feasible_ring_count: int | None

class EvaluationCache:

    def __init__(self) -> None:
        self._values: dict[str, Question1Solution] = {}

    def get(self, key: str) -> Question1Solution | None:
        return self._values.get(key)

    def put(self, key: str, value: Question1Solution) -> None:
        self._values[key] = value

    def __len__(self) -> int:
        return len(self._values)

def exploration_profile() -> EvaluationProfile:
    return EvaluationProfile(name='exploration', solver=SolverConfig(shadow_grid_size=5, truncation_rays=64, neighbor_radius_m=60.0, truncation_chunk_size=128, sobol_seed=2023))

def refinement_profile() -> EvaluationProfile:
    return EvaluationProfile(name='refinement', solver=SolverConfig(shadow_grid_size=10, truncation_rays=128, neighbor_radius_m=60.0, truncation_chunk_size=128, sobol_seed=2023))

def final_profile() -> EvaluationProfile:
    return EvaluationProfile(name='final', solver=SolverConfig(shadow_grid_size=15, truncation_rays=256, neighbor_radius_m=60.0, truncation_chunk_size=128, sobol_seed=2023))

def _field_config(parameters: CommonParameters, base: FieldConfig | None=None) -> FieldConfig:
    config = base or FieldConfig()
    return replace(config, field_radius=parameters.field_radius, exclusion_radius=parameters.exclusion_radius, tower_x=parameters.tower_x, tower_y=parameters.tower_y, mirror_width=parameters.mirror_width, mirror_height=parameters.mirror_height, mirror_center_z=parameters.installation_height)

def _cache_key(coordinates: np.ndarray, config: FieldConfig, profile: EvaluationProfile) -> str:
    digest = hashlib.sha256()
    rounded = np.round(np.asarray(coordinates, dtype='<f8'), decimals=9)
    digest.update(rounded.tobytes(order='C'))
    digest.update(repr(config.to_dict()).encode('utf-8'))
    digest.update(repr(profile.solver.to_dict()).encode('utf-8'))
    digest.update(repr(profile.months).encode('ascii'))
    digest.update(repr(profile.solar_times).encode('ascii'))
    return digest.hexdigest()

def evaluate_coordinates(*, layout_kind: str, ring_count: int, coordinates: np.ndarray, parameters: LayoutParameters, profile: EvaluationProfile, cache: EvaluationCache | None=None, base_field_config: FieldConfig | None=None) -> FieldEvaluation:
    xy = np.asarray(coordinates, dtype=float)
    if xy.ndim != 2 or xy.shape[1] != 2 or xy.shape[0] == 0:
        raise ValueError('候选镜场坐标必须为非空 N×2 数组。')
    config = _field_config(parameters, base_field_config)
    key = _cache_key(xy, config, profile)
    solution = cache.get(key) if cache is not None else None
    if solution is None:
        prepared = prepare_field(xy, config)
        solution = solve_question1(prepared=prepared, solver=profile.solver, months=profile.months, solar_times=profile.solar_times)
        if cache is not None:
            cache.put(key, solution)
    mirror_area = parameters.mirror_width * parameters.mirror_height
    return FieldEvaluation(layout_kind=layout_kind, ring_count=ring_count, mirror_count=int(xy.shape[0]), mirror_area_m2=mirror_area, total_area_m2=float(xy.shape[0] * mirror_area), coordinates=xy, solution=solution)

def better_evaluation(left: FieldEvaluation, right: FieldEvaluation, *, target_power_mw: float=42.0) -> FieldEvaluation:
    left_feasible = left.is_feasible(target_power_mw)
    right_feasible = right.is_feasible(target_power_mw)
    if left_feasible != right_feasible:
        return left if left_feasible else right
    if left_feasible:
        if left.unit_area_power_kw_m2 != right.unit_area_power_kw_m2:
            return left if left.unit_area_power_kw_m2 > right.unit_area_power_kw_m2 else right
        return left if left.annual_power_mw <= right.annual_power_mw else right
    return left if left.annual_power_mw >= right.annual_power_mw else right

def _unique_ring_counts(values: Sequence[int], total: int) -> tuple[int, ...]:
    return tuple(sorted({value for value in values if 1 <= value <= total}))

def scan_layout_extents(layout: GeneratedLayout, parameters: LayoutParameters, profile: EvaluationProfile, *, target_power_mw: float=42.0, coarse_stride: int=4, window: int=2, cache: EvaluationCache | None=None, base_field_config: FieldConfig | None=None) -> ExtentScanResult:
    if not layout.rings:
        raise ValueError('布局中没有可用于评价的圆环。')
    if coarse_stride < 1:
        raise ValueError('coarse_stride 必须大于等于 1。')
    if window < 0:
        raise ValueError('window 不能小于 0。')
    total_rings = len(layout.rings)
    coarse_counts = list(range(coarse_stride, total_rings + 1, coarse_stride))
    if not coarse_counts or coarse_counts[-1] != total_rings:
        coarse_counts.append(total_rings)
    evaluated: dict[int, FieldEvaluation] = {}

    def evaluate(ring_count: int) -> FieldEvaluation:
        previous = evaluated.get(ring_count)
        if previous is not None:
            return previous
        value = evaluate_coordinates(layout_kind=layout.kind, ring_count=ring_count, coordinates=layout.prefix(ring_count), parameters=parameters, profile=profile, cache=cache, base_field_config=base_field_config)
        evaluated[ring_count] = value
        return value
    first_feasible: int | None = None
    previous_coarse = 0
    for ring_count in coarse_counts:
        value = evaluate(ring_count)
        if value.is_feasible(target_power_mw):
            for refined_count in range(previous_coarse + 1, ring_count + 1):
                refined = evaluate(refined_count)
                if refined.is_feasible(target_power_mw):
                    first_feasible = refined_count
                    break
            break
        previous_coarse = ring_count
    center = first_feasible if first_feasible is not None else total_rings
    local_counts = _unique_ring_counts(range(center - window, center + window + 1), total_rings)
    for ring_count in local_counts:
        evaluate(ring_count)
    values = tuple((evaluated[key] for key in sorted(evaluated)))
    best = values[0]
    for value in values[1:]:
        best = better_evaluation(best, value, target_power_mw=target_power_mw)
    return ExtentScanResult(best=best, evaluations=values, first_feasible_ring_count=first_feasible)

# ---- search.py ----

import math
from dataclasses import dataclass, replace
from typing import Callable, Generic, Iterable, TypeVar
from scipy.stats import qmc
ParametersT = TypeVar('ParametersT', PartitionedRingParameters, CampoParameters)

@dataclass(frozen=True)
class Interval:
    lower: float
    upper: float

    def __post_init__(self) -> None:
        if not math.isfinite(self.lower) or not math.isfinite(self.upper):
            raise ValueError('参数边界必须为有限数。')
        if self.lower > self.upper:
            raise ValueError('参数下界不能大于上界。')

    def contains(self, value: float) -> bool:
        return self.lower <= value <= self.upper

    def map_unit(self, value: float) -> float:
        return self.lower + value * (self.upper - self.lower)

@dataclass(frozen=True)
class IntegerInterval:
    lower: int
    upper: int

    def __post_init__(self) -> None:
        if self.lower > self.upper:
            raise ValueError('整数参数下界不能大于上界。')

    def contains(self, value: int) -> bool:
        return self.lower <= value <= self.upper

    def map_unit(self, value: float) -> int:
        mapped = self.lower + value * (self.upper - self.lower)
        return min(self.upper, max(self.lower, int(round(mapped))))

@dataclass(frozen=True)
class PartitionedBounds:
    tower_y: Interval = Interval(-220.0, 0.0)
    mirror_width: Interval = Interval(4.5, 7.5)
    mirror_height: Interval = Interval(4.0, 7.5)
    installation_height: Interval = Interval(2.0, 6.0)
    split_radius: Interval = Interval(150.0, 300.0)
    near_spacing: Interval = Interval(8.0, 22.0)
    far_spacing: Interval = Interval(10.0, 32.0)

@dataclass(frozen=True)
class CampoBounds:
    tower_y: Interval = Interval(-220.0, 0.0)
    mirror_width: Interval = Interval(4.5, 7.5)
    mirror_height: Interval = Interval(4.0, 7.5)
    installation_height: Interval = Interval(2.0, 6.0)
    first_ring_count: IntegerInterval = IntegerInterval(44, 76)
    initial_spacing: Interval = Interval(8.0, 22.0)
    spacing_growth: Interval = Interval(0.0, 0.6)

@dataclass(frozen=True)
class StepLevel:
    steps: tuple[tuple[str, float], ...]

    def as_dict(self) -> dict[str, float]:
        return dict(self.steps)
PARTITIONED_STEP_LEVELS = (StepLevel((('tower_y', 20.0), ('mirror_width', 0.5), ('mirror_height', 0.5), ('installation_height', 0.5), ('split_radius', 20.0), ('near_spacing', 2.0), ('far_spacing', 2.0))), StepLevel((('tower_y', 10.0), ('mirror_width', 0.2), ('mirror_height', 0.2), ('installation_height', 0.2), ('split_radius', 10.0), ('near_spacing', 1.0), ('far_spacing', 1.0))), StepLevel((('tower_y', 5.0), ('mirror_width', 0.1), ('mirror_height', 0.1), ('installation_height', 0.1), ('split_radius', 5.0), ('near_spacing', 0.5), ('far_spacing', 0.5))))
CAMPO_STEP_LEVELS = (StepLevel((('tower_y', 20.0), ('mirror_width', 0.5), ('mirror_height', 0.5), ('installation_height', 0.5), ('first_ring_count', 4.0), ('initial_spacing', 2.0), ('spacing_growth', 0.1))), StepLevel((('tower_y', 10.0), ('mirror_width', 0.2), ('mirror_height', 0.2), ('installation_height', 0.2), ('first_ring_count', 2.0), ('initial_spacing', 1.0), ('spacing_growth', 0.05))), StepLevel((('tower_y', 5.0), ('mirror_width', 0.1), ('mirror_height', 0.1), ('installation_height', 0.1), ('first_ring_count', 1.0), ('initial_spacing', 0.5), ('spacing_growth', 0.02))))

@dataclass(frozen=True)
class SearchOutcome(Generic[ParametersT]):
    parameters: ParametersT
    extent_scan: ExtentScanResult

    @property
    def feasible(self) -> bool:
        return self.extent_scan.best.is_feasible()

    @property
    def annual_power_mw(self) -> float:
        return self.extent_scan.best.annual_power_mw

    @property
    def unit_area_power_kw_m2(self) -> float:
        return self.extent_scan.best.unit_area_power_kw_m2

@dataclass(frozen=True)
class SearchTrace(Generic[ParametersT]):
    step_level: int
    cycle: int
    block: str
    outcome: SearchOutcome[ParametersT]

@dataclass(frozen=True)
class OptimizationResult(Generic[ParametersT]):
    layout_kind: str
    best: SearchOutcome[ParametersT]
    starts: tuple[SearchOutcome[ParametersT], ...]
    local_results: tuple[SearchOutcome[ParametersT], ...]
    trace: tuple[SearchTrace[ParametersT], ...]

def _sobol_points(dimension: int, count: int, seed: int) -> list[list[float]]:
    if count < 1:
        raise ValueError('Sobol 样本数必须大于等于 1。')
    exponent = int(math.ceil(math.log2(count)))
    sampler = qmc.Sobol(d=dimension, scramble=True, seed=seed)
    return sampler.random_base2(exponent)[:count].tolist()

def sample_partitioned_parameters(count: int, *, seed: int=2023, bounds: PartitionedBounds=PartitionedBounds()) -> tuple[PartitionedRingParameters, ...]:
    samples: list[PartitionedRingParameters] = []
    for point in _sobol_points(7, count, seed):
        width = bounds.mirror_width.map_unit(point[1])
        height_upper = min(bounds.mirror_height.upper, width)
        height_lower = min(bounds.mirror_height.lower, height_upper)
        height = height_lower + point[2] * (height_upper - height_lower)
        installation_lower = max(bounds.installation_height.lower, height / 2.0)
        installation = installation_lower + point[3] * (bounds.installation_height.upper - installation_lower)
        near = bounds.near_spacing.map_unit(point[5])
        far_lower = max(bounds.far_spacing.lower, near)
        far = far_lower + point[6] * (bounds.far_spacing.upper - far_lower)
        samples.append(PartitionedRingParameters(tower_y=bounds.tower_y.map_unit(point[0]), mirror_width=width, mirror_height=height, installation_height=installation, split_radius=bounds.split_radius.map_unit(point[4]), near_spacing=near, far_spacing=far))
    return tuple(samples)

def sample_campo_parameters(count: int, *, seed: int=2023, bounds: CampoBounds=CampoBounds()) -> tuple[CampoParameters, ...]:
    samples: list[CampoParameters] = []
    for point in _sobol_points(7, count, seed):
        width = bounds.mirror_width.map_unit(point[1])
        height_upper = min(bounds.mirror_height.upper, width)
        height_lower = min(bounds.mirror_height.lower, height_upper)
        height = height_lower + point[2] * (height_upper - height_lower)
        installation_lower = max(bounds.installation_height.lower, height / 2.0)
        installation = installation_lower + point[3] * (bounds.installation_height.upper - installation_lower)
        samples.append(CampoParameters(tower_y=bounds.tower_y.map_unit(point[0]), mirror_width=width, mirror_height=height, installation_height=installation, first_ring_count=bounds.first_ring_count.map_unit(point[4]), initial_spacing=bounds.initial_spacing.map_unit(point[5]), spacing_growth=bounds.spacing_growth.map_unit(point[6])))
    return tuple(samples)

def _partitioned_within_bounds(parameters: PartitionedRingParameters, bounds: PartitionedBounds) -> bool:
    return bounds.tower_y.contains(parameters.tower_y) and bounds.mirror_width.contains(parameters.mirror_width) and bounds.mirror_height.contains(parameters.mirror_height) and (parameters.mirror_height <= parameters.mirror_width) and bounds.installation_height.contains(parameters.installation_height) and (parameters.installation_height >= parameters.mirror_height / 2.0) and bounds.split_radius.contains(parameters.split_radius) and bounds.near_spacing.contains(parameters.near_spacing) and bounds.far_spacing.contains(parameters.far_spacing) and (parameters.far_spacing >= parameters.near_spacing)

def _campo_within_bounds(parameters: CampoParameters, bounds: CampoBounds) -> bool:
    return bounds.tower_y.contains(parameters.tower_y) and bounds.mirror_width.contains(parameters.mirror_width) and bounds.mirror_height.contains(parameters.mirror_height) and (parameters.mirror_height <= parameters.mirror_width) and bounds.installation_height.contains(parameters.installation_height) and (parameters.installation_height >= parameters.mirror_height / 2.0) and bounds.first_ring_count.contains(parameters.first_ring_count) and bounds.initial_spacing.contains(parameters.initial_spacing) and bounds.spacing_growth.contains(parameters.spacing_growth)

def _better_outcome(left: SearchOutcome[ParametersT], right: SearchOutcome[ParametersT]) -> SearchOutcome[ParametersT]:
    best_evaluation = better_evaluation(left.extent_scan.best, right.extent_scan.best)
    return left if best_evaluation is left.extent_scan.best else right

def _rank_outcomes(outcomes: Iterable[SearchOutcome[ParametersT]]) -> list[SearchOutcome[ParametersT]]:
    ranked: list[SearchOutcome[ParametersT]] = []
    for outcome in outcomes:
        inserted = False
        for (index, current) in enumerate(ranked):
            if _better_outcome(outcome, current) is outcome:
                ranked.insert(index, outcome)
                inserted = True
                break
        if not inserted:
            ranked.append(outcome)
    return ranked

def evaluate_partitioned_parameters(parameters: PartitionedRingParameters, profile: EvaluationProfile, *, cache: EvaluationCache | None=None, coarse_stride: int=4, window: int=2) -> SearchOutcome[PartitionedRingParameters]:
    layout = generate_partitioned_layout(parameters)
    scan = scan_layout_extents(layout, parameters, profile, coarse_stride=coarse_stride, window=window, cache=cache)
    return SearchOutcome(parameters, scan)

def evaluate_campo_parameters(parameters: CampoParameters, profile: EvaluationProfile, *, cache: EvaluationCache | None=None, coarse_stride: int=4, window: int=2) -> SearchOutcome[CampoParameters]:
    layout = generate_campo_layout(parameters)
    scan = scan_layout_extents(layout, parameters, profile, coarse_stride=coarse_stride, window=window, cache=cache)
    return SearchOutcome(parameters, scan)

def _coordinate_candidates(parameters: ParametersT, fields: tuple[str, ...], steps: dict[str, float]) -> tuple[ParametersT, ...]:
    candidates: list[ParametersT] = []
    for field in fields:
        step = steps[field]
        value = getattr(parameters, field)
        for direction in (-1.0, 1.0):
            new_value: float | int = value + direction * step
            if isinstance(value, int):
                new_value = int(round(new_value))
            candidates.append(replace(parameters, **{field: new_value}))
    return tuple(candidates)

def _cyclic_search(initial: SearchOutcome[ParametersT], *, evaluator: Callable[[ParametersT], SearchOutcome[ParametersT]], within_bounds: Callable[[ParametersT], bool], blocks: tuple[tuple[str, tuple[str, ...]], ...], step_levels: tuple[StepLevel, ...], maximum_cycles_per_level: int) -> tuple[SearchOutcome[ParametersT], tuple[SearchTrace[ParametersT], ...]]:
    current = initial
    trace: list[SearchTrace[ParametersT]] = []
    for (level_index, level) in enumerate(step_levels):
        steps = level.as_dict()
        for cycle in range(maximum_cycles_per_level):
            cycle_improved = False
            for (block_name, fields) in blocks:
                block_best = current
                for candidate in _coordinate_candidates(current.parameters, fields, steps):
                    if not within_bounds(candidate):
                        continue
                    try:
                        outcome = evaluator(candidate)
                    except LayoutError:
                        continue
                    block_best = _better_outcome(block_best, outcome)
                if block_best is not current:
                    current = block_best
                    cycle_improved = True
                    trace.append(SearchTrace(step_level=level_index, cycle=cycle, block=block_name, outcome=current))
            if not cycle_improved:
                break
    return (current, tuple(trace))

def optimize_partitioned(*, profile: EvaluationProfile, initial_sample_count: int=16, retained_starts: int=3, seed: int=2023, bounds: PartitionedBounds=PartitionedBounds(), step_levels: tuple[StepLevel, ...]=PARTITIONED_STEP_LEVELS, maximum_cycles_per_level: int=4, coarse_stride: int=4, window: int=2, cache: EvaluationCache | None=None, progress: Callable[[str], None] | None=None) -> OptimizationResult[PartitionedRingParameters]:
    cache = cache or EvaluationCache()
    outcomes: list[SearchOutcome[PartitionedRingParameters]] = []
    samples = sample_partitioned_parameters(initial_sample_count, seed=seed, bounds=bounds)
    for (index, parameters) in enumerate(samples, start=1):
        try:
            outcome = evaluate_partitioned_parameters(parameters, profile, cache=cache, coarse_stride=coarse_stride, window=window)
            outcomes.append(outcome)
            if progress is not None:
                progress(f'方案 A 初值 {index}/{len(samples)}：P={outcome.annual_power_mw:.4f} MW，q={outcome.unit_area_power_kw_m2:.6f}')
        except LayoutError as exc:
            if progress is not None:
                progress(f'方案 A 初值 {index}/{len(samples)}：几何不可行（{exc}）')
            continue
    if not outcomes:
        raise RuntimeError('分区圆环的分散初值中没有几何合法方案。')
    starts = tuple(_rank_outcomes(outcomes)[:retained_starts])
    local_results: list[SearchOutcome[PartitionedRingParameters]] = []
    trace: list[SearchTrace[PartitionedRingParameters]] = []

    def evaluator(parameters: PartitionedRingParameters) -> SearchOutcome[PartitionedRingParameters]:
        return evaluate_partitioned_parameters(parameters, profile, cache=cache, coarse_stride=coarse_stride, window=window)
    blocks = (('tower', ('tower_y',)), ('mirror', ('mirror_width', 'mirror_height', 'installation_height')), ('layout', ('split_radius', 'near_spacing', 'far_spacing')))
    for (index, start) in enumerate(starts, start=1):
        if progress is not None:
            progress(f'方案 A 开始局部搜索 {index}/{len(starts)}')
        (local, local_trace) = _cyclic_search(start, evaluator=evaluator, within_bounds=lambda value: _partitioned_within_bounds(value, bounds), blocks=blocks, step_levels=step_levels, maximum_cycles_per_level=maximum_cycles_per_level)
        local_results.append(local)
        trace.extend(local_trace)
    best = _rank_outcomes(local_results)[0]
    return OptimizationResult('partitioned', best, starts, tuple(local_results), tuple(trace))

def refine_partitioned(initial_parameters: PartitionedRingParameters, *, profile: EvaluationProfile, bounds: PartitionedBounds=PartitionedBounds(), step_levels: tuple[StepLevel, ...]=PARTITIONED_STEP_LEVELS, maximum_cycles_per_level: int=4, coarse_stride: int=4, window: int=2, cache: EvaluationCache | None=None, progress: Callable[[str], None] | None=None) -> OptimizationResult[PartitionedRingParameters]:
    if not _partitioned_within_bounds(initial_parameters, bounds):
        raise ValueError('方案 A 的恢复参数超出当前搜索边界。')
    cache = cache or EvaluationCache()

    def evaluator(parameters: PartitionedRingParameters) -> SearchOutcome[PartitionedRingParameters]:
        return evaluate_partitioned_parameters(parameters, profile, cache=cache, coarse_stride=coarse_stride, window=window)
    initial = evaluator(initial_parameters)
    if progress is not None:
        progress(f'方案 A 恢复起点：P={initial.annual_power_mw:.4f} MW，q={initial.unit_area_power_kw_m2:.6f}')
    blocks = (('tower', ('tower_y',)), ('mirror', ('mirror_width', 'mirror_height', 'installation_height')), ('layout', ('split_radius', 'near_spacing', 'far_spacing')))
    (best, trace) = _cyclic_search(initial, evaluator=evaluator, within_bounds=lambda value: _partitioned_within_bounds(value, bounds), blocks=blocks, step_levels=step_levels, maximum_cycles_per_level=maximum_cycles_per_level)
    return OptimizationResult('partitioned', best, (initial,), (best,), trace)

def optimize_campo(*, profile: EvaluationProfile, initial_sample_count: int=16, retained_starts: int=3, seed: int=2023, bounds: CampoBounds=CampoBounds(), step_levels: tuple[StepLevel, ...]=CAMPO_STEP_LEVELS, maximum_cycles_per_level: int=4, coarse_stride: int=4, window: int=2, cache: EvaluationCache | None=None, progress: Callable[[str], None] | None=None) -> OptimizationResult[CampoParameters]:
    cache = cache or EvaluationCache()
    outcomes: list[SearchOutcome[CampoParameters]] = []
    samples = sample_campo_parameters(initial_sample_count, seed=seed, bounds=bounds)
    for (index, parameters) in enumerate(samples, start=1):
        try:
            outcome = evaluate_campo_parameters(parameters, profile, cache=cache, coarse_stride=coarse_stride, window=window)
            outcomes.append(outcome)
            if progress is not None:
                progress(f'方案 B 初值 {index}/{len(samples)}：P={outcome.annual_power_mw:.4f} MW，q={outcome.unit_area_power_kw_m2:.6f}')
        except LayoutError as exc:
            if progress is not None:
                progress(f'方案 B 初值 {index}/{len(samples)}：几何不可行（{exc}）')
            continue
    if not outcomes:
        raise RuntimeError('Campo 的分散初值中没有几何合法方案。')
    starts = tuple(_rank_outcomes(outcomes)[:retained_starts])
    local_results: list[SearchOutcome[CampoParameters]] = []
    trace: list[SearchTrace[CampoParameters]] = []

    def evaluator(parameters: CampoParameters) -> SearchOutcome[CampoParameters]:
        return evaluate_campo_parameters(parameters, profile, cache=cache, coarse_stride=coarse_stride, window=window)
    blocks = (('tower', ('tower_y',)), ('mirror', ('mirror_width', 'mirror_height', 'installation_height')), ('layout', ('first_ring_count', 'initial_spacing', 'spacing_growth')))
    for start in starts:
        (local, local_trace) = _cyclic_search(start, evaluator=evaluator, within_bounds=lambda value: _campo_within_bounds(value, bounds), blocks=blocks, step_levels=step_levels, maximum_cycles_per_level=maximum_cycles_per_level)
        local_results.append(local)
        trace.extend(local_trace)
    best = _rank_outcomes(local_results)[0]
    return OptimizationResult('campo', best, starts, tuple(local_results), tuple(trace))

def refine_campo(initial_parameters: CampoParameters, *, profile: EvaluationProfile, bounds: CampoBounds=CampoBounds(), step_levels: tuple[StepLevel, ...]=CAMPO_STEP_LEVELS, maximum_cycles_per_level: int=4, coarse_stride: int=4, window: int=2, cache: EvaluationCache | None=None, progress: Callable[[str], None] | None=None) -> OptimizationResult[CampoParameters]:
    if not _campo_within_bounds(initial_parameters, bounds):
        raise ValueError('方案 B 的恢复参数超出当前搜索边界。')
    cache = cache or EvaluationCache()

    def evaluator(parameters: CampoParameters) -> SearchOutcome[CampoParameters]:
        return evaluate_campo_parameters(parameters, profile, cache=cache, coarse_stride=coarse_stride, window=window)
    initial = evaluator(initial_parameters)
    if progress is not None:
        progress(f'方案 B 恢复起点：P={initial.annual_power_mw:.4f} MW，q={initial.unit_area_power_kw_m2:.6f}')
    blocks = (('tower', ('tower_y',)), ('mirror', ('mirror_width', 'mirror_height', 'installation_height')), ('layout', ('first_ring_count', 'initial_spacing', 'spacing_growth')))
    (best, trace) = _cyclic_search(initial, evaluator=evaluator, within_bounds=lambda value: _campo_within_bounds(value, bounds), blocks=blocks, step_levels=step_levels, maximum_cycles_per_level=maximum_cycles_per_level)
    return OptimizationResult('campo', best, (initial,), (best,), trace)

# ---- prune.py ----

from dataclasses import dataclass
import numpy as np

@dataclass(frozen=True)
class PruneStep:
    removed_indices: tuple[int, int]
    evaluation: FieldEvaluation

@dataclass(frozen=True)
class PruneResult:
    initial: FieldEvaluation
    best: FieldEvaluation
    steps: tuple[PruneStep, ...]

def _outer_symmetric_pairs(layout: GeneratedLayout, ring_count: int, *, ring_depth: int, tolerance: float=1e-07) -> tuple[tuple[int, int], ...]:
    if ring_depth < 1:
        raise ValueError('ring_depth 必须大于等于 1。')
    selected_start = max(0, ring_count - ring_depth)
    offsets: list[int] = []
    running = 0
    for ring in layout.rings[:ring_count]:
        offsets.append(running)
        running += ring.mirror_count
    pairs: list[tuple[int, int]] = []
    for ring_index in range(selected_start, ring_count):
        ring = layout.rings[ring_index]
        offset = offsets[ring_index]
        coordinates = ring.coordinates
        unused = set(range(coordinates.shape[0]))
        while unused:
            local = min(unused)
            unused.remove(local)
            (x, y) = coordinates[local]
            if abs(float(x)) <= tolerance:
                continue
            candidates = [other for other in unused if abs(float(coordinates[other, 0] + x)) <= tolerance and abs(float(coordinates[other, 1] - y)) <= tolerance]
            if not candidates:
                continue
            partner = min(candidates)
            unused.remove(partner)
            pairs.append((offset + local, offset + partner))
    return tuple(pairs)

def prune_outer_symmetric_pairs(*, layout: GeneratedLayout, parameters: LayoutParameters, initial: FieldEvaluation, profile: EvaluationProfile, target_power_mw: float=42.0, ring_depth: int=2, maximum_rounds: int=10, maximum_pairs_per_round: int | None=None, cache: EvaluationCache | None=None) -> PruneResult:
    if maximum_rounds < 0:
        raise ValueError('maximum_rounds 不能小于 0。')
    if not initial.is_feasible(target_power_mw):
        raise ValueError('结构化修剪要求初始镜场已经满足功率约束。')
    original = layout.prefix(initial.ring_count)
    if original.shape != initial.coordinates.shape or not np.allclose(original, initial.coordinates, atol=1e-09):
        raise ValueError('initial 坐标与指定布局前缀不一致。')
    pairs = _outer_symmetric_pairs(layout, initial.ring_count, ring_depth=ring_depth)
    active = np.ones(original.shape[0], dtype=bool)
    current = initial
    steps: list[PruneStep] = []
    for _ in range(maximum_rounds):
        remaining_pairs = [pair for pair in pairs if active[pair[0]] and active[pair[1]]]
        if maximum_pairs_per_round is not None and len(remaining_pairs) > maximum_pairs_per_round:
            sampled_indices = np.linspace(0, len(remaining_pairs) - 1, maximum_pairs_per_round, dtype=int)
            remaining_pairs = [remaining_pairs[index] for index in sampled_indices]
        best_pair: tuple[int, int] | None = None
        best_evaluation: FieldEvaluation | None = None
        for pair in remaining_pairs:
            candidate_active = active.copy()
            candidate_active[list(pair)] = False
            candidate = evaluate_coordinates(layout_kind=layout.kind, ring_count=initial.ring_count, coordinates=original[candidate_active], parameters=parameters, profile=profile, cache=cache)
            if not candidate.is_feasible(target_power_mw):
                continue
            if candidate.unit_area_power_kw_m2 <= current.unit_area_power_kw_m2:
                continue
            if best_evaluation is None or candidate.unit_area_power_kw_m2 > best_evaluation.unit_area_power_kw_m2:
                best_pair = pair
                best_evaluation = candidate
        if best_pair is None or best_evaluation is None:
            break
        active[list(best_pair)] = False
        current = best_evaluation
        steps.append(PruneStep(best_pair, current))
    return PruneResult(initial, current, tuple(steps))

# ---- export.py ----

import csv
import json
from copy import copy
from dataclasses import asdict
from pathlib import Path
from typing import Any
from openpyxl import load_workbook
TARGET_ANNUAL_POWER_MW = 42.0

def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        raise ValueError(f'没有可写入 {path.name} 的结果。')
    with path.open('w', encoding='utf-8-sig', newline='') as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]), lineterminator='\n')
        writer.writeheader()
        writer.writerows(rows)

def write_result2_workbook(*, template_path: str | Path, output_path: str | Path, evaluation: FieldEvaluation, parameters: LayoutParameters) -> Path:
    template = Path(template_path)
    if not template.exists():
        raise FileNotFoundError(f'找不到 result2.xlsx 模板：{template}')
    destination = Path(output_path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(template)
    sheet = workbook.active
    if sheet.max_column < 8:
        workbook.close()
        raise ValueError('result2.xlsx 模板列数不足 8 列。')
    style_source = [copy(sheet.cell(2, column)._style) for column in range(1, 9)]
    number_formats = [sheet.cell(2, column).number_format for column in range(1, 9)]
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)
    for (row_index, (x_m, y_m)) in enumerate(evaluation.coordinates, start=2):
        values = (parameters.tower_x, parameters.tower_y, row_index - 1, parameters.mirror_width, parameters.mirror_height, float(x_m), float(y_m), parameters.installation_height)
        for (column, value) in enumerate(values, start=1):
            cell = sheet.cell(row_index, column, value)
            cell._style = copy(style_source[column - 1])
            cell.number_format = number_formats[column - 1]
    workbook.save(destination)
    workbook.close()
    return destination

def write_question2_results(*, output_dir: str | Path, layout_name: str, parameters: LayoutParameters, evaluation: FieldEvaluation, result2_template: str | Path, comparison: dict[str, Any] | None=None) -> dict[str, Path]:
    destination = Path(output_dir)
    destination.mkdir(parents=True, exist_ok=True)
    coordinate_rows = [{'mirror_id': index + 1, 'mirror_width_m': parameters.mirror_width, 'mirror_height_m': parameters.mirror_height, 'x_m': float(x_m), 'y_m': float(y_m), 'z_m': parameters.installation_height} for (index, (x_m, y_m)) in enumerate(evaluation.coordinates)]
    monthly_rows = [asdict(record) for record in evaluation.solution.monthly_results]
    mirror_rows = [asdict(record) for record in evaluation.solution.mirror_annual_results]
    annual = asdict(evaluation.solution.annual_result)
    coordinates_path = destination / '03_最终镜位坐标.csv'
    monthly_path = destination / '04_月平均计算结果.csv'
    annual_path = destination / '05_年平均计算结果.json'
    mirror_path = destination / '06_单镜年平均结果.csv'
    summary_path = destination / '07_最终方案摘要.json'
    table_path = destination / '08_论文结果与验证表.md'
    workbook_path = destination / 'result2.xlsx'
    _write_csv(coordinates_path, coordinate_rows)
    _write_csv(monthly_path, monthly_rows)
    _write_csv(mirror_path, mirror_rows)
    annual_path.write_text(json.dumps(annual, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    summary = {'layout': layout_name, 'annual_power_constraint_mw': TARGET_ANNUAL_POWER_MW, 'annual_power_margin_mw': evaluation.annual_power_mw - TARGET_ANNUAL_POWER_MW, 'constraint_satisfied': evaluation.annual_power_mw >= TARGET_ANNUAL_POWER_MW, 'parameters': asdict(parameters), 'ring_count': evaluation.ring_count, 'mirror_count': evaluation.mirror_count, 'mirror_area_m2': evaluation.mirror_area_m2, 'total_area_m2': evaluation.total_area_m2, 'annual': annual}
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    lines = ['# 第二问结果与验证表', '', '本文档汇总第二问的功率约束、最终参数、双布局比较、月平均结果和加密验证。', '', '## 表 1 功率约束与优化目标', '', '| 年平均输出热功率下限 (MW) | 最终年平均输出热功率 (MW) | 功率余量 (MW) | 是否满足约束 | 单位镜面面积年平均输出热功率 (kW/m²) |', '| ---: | ---: | ---: | :---: | ---: |', f"| {TARGET_ANNUAL_POWER_MW:.6f} | {evaluation.annual_power_mw:.6f} | {evaluation.annual_power_mw - TARGET_ANNUAL_POWER_MW:.6f} | {('是' if evaluation.annual_power_mw >= TARGET_ANNUAL_POWER_MW else '否')} | {evaluation.unit_area_power_kw_m2:.6f} |", '', '> 本题中的 42 MW 是年平均输出热功率下限；优化目标是在满足该下限后最大化单位镜面面积年平均输出热功率。', '', '## 表 2 最终设计参数', '', '| 布局 | 塔坐标 | 镜面尺寸 | 安装高度 | 镜子数 | 总镜面面积 (m²) | 年平均功率 (MW) | 单位面积功率 (kW/m²) |', '| --- | --- | ---: | ---: | ---: | ---: | ---: | ---: |', f'| {layout_name} | ({parameters.tower_x:.3f}, {parameters.tower_y:.3f}) | {parameters.mirror_width:.3f}×{parameters.mirror_height:.3f} | {parameters.installation_height:.3f} | {evaluation.mirror_count} | {evaluation.total_area_m2:.3f} | {evaluation.annual_power_mw:.6f} | {evaluation.unit_area_power_kw_m2:.6f} |']
    if comparison is not None and {'partitioned', 'campo'}.issubset(comparison):
        lines.extend(['', '## 表 3 两种候选布局的正式精度对比', '', '| 布局 | 安全余量 (m) | 镜子数 | 总镜面面积 (m²) | 年平均功率 (MW) | 单位面积功率 (kW/m²) |', '| --- | ---: | ---: | ---: | ---: | ---: |'])
        for (kind, label) in (('partitioned', '分区交错同心圆'), ('campo', '改进 Campo')):
            record = comparison[kind]
            lines.append(f"| {label} | {record['parameters']['safety_epsilon']:.6f} | {record['mirror_count']} | {record['total_area_m2']:.3f} | {record['annual_power_mw']:.6f} | {record['unit_area_power_kw_m2']:.6f} |")
    geometry = validate_layout(evaluation.coordinates, parameters)
    lines.extend(['', '## 表 4 几何约束复核', '', '| 检查项 | 实际值 | 约束 | 结果 |', '| --- | ---: | ---: | :---: |', f"| 最小镜心距离 (m) | {geometry.minimum_center_distance:.9f} | > {parameters.mirror_width + 5.0:.9f} | {('通过' if geometry.valid else '未通过')} |", f"| 镜心距离安全余量 (m) | {geometry.minimum_center_distance - parameters.mirror_width - 5.0:.9f} | > 0 | {('通过' if geometry.minimum_center_distance > parameters.mirror_width + 5.0 else '未通过')} |", f"| 最大场地半径 (m) | {geometry.maximum_field_radius:.6f} | ≤ {parameters.field_radius:.3f} | {('通过' if geometry.maximum_field_radius <= parameters.field_radius + 1e-09 else '未通过')} |", f"| 最小塔距 (m) | {geometry.minimum_tower_distance:.6f} | ≥ {parameters.exclusion_radius:.3f} | {('通过' if geometry.minimum_tower_distance >= parameters.exclusion_radius - 1e-09 else '未通过')} |", f"| 不触地高度余量 (m) | {parameters.installation_height - parameters.mirror_height / 2.0:.6f} | ≥ 0 | {('通过' if parameters.installation_height >= parameters.mirror_height / 2.0 else '未通过')} |", '', '## 表 5 月平均光学效率及输出热功率', '', '| 月份 | 光学效率 | 余弦效率 | 阴影遮挡效率 | 截断效率 | 输出热功率 (MW) | 单位面积功率 (kW/m²) |', '| ---: | ---: | ---: | ---: | ---: | ---: | ---: |'])
    for record in evaluation.solution.monthly_results:
        lines.append(f'| {record.month} | {record.average_optical_efficiency:.6f} | {record.average_cosine_efficiency:.6f} | {record.average_shadow_blocking_efficiency:.6f} | {record.average_truncation_efficiency:.6f} | {record.field_output_mw:.6f} | {record.unit_area_output_kw_m2:.6f} |')
    table_path.write_text('\n'.join(lines) + '\n', encoding='utf-8')
    write_result2_workbook(template_path=result2_template, output_path=workbook_path, evaluation=evaluation, parameters=parameters)
    return {'coordinates': coordinates_path, 'monthly': monthly_path, 'annual': annual_path, 'mirror_annual': mirror_path, 'summary': summary_path, 'paper_table': table_path, 'result2': workbook_path}

def write_high_precision_validation(*, output_dir: str | Path, evaluation: FieldEvaluation, profile: EvaluationProfile) -> Path:
    destination = Path(output_dir)
    validation_path = destination / '09_高精度加密验证.json'
    annual = evaluation.solution.annual_result
    payload = {'profile': {'months': len(profile.months), 'solar_times_per_month': len(profile.solar_times), 'shadow_grid_size': profile.solver.shadow_grid_size, 'truncation_rays': profile.solver.truncation_rays, 'neighbor_radius_m': profile.solver.neighbor_radius_m}, 'mirror_count': evaluation.mirror_count, 'annual_power_constraint_mw': TARGET_ANNUAL_POWER_MW, 'annual_power_mw': evaluation.annual_power_mw, 'annual_power_margin_mw': evaluation.annual_power_mw - TARGET_ANNUAL_POWER_MW, 'unit_area_power_kw_m2': evaluation.unit_area_power_kw_m2, 'average_optical_efficiency': annual.average_optical_efficiency, 'average_shadow_blocking_efficiency': annual.average_shadow_blocking_efficiency, 'average_truncation_efficiency': annual.average_truncation_efficiency, 'constraint_satisfied': evaluation.is_feasible()}
    validation_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    table_path = destination / '08_论文结果与验证表.md'
    lines = ['', '## 表 6 高精度加密验证', '', '| 阴影网格 | 截断光线 | 邻镜半径 (m) | 年平均功率 (MW) | 功率余量 (MW) | 单位面积功率 (kW/m²) | 是否满足约束 |', '| ---: | ---: | ---: | ---: | ---: | ---: | :---: |', f"| {profile.solver.shadow_grid_size}×{profile.solver.shadow_grid_size} | {profile.solver.truncation_rays} | {profile.solver.neighbor_radius_m:.0f} | {evaluation.annual_power_mw:.6f} | {evaluation.annual_power_mw - TARGET_ANNUAL_POWER_MW:.6f} | {evaluation.unit_area_power_kw_m2:.6f} | {('是' if evaluation.is_feasible() else '否')} |"]
    with table_path.open('a', encoding='utf-8') as handle:
        handle.write('\n'.join(lines) + '\n')
    return validation_path

# ---- solve.py ----

import argparse
import json
from dataclasses import asdict, replace
from pathlib import Path
from typing import Sequence
PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_OUTPUT = PROJECT_ROOT / 'outputs' / 'q2'
DEFAULT_TEMPLATE = PROJECT_ROOT / 'task' / 'A' / 'result2.xlsx'

def _smoke_profile() -> EvaluationProfile:
    return EvaluationProfile(name='smoke', solver=SolverConfig(shadow_grid_size=3, truncation_rays=8, neighbor_radius_m=60.0, truncation_chunk_size=64, sobol_seed=2023), months=(6,), solar_times=(12.0,))

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description='独立优化分区圆环和改进 Campo 两种问题二镜场')
    parser.add_argument('--layout', choices=('both', 'partitioned', 'campo'), default='both')
    parser.add_argument('--output', type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument('--template', type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument('--initial-samples', type=int, default=16)
    parser.add_argument('--retained-starts', type=int, default=3)
    parser.add_argument('--max-cycles', type=int, default=4)
    parser.add_argument('--coarse-stride', type=int, default=4)
    parser.add_argument('--extent-window', type=int, default=2)
    parser.add_argument('--seed', type=int, default=2023)
    parser.add_argument('--resume-comparison', type=Path, default=None, help='从上一阶段的 02_双布局比较.json 参数继续局部搜索')
    parser.add_argument('--search-profile', choices=('exploration', 'refinement'), default='exploration', help='非烟雾搜索使用的数值离散精度')
    parser.add_argument('--step-level-count', type=int, choices=(1, 2, 3), default=3, help='本阶段连续使用几档步长')
    parser.add_argument('--step-level-start', type=int, choices=(1, 2, 3), default=1, help='本阶段从第几档步长开始（1 为粗、3 为细）')
    parser.add_argument('--prune-rounds', type=int, default=10, help='胜出布局结构化删镜的最大轮数；0 表示跳过')
    parser.add_argument('--prune-pairs-per-round', type=int, default=None, help='每轮最多复算的外层对称镜位对；默认全部')
    parser.add_argument('--smoke', action='store_true', help='仅用 6 月正午、3×3 阴影网格和 8 条截断光线验证流程')
    parser.add_argument('--skip-x-check', action='store_true', help='跳过塔东西坐标 {-10,-5,0,5,10} m 的少量复核')
    parser.add_argument('--skip-figures', action='store_true', help='不生成四张正式结果图')
    parser.add_argument('--figures-only', action='store_true', help='读取输出目录中的正式结果并重新生成四张图')
    parser.add_argument('--run-validation', action='store_true', help='额外运行 20×20 阴影网格、512 条截断光线的加密复算')
    return parser

def _validate_args(args: argparse.Namespace) -> None:
    for name in ('initial_samples', 'retained_starts', 'coarse_stride'):
        if getattr(args, name) < 1:
            raise SystemExit(f"--{name.replace('_', '-')} 必须大于等于 1。")
    if args.extent_window < 0:
        raise SystemExit('--extent-window 不能小于 0。')
    if args.max_cycles < 0:
        raise SystemExit('--max-cycles 不能小于 0。')
    if args.step_level_start + args.step_level_count - 1 > 3:
        raise SystemExit('--step-level-start 与 --step-level-count 超出三档步长。')
    if args.prune_rounds < 0:
        raise SystemExit('--prune-rounds 不能小于 0。')
    if args.prune_pairs_per_round is not None and args.prune_pairs_per_round < 1:
        raise SystemExit('--prune-pairs-per-round 必须大于等于 1。')

def run(argv: Sequence[str] | None=None) -> int:
    args = build_parser().parse_args(argv)
    _validate_args(args)
    if args.figures_only:
        for path in build_question2_figures_from_output(args.output):
            print(f'输出：{path}')
        return 0
    if args.smoke:
        search_profile = _smoke_profile()
    elif args.search_profile == 'refinement':
        search_profile = refinement_profile()
    else:
        search_profile = exploration_profile()
    verification_profile = _smoke_profile() if args.smoke else final_profile()
    cache = EvaluationCache()
    optimized: dict[str, object] = {}
    step_start = args.step_level_start - 1
    step_stop = step_start + args.step_level_count
    partitioned_steps = PARTITIONED_STEP_LEVELS[step_start:step_stop]
    campo_steps = CAMPO_STEP_LEVELS[step_start:step_stop]
    resumed = None
    if args.resume_comparison is not None:
        if not args.resume_comparison.exists():
            raise SystemExit(f'找不到恢复文件：{args.resume_comparison}')
        resumed = json.loads(args.resume_comparison.read_text(encoding='utf-8'))
    if args.layout in ('both', 'partitioned'):
        print('开始独立优化方案 A：分区交错同心圆')
        if resumed is not None:
            optimized['partitioned'] = refine_partitioned(PartitionedRingParameters(**resumed['partitioned']['parameters']), profile=search_profile, step_levels=partitioned_steps, maximum_cycles_per_level=args.max_cycles, coarse_stride=args.coarse_stride, window=args.extent_window, cache=cache, progress=print)
        else:
            optimized['partitioned'] = optimize_partitioned(profile=search_profile, initial_sample_count=args.initial_samples, retained_starts=args.retained_starts, seed=args.seed, step_levels=partitioned_steps, maximum_cycles_per_level=args.max_cycles, coarse_stride=args.coarse_stride, window=args.extent_window, cache=cache, progress=print)
    if args.layout in ('both', 'campo'):
        print('开始独立优化方案 B：改进 Campo 径向交错')
        if resumed is not None:
            optimized['campo'] = refine_campo(CampoParameters(**resumed['campo']['parameters']), profile=search_profile, step_levels=campo_steps, maximum_cycles_per_level=args.max_cycles, coarse_stride=args.coarse_stride, window=args.extent_window, cache=cache, progress=print)
        else:
            optimized['campo'] = optimize_campo(profile=search_profile, initial_sample_count=args.initial_samples, retained_starts=args.retained_starts, seed=args.seed + 1, step_levels=campo_steps, maximum_cycles_per_level=args.max_cycles, coarse_stride=args.coarse_stride, window=args.extent_window, cache=cache, progress=print)
    verified: dict[str, tuple[object, object, object, object]] = {}
    for (kind, result) in optimized.items():
        parameters = result.best.parameters
        x_check_scan = None
        if not args.smoke and (not args.skip_x_check):
            center_layout = generate_partitioned_layout(parameters) if kind == 'partitioned' else generate_campo_layout(parameters)
            center_scan = scan_layout_extents(center_layout, parameters, verification_profile, coarse_stride=args.coarse_stride, window=args.extent_window, cache=cache)
            selected_parameters = parameters
            selected_scan = center_scan
            for tower_x in (-10.0, -5.0, 5.0, 10.0):
                candidate_parameters = replace(parameters, tower_x=tower_x)
                candidate_layout = generate_partitioned_layout(candidate_parameters) if kind == 'partitioned' else generate_campo_layout(candidate_parameters)
                candidate_scan = scan_layout_extents(candidate_layout, candidate_parameters, verification_profile, coarse_stride=args.coarse_stride, window=args.extent_window, cache=cache)
                selected = better_evaluation(selected_scan.best, candidate_scan.best)
                selected_feasible = selected_scan.best.is_feasible()
                candidate_feasible = candidate_scan.best.is_feasible()
                if candidate_feasible != selected_feasible:
                    stable_gain = candidate_feasible
                elif candidate_feasible:
                    stable_gain = candidate_scan.best.unit_area_power_kw_m2 - selected_scan.best.unit_area_power_kw_m2 > 0.0001
                else:
                    stable_gain = candidate_scan.best.annual_power_mw - selected_scan.best.annual_power_mw > 0.001
                if selected is candidate_scan.best and stable_gain:
                    selected_parameters = candidate_parameters
                    selected_scan = candidate_scan
            parameters = selected_parameters
            x_check_scan = selected_scan
        layout = generate_partitioned_layout(parameters) if kind == 'partitioned' else generate_campo_layout(parameters)
        precision_label = '烟雾测试精度' if args.smoke else '问题一最终精度'
        print(f'使用统一{precision_label}复算 {kind}')
        if x_check_scan is None:
            scan = scan_layout_extents(layout, parameters, verification_profile, coarse_stride=args.coarse_stride, window=args.extent_window, cache=cache)
        else:
            scan = x_check_scan
        verified[kind] = (parameters, result, layout, scan)
    verified_values = list(verified.items())
    (winner_kind, winner_bundle) = verified_values[0]
    for (kind, bundle) in verified_values[1:]:
        if better_evaluation(winner_bundle[3].best, bundle[3].best) is bundle[3].best:
            (winner_kind, winner_bundle) = (kind, bundle)
    (winner_parameters, _, winner_layout, winner_scan) = winner_bundle
    winner_evaluation = winner_scan.best
    if args.prune_rounds and abs(winner_parameters.tower_x) <= 1e-09:
        print('对胜出布局执行外层东西对称镜位修剪')
        prune = prune_outer_symmetric_pairs(layout=winner_layout, parameters=winner_parameters, initial=winner_evaluation, profile=verification_profile, maximum_rounds=args.prune_rounds, maximum_pairs_per_round=args.prune_pairs_per_round, cache=cache)
        winner_evaluation = prune.best
    elif args.prune_rounds:
        print('塔东西坐标不为 0，跳过要求南北轴对称的外层镜位修剪')
    args.output.mkdir(parents=True, exist_ok=True)
    comparison = {kind: {'parameters': asdict(bundle[0]), 'ring_count': bundle[3].best.ring_count, 'mirror_count': bundle[3].best.mirror_count, 'total_area_m2': bundle[3].best.total_area_m2, 'annual_power_mw': bundle[3].best.annual_power_mw, 'unit_area_power_kw_m2': bundle[3].best.unit_area_power_kw_m2} for (kind, bundle) in verified.items()}
    comparison[winner_kind].update({'ring_count': winner_evaluation.ring_count, 'mirror_count': winner_evaluation.mirror_count, 'total_area_m2': winner_evaluation.total_area_m2, 'annual_power_mw': winner_evaluation.annual_power_mw, 'unit_area_power_kw_m2': winner_evaluation.unit_area_power_kw_m2})
    comparison['winner'] = winner_kind
    comparison_path = args.output / '02_双布局比较.json'
    comparison_path.write_text(json.dumps(comparison, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    written = write_question2_results(output_dir=args.output, layout_name=winner_kind, parameters=winner_parameters, evaluation=winner_evaluation, result2_template=args.template, comparison=comparison)
    if args.run_validation and (not args.smoke):
        dense_profile = EvaluationProfile(name='dense-validation', solver=SolverConfig(shadow_grid_size=20, truncation_rays=512, neighbor_radius_m=80.0, truncation_chunk_size=128, sobol_seed=2023))
        dense_evaluation = evaluate_coordinates(layout_kind=winner_kind, ring_count=winner_evaluation.ring_count, coordinates=winner_evaluation.coordinates, parameters=winner_parameters, profile=dense_profile)
        written['dense_validation'] = write_high_precision_validation(output_dir=args.output, evaluation=dense_evaluation, profile=dense_profile)
    if not args.skip_figures and len(verified) == 2:
        figure_evaluations = {kind: bundle[3].best for (kind, bundle) in verified.items()}
        figure_evaluations[winner_kind] = winner_evaluation
        figure_parameters = {kind: bundle[0] for (kind, bundle) in verified.items()}
        for path in build_question2_figures(output_dir=args.output, comparison=comparison, parameters=figure_parameters, evaluations=figure_evaluations):
            written[path.stem] = path
    elif not args.skip_figures:
        print('仅优化一种布局，跳过双布局对比图。')
    print('\n第二问烟雾测试结果（不可作为正式年平均结论）' if args.smoke else '\n第二问结果')
    print(f'胜出布局：{winner_kind}')
    print(f'镜子数：{winner_evaluation.mirror_count}')
    print(f'总镜面面积：{winner_evaluation.total_area_m2:.3f} m²')
    target_power_mw = 42.0
    print(f'年平均输出热功率约束下限：{target_power_mw:.6f} MW')
    print(f'最终年平均输出热功率：{winner_evaluation.annual_power_mw:.6f} MW')
    print(f'相对约束下限的功率余量：{winner_evaluation.annual_power_mw - target_power_mw:.6f} MW')
    print(f'单位面积年平均输出热功率：{winner_evaluation.unit_area_power_kw_m2:.6f} kW/m²')
    print(f'双布局比较：{comparison_path}')
    for path in written.values():
        print(f'输出：{path}')
    return 0

def main() -> None:
    raise SystemExit(run())

if __name__ == "__main__":
    raise SystemExit(run(["--skip-figures", *sys.argv[1:]]))
