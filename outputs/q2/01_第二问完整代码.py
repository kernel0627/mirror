"""第二问完整代码展示稿。

本文件把共享光学核心、两种布局、搜索、验证、输出和绘图流程合并为单文件，可直接运行。
"""

from __future__ import annotations

# ruff: noqa: E402,F401,F811


# ========================================================================
# 参数配置
# ========================================================================

from dataclasses import asdict, dataclass


@dataclass(frozen=True)
class FieldConfig:
    """第一问给定的镜场和环境参数。"""

    latitude_deg: float = 39.4
    altitude_km: float = 3.0
    field_radius: float = 350.0
    exclusion_radius: float = 100.0
    tower_x: float = 0.0
    tower_y: float = 0.0
    receiver_center_z: float = 86.0
    receiver_radius: float = 4.0
    receiver_height: float = 8.0
    mirror_width: float = 6.2
    mirror_height: float = 6.2
    mirror_center_z: float = 4.5
    reflectivity: float = 0.92
    solar_angular_radius_rad: float = 4.65e-3

    @property
    def receiver_z_min(self) -> float:
        return self.receiver_center_z - self.receiver_height / 2.0

    @property
    def receiver_z_max(self) -> float:
        return self.receiver_center_z + self.receiver_height / 2.0

    @property
    def mirror_area(self) -> float:
        return self.mirror_width * self.mirror_height

    def to_dict(self) -> dict[str, float]:
        return asdict(self)


@dataclass(frozen=True)
class SolverConfig:
    """可调的采样精度和加速参数。"""

    shadow_grid_size: int = 15
    truncation_rays: int = 256
    neighbor_radius_m: float = 60.0
    candidate_margin: float = 1.05
    ray_epsilon: float = 1e-7
    truncation_chunk_size: int = 128
    sobol_seed: int = 2023
    calculate_shadow: bool = True
    calculate_truncation: bool = True

    def __post_init__(self) -> None:
        if self.shadow_grid_size < 1:
            raise ValueError("shadow_grid_size 必须大于等于 1。")
        if self.truncation_rays < 1:
            raise ValueError("truncation_rays 必须大于等于 1。")
        if self.neighbor_radius_m <= 0.0:
            raise ValueError("neighbor_radius_m 必须大于 0。")
        if self.candidate_margin < 1.0:
            raise ValueError("candidate_margin 不能小于 1。")
        if self.ray_epsilon <= 0.0:
            raise ValueError("ray_epsilon 必须大于 0。")
        if self.truncation_chunk_size < 1:
            raise ValueError("truncation_chunk_size 必须大于等于 1。")

    def to_dict(self) -> dict[str, int | float | bool]:
        return asdict(self)


# ========================================================================
# 太阳位置与 DNI
# ========================================================================

import math
from dataclasses import dataclass

import numpy as np
from numpy.typing import NDArray


FloatArray = NDArray[np.float64]


@dataclass(frozen=True)
class SolarState:
    month: int
    solar_time: float
    direction: FloatArray
    altitude_rad: float
    azimuth_rad: float
    declination_rad: float
    dni_kw_m2: float


MONTH_DAYS = (31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31)


def day_from_spring_equinox(month: int, day: int = 21) -> int:
    """以 3 月 21 日为第 0 天，返回题面赤纬公式所需的 D。"""

    if not 1 <= month <= 12:
        raise ValueError("month 必须位于 1 到 12。")
    if not 1 <= day <= MONTH_DAYS[month - 1]:
        raise ValueError("day 不在指定月份的有效范围内。")
    day_of_year = sum(MONTH_DAYS[: month - 1]) + day
    return day_of_year - 80


def calculate_solar_state(
    month: int,
    solar_time: float,
    latitude_deg: float,
    altitude_km: float,
) -> SolarState:
    """按题面附录计算东-北-天坐标下的太阳单位方向和 DNI。"""

    if not 0.0 <= solar_time <= 24.0:
        raise ValueError("solar_time 必须位于 0 到 24 小时。")

    d = day_from_spring_equinox(month)
    declination = math.asin(
        math.sin(2.0 * math.pi * d / 365.0) * math.sin(math.radians(23.45))
    )
    latitude = math.radians(latitude_deg)
    hour_angle = math.pi / 12.0 * (solar_time - 12.0)

    direction = np.array(
        [
            -math.cos(declination) * math.sin(hour_angle),
            math.cos(latitude) * math.sin(declination)
            - math.sin(latitude) * math.cos(declination) * math.cos(hour_angle),
            math.sin(latitude) * math.sin(declination)
            + math.cos(latitude) * math.cos(declination) * math.cos(hour_angle),
        ],
        dtype=float,
    )
    direction /= np.linalg.norm(direction)

    altitude = math.asin(float(np.clip(direction[2], -1.0, 1.0)))
    azimuth = math.atan2(float(direction[0]), float(direction[1])) % (2.0 * math.pi)

    h = altitude_km
    a = 0.4237 - 0.00821 * (6.0 - h) ** 2
    b = 0.5055 + 0.00595 * (6.5 - h) ** 2
    c = 0.2711 + 0.01858 * (2.5 - h) ** 2
    if altitude <= 0.0:
        dni = 0.0
    else:
        dni = 1.366 * (a + b * math.exp(-c / math.sin(altitude)))

    return SolarState(
        month=month,
        solar_time=solar_time,
        direction=direction,
        altitude_rad=altitude,
        azimuth_rad=azimuth,
        declination_rad=declination,
        dni_kw_m2=dni,
    )


# ========================================================================
# 镜场几何与镜面姿态
# ========================================================================

from dataclasses import dataclass

import numpy as np
from numpy.typing import NDArray


FloatArray = NDArray[np.float64]


def normalize_rows(vectors: FloatArray) -> FloatArray:
    norms = np.linalg.norm(vectors, axis=1, keepdims=True)
    if np.any(norms <= 1e-15):
        raise ValueError("不能归一化长度为零的向量。")
    return vectors / norms


def reflect(directions: FloatArray, normals: FloatArray) -> FloatArray:
    """根据 d - 2(d·n)n 计算反射方向。"""

    dot = np.sum(directions * normals, axis=-1, keepdims=True)
    return directions - 2.0 * dot * normals


@dataclass(frozen=True)
class PreparedField:
    config: FieldConfig
    centers: FloatArray
    receiver_center: FloatArray
    receiver_directions: FloatArray
    receiver_distances: FloatArray
    atmospheric_efficiency: FloatArray

    @property
    def mirror_count(self) -> int:
        return int(self.centers.shape[0])

    @property
    def total_mirror_area(self) -> float:
        return self.mirror_count * self.config.mirror_area


@dataclass(frozen=True)
class MirrorOrientation:
    normals: FloatArray
    width_axes: FloatArray
    height_axes: FloatArray
    cosine_efficiency: FloatArray


def prepare_field(mirror_xy: FloatArray, config: FieldConfig) -> PreparedField:
    mirror_count = int(mirror_xy.shape[0])
    centers = np.column_stack(
        (
            mirror_xy,
            np.full(mirror_count, config.mirror_center_z, dtype=float),
        )
    )
    receiver_center = np.array(
        [config.tower_x, config.tower_y, config.receiver_center_z],
        dtype=float,
    )
    receiver_vectors = receiver_center[None, :] - centers
    distances = np.linalg.norm(receiver_vectors, axis=1)
    receiver_directions = receiver_vectors / distances[:, None]
    atmospheric = 0.99321 - 0.0001176 * distances + 1.97e-8 * distances**2
    atmospheric = np.clip(atmospheric, 0.0, 1.0)
    return PreparedField(
        config=config,
        centers=centers,
        receiver_center=receiver_center,
        receiver_directions=receiver_directions,
        receiver_distances=distances,
        atmospheric_efficiency=atmospheric,
    )


def calculate_orientation(
    prepared: PreparedField,
    sun_direction: FloatArray,
) -> MirrorOrientation:
    sun_rows = np.broadcast_to(sun_direction, prepared.receiver_directions.shape)
    normals = normalize_rows(sun_rows + prepared.receiver_directions)

    upward = np.broadcast_to(np.array([0.0, 0.0, 1.0]), normals.shape)
    width_axes = np.cross(upward, normals)
    weak = np.linalg.norm(width_axes, axis=1) < 1e-10
    width_axes[weak] = np.array([1.0, 0.0, 0.0])
    width_axes = normalize_rows(width_axes)
    height_axes = normalize_rows(np.cross(normals, width_axes))

    cosine = np.clip(normals @ sun_direction, 0.0, 1.0)
    return MirrorOrientation(
        normals=normals,
        width_axes=width_axes,
        height_axes=height_axes,
        cosine_efficiency=cosine,
    )


def maximum_reflection_error(
    prepared: PreparedField,
    orientation: MirrorOrientation,
    sun_direction: FloatArray,
) -> float:
    incoming = np.broadcast_to(-sun_direction, orientation.normals.shape)
    reflected = reflect(incoming, orientation.normals)
    errors = np.linalg.norm(reflected - prepared.receiver_directions, axis=1)
    return float(np.max(errors))


# ========================================================================
# 阴影遮挡效率
# ========================================================================

import numpy as np
from numpy.typing import NDArray
from scipy.spatial import cKDTree


IntArray = NDArray[np.int64]


def mirror_grid_offsets(
    grid_size: int,
    mirror_width: float,
    mirror_height: float,
) -> FloatArray:
    """返回位于等面积小格中心的局部二维坐标。"""

    width_step = mirror_width / grid_size
    height_step = mirror_height / grid_size
    width_values = np.linspace(
        -mirror_width / 2.0 + width_step / 2.0,
        mirror_width / 2.0 - width_step / 2.0,
        grid_size,
    )
    height_values = np.linspace(
        -mirror_height / 2.0 + height_step / 2.0,
        mirror_height / 2.0 - height_step / 2.0,
        grid_size,
    )
    width_grid, height_grid = np.meshgrid(
        width_values,
        height_values,
        indexing="xy",
    )
    return np.column_stack((width_grid.ravel(), height_grid.ravel()))


def _direction_candidates(
    target_index: int,
    neighbors: IntArray,
    centers: FloatArray,
    direction: FloatArray,
    reach: float,
    maximum_distance: float | None = None,
) -> IntArray:
    if neighbors.size == 0:
        return neighbors

    delta = centers[neighbors] - centers[target_index]
    projection = delta @ direction
    perpendicular = delta - projection[:, None] * direction[None, :]
    perpendicular_distance = np.linalg.norm(perpendicular, axis=1)

    keep = (projection > -reach) & (perpendicular_distance <= reach)
    if maximum_distance is not None:
        keep &= projection < maximum_distance + reach
    return neighbors[keep]


def ray_rectangle_hits(
    origins: FloatArray,
    direction: FloatArray,
    rectangle_center: FloatArray,
    rectangle_normal: FloatArray,
    rectangle_width_axis: FloatArray,
    rectangle_height_axis: FloatArray,
    half_width: float,
    half_height: float,
    epsilon: float,
    maximum_distance: float | None = None,
) -> NDArray[np.bool_]:
    """判断一组同向射线是否与一面有限矩形相交。"""

    denominator = float(direction @ rectangle_normal)
    if abs(denominator) <= epsilon:
        return np.zeros(origins.shape[0], dtype=bool)

    distance = (rectangle_center - origins) @ rectangle_normal / denominator
    active = distance > epsilon
    if maximum_distance is not None:
        active &= distance < maximum_distance - epsilon
    if not np.any(active):
        return active

    intersections = origins + distance[:, None] * direction[None, :]
    relative = intersections - rectangle_center[None, :]
    local_width = relative @ rectangle_width_axis
    local_height = relative @ rectangle_height_axis
    active &= np.abs(local_width) <= half_width + epsilon
    active &= np.abs(local_height) <= half_height + epsilon
    return active


def _blocked_by_candidates(
    origins: FloatArray,
    direction: FloatArray,
    candidates: IntArray,
    prepared: PreparedField,
    orientation: MirrorOrientation,
    solver: SolverConfig,
    maximum_distance: float | None = None,
) -> NDArray[np.bool_]:
    blocked = np.zeros(origins.shape[0], dtype=bool)
    config = prepared.config

    for candidate in candidates:
        active_indices = np.flatnonzero(~blocked)
        if active_indices.size == 0:
            break
        hits = ray_rectangle_hits(
            origins=origins[active_indices],
            direction=direction,
            rectangle_center=prepared.centers[candidate],
            rectangle_normal=orientation.normals[candidate],
            rectangle_width_axis=orientation.width_axes[candidate],
            rectangle_height_axis=orientation.height_axes[candidate],
            half_width=config.mirror_width / 2.0,
            half_height=config.mirror_height / 2.0,
            epsilon=solver.ray_epsilon,
            maximum_distance=maximum_distance,
        )
        blocked[active_indices[hits]] = True

    return blocked


def calculate_shadow_blocking_efficiency(
    prepared: PreparedField,
    orientation: MirrorOrientation,
    sun_direction: FloatArray,
    solver: SolverConfig,
) -> FloatArray:
    """逐镜计算入射阴影和反射遮挡损失的采样点并集。"""

    mirror_count = prepared.mirror_count
    if mirror_count == 1:
        return np.ones(1, dtype=float)

    config = prepared.config
    offsets = mirror_grid_offsets(
        solver.shadow_grid_size,
        config.mirror_width,
        config.mirror_height,
    )
    sample_count = offsets.shape[0]
    tree = cKDTree(prepared.centers[:, :2])
    bounding_radius = 0.5 * np.hypot(
        config.mirror_width,
        config.mirror_height,
    )
    reach = 2.0 * bounding_radius * solver.candidate_margin
    efficiencies = np.empty(mirror_count, dtype=float)

    for index in range(mirror_count):
        points = (
            prepared.centers[index][None, :]
            + offsets[:, :1] * orientation.width_axes[index][None, :]
            + offsets[:, 1:] * orientation.height_axes[index][None, :]
        )
        neighbors = np.asarray(
            tree.query_ball_point(
                prepared.centers[index, :2],
                solver.neighbor_radius_m,
            ),
            dtype=np.int64,
        )
        neighbors = neighbors[neighbors != index]

        incoming_candidates = _direction_candidates(
            target_index=index,
            neighbors=neighbors,
            centers=prepared.centers,
            direction=sun_direction,
            reach=reach,
        )
        incoming_blocked = _blocked_by_candidates(
            origins=points,
            direction=sun_direction,
            candidates=incoming_candidates,
            prepared=prepared,
            orientation=orientation,
            solver=solver,
        )

        reflected_direction = prepared.receiver_directions[index]
        reflected_candidates = _direction_candidates(
            target_index=index,
            neighbors=neighbors,
            centers=prepared.centers,
            direction=reflected_direction,
            reach=reach,
            maximum_distance=prepared.receiver_distances[index],
        )
        reflected_blocked = _blocked_by_candidates(
            origins=points,
            direction=reflected_direction,
            candidates=reflected_candidates,
            prepared=prepared,
            orientation=orientation,
            solver=solver,
            maximum_distance=prepared.receiver_distances[index],
        )

        blocked = incoming_blocked | reflected_blocked
        efficiencies[index] = 1.0 - np.count_nonzero(blocked) / sample_count

    return np.clip(efficiencies, 0.0, 1.0)


# ========================================================================
# 截断效率
# ========================================================================

import math

import numpy as np
from numpy.typing import NDArray
from scipy.stats import qmc


def build_sobol_samples(sample_count: int, seed: int) -> FloatArray:
    """生成固定、可复现的四维 Sobol 样本。"""

    exponent = int(math.ceil(math.log2(sample_count)))
    sampler = qmc.Sobol(d=4, scramble=True, seed=seed)
    return sampler.random_base2(exponent)[:sample_count]


def _sun_disk_directions(
    sun_direction: FloatArray,
    samples: FloatArray,
    angular_radius: float,
) -> FloatArray:
    reference = (
        np.array([1.0, 0.0, 0.0])
        if abs(float(sun_direction[2])) > 0.9
        else np.array([0.0, 0.0, 1.0])
    )
    tangent_one = np.cross(reference, sun_direction)
    tangent_one /= np.linalg.norm(tangent_one)
    tangent_two = np.cross(sun_direction, tangent_one)

    radial_angle = angular_radius * np.sqrt(samples[:, 2])
    polar_angle = 2.0 * math.pi * samples[:, 3]
    tangent = (
        np.cos(polar_angle)[:, None] * tangent_one[None, :]
        + np.sin(polar_angle)[:, None] * tangent_two[None, :]
    )
    directions = (
        np.cos(radial_angle)[:, None] * sun_direction[None, :]
        + np.sin(radial_angle)[:, None] * tangent
    )
    directions /= np.linalg.norm(directions, axis=1, keepdims=True)
    return directions


def ray_cylinder_side_hits(
    origins: FloatArray,
    directions: FloatArray,
    tower_x: float,
    tower_y: float,
    radius: float,
    z_min: float,
    z_max: float,
    epsilon: float,
) -> NDArray[np.bool_]:
    """判断任意形状批量射线的两个正根中是否有有限圆柱侧面交点。"""

    origin_x = origins[..., 0] - tower_x
    origin_y = origins[..., 1] - tower_y
    direction_x = directions[..., 0]
    direction_y = directions[..., 1]

    a = direction_x**2 + direction_y**2
    b = 2.0 * (origin_x * direction_x + origin_y * direction_y)
    c = origin_x**2 + origin_y**2 - radius**2
    discriminant = b**2 - 4.0 * a * c

    valid = (a > epsilon) & (discriminant >= 0.0)
    square_root = np.sqrt(np.maximum(discriminant, 0.0))
    denominator = np.where(valid, 2.0 * a, 1.0)
    near = (-b - square_root) / denominator
    far = (-b + square_root) / denominator

    near_z = origins[..., 2] + near * directions[..., 2]
    far_z = origins[..., 2] + far * directions[..., 2]
    near_hit = (
        (near > epsilon) & (near_z >= z_min - epsilon) & (near_z <= z_max + epsilon)
    )
    far_hit = (far > epsilon) & (far_z >= z_min - epsilon) & (far_z <= z_max + epsilon)
    return valid & (near_hit | far_hit)


def calculate_truncation_efficiency(
    prepared: PreparedField,
    orientation: MirrorOrientation,
    sun_direction: FloatArray,
    solver: SolverConfig,
) -> FloatArray:
    """联合采样镜面位置和太阳圆盘方向，计算集热器截断效率。"""

    config = prepared.config
    samples = build_sobol_samples(solver.truncation_rays, solver.sobol_seed)
    local_width = (samples[:, 0] - 0.5) * config.mirror_width
    local_height = (samples[:, 1] - 0.5) * config.mirror_height
    sampled_sun = _sun_disk_directions(
        sun_direction,
        samples,
        config.solar_angular_radius_rad,
    )
    incoming = -sampled_sun

    efficiencies = np.empty(prepared.mirror_count, dtype=float)
    chunk_size = solver.truncation_chunk_size
    for start in range(0, prepared.mirror_count, chunk_size):
        stop = min(start + chunk_size, prepared.mirror_count)
        centers = prepared.centers[start:stop]
        normals = orientation.normals[start:stop]
        width_axes = orientation.width_axes[start:stop]
        height_axes = orientation.height_axes[start:stop]

        origins = (
            centers[:, None, :]
            + local_width[None, :, None] * width_axes[:, None, :]
            + local_height[None, :, None] * height_axes[:, None, :]
        )
        incoming_chunk = np.broadcast_to(
            incoming[None, :, :],
            origins.shape,
        )
        dot = np.einsum("csj,cj->cs", incoming_chunk, normals)
        reflected = incoming_chunk - 2.0 * dot[:, :, None] * normals[:, None, :]
        hits = ray_cylinder_side_hits(
            origins=origins,
            directions=reflected,
            tower_x=config.tower_x,
            tower_y=config.tower_y,
            radius=config.receiver_radius,
            z_min=config.receiver_z_min,
            z_max=config.receiver_z_max,
            epsilon=solver.ray_epsilon,
        )
        efficiencies[start:stop] = np.mean(hits, axis=1)

    return np.clip(efficiencies, 0.0, 1.0)


# ========================================================================
# 月平均、年平均与单镜年平均汇总
# ========================================================================

from dataclasses import dataclass
from typing import Any, Sequence

import numpy as np


@dataclass(frozen=True)
class MonthlyResult:
    month: int
    average_optical_efficiency: float
    average_cosine_efficiency: float
    average_shadow_blocking_efficiency: float
    average_atmospheric_efficiency: float
    average_truncation_efficiency: float
    field_output_mw: float
    unit_area_output_kw_m2: float


@dataclass(frozen=True)
class AnnualResult:
    average_optical_efficiency: float
    average_cosine_efficiency: float
    average_shadow_blocking_efficiency: float
    average_atmospheric_efficiency: float
    average_truncation_efficiency: float
    field_output_mw: float
    unit_area_output_kw_m2: float


@dataclass(frozen=True)
class MirrorAnnualResult:
    mirror_id: int
    x_m: float
    y_m: float
    radius_to_tower_m: float
    average_optical_efficiency: float
    average_cosine_efficiency: float
    average_shadow_blocking_efficiency: float
    average_atmospheric_efficiency: float
    average_truncation_efficiency: float
    average_output_power_kw: float


@dataclass(frozen=True)
class Question1Solution:
    time_results: tuple[Any, ...]
    monthly_results: tuple[MonthlyResult, ...]
    annual_result: AnnualResult
    mirror_annual_results: tuple[MirrorAnnualResult, ...]


_MEAN_FIELDS = (
    "average_optical_efficiency",
    "average_cosine_efficiency",
    "average_shadow_blocking_efficiency",
    "average_atmospheric_efficiency",
    "average_truncation_efficiency",
    "field_output_mw",
    "unit_area_output_kw_m2",
)


def _means(records: Sequence[Any]) -> tuple[float, ...]:
    if not records:
        raise ValueError("汇总记录不能为空。")
    return tuple(
        float(np.mean([getattr(record, field) for record in records]))
        for field in _MEAN_FIELDS
    )


def summarize_monthly(records: Sequence[Any]) -> tuple[MonthlyResult, ...]:
    """对每月规定的五个时刻等权平均。"""

    results: list[MonthlyResult] = []
    for month in sorted({record.month for record in records}):
        monthly = [record for record in records if record.month == month]
        results.append(MonthlyResult(month, *_means(monthly)))
    return tuple(results)


def summarize_annual(records: Sequence[Any]) -> AnnualResult:
    """对题目规定的全部时刻等权平均。"""

    return AnnualResult(*_means(records))


def summarize_mirror_annual(
    mirror_xy: np.ndarray,
    tower_x: float,
    tower_y: float,
    state_count: int,
    optical_efficiency_sum: np.ndarray,
    cosine_efficiency_sum: np.ndarray,
    shadow_blocking_efficiency_sum: np.ndarray,
    atmospheric_efficiency_sum: np.ndarray,
    truncation_efficiency_sum: np.ndarray,
    output_power_kw_sum: np.ndarray,
) -> tuple[MirrorAnnualResult, ...]:
    """由逐时刻运行和生成单镜年平均结果，不保留单镜逐时刻明细。"""

    if state_count < 1:
        raise ValueError("state_count 必须大于等于 1。")
    radius = np.hypot(mirror_xy[:, 0] - tower_x, mirror_xy[:, 1] - tower_y)
    means = {
        "optical": optical_efficiency_sum / state_count,
        "cosine": cosine_efficiency_sum / state_count,
        "shadow": shadow_blocking_efficiency_sum / state_count,
        "atmospheric": atmospheric_efficiency_sum / state_count,
        "truncation": truncation_efficiency_sum / state_count,
        "power": output_power_kw_sum / state_count,
    }
    return tuple(
        MirrorAnnualResult(
            mirror_id=index + 1,
            x_m=float(mirror_xy[index, 0]),
            y_m=float(mirror_xy[index, 1]),
            radius_to_tower_m=float(radius[index]),
            average_optical_efficiency=float(means["optical"][index]),
            average_cosine_efficiency=float(means["cosine"][index]),
            average_shadow_blocking_efficiency=float(means["shadow"][index]),
            average_atmospheric_efficiency=float(means["atmospheric"][index]),
            average_truncation_efficiency=float(means["truncation"][index]),
            average_output_power_kw=float(means["power"][index]),
        )
        for index in range(mirror_xy.shape[0])
    )


# ========================================================================
# 问题一光学评价核心
# ========================================================================

import argparse
import time
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Callable, Sequence

import numpy as np


SOLAR_TIMES = (9.0, 10.5, 12.0, 13.5, 15.0)
PROJECT_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_INPUT = PROJECT_ROOT / "task" / "A" / "fj.xlsx"
DEFAULT_OUTPUT = PROJECT_ROOT / "outputs" / "q1"


@dataclass(frozen=True)
class TimeResult:
    month: int
    solar_time: float
    dni_kw_m2: float
    average_optical_efficiency: float
    average_cosine_efficiency: float
    average_shadow_blocking_efficiency: float
    average_atmospheric_efficiency: float
    average_truncation_efficiency: float
    field_output_mw: float
    unit_area_output_kw_m2: float
    maximum_reflection_error: float


@dataclass(frozen=True)
class ValidationResult:
    category: str
    parameter: str
    metric: str
    value: float
    relative_difference_percent: float
    runtime_seconds: float


ProgressCallback = Callable[[int, int, TimeResult], None]


def _check_efficiency(name: str, values: np.ndarray) -> None:
    tolerance = 1e-12
    if np.any(values < -tolerance) or np.any(values > 1.0 + tolerance):
        minimum = float(np.min(values))
        maximum = float(np.max(values))
        raise RuntimeError(f"{name} 超出 [0, 1]：min={minimum:.6g}, max={maximum:.6g}")


def evaluate_time(
    prepared: PreparedField,
    month: int,
    solar_time: float,
    solver: SolverConfig,
    mirror_sums: dict[str, np.ndarray] | None = None,
) -> TimeResult:
    """计算一个月份、一个规定时刻的全场平均结果。"""

    solar = calculate_solar_state(
        month=month,
        solar_time=solar_time,
        latitude_deg=prepared.config.latitude_deg,
        altitude_km=prepared.config.altitude_km,
    )
    orientation = calculate_orientation(prepared, solar.direction)
    reflection_error = maximum_reflection_error(
        prepared,
        orientation,
        solar.direction,
    )
    if reflection_error >= 1e-8:
        raise RuntimeError(f"中心光线反射误差过大：{reflection_error:.3e}")

    if solver.calculate_shadow:
        shadow = calculate_shadow_blocking_efficiency(
            prepared,
            orientation,
            solar.direction,
            solver,
        )
    else:
        shadow = np.ones(prepared.mirror_count, dtype=float)

    if solver.calculate_truncation:
        truncation = calculate_truncation_efficiency(
            prepared,
            orientation,
            solar.direction,
            solver,
        )
    else:
        truncation = np.ones(prepared.mirror_count, dtype=float)

    cosine = orientation.cosine_efficiency
    atmospheric = prepared.atmospheric_efficiency
    optical = cosine * shadow * atmospheric * truncation * prepared.config.reflectivity
    for name, values in (
        ("余弦效率", cosine),
        ("阴影遮挡效率", shadow),
        ("大气透射率", atmospheric),
        ("截断效率", truncation),
        ("光学效率", optical),
    ):
        _check_efficiency(name, values)

    mirror_power_kw = solar.dni_kw_m2 * prepared.config.mirror_area * optical
    if mirror_sums is not None:
        mirror_sums["optical_efficiency_sum"] += optical
        mirror_sums["cosine_efficiency_sum"] += cosine
        mirror_sums["shadow_blocking_efficiency_sum"] += shadow
        mirror_sums["atmospheric_efficiency_sum"] += atmospheric
        mirror_sums["truncation_efficiency_sum"] += truncation
        mirror_sums["output_power_kw_sum"] += mirror_power_kw

    field_power_kw = float(np.sum(mirror_power_kw))
    return TimeResult(
        month=month,
        solar_time=solar_time,
        dni_kw_m2=solar.dni_kw_m2,
        average_optical_efficiency=float(np.mean(optical)),
        average_cosine_efficiency=float(np.mean(cosine)),
        average_shadow_blocking_efficiency=float(np.mean(shadow)),
        average_atmospheric_efficiency=float(np.mean(atmospheric)),
        average_truncation_efficiency=float(np.mean(truncation)),
        field_output_mw=field_power_kw / 1000.0,
        unit_area_output_kw_m2=field_power_kw / prepared.total_mirror_area,
        maximum_reflection_error=reflection_error,
    )


def solve_question1(
    prepared: PreparedField,
    solver: SolverConfig,
    months: Sequence[int] = tuple(range(1, 13)),
    solar_times: Sequence[float] = SOLAR_TIMES,
    progress: ProgressCallback | None = None,
) -> Question1Solution:
    """执行所选月份和时刻；默认即题目规定的 60 个状态。"""

    if not months or not solar_times:
        raise ValueError("months 和 solar_times 不能为空。")
    if any(month < 1 or month > 12 for month in months):
        raise ValueError("months 必须位于 1 到 12。")

    records: list[TimeResult] = []
    mirror_sums = {
        name: np.zeros(prepared.mirror_count, dtype=float)
        for name in (
            "optical_efficiency_sum",
            "cosine_efficiency_sum",
            "shadow_blocking_efficiency_sum",
            "atmospheric_efficiency_sum",
            "truncation_efficiency_sum",
            "output_power_kw_sum",
        )
    }
    total = len(months) * len(solar_times)
    for month in months:
        for solar_time in solar_times:
            record = evaluate_time(
                prepared,
                month,
                solar_time,
                solver,
                mirror_sums=mirror_sums,
            )
            records.append(record)
            if progress is not None:
                progress(len(records), total, record)

    time_results = tuple(records)
    return Question1Solution(
        time_results=time_results,
        monthly_results=summarize_monthly(time_results),
        annual_result=summarize_annual(time_results),
        mirror_annual_results=summarize_mirror_annual(
            mirror_xy=prepared.centers[:, :2],
            tower_x=prepared.config.tower_x,
            tower_y=prepared.config.tower_y,
            state_count=len(time_results),
            **mirror_sums,
        ),
    )


# ========================================================================
# 两种参数化镜场布局与几何检查
# ========================================================================

import math
from dataclasses import dataclass
from typing import Protocol

import numpy as np
from numpy.typing import NDArray
from scipy.spatial import cKDTree


FloatArray = NDArray[np.float64]


class LayoutError(ValueError):
    """布局参数或生成结果不可行。"""


class CommonParameters(Protocol):
    tower_x: float
    tower_y: float
    mirror_width: float
    mirror_height: float
    installation_height: float
    field_radius: float
    exclusion_radius: float
    safety_epsilon: float


@dataclass(frozen=True)
class PartitionedRingParameters:
    """分区交错同心圆布局参数。"""

    tower_y: float
    mirror_width: float
    mirror_height: float
    installation_height: float
    split_radius: float
    near_spacing: float
    far_spacing: float
    tower_x: float = 0.0
    field_radius: float = 350.0
    exclusion_radius: float = 100.0
    safety_epsilon: float = 0.01

    @property
    def safe_distance(self) -> float:
        return self.mirror_width + 5.0 + self.safety_epsilon


@dataclass(frozen=True)
class CampoParameters:
    """带渐增径向行距的 Campo 径向交错布局参数。"""

    tower_y: float
    mirror_width: float
    mirror_height: float
    installation_height: float
    first_ring_count: int
    initial_spacing: float
    spacing_growth: float
    tower_x: float = 0.0
    field_radius: float = 350.0
    exclusion_radius: float = 100.0
    safety_epsilon: float = 0.01

    @property
    def safe_distance(self) -> float:
        return self.mirror_width + 5.0 + self.safety_epsilon


@dataclass(frozen=True)
class LayoutRing:
    """一条生成并经过场地裁剪的圆环或圆弧。"""

    index: int
    radius: float
    zone: int
    nominal_count: int
    coordinates: FloatArray

    @property
    def mirror_count(self) -> int:
        return int(self.coordinates.shape[0])


@dataclass(frozen=True)
class GeneratedLayout:
    """由若干按半径排序的圆环或圆弧组成的镜场。"""

    kind: str
    rings: tuple[LayoutRing, ...]

    @property
    def mirror_count(self) -> int:
        return sum(ring.mirror_count for ring in self.rings)

    @property
    def coordinates(self) -> FloatArray:
        if not self.rings:
            return np.empty((0, 2), dtype=float)
        return np.concatenate(
            [ring.coordinates for ring in self.rings],
            axis=0,
        )

    def prefix(self, ring_count: int) -> FloatArray:
        if ring_count < 1 or ring_count > len(self.rings):
            raise ValueError(
                f"ring_count 应位于 1 到 {len(self.rings)}，实际为 {ring_count}。"
            )
        return np.concatenate(
            [ring.coordinates for ring in self.rings[:ring_count]],
            axis=0,
        )

    def prefix_mirror_count(self, ring_count: int) -> int:
        if ring_count < 1 or ring_count > len(self.rings):
            raise ValueError(
                f"ring_count 应位于 1 到 {len(self.rings)}，实际为 {ring_count}。"
            )
        return sum(ring.mirror_count for ring in self.rings[:ring_count])


@dataclass(frozen=True)
class GeometryCheck:
    valid: bool
    reason: str | None
    mirror_count: int
    minimum_center_distance: float
    maximum_field_radius: float
    minimum_tower_distance: float


def _validate_common_parameters(parameters: CommonParameters) -> None:
    values = (
        parameters.tower_x,
        parameters.tower_y,
        parameters.mirror_width,
        parameters.mirror_height,
        parameters.installation_height,
        parameters.field_radius,
        parameters.exclusion_radius,
        parameters.safety_epsilon,
    )
    if not all(math.isfinite(value) for value in values):
        raise LayoutError("布局参数必须全部为有限数。")
    if not 2.0 <= parameters.mirror_height <= parameters.mirror_width <= 8.0:
        raise LayoutError("镜面尺寸必须满足 2 ≤ h ≤ w ≤ 8。")
    if not 2.0 <= parameters.installation_height <= 6.0:
        raise LayoutError("安装高度必须位于 2 m 到 6 m。")
    if parameters.installation_height < parameters.mirror_height / 2.0:
        raise LayoutError("安装高度不足，镜面转动时可能触地。")
    if parameters.field_radius <= 0.0:
        raise LayoutError("场地半径必须大于 0。")
    if parameters.exclusion_radius <= 0.0:
        raise LayoutError("塔周禁区半径必须大于 0。")
    if parameters.safety_epsilon <= 0.0:
        raise LayoutError("安全距离余量必须大于 0。")


def _maximum_tower_centered_radius(parameters: CommonParameters) -> float:
    return parameters.field_radius + math.hypot(
        parameters.tower_x,
        parameters.tower_y,
    )


def _ring_coordinates(
    parameters: CommonParameters,
    radius: float,
    mirror_count: int,
    phase: float,
) -> FloatArray:
    if mirror_count < 2:
        raise LayoutError("单圈镜子数必须大于等于 2。")
    angles = 2.0 * math.pi * np.arange(mirror_count, dtype=float) / mirror_count + phase
    coordinates = np.column_stack(
        (
            parameters.tower_x + radius * np.sin(angles),
            parameters.tower_y + radius * np.cos(angles),
        )
    )
    field_radius = np.hypot(coordinates[:, 0], coordinates[:, 1])
    keep = field_radius <= parameters.field_radius + 1e-9
    clipped = np.asarray(coordinates[keep], dtype=float)
    clipped.setflags(write=False)
    return clipped


def _within_ring_count(radius: float, safe_distance: float) -> int:
    ratio = safe_distance / (2.0 * radius)
    if ratio >= 1.0:
        raise LayoutError("圆环半径过小，无法放置满足安全距离的镜子。")
    return int(math.floor(math.pi / math.asin(ratio)))


def generate_partitioned_layout(
    parameters: PartitionedRingParameters,
    *,
    maximum_rings: int = 256,
) -> GeneratedLayout:
    """生成分区交错同心圆，并拒绝跨环距离冲突。"""

    _validate_common_parameters(parameters)
    if not math.isfinite(parameters.split_radius):
        raise LayoutError("分区半径必须为有限数。")
    if parameters.split_radius <= parameters.exclusion_radius:
        raise LayoutError("分区半径必须位于塔周禁区之外。")
    if parameters.near_spacing <= 0.0:
        raise LayoutError("近区行距必须大于 0。")
    if parameters.far_spacing < parameters.near_spacing:
        raise LayoutError("远区行距必须大于等于近区行距。")

    rings: list[LayoutRing] = []
    radius = parameters.exclusion_radius
    maximum_radius = _maximum_tower_centered_radius(parameters)
    for ring_index in range(maximum_rings):
        if radius > maximum_radius + 1e-9:
            break
        count = _within_ring_count(radius, parameters.safe_distance)
        phase = 0.0 if ring_index % 2 == 0 else math.pi / count
        coordinates = _ring_coordinates(
            parameters,
            radius,
            count,
            phase,
        )
        if coordinates.size:
            rings.append(
                LayoutRing(
                    index=ring_index,
                    radius=radius,
                    zone=1 if radius < parameters.split_radius else 2,
                    nominal_count=count,
                    coordinates=coordinates,
                )
            )
        spacing = (
            parameters.near_spacing
            if radius < parameters.split_radius
            else parameters.far_spacing
        )
        radius += spacing
    else:
        raise LayoutError("达到 maximum_rings，圆环生成未正常终止。")

    layout = GeneratedLayout("partitioned", tuple(rings))
    check = validate_layout(layout.coordinates, parameters)
    if not check.valid:
        raise LayoutError(check.reason or "分区圆环布局不满足几何约束。")
    return layout


def _campo_zone(radius: float, first_radius: float) -> tuple[int, int]:
    if radius < 2.0 * first_radius:
        return 1, 1
    if radius < 4.0 * first_radius:
        return 2, 2
    return 3, 4


def generate_campo_layout(
    parameters: CampoParameters,
    *,
    maximum_rings: int = 256,
) -> GeneratedLayout:
    """生成三分区、渐增径向行距的 Campo 径向交错镜场。"""

    _validate_common_parameters(parameters)
    if parameters.first_ring_count < 2:
        raise LayoutError("Campo 首环镜子数必须大于等于 2。")
    if parameters.initial_spacing <= 0.0:
        raise LayoutError("Campo 初始行距必须大于 0。")
    if parameters.spacing_growth < 0.0:
        raise LayoutError("Campo 行距增长量不能小于 0。")

    first_radius = max(
        parameters.exclusion_radius,
        parameters.safe_distance
        / (2.0 * math.sin(math.pi / parameters.first_ring_count)),
    )
    maximum_radius = _maximum_tower_centered_radius(parameters)
    rings: list[LayoutRing] = []
    zone_rows = {1: 0, 2: 0, 3: 0}
    radius = first_radius

    for ring_index in range(maximum_rings):
        if radius > maximum_radius + 1e-9:
            break
        zone, multiplier = _campo_zone(radius, first_radius)
        count = parameters.first_ring_count * multiplier
        zone_index = zone_rows[zone]
        phase = 0.0 if zone_index % 2 == 0 else math.pi / count
        coordinates = _ring_coordinates(
            parameters,
            radius,
            count,
            phase,
        )
        if coordinates.size:
            rings.append(
                LayoutRing(
                    index=ring_index,
                    radius=radius,
                    zone=zone,
                    nominal_count=count,
                    coordinates=coordinates,
                )
            )
        zone_rows[zone] += 1
        radius += parameters.initial_spacing + parameters.spacing_growth * ring_index
    else:
        raise LayoutError("达到 maximum_rings，Campo 圆环生成未正常终止。")

    layout = GeneratedLayout("campo", tuple(rings))
    check = validate_layout(layout.coordinates, parameters)
    if not check.valid:
        raise LayoutError(check.reason or "Campo 布局不满足几何约束。")
    return layout


def validate_layout(
    coordinates: FloatArray,
    parameters: CommonParameters,
) -> GeometryCheck:
    """按题目口径检查尺寸、场地、禁区和严格中心距离约束。"""

    try:
        _validate_common_parameters(parameters)
    except LayoutError as exc:
        return GeometryCheck(False, str(exc), 0, math.inf, math.inf, math.inf)

    xy = np.asarray(coordinates, dtype=float)
    if xy.ndim != 2 or xy.shape[1] != 2:
        return GeometryCheck(
            False,
            f"镜位坐标应为 N×2，实际形状为 {xy.shape}。",
            0,
            math.inf,
            math.inf,
            math.inf,
        )
    if xy.shape[0] == 0:
        return GeometryCheck(False, "镜场中没有定日镜。", 0, math.inf, 0.0, 0.0)
    if not np.all(np.isfinite(xy)):
        return GeometryCheck(
            False,
            "镜位坐标包含 NaN 或无穷值。",
            int(xy.shape[0]),
            math.inf,
            math.inf,
            math.inf,
        )

    field_radii = np.hypot(xy[:, 0], xy[:, 1])
    tower_distances = np.hypot(
        xy[:, 0] - parameters.tower_x,
        xy[:, 1] - parameters.tower_y,
    )
    maximum_field_radius = float(np.max(field_radii))
    minimum_tower_distance = float(np.min(tower_distances))
    if maximum_field_radius > parameters.field_radius + 1e-9:
        return GeometryCheck(
            False,
            "存在越过 350 m 场地边界的镜位。",
            int(xy.shape[0]),
            math.inf,
            maximum_field_radius,
            minimum_tower_distance,
        )
    if minimum_tower_distance < parameters.exclusion_radius - 1e-9:
        return GeometryCheck(
            False,
            "存在进入塔周 100 m 禁区的镜位。",
            int(xy.shape[0]),
            math.inf,
            maximum_field_radius,
            minimum_tower_distance,
        )

    if xy.shape[0] == 1:
        minimum_distance = math.inf
    else:
        distances, _ = cKDTree(xy).query(xy, k=2)
        minimum_distance = float(np.min(distances[:, 1]))
    if minimum_distance <= parameters.mirror_width + 5.0:
        return GeometryCheck(
            False,
            (
                "最小中心距离不满足严格约束："
                f"{minimum_distance:.9f} m ≤ "
                f"{parameters.mirror_width + 5.0:.9f} m。"
            ),
            int(xy.shape[0]),
            minimum_distance,
            maximum_field_radius,
            minimum_tower_distance,
        )

    return GeometryCheck(
        True,
        None,
        int(xy.shape[0]),
        minimum_distance,
        maximum_field_radius,
        minimum_tower_distance,
    )


# ========================================================================
# 候选镜场评价、缓存与外边界扫描
# ========================================================================

import hashlib
from dataclasses import dataclass, replace
from typing import Sequence

import numpy as np


LayoutParameters = PartitionedRingParameters | CampoParameters


@dataclass(frozen=True)
class EvaluationProfile:
    """同一物理模型下的一组数值离散精度。"""

    name: str
    solver: SolverConfig
    months: tuple[int, ...] = tuple(range(1, 13))
    solar_times: tuple[float, ...] = SOLAR_TIMES


@dataclass(frozen=True)
class FieldEvaluation:
    """一个确定镜场外边界下的完整评价结果。"""

    layout_kind: str
    ring_count: int
    mirror_count: int
    mirror_area_m2: float
    total_area_m2: float
    coordinates: np.ndarray
    solution: Question1Solution

    @property
    def annual_power_mw(self) -> float:
        return self.solution.annual_result.field_output_mw

    @property
    def unit_area_power_kw_m2(self) -> float:
        return self.solution.annual_result.unit_area_output_kw_m2

    def is_feasible(self, target_power_mw: float = 42.0) -> bool:
        return self.annual_power_mw >= target_power_mw


@dataclass(frozen=True)
class ExtentScanResult:
    """一组布局参数在若干镜场外边界中的最好结果。"""

    best: FieldEvaluation
    evaluations: tuple[FieldEvaluation, ...]
    first_feasible_ring_count: int | None


class EvaluationCache:
    """按坐标、塔和数值精度缓存昂贵的光学评价。"""

    def __init__(self) -> None:
        self._values: dict[str, Question1Solution] = {}

    def get(self, key: str) -> Question1Solution | None:
        return self._values.get(key)

    def put(self, key: str, value: Question1Solution) -> None:
        self._values[key] = value

    def __len__(self) -> int:
        return len(self._values)


def exploration_profile() -> EvaluationProfile:
    return EvaluationProfile(
        name="exploration",
        solver=SolverConfig(
            shadow_grid_size=5,
            truncation_rays=64,
            neighbor_radius_m=60.0,
            truncation_chunk_size=128,
            sobol_seed=2023,
        ),
    )


def refinement_profile() -> EvaluationProfile:
    return EvaluationProfile(
        name="refinement",
        solver=SolverConfig(
            shadow_grid_size=10,
            truncation_rays=128,
            neighbor_radius_m=60.0,
            truncation_chunk_size=128,
            sobol_seed=2023,
        ),
    )


def final_profile() -> EvaluationProfile:
    return EvaluationProfile(
        name="final",
        solver=SolverConfig(
            shadow_grid_size=15,
            truncation_rays=256,
            neighbor_radius_m=60.0,
            truncation_chunk_size=128,
            sobol_seed=2023,
        ),
    )


def _field_config(
    parameters: CommonParameters,
    base: FieldConfig | None = None,
) -> FieldConfig:
    config = base or FieldConfig()
    return replace(
        config,
        field_radius=parameters.field_radius,
        exclusion_radius=parameters.exclusion_radius,
        tower_x=parameters.tower_x,
        tower_y=parameters.tower_y,
        mirror_width=parameters.mirror_width,
        mirror_height=parameters.mirror_height,
        mirror_center_z=parameters.installation_height,
    )


def _cache_key(
    coordinates: np.ndarray,
    config: FieldConfig,
    profile: EvaluationProfile,
) -> str:
    digest = hashlib.sha256()
    rounded = np.round(np.asarray(coordinates, dtype="<f8"), decimals=9)
    digest.update(rounded.tobytes(order="C"))
    digest.update(repr(config.to_dict()).encode("utf-8"))
    digest.update(repr(profile.solver.to_dict()).encode("utf-8"))
    digest.update(repr(profile.months).encode("ascii"))
    digest.update(repr(profile.solar_times).encode("ascii"))
    return digest.hexdigest()


def evaluate_coordinates(
    *,
    layout_kind: str,
    ring_count: int,
    coordinates: np.ndarray,
    parameters: LayoutParameters,
    profile: EvaluationProfile,
    cache: EvaluationCache | None = None,
    base_field_config: FieldConfig | None = None,
) -> FieldEvaluation:
    """直接复用问题一模型评价一套确定坐标。"""

    xy = np.asarray(coordinates, dtype=float)
    if xy.ndim != 2 or xy.shape[1] != 2 or xy.shape[0] == 0:
        raise ValueError("候选镜场坐标必须为非空 N×2 数组。")

    config = _field_config(parameters, base_field_config)
    key = _cache_key(xy, config, profile)
    solution = cache.get(key) if cache is not None else None
    if solution is None:
        prepared = prepare_field(xy, config)
        solution = solve_question1(
            prepared=prepared,
            solver=profile.solver,
            months=profile.months,
            solar_times=profile.solar_times,
        )
        if cache is not None:
            cache.put(key, solution)

    mirror_area = parameters.mirror_width * parameters.mirror_height
    return FieldEvaluation(
        layout_kind=layout_kind,
        ring_count=ring_count,
        mirror_count=int(xy.shape[0]),
        mirror_area_m2=mirror_area,
        total_area_m2=float(xy.shape[0] * mirror_area),
        coordinates=xy,
        solution=solution,
    )


def better_evaluation(
    left: FieldEvaluation,
    right: FieldEvaluation,
    *,
    target_power_mw: float = 42.0,
) -> FieldEvaluation:
    """按可行性优先规则返回较优结果。"""

    left_feasible = left.is_feasible(target_power_mw)
    right_feasible = right.is_feasible(target_power_mw)
    if left_feasible != right_feasible:
        return left if left_feasible else right
    if left_feasible:
        if left.unit_area_power_kw_m2 != right.unit_area_power_kw_m2:
            return (
                left
                if left.unit_area_power_kw_m2 > right.unit_area_power_kw_m2
                else right
            )
        return left if left.annual_power_mw <= right.annual_power_mw else right
    return left if left.annual_power_mw >= right.annual_power_mw else right


def _unique_ring_counts(values: Sequence[int], total: int) -> tuple[int, ...]:
    return tuple(sorted({value for value in values if 1 <= value <= total}))


def scan_layout_extents(
    layout: GeneratedLayout,
    parameters: LayoutParameters,
    profile: EvaluationProfile,
    *,
    target_power_mw: float = 42.0,
    coarse_stride: int = 4,
    window: int = 2,
    cache: EvaluationCache | None = None,
    base_field_config: FieldConfig | None = None,
) -> ExtentScanResult:
    """先粗定位功率阈值，再评价阈值附近的连续圆环外边界。"""

    if not layout.rings:
        raise ValueError("布局中没有可用于评价的圆环。")
    if coarse_stride < 1:
        raise ValueError("coarse_stride 必须大于等于 1。")
    if window < 0:
        raise ValueError("window 不能小于 0。")

    total_rings = len(layout.rings)
    coarse_counts = list(range(coarse_stride, total_rings + 1, coarse_stride))
    if not coarse_counts or coarse_counts[-1] != total_rings:
        coarse_counts.append(total_rings)

    evaluated: dict[int, FieldEvaluation] = {}

    def evaluate(ring_count: int) -> FieldEvaluation:
        previous = evaluated.get(ring_count)
        if previous is not None:
            return previous
        value = evaluate_coordinates(
            layout_kind=layout.kind,
            ring_count=ring_count,
            coordinates=layout.prefix(ring_count),
            parameters=parameters,
            profile=profile,
            cache=cache,
            base_field_config=base_field_config,
        )
        evaluated[ring_count] = value
        return value

    first_feasible: int | None = None
    previous_coarse = 0
    for ring_count in coarse_counts:
        value = evaluate(ring_count)
        if value.is_feasible(target_power_mw):
            for refined_count in range(previous_coarse + 1, ring_count + 1):
                refined = evaluate(refined_count)
                if refined.is_feasible(target_power_mw):
                    first_feasible = refined_count
                    break
            break
        previous_coarse = ring_count

    center = first_feasible if first_feasible is not None else total_rings
    local_counts = _unique_ring_counts(
        range(center - window, center + window + 1),
        total_rings,
    )
    for ring_count in local_counts:
        evaluate(ring_count)

    values = tuple(evaluated[key] for key in sorted(evaluated))
    best = values[0]
    for value in values[1:]:
        best = better_evaluation(
            best,
            value,
            target_power_mw=target_power_mw,
        )
    return ExtentScanResult(
        best=best,
        evaluations=values,
        first_feasible_ring_count=first_feasible,
    )


# ========================================================================
# 分散初值与循环变步长搜索
# ========================================================================

import math
from dataclasses import dataclass, replace
from typing import Callable, Generic, Iterable, TypeVar

from scipy.stats import qmc


ParametersT = TypeVar(
    "ParametersT",
    PartitionedRingParameters,
    CampoParameters,
)


@dataclass(frozen=True)
class Interval:
    lower: float
    upper: float

    def __post_init__(self) -> None:
        if not math.isfinite(self.lower) or not math.isfinite(self.upper):
            raise ValueError("参数边界必须为有限数。")
        if self.lower > self.upper:
            raise ValueError("参数下界不能大于上界。")

    def contains(self, value: float) -> bool:
        return self.lower <= value <= self.upper

    def map_unit(self, value: float) -> float:
        return self.lower + value * (self.upper - self.lower)


@dataclass(frozen=True)
class IntegerInterval:
    lower: int
    upper: int

    def __post_init__(self) -> None:
        if self.lower > self.upper:
            raise ValueError("整数参数下界不能大于上界。")

    def contains(self, value: int) -> bool:
        return self.lower <= value <= self.upper

    def map_unit(self, value: float) -> int:
        mapped = self.lower + value * (self.upper - self.lower)
        return min(self.upper, max(self.lower, int(round(mapped))))


@dataclass(frozen=True)
class PartitionedBounds:
    tower_y: Interval = Interval(-220.0, 0.0)
    mirror_width: Interval = Interval(4.5, 7.5)
    mirror_height: Interval = Interval(4.0, 7.5)
    installation_height: Interval = Interval(2.0, 6.0)
    split_radius: Interval = Interval(150.0, 300.0)
    near_spacing: Interval = Interval(8.0, 22.0)
    far_spacing: Interval = Interval(10.0, 32.0)


@dataclass(frozen=True)
class CampoBounds:
    tower_y: Interval = Interval(-220.0, 0.0)
    mirror_width: Interval = Interval(4.5, 7.5)
    mirror_height: Interval = Interval(4.0, 7.5)
    installation_height: Interval = Interval(2.0, 6.0)
    first_ring_count: IntegerInterval = IntegerInterval(44, 76)
    initial_spacing: Interval = Interval(8.0, 22.0)
    spacing_growth: Interval = Interval(0.0, 0.6)


@dataclass(frozen=True)
class StepLevel:
    steps: tuple[tuple[str, float], ...]

    def as_dict(self) -> dict[str, float]:
        return dict(self.steps)


PARTITIONED_STEP_LEVELS = (
    StepLevel(
        (
            ("tower_y", 20.0),
            ("mirror_width", 0.5),
            ("mirror_height", 0.5),
            ("installation_height", 0.5),
            ("split_radius", 20.0),
            ("near_spacing", 2.0),
            ("far_spacing", 2.0),
        )
    ),
    StepLevel(
        (
            ("tower_y", 10.0),
            ("mirror_width", 0.2),
            ("mirror_height", 0.2),
            ("installation_height", 0.2),
            ("split_radius", 10.0),
            ("near_spacing", 1.0),
            ("far_spacing", 1.0),
        )
    ),
    StepLevel(
        (
            ("tower_y", 5.0),
            ("mirror_width", 0.1),
            ("mirror_height", 0.1),
            ("installation_height", 0.1),
            ("split_radius", 5.0),
            ("near_spacing", 0.5),
            ("far_spacing", 0.5),
        )
    ),
)


CAMPO_STEP_LEVELS = (
    StepLevel(
        (
            ("tower_y", 20.0),
            ("mirror_width", 0.5),
            ("mirror_height", 0.5),
            ("installation_height", 0.5),
            ("first_ring_count", 4.0),
            ("initial_spacing", 2.0),
            ("spacing_growth", 0.1),
        )
    ),
    StepLevel(
        (
            ("tower_y", 10.0),
            ("mirror_width", 0.2),
            ("mirror_height", 0.2),
            ("installation_height", 0.2),
            ("first_ring_count", 2.0),
            ("initial_spacing", 1.0),
            ("spacing_growth", 0.05),
        )
    ),
    StepLevel(
        (
            ("tower_y", 5.0),
            ("mirror_width", 0.1),
            ("mirror_height", 0.1),
            ("installation_height", 0.1),
            ("first_ring_count", 1.0),
            ("initial_spacing", 0.5),
            ("spacing_growth", 0.02),
        )
    ),
)


@dataclass(frozen=True)
class SearchOutcome(Generic[ParametersT]):
    parameters: ParametersT
    extent_scan: ExtentScanResult

    @property
    def feasible(self) -> bool:
        return self.extent_scan.best.is_feasible()

    @property
    def annual_power_mw(self) -> float:
        return self.extent_scan.best.annual_power_mw

    @property
    def unit_area_power_kw_m2(self) -> float:
        return self.extent_scan.best.unit_area_power_kw_m2


@dataclass(frozen=True)
class SearchTrace(Generic[ParametersT]):
    step_level: int
    cycle: int
    block: str
    outcome: SearchOutcome[ParametersT]


@dataclass(frozen=True)
class OptimizationResult(Generic[ParametersT]):
    layout_kind: str
    best: SearchOutcome[ParametersT]
    starts: tuple[SearchOutcome[ParametersT], ...]
    local_results: tuple[SearchOutcome[ParametersT], ...]
    trace: tuple[SearchTrace[ParametersT], ...]


def _sobol_points(dimension: int, count: int, seed: int) -> list[list[float]]:
    if count < 1:
        raise ValueError("Sobol 样本数必须大于等于 1。")
    exponent = int(math.ceil(math.log2(count)))
    sampler = qmc.Sobol(d=dimension, scramble=True, seed=seed)
    return sampler.random_base2(exponent)[:count].tolist()


def sample_partitioned_parameters(
    count: int,
    *,
    seed: int = 2023,
    bounds: PartitionedBounds = PartitionedBounds(),
) -> tuple[PartitionedRingParameters, ...]:
    samples: list[PartitionedRingParameters] = []
    for point in _sobol_points(7, count, seed):
        width = bounds.mirror_width.map_unit(point[1])
        height_upper = min(bounds.mirror_height.upper, width)
        height_lower = min(bounds.mirror_height.lower, height_upper)
        height = height_lower + point[2] * (height_upper - height_lower)
        installation_lower = max(
            bounds.installation_height.lower,
            height / 2.0,
        )
        installation = installation_lower + point[3] * (
            bounds.installation_height.upper - installation_lower
        )
        near = bounds.near_spacing.map_unit(point[5])
        far_lower = max(bounds.far_spacing.lower, near)
        far = far_lower + point[6] * (bounds.far_spacing.upper - far_lower)
        samples.append(
            PartitionedRingParameters(
                tower_y=bounds.tower_y.map_unit(point[0]),
                mirror_width=width,
                mirror_height=height,
                installation_height=installation,
                split_radius=bounds.split_radius.map_unit(point[4]),
                near_spacing=near,
                far_spacing=far,
            )
        )
    return tuple(samples)


def sample_campo_parameters(
    count: int,
    *,
    seed: int = 2023,
    bounds: CampoBounds = CampoBounds(),
) -> tuple[CampoParameters, ...]:
    samples: list[CampoParameters] = []
    for point in _sobol_points(7, count, seed):
        width = bounds.mirror_width.map_unit(point[1])
        height_upper = min(bounds.mirror_height.upper, width)
        height_lower = min(bounds.mirror_height.lower, height_upper)
        height = height_lower + point[2] * (height_upper - height_lower)
        installation_lower = max(
            bounds.installation_height.lower,
            height / 2.0,
        )
        installation = installation_lower + point[3] * (
            bounds.installation_height.upper - installation_lower
        )
        samples.append(
            CampoParameters(
                tower_y=bounds.tower_y.map_unit(point[0]),
                mirror_width=width,
                mirror_height=height,
                installation_height=installation,
                first_ring_count=bounds.first_ring_count.map_unit(point[4]),
                initial_spacing=bounds.initial_spacing.map_unit(point[5]),
                spacing_growth=bounds.spacing_growth.map_unit(point[6]),
            )
        )
    return tuple(samples)


def _partitioned_within_bounds(
    parameters: PartitionedRingParameters,
    bounds: PartitionedBounds,
) -> bool:
    return (
        bounds.tower_y.contains(parameters.tower_y)
        and bounds.mirror_width.contains(parameters.mirror_width)
        and bounds.mirror_height.contains(parameters.mirror_height)
        and parameters.mirror_height <= parameters.mirror_width
        and bounds.installation_height.contains(parameters.installation_height)
        and parameters.installation_height >= parameters.mirror_height / 2.0
        and bounds.split_radius.contains(parameters.split_radius)
        and bounds.near_spacing.contains(parameters.near_spacing)
        and bounds.far_spacing.contains(parameters.far_spacing)
        and parameters.far_spacing >= parameters.near_spacing
    )


def _campo_within_bounds(
    parameters: CampoParameters,
    bounds: CampoBounds,
) -> bool:
    return (
        bounds.tower_y.contains(parameters.tower_y)
        and bounds.mirror_width.contains(parameters.mirror_width)
        and bounds.mirror_height.contains(parameters.mirror_height)
        and parameters.mirror_height <= parameters.mirror_width
        and bounds.installation_height.contains(parameters.installation_height)
        and parameters.installation_height >= parameters.mirror_height / 2.0
        and bounds.first_ring_count.contains(parameters.first_ring_count)
        and bounds.initial_spacing.contains(parameters.initial_spacing)
        and bounds.spacing_growth.contains(parameters.spacing_growth)
    )


def _better_outcome(
    left: SearchOutcome[ParametersT],
    right: SearchOutcome[ParametersT],
) -> SearchOutcome[ParametersT]:
    best_evaluation = better_evaluation(
        left.extent_scan.best,
        right.extent_scan.best,
    )
    return left if best_evaluation is left.extent_scan.best else right


def _rank_outcomes(
    outcomes: Iterable[SearchOutcome[ParametersT]],
) -> list[SearchOutcome[ParametersT]]:
    ranked: list[SearchOutcome[ParametersT]] = []
    for outcome in outcomes:
        inserted = False
        for index, current in enumerate(ranked):
            if _better_outcome(outcome, current) is outcome:
                ranked.insert(index, outcome)
                inserted = True
                break
        if not inserted:
            ranked.append(outcome)
    return ranked


def evaluate_partitioned_parameters(
    parameters: PartitionedRingParameters,
    profile: EvaluationProfile,
    *,
    cache: EvaluationCache | None = None,
    coarse_stride: int = 4,
    window: int = 2,
) -> SearchOutcome[PartitionedRingParameters]:
    layout = generate_partitioned_layout(parameters)
    scan = scan_layout_extents(
        layout,
        parameters,
        profile,
        coarse_stride=coarse_stride,
        window=window,
        cache=cache,
    )
    return SearchOutcome(parameters, scan)


def evaluate_campo_parameters(
    parameters: CampoParameters,
    profile: EvaluationProfile,
    *,
    cache: EvaluationCache | None = None,
    coarse_stride: int = 4,
    window: int = 2,
) -> SearchOutcome[CampoParameters]:
    layout = generate_campo_layout(parameters)
    scan = scan_layout_extents(
        layout,
        parameters,
        profile,
        coarse_stride=coarse_stride,
        window=window,
        cache=cache,
    )
    return SearchOutcome(parameters, scan)


def _coordinate_candidates(
    parameters: ParametersT,
    fields: tuple[str, ...],
    steps: dict[str, float],
) -> tuple[ParametersT, ...]:
    candidates: list[ParametersT] = []
    for field in fields:
        step = steps[field]
        value = getattr(parameters, field)
        for direction in (-1.0, 1.0):
            new_value: float | int = value + direction * step
            if isinstance(value, int):
                new_value = int(round(new_value))
            candidates.append(replace(parameters, **{field: new_value}))
    return tuple(candidates)


def _cyclic_search(
    initial: SearchOutcome[ParametersT],
    *,
    evaluator: Callable[[ParametersT], SearchOutcome[ParametersT]],
    within_bounds: Callable[[ParametersT], bool],
    blocks: tuple[tuple[str, tuple[str, ...]], ...],
    step_levels: tuple[StepLevel, ...],
    maximum_cycles_per_level: int,
) -> tuple[SearchOutcome[ParametersT], tuple[SearchTrace[ParametersT], ...]]:
    current = initial
    trace: list[SearchTrace[ParametersT]] = []

    for level_index, level in enumerate(step_levels):
        steps = level.as_dict()
        for cycle in range(maximum_cycles_per_level):
            cycle_improved = False
            for block_name, fields in blocks:
                block_best = current
                for candidate in _coordinate_candidates(
                    current.parameters,
                    fields,
                    steps,
                ):
                    if not within_bounds(candidate):
                        continue
                    try:
                        outcome = evaluator(candidate)
                    except LayoutError:
                        continue
                    block_best = _better_outcome(block_best, outcome)
                if block_best is not current:
                    current = block_best
                    cycle_improved = True
                    trace.append(
                        SearchTrace(
                            step_level=level_index,
                            cycle=cycle,
                            block=block_name,
                            outcome=current,
                        )
                    )
            if not cycle_improved:
                break
    return current, tuple(trace)


def optimize_partitioned(
    *,
    profile: EvaluationProfile,
    initial_sample_count: int = 16,
    retained_starts: int = 3,
    seed: int = 2023,
    bounds: PartitionedBounds = PartitionedBounds(),
    step_levels: tuple[StepLevel, ...] = PARTITIONED_STEP_LEVELS,
    maximum_cycles_per_level: int = 4,
    coarse_stride: int = 4,
    window: int = 2,
    cache: EvaluationCache | None = None,
    progress: Callable[[str], None] | None = None,
) -> OptimizationResult[PartitionedRingParameters]:
    cache = cache or EvaluationCache()
    outcomes: list[SearchOutcome[PartitionedRingParameters]] = []
    samples = sample_partitioned_parameters(
        initial_sample_count,
        seed=seed,
        bounds=bounds,
    )
    for index, parameters in enumerate(samples, start=1):
        try:
            outcome = evaluate_partitioned_parameters(
                parameters,
                profile,
                cache=cache,
                coarse_stride=coarse_stride,
                window=window,
            )
            outcomes.append(outcome)
            if progress is not None:
                progress(
                    f"方案 A 初值 {index}/{len(samples)}："
                    f"P={outcome.annual_power_mw:.4f} MW，"
                    f"q={outcome.unit_area_power_kw_m2:.6f}"
                )
        except LayoutError as exc:
            if progress is not None:
                progress(f"方案 A 初值 {index}/{len(samples)}：几何不可行（{exc}）")
            continue
    if not outcomes:
        raise RuntimeError("分区圆环的分散初值中没有几何合法方案。")

    starts = tuple(_rank_outcomes(outcomes)[:retained_starts])
    local_results: list[SearchOutcome[PartitionedRingParameters]] = []
    trace: list[SearchTrace[PartitionedRingParameters]] = []

    def evaluator(
        parameters: PartitionedRingParameters,
    ) -> SearchOutcome[PartitionedRingParameters]:
        return evaluate_partitioned_parameters(
            parameters,
            profile,
            cache=cache,
            coarse_stride=coarse_stride,
            window=window,
        )

    blocks = (
        ("tower", ("tower_y",)),
        (
            "mirror",
            ("mirror_width", "mirror_height", "installation_height"),
        ),
        ("layout", ("split_radius", "near_spacing", "far_spacing")),
    )
    for index, start in enumerate(starts, start=1):
        if progress is not None:
            progress(f"方案 A 开始局部搜索 {index}/{len(starts)}")
        local, local_trace = _cyclic_search(
            start,
            evaluator=evaluator,
            within_bounds=lambda value: _partitioned_within_bounds(
                value,
                bounds,
            ),
            blocks=blocks,
            step_levels=step_levels,
            maximum_cycles_per_level=maximum_cycles_per_level,
        )
        local_results.append(local)
        trace.extend(local_trace)
    best = _rank_outcomes(local_results)[0]
    return OptimizationResult(
        "partitioned",
        best,
        starts,
        tuple(local_results),
        tuple(trace),
    )


def refine_partitioned(
    initial_parameters: PartitionedRingParameters,
    *,
    profile: EvaluationProfile,
    bounds: PartitionedBounds = PartitionedBounds(),
    step_levels: tuple[StepLevel, ...] = PARTITIONED_STEP_LEVELS,
    maximum_cycles_per_level: int = 4,
    coarse_stride: int = 4,
    window: int = 2,
    cache: EvaluationCache | None = None,
    progress: Callable[[str], None] | None = None,
) -> OptimizationResult[PartitionedRingParameters]:
    """从已落盘的方案 A 参数继续循环变步长搜索。"""

    if not _partitioned_within_bounds(initial_parameters, bounds):
        raise ValueError("方案 A 的恢复参数超出当前搜索边界。")
    cache = cache or EvaluationCache()

    def evaluator(
        parameters: PartitionedRingParameters,
    ) -> SearchOutcome[PartitionedRingParameters]:
        return evaluate_partitioned_parameters(
            parameters,
            profile,
            cache=cache,
            coarse_stride=coarse_stride,
            window=window,
        )

    initial = evaluator(initial_parameters)
    if progress is not None:
        progress(
            "方案 A 恢复起点："
            f"P={initial.annual_power_mw:.4f} MW，"
            f"q={initial.unit_area_power_kw_m2:.6f}"
        )
    blocks = (
        ("tower", ("tower_y",)),
        (
            "mirror",
            ("mirror_width", "mirror_height", "installation_height"),
        ),
        ("layout", ("split_radius", "near_spacing", "far_spacing")),
    )
    best, trace = _cyclic_search(
        initial,
        evaluator=evaluator,
        within_bounds=lambda value: _partitioned_within_bounds(value, bounds),
        blocks=blocks,
        step_levels=step_levels,
        maximum_cycles_per_level=maximum_cycles_per_level,
    )
    return OptimizationResult(
        "partitioned",
        best,
        (initial,),
        (best,),
        trace,
    )


def optimize_campo(
    *,
    profile: EvaluationProfile,
    initial_sample_count: int = 16,
    retained_starts: int = 3,
    seed: int = 2023,
    bounds: CampoBounds = CampoBounds(),
    step_levels: tuple[StepLevel, ...] = CAMPO_STEP_LEVELS,
    maximum_cycles_per_level: int = 4,
    coarse_stride: int = 4,
    window: int = 2,
    cache: EvaluationCache | None = None,
    progress: Callable[[str], None] | None = None,
) -> OptimizationResult[CampoParameters]:
    cache = cache or EvaluationCache()
    outcomes: list[SearchOutcome[CampoParameters]] = []
    samples = sample_campo_parameters(
        initial_sample_count,
        seed=seed,
        bounds=bounds,
    )
    for index, parameters in enumerate(samples, start=1):
        try:
            outcome = evaluate_campo_parameters(
                parameters,
                profile,
                cache=cache,
                coarse_stride=coarse_stride,
                window=window,
            )
            outcomes.append(outcome)
            if progress is not None:
                progress(
                    f"方案 B 初值 {index}/{len(samples)}："
                    f"P={outcome.annual_power_mw:.4f} MW，"
                    f"q={outcome.unit_area_power_kw_m2:.6f}"
                )
        except LayoutError as exc:
            if progress is not None:
                progress(f"方案 B 初值 {index}/{len(samples)}：几何不可行（{exc}）")
            continue
    if not outcomes:
        raise RuntimeError("Campo 的分散初值中没有几何合法方案。")

    starts = tuple(_rank_outcomes(outcomes)[:retained_starts])
    local_results: list[SearchOutcome[CampoParameters]] = []
    trace: list[SearchTrace[CampoParameters]] = []

    def evaluator(
        parameters: CampoParameters,
    ) -> SearchOutcome[CampoParameters]:
        return evaluate_campo_parameters(
            parameters,
            profile,
            cache=cache,
            coarse_stride=coarse_stride,
            window=window,
        )

    blocks = (
        ("tower", ("tower_y",)),
        (
            "mirror",
            ("mirror_width", "mirror_height", "installation_height"),
        ),
        ("layout", ("first_ring_count", "initial_spacing", "spacing_growth")),
    )
    for start in starts:
        local, local_trace = _cyclic_search(
            start,
            evaluator=evaluator,
            within_bounds=lambda value: _campo_within_bounds(value, bounds),
            blocks=blocks,
            step_levels=step_levels,
            maximum_cycles_per_level=maximum_cycles_per_level,
        )
        local_results.append(local)
        trace.extend(local_trace)
    best = _rank_outcomes(local_results)[0]
    return OptimizationResult(
        "campo",
        best,
        starts,
        tuple(local_results),
        tuple(trace),
    )


def refine_campo(
    initial_parameters: CampoParameters,
    *,
    profile: EvaluationProfile,
    bounds: CampoBounds = CampoBounds(),
    step_levels: tuple[StepLevel, ...] = CAMPO_STEP_LEVELS,
    maximum_cycles_per_level: int = 4,
    coarse_stride: int = 4,
    window: int = 2,
    cache: EvaluationCache | None = None,
    progress: Callable[[str], None] | None = None,
) -> OptimizationResult[CampoParameters]:
    """从已落盘的方案 B 参数继续循环变步长搜索。"""

    if not _campo_within_bounds(initial_parameters, bounds):
        raise ValueError("方案 B 的恢复参数超出当前搜索边界。")
    cache = cache or EvaluationCache()

    def evaluator(
        parameters: CampoParameters,
    ) -> SearchOutcome[CampoParameters]:
        return evaluate_campo_parameters(
            parameters,
            profile,
            cache=cache,
            coarse_stride=coarse_stride,
            window=window,
        )

    initial = evaluator(initial_parameters)
    if progress is not None:
        progress(
            "方案 B 恢复起点："
            f"P={initial.annual_power_mw:.4f} MW，"
            f"q={initial.unit_area_power_kw_m2:.6f}"
        )
    blocks = (
        ("tower", ("tower_y",)),
        (
            "mirror",
            ("mirror_width", "mirror_height", "installation_height"),
        ),
        ("layout", ("first_ring_count", "initial_spacing", "spacing_growth")),
    )
    best, trace = _cyclic_search(
        initial,
        evaluator=evaluator,
        within_bounds=lambda value: _campo_within_bounds(value, bounds),
        blocks=blocks,
        step_levels=step_levels,
        maximum_cycles_per_level=maximum_cycles_per_level,
    )
    return OptimizationResult(
        "campo",
        best,
        (initial,),
        (best,),
        trace,
    )


# ========================================================================
# 胜出布局的外层结构化修剪
# ========================================================================

from dataclasses import dataclass

import numpy as np


@dataclass(frozen=True)
class PruneStep:
    removed_indices: tuple[int, int]
    evaluation: FieldEvaluation


@dataclass(frozen=True)
class PruneResult:
    initial: FieldEvaluation
    best: FieldEvaluation
    steps: tuple[PruneStep, ...]


def _outer_symmetric_pairs(
    layout: GeneratedLayout,
    ring_count: int,
    *,
    ring_depth: int,
    tolerance: float = 1e-7,
) -> tuple[tuple[int, int], ...]:
    if ring_depth < 1:
        raise ValueError("ring_depth 必须大于等于 1。")
    selected_start = max(0, ring_count - ring_depth)
    offsets: list[int] = []
    running = 0
    for ring in layout.rings[:ring_count]:
        offsets.append(running)
        running += ring.mirror_count

    pairs: list[tuple[int, int]] = []
    for ring_index in range(selected_start, ring_count):
        ring = layout.rings[ring_index]
        offset = offsets[ring_index]
        coordinates = ring.coordinates
        unused = set(range(coordinates.shape[0]))
        while unused:
            local = min(unused)
            unused.remove(local)
            x, y = coordinates[local]
            if abs(float(x)) <= tolerance:
                continue
            candidates = [
                other
                for other in unused
                if abs(float(coordinates[other, 0] + x)) <= tolerance
                and abs(float(coordinates[other, 1] - y)) <= tolerance
            ]
            if not candidates:
                continue
            partner = min(candidates)
            unused.remove(partner)
            pairs.append((offset + local, offset + partner))
    return tuple(pairs)


def prune_outer_symmetric_pairs(
    *,
    layout: GeneratedLayout,
    parameters: LayoutParameters,
    initial: FieldEvaluation,
    profile: EvaluationProfile,
    target_power_mw: float = 42.0,
    ring_depth: int = 2,
    maximum_rounds: int = 10,
    maximum_pairs_per_round: int | None = None,
    cache: EvaluationCache | None = None,
) -> PruneResult:
    """逐轮全场复算，仅接受保持可行且提高单位面积功率的删镜。"""

    if maximum_rounds < 0:
        raise ValueError("maximum_rounds 不能小于 0。")
    if not initial.is_feasible(target_power_mw):
        raise ValueError("结构化修剪要求初始镜场已经满足功率约束。")

    original = layout.prefix(initial.ring_count)
    if original.shape != initial.coordinates.shape or not np.allclose(
        original,
        initial.coordinates,
        atol=1e-9,
    ):
        raise ValueError("initial 坐标与指定布局前缀不一致。")

    pairs = _outer_symmetric_pairs(
        layout,
        initial.ring_count,
        ring_depth=ring_depth,
    )
    active = np.ones(original.shape[0], dtype=bool)
    current = initial
    steps: list[PruneStep] = []

    for _ in range(maximum_rounds):
        remaining_pairs = [
            pair for pair in pairs if active[pair[0]] and active[pair[1]]
        ]
        if (
            maximum_pairs_per_round is not None
            and len(remaining_pairs) > maximum_pairs_per_round
        ):
            sampled_indices = np.linspace(
                0,
                len(remaining_pairs) - 1,
                maximum_pairs_per_round,
                dtype=int,
            )
            remaining_pairs = [remaining_pairs[index] for index in sampled_indices]
        best_pair: tuple[int, int] | None = None
        best_evaluation: FieldEvaluation | None = None

        for pair in remaining_pairs:
            candidate_active = active.copy()
            candidate_active[list(pair)] = False
            candidate = evaluate_coordinates(
                layout_kind=layout.kind,
                ring_count=initial.ring_count,
                coordinates=original[candidate_active],
                parameters=parameters,
                profile=profile,
                cache=cache,
            )
            if not candidate.is_feasible(target_power_mw):
                continue
            if candidate.unit_area_power_kw_m2 <= current.unit_area_power_kw_m2:
                continue
            if (
                best_evaluation is None
                or candidate.unit_area_power_kw_m2
                > best_evaluation.unit_area_power_kw_m2
            ):
                best_pair = pair
                best_evaluation = candidate

        if best_pair is None or best_evaluation is None:
            break
        active[list(best_pair)] = False
        current = best_evaluation
        steps.append(PruneStep(best_pair, current))

    return PruneResult(initial, current, tuple(steps))


# ========================================================================
# 结果、论文表格与 Excel 输出
# ========================================================================

import csv
import json
from copy import copy
from dataclasses import asdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


TARGET_ANNUAL_POWER_MW = 42.0


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        raise ValueError(f"没有可写入 {path.name} 的结果。")
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=list(rows[0]),
            lineterminator="\n",
        )
        writer.writeheader()
        writer.writerows(rows)


def write_result2_workbook(
    *,
    template_path: str | Path,
    output_path: str | Path,
    evaluation: FieldEvaluation,
    parameters: LayoutParameters,
) -> Path:
    """按题目模板写出塔坐标、统一尺寸、高度和全部镜位。"""

    template = Path(template_path)
    if not template.exists():
        raise FileNotFoundError(f"找不到 result2.xlsx 模板：{template}")
    destination = Path(output_path)
    destination.parent.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(template)
    sheet = workbook.active
    if sheet.max_column < 8:
        workbook.close()
        raise ValueError("result2.xlsx 模板列数不足 8 列。")

    style_source = [copy(sheet.cell(2, column)._style) for column in range(1, 9)]
    number_formats = [sheet.cell(2, column).number_format for column in range(1, 9)]
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)

    for row_index, (x_m, y_m) in enumerate(
        evaluation.coordinates,
        start=2,
    ):
        values = (
            parameters.tower_x,
            parameters.tower_y,
            row_index - 1,
            parameters.mirror_width,
            parameters.mirror_height,
            float(x_m),
            float(y_m),
            parameters.installation_height,
        )
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column, value)
            cell._style = copy(style_source[column - 1])
            cell.number_format = number_formats[column - 1]

    workbook.save(destination)
    workbook.close()
    return destination


def write_question2_results(
    *,
    output_dir: str | Path,
    layout_name: str,
    parameters: LayoutParameters,
    evaluation: FieldEvaluation,
    result2_template: str | Path,
    comparison: dict[str, Any] | None = None,
) -> dict[str, Path]:
    """写出坐标、月年平均结果、配置摘要、论文表和 result2.xlsx。"""

    destination = Path(output_dir)
    destination.mkdir(parents=True, exist_ok=True)

    coordinate_rows = [
        {
            "mirror_id": index + 1,
            "mirror_width_m": parameters.mirror_width,
            "mirror_height_m": parameters.mirror_height,
            "x_m": float(x_m),
            "y_m": float(y_m),
            "z_m": parameters.installation_height,
        }
        for index, (x_m, y_m) in enumerate(evaluation.coordinates)
    ]
    monthly_rows = [asdict(record) for record in evaluation.solution.monthly_results]
    mirror_rows = [
        asdict(record) for record in evaluation.solution.mirror_annual_results
    ]
    annual = asdict(evaluation.solution.annual_result)

    coordinates_path = destination / "03_最终镜位坐标.csv"
    monthly_path = destination / "04_月平均计算结果.csv"
    annual_path = destination / "05_年平均计算结果.json"
    mirror_path = destination / "06_单镜年平均结果.csv"
    summary_path = destination / "07_最终方案摘要.json"
    table_path = destination / "08_论文结果与验证表.md"
    workbook_path = destination / "result2.xlsx"

    _write_csv(coordinates_path, coordinate_rows)
    _write_csv(monthly_path, monthly_rows)
    _write_csv(mirror_path, mirror_rows)
    annual_path.write_text(
        json.dumps(annual, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    summary = {
        "layout": layout_name,
        "annual_power_constraint_mw": TARGET_ANNUAL_POWER_MW,
        "annual_power_margin_mw": (evaluation.annual_power_mw - TARGET_ANNUAL_POWER_MW),
        "constraint_satisfied": (evaluation.annual_power_mw >= TARGET_ANNUAL_POWER_MW),
        "parameters": asdict(parameters),
        "ring_count": evaluation.ring_count,
        "mirror_count": evaluation.mirror_count,
        "mirror_area_m2": evaluation.mirror_area_m2,
        "total_area_m2": evaluation.total_area_m2,
        "annual": annual,
    }
    summary_path.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    lines = [
        "# 第二问结果与验证表",
        "",
        "## 表 1 功率约束与优化目标",
        "",
        "| 年平均输出热功率下限 (MW) | 最终年平均输出热功率 (MW) | 功率余量 (MW) | 是否满足约束 | 单位镜面面积年平均输出热功率 (kW/m²) |",
        "| ---: | ---: | ---: | :---: | ---: |",
        (
            f"| {TARGET_ANNUAL_POWER_MW:.6f} "
            f"| {evaluation.annual_power_mw:.6f} "
            f"| {evaluation.annual_power_mw - TARGET_ANNUAL_POWER_MW:.6f} "
            f"| {'是' if evaluation.annual_power_mw >= TARGET_ANNUAL_POWER_MW else '否'} "
            f"| {evaluation.unit_area_power_kw_m2:.6f} |"
        ),
        "",
        "> 本题中的 42 MW 是年平均输出热功率下限；优化目标是在满足该下限后最大化单位镜面面积年平均输出热功率。",
        "",
        "## 表 2 最终设计参数",
        "",
        "| 布局 | 塔坐标 | 镜面尺寸 | 安装高度 | 镜子数 | 总镜面面积 (m²) | 年平均功率 (MW) | 单位面积功率 (kW/m²) |",
        "| --- | --- | ---: | ---: | ---: | ---: | ---: | ---: |",
        (
            f"| {layout_name} "
            f"| ({parameters.tower_x:.3f}, {parameters.tower_y:.3f}) "
            f"| {parameters.mirror_width:.3f}×{parameters.mirror_height:.3f} "
            f"| {parameters.installation_height:.3f} "
            f"| {evaluation.mirror_count} "
            f"| {evaluation.total_area_m2:.3f} "
            f"| {evaluation.annual_power_mw:.6f} "
            f"| {evaluation.unit_area_power_kw_m2:.6f} |"
        ),
    ]
    if comparison is not None and {
        "partitioned",
        "campo",
    }.issubset(comparison):
        lines.extend(
            [
                "",
                "## 表 3 两种候选布局的正式精度对比",
                "",
                "| 布局 | 安全余量 (m) | 镜子数 | 总镜面面积 (m²) | 年平均功率 (MW) | 单位面积功率 (kW/m²) |",
                "| --- | ---: | ---: | ---: | ---: | ---: |",
            ]
        )
        for kind, label in (
            ("partitioned", "分区交错同心圆"),
            ("campo", "改进 Campo"),
        ):
            record = comparison[kind]
            lines.append(
                f"| {label} "
                f"| {record['parameters']['safety_epsilon']:.6f} "
                f"| {record['mirror_count']} "
                f"| {record['total_area_m2']:.3f} "
                f"| {record['annual_power_mw']:.6f} "
                f"| {record['unit_area_power_kw_m2']:.6f} |"
            )

    geometry = validate_layout(evaluation.coordinates, parameters)
    lines.extend(
        [
            "",
            "## 表 4 几何约束复核",
            "",
            "| 检查项 | 实际值 | 约束 | 结果 |",
            "| --- | ---: | ---: | :---: |",
            (
                "| 最小镜心距离 (m) "
                f"| {geometry.minimum_center_distance:.9f} "
                f"| > {parameters.mirror_width + 5.0:.9f} "
                f"| {'通过' if geometry.valid else '未通过'} |"
            ),
            (
                "| 镜心距离安全余量 (m) "
                f"| {geometry.minimum_center_distance - parameters.mirror_width - 5.0:.9f} "
                "| > 0 | "
                f"{'通过' if geometry.minimum_center_distance > parameters.mirror_width + 5.0 else '未通过'} |"
            ),
            (
                "| 最大场地半径 (m) "
                f"| {geometry.maximum_field_radius:.6f} "
                f"| ≤ {parameters.field_radius:.3f} "
                f"| {'通过' if geometry.maximum_field_radius <= parameters.field_radius + 1e-9 else '未通过'} |"
            ),
            (
                "| 最小塔距 (m) "
                f"| {geometry.minimum_tower_distance:.6f} "
                f"| ≥ {parameters.exclusion_radius:.3f} "
                f"| {'通过' if geometry.minimum_tower_distance >= parameters.exclusion_radius - 1e-9 else '未通过'} |"
            ),
            (
                "| 不触地高度余量 (m) "
                f"| {parameters.installation_height - parameters.mirror_height / 2.0:.6f} "
                "| ≥ 0 "
                f"| {'通过' if parameters.installation_height >= parameters.mirror_height / 2.0 else '未通过'} |"
            ),
            "",
            "## 表 5 月平均光学效率及输出热功率",
            "",
            "| 月份 | 光学效率 | 余弦效率 | 阴影遮挡效率 | 截断效率 | 输出热功率 (MW) | 单位面积功率 (kW/m²) |",
            "| ---: | ---: | ---: | ---: | ---: | ---: | ---: |",
        ]
    )
    for record in evaluation.solution.monthly_results:
        lines.append(
            f"| {record.month} "
            f"| {record.average_optical_efficiency:.6f} "
            f"| {record.average_cosine_efficiency:.6f} "
            f"| {record.average_shadow_blocking_efficiency:.6f} "
            f"| {record.average_truncation_efficiency:.6f} "
            f"| {record.field_output_mw:.6f} "
            f"| {record.unit_area_output_kw_m2:.6f} |"
        )
    table_path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    write_result2_workbook(
        template_path=result2_template,
        output_path=workbook_path,
        evaluation=evaluation,
        parameters=parameters,
    )
    return {
        "coordinates": coordinates_path,
        "monthly": monthly_path,
        "annual": annual_path,
        "mirror_annual": mirror_path,
        "summary": summary_path,
        "paper_table": table_path,
        "result2": workbook_path,
    }


def write_high_precision_validation(
    *,
    output_dir: str | Path,
    evaluation: FieldEvaluation,
    profile: EvaluationProfile,
) -> Path:
    """写出并追加 20×20、512 条光线的高精度可行性复核。"""

    destination = Path(output_dir)
    validation_path = destination / "09_高精度加密验证.json"
    annual = evaluation.solution.annual_result
    payload = {
        "profile": {
            "months": len(profile.months),
            "solar_times_per_month": len(profile.solar_times),
            "shadow_grid_size": profile.solver.shadow_grid_size,
            "truncation_rays": profile.solver.truncation_rays,
            "neighbor_radius_m": profile.solver.neighbor_radius_m,
        },
        "mirror_count": evaluation.mirror_count,
        "annual_power_constraint_mw": TARGET_ANNUAL_POWER_MW,
        "annual_power_mw": evaluation.annual_power_mw,
        "annual_power_margin_mw": (evaluation.annual_power_mw - TARGET_ANNUAL_POWER_MW),
        "unit_area_power_kw_m2": evaluation.unit_area_power_kw_m2,
        "average_optical_efficiency": annual.average_optical_efficiency,
        "average_shadow_blocking_efficiency": (
            annual.average_shadow_blocking_efficiency
        ),
        "average_truncation_efficiency": (annual.average_truncation_efficiency),
        "constraint_satisfied": evaluation.is_feasible(),
    }
    validation_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    table_path = destination / "08_论文结果与验证表.md"
    lines = [
        "",
        "## 表 6 高精度加密验证",
        "",
        "| 阴影网格 | 截断光线 | 邻镜半径 (m) | 年平均功率 (MW) | 功率余量 (MW) | 单位面积功率 (kW/m²) | 是否满足约束 |",
        "| ---: | ---: | ---: | ---: | ---: | ---: | :---: |",
        (
            f"| {profile.solver.shadow_grid_size}×{profile.solver.shadow_grid_size} "
            f"| {profile.solver.truncation_rays} "
            f"| {profile.solver.neighbor_radius_m:.0f} "
            f"| {evaluation.annual_power_mw:.6f} "
            f"| {evaluation.annual_power_mw - TARGET_ANNUAL_POWER_MW:.6f} "
            f"| {evaluation.unit_area_power_kw_m2:.6f} "
            f"| {'是' if evaluation.is_feasible() else '否'} |"
        ),
    ]
    with table_path.open("a", encoding="utf-8") as handle:
        handle.write("\n".join(lines) + "\n")
    return validation_path


# ========================================================================
# 四张正式结果图
# ========================================================================

import csv
import json
from dataclasses import replace
from pathlib import Path

import matplotlib
import matplotlib.pyplot as plt
import numpy as np
from matplotlib import colors
from matplotlib.lines import Line2D
from matplotlib.patches import Circle


PARTITIONED_COLOR = "#2563EB"
CAMPO_COLOR = "#D97706"
TARGET_COLOR = "#C2413B"
GRID_COLOR = "#D9DEE7"
TEXT_COLOR = "#172033"
RECEIVER_COLOR = "#C2410C"
RAY_COLOR = "#E76F51"
POWER_CMAP = "viridis"


def configure_matplotlib() -> None:
    font_candidates = (
        Path("/System/Library/Fonts/STHeiti Medium.ttc"),
        Path("/System/Library/Fonts/PingFang.ttc"),
        Path("C:/Windows/Fonts/msyh.ttc"),
        Path("C:/Windows/Fonts/simhei.ttf"),
        Path("/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"),
        Path("/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc"),
    )
    for font_path in font_candidates:
        if not font_path.exists():
            continue
        matplotlib.font_manager.fontManager.addfont(font_path)
        font_name = matplotlib.font_manager.FontProperties(
            fname=font_path
        ).get_name()
        plt.rcParams["font.family"] = font_name
        break
    plt.rcParams.update(
        {
            "font.sans-serif": [
                "PingFang SC",
                "Microsoft YaHei",
                "SimHei",
                "Noto Sans CJK SC",
                "WenQuanYi Zen Hei",
                "DejaVu Sans",
            ],
            "axes.unicode_minus": False,
            "figure.facecolor": "white",
            "axes.facecolor": "white",
            "axes.edgecolor": "#8A93A3",
            "axes.labelcolor": TEXT_COLOR,
            "axes.titlecolor": TEXT_COLOR,
            "xtick.color": "#465166",
            "ytick.color": "#465166",
            "text.color": TEXT_COLOR,
            "grid.color": GRID_COLOR,
            "grid.linewidth": 0.7,
            "grid.alpha": 0.72,
            "savefig.facecolor": "white",
            "savefig.bbox": "tight",
        }
    )


def load_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def load_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def build_partitioned_result(comparison: dict):
    record = comparison["partitioned"]
    parameters = PartitionedRingParameters(**record["parameters"])
    layout = generate_partitioned_layout(parameters)
    coordinates = layout.prefix(int(record["ring_count"]))
    if coordinates.shape[0] != int(record["mirror_count"]):
        raise RuntimeError(
            "分区交错同心圆镜数不一致："
            f"{coordinates.shape[0]} != {record['mirror_count']}"
        )
    evaluation = evaluate_coordinates(
        layout_kind="partitioned",
        ring_count=int(record["ring_count"]),
        coordinates=coordinates,
        parameters=parameters,
        profile=final_profile(),
    )
    if abs(evaluation.annual_power_mw - float(record["annual_power_mw"])) > 1e-9:
        raise RuntimeError("分区交错同心圆正式精度复算未复现 final 结果。")
    return parameters, coordinates, evaluation


def build_campo_result(comparison: dict, output_dir: Path):
    record = comparison["campo"]
    parameters = CampoParameters(**record["parameters"])
    coordinate_rows = load_csv(output_dir / "03_最终镜位坐标.csv")
    coordinates = np.asarray(
        [[float(row["x_m"]), float(row["y_m"])] for row in coordinate_rows],
        dtype=float,
    )
    evaluation = evaluate_coordinates(
        layout_kind="campo",
        ring_count=int(record["ring_count"]),
        coordinates=coordinates,
        parameters=parameters,
        profile=final_profile(),
    )
    if coordinates.shape[0] != int(record["mirror_count"]):
        raise RuntimeError("Campo 坐标镜数与正式摘要不一致。")
    if abs(evaluation.annual_power_mw - float(record["annual_power_mw"])) > 1e-8:
        raise RuntimeError("Campo 正式精度复算未复现交付结果。")
    return parameters, coordinates, evaluation


def add_layout_frame(
    ax,
    *,
    tower_x: float,
    tower_y: float,
    field_radius: float,
    exclusion_radius: float,
) -> None:
    ax.add_patch(
        Circle(
            (0.0, 0.0),
            field_radius,
            fill=False,
            color="#111827",
            linewidth=1.7,
            zorder=4,
        )
    )
    ax.add_patch(
        Circle(
            (tower_x, tower_y),
            exclusion_radius,
            fill=False,
            color=TARGET_COLOR,
            linestyle="--",
            linewidth=1.4,
            zorder=4,
        )
    )
    ax.scatter(
        [tower_x],
        [tower_y],
        marker="*",
        s=210,
        color=TARGET_COLOR,
        edgecolor="white",
        linewidth=0.8,
        zorder=6,
    )
    ax.set_xlim(-370, 370)
    ax.set_ylim(-370, 370)
    ax.set_aspect("equal", adjustable="box")
    ax.set_xlabel("x / East (m)")
    ax.set_ylabel("y / North (m)")
    ax.grid(True)


def plot_layout_comparison(
    *,
    figure_dir: Path,
    comparison: dict,
    partitioned_parameters: PartitionedRingParameters,
    partitioned_coordinates: np.ndarray,
    partitioned_powers: np.ndarray,
    campo_parameters: CampoParameters,
    campo_coordinates: np.ndarray,
    campo_powers: np.ndarray,
) -> Path:
    all_powers = np.concatenate((partitioned_powers, campo_powers))
    norm = colors.Normalize(
        vmin=float(np.percentile(all_powers, 1.0)),
        vmax=float(np.percentile(all_powers, 99.0)),
    )
    figure, axes = plt.subplots(1, 2, figsize=(15.8, 7.4), constrained_layout=True)
    layouts = (
        (
            axes[0],
            "方案A：分区交错同心圆",
            comparison["partitioned"],
            partitioned_parameters,
            partitioned_coordinates,
            partitioned_powers,
        ),
        (
            axes[1],
            "方案B：改进 Campo",
            comparison["campo"],
            campo_parameters,
            campo_coordinates,
            campo_powers,
        ),
    )
    scatter = None
    for ax, title, record, parameters, coordinates, powers in layouts:
        scatter = ax.scatter(
            coordinates[:, 0],
            coordinates[:, 1],
            c=powers,
            cmap=POWER_CMAP,
            norm=norm,
            s=8,
            linewidths=0,
            rasterized=True,
            zorder=3,
        )
        add_layout_frame(
            ax,
            tower_x=parameters.tower_x,
            tower_y=parameters.tower_y,
            field_radius=parameters.field_radius,
            exclusion_radius=parameters.exclusion_radius,
        )
        ax.set_title(
            f"{title}\n"
            f"N = {record['mirror_count']}，"
            f"q = {record['unit_area_power_kw_m2']:.6f} kW/m²",
            fontsize=13,
            fontweight="bold",
            pad=10,
        )
        ax.text(
            0.02,
            0.025,
            f"年平均功率 {record['annual_power_mw']:.6f} MW\n"
            f"总镜面面积 {record['total_area_m2']:.3f} m²",
            transform=ax.transAxes,
            fontsize=9.3,
            va="bottom",
            ha="left",
            bbox={
                "boxstyle": "round,pad=0.4",
                "facecolor": "white",
                "edgecolor": "#CBD2DF",
                "alpha": 0.92,
            },
        )
    legend_handles = [
        Line2D(
            [0],
            [0],
            marker="*",
            color="none",
            markerfacecolor=TARGET_COLOR,
            markeredgecolor="white",
            markersize=13,
            label="吸收塔",
        ),
        Line2D(
            [0],
            [0],
            color=TARGET_COLOR,
            linestyle="--",
            linewidth=1.4,
            label="塔周禁布边界（100 m）",
        ),
        Line2D(
            [0],
            [0],
            color="#111827",
            linewidth=1.7,
            label="镜场边界（350 m）",
        ),
    ]
    axes[0].legend(handles=legend_handles, loc="upper right", fontsize=8.8)
    axes[1].legend(handles=legend_handles, loc="upper right", fontsize=8.8)
    if scatter is not None:
        colorbar = figure.colorbar(
            scatter,
            ax=axes,
            fraction=0.026,
            pad=0.025,
            shrink=0.88,
        )
        colorbar.set_label("单镜年平均输出热功率 (kW)")
    figure.suptitle(
        "图2-1  两种候选布局的平面分布与单镜年平均输出",
        fontsize=16,
        fontweight="bold",
    )
    path = figure_dir / "11_图2-1_两种候选布局平面分布与单镜年平均输出.png"
    figure.savefig(path, dpi=220)
    plt.close(figure)
    return path


def plot_metric_comparison(
    *,
    figure_dir: Path,
    comparison: dict,
    partitioned_optical: float,
    campo_optical: float,
) -> Path:
    names = ("分区交错同心圆", "改进 Campo")
    palette = (PARTITIONED_COLOR, CAMPO_COLOR)
    metrics = (
        (
            "年平均输出热功率",
            (
                comparison["partitioned"]["annual_power_mw"],
                comparison["campo"]["annual_power_mw"],
            ),
            "MW",
            48.0,
            42.0,
            "{:.3f}",
        ),
        (
            "单位镜面面积年平均输出",
            (
                comparison["partitioned"]["unit_area_power_kw_m2"],
                comparison["campo"]["unit_area_power_kw_m2"],
            ),
            "kW/m²",
            0.75,
            None,
            "{:.4f}",
        ),
        (
            "总镜面面积",
            (
                comparison["partitioned"]["total_area_m2"],
                comparison["campo"]["total_area_m2"],
            ),
            "m²",
            72000.0,
            None,
            "{:,.0f}",
        ),
        (
            "年平均综合光学效率",
            (partitioned_optical, campo_optical),
            "",
            0.78,
            None,
            "{:.4f}",
        ),
    )
    figure, axes = plt.subplots(2, 2, figsize=(12.6, 8.6))
    x = np.arange(2)
    for ax, (title, values, unit, upper, benchmark, label_format) in zip(
        axes.flat, metrics, strict=True
    ):
        bars = ax.bar(
            x,
            values,
            width=0.56,
            color=palette,
            edgecolor=("#1D4ED8", "#B45309"),
            linewidth=0.9,
        )
        ax.set_title(title, fontsize=12.5, fontweight="bold")
        ax.set_xticks(x, names)
        ax.set_ylabel(unit)
        ax.set_ylim(0.0, upper)
        ax.grid(axis="y")
        if benchmark is not None:
            ax.axhline(
                benchmark,
                color=TARGET_COLOR,
                linestyle="--",
                linewidth=1.5,
                label="42 MW约束",
            )
            ax.legend(loc="lower right", frameon=False)
        for bar, value in zip(bars, values, strict=True):
            ax.text(
                bar.get_x() + bar.get_width() / 2.0,
                bar.get_height() + upper * 0.025,
                label_format.format(value),
                ha="center",
                va="bottom",
                fontsize=10,
                fontweight="bold",
            )
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
    figure.suptitle(
        "图2-2  两种候选布局的主要性能指标对比",
        fontsize=16,
        fontweight="bold",
        y=0.98,
    )
    figure.text(
        0.5,
        0.94,
        "两种布局均采用60个规定时刻与相同正式计算精度",
        ha="center",
        fontsize=10.5,
        color="#526075",
    )
    figure.tight_layout(rect=(0.02, 0.02, 0.98, 0.92), h_pad=2.4, w_pad=2.0)
    path = figure_dir / "12_图2-2_两种候选布局主要性能指标对比.png"
    figure.savefig(path, dpi=220)
    plt.close(figure)
    return path


def plot_monthly_comparison(
    *,
    figure_dir: Path,
    partitioned_evaluation,
    campo_evaluation,
    comparison: dict,
) -> Path:
    months = np.arange(1, 13)
    partitioned_monthly = partitioned_evaluation.solution.monthly_results
    partitioned_power = np.asarray([row.field_output_mw for row in partitioned_monthly])
    campo_monthly = campo_evaluation.solution.monthly_results
    campo_power = np.asarray([row.field_output_mw for row in campo_monthly])
    partitioned_unit = np.asarray(
        [row.unit_area_output_kw_m2 for row in partitioned_monthly]
    )
    campo_unit = np.asarray([row.unit_area_output_kw_m2 for row in campo_monthly])
    partitioned_optical = np.asarray(
        [row.average_optical_efficiency for row in partitioned_monthly]
    )
    campo_optical = np.asarray(
        [row.average_optical_efficiency for row in campo_monthly]
    )

    figure, axes = plt.subplots(
        3,
        1,
        figsize=(12.6, 10.6),
        sharex=True,
        gridspec_kw={"height_ratios": (1.35, 1.0, 1.0)},
    )
    width = 0.36
    axes[0].bar(
        months - width / 2,
        partitioned_power,
        width,
        color=PARTITIONED_COLOR,
        edgecolor="#1D4ED8",
        linewidth=0.7,
        label=(
            "分区交错同心圆"
            f"（年均 {comparison['partitioned']['annual_power_mw']:.3f} MW）"
        ),
    )
    axes[0].bar(
        months + width / 2,
        campo_power,
        width,
        color=CAMPO_COLOR,
        edgecolor="#B45309",
        linewidth=0.7,
        label=(f"改进 Campo（年均 {comparison['campo']['annual_power_mw']:.3f} MW）"),
    )
    axes[0].axhline(
        42.0,
        color=TARGET_COLOR,
        linestyle="--",
        linewidth=1.5,
        label="42 MW约束",
    )
    axes[0].set_ylabel("热功率 (MW)")
    axes[0].set_title("月平均输出热功率", fontsize=12.5, fontweight="bold")
    axes[0].set_ylim(0, max(partitioned_power.max(), campo_power.max()) * 1.18)
    axes[0].legend(ncol=3, loc="upper center", fontsize=9)
    axes[0].grid(axis="y")

    axes[1].plot(
        months,
        partitioned_unit,
        color=PARTITIONED_COLOR,
        marker="o",
        linewidth=2.0,
        markersize=5,
        label="分区交错同心圆",
    )
    axes[1].plot(
        months,
        campo_unit,
        color=CAMPO_COLOR,
        marker="s",
        linestyle="--",
        linewidth=2.0,
        markersize=5,
        label="改进 Campo",
    )
    axes[1].set_ylabel("单位面积功率\n(kW/m²)")
    axes[1].set_title(
        "单位镜面面积月平均输出热功率",
        fontsize=12.5,
        fontweight="bold",
    )
    axes[1].grid(True)
    axes[1].legend(loc="lower center", ncol=2)

    axes[2].plot(
        months,
        partitioned_optical,
        color=PARTITIONED_COLOR,
        marker="o",
        linewidth=2.0,
        markersize=5,
        label="分区交错同心圆",
    )
    axes[2].plot(
        months,
        campo_optical,
        color=CAMPO_COLOR,
        marker="s",
        linestyle="--",
        linewidth=2.0,
        markersize=5,
        label="改进 Campo",
    )
    axes[2].set_xlabel("月份")
    axes[2].set_ylabel("综合光学效率")
    axes[2].set_title("月平均综合光学效率", fontsize=12.5, fontweight="bold")
    axes[2].set_xticks(months)
    optical_min = min(partitioned_optical.min(), campo_optical.min())
    optical_max = max(partitioned_optical.max(), campo_optical.max())
    axes[2].set_ylim(
        max(0.0, optical_min - 0.02),
        min(1.0, optical_max + 0.02),
    )
    axes[2].grid(True)
    axes[2].legend(loc="lower center", ncol=2)
    for ax in axes:
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
    figure.suptitle(
        "图2-3  两种候选布局的月平均性能对比",
        fontsize=16,
        fontweight="bold",
        y=0.99,
    )
    figure.tight_layout(rect=(0.02, 0.02, 0.98, 0.96), h_pad=1.6)
    path = figure_dir / "13_图2-3_两种候选布局月平均性能对比.png"
    figure.savefig(path, dpi=220)
    plt.close(figure)
    return path


def representative_indices(
    coordinates: np.ndarray,
    tower_x: float,
    tower_y: float,
    count: int = 20,
) -> np.ndarray:
    angles = np.arctan2(
        coordinates[:, 1] - tower_y,
        coordinates[:, 0] - tower_x,
    )
    order = np.argsort(angles)
    picks = np.linspace(0, len(order) - 1, count, dtype=int)
    return order[picks]


def draw_receiver(ax, config: FieldConfig) -> None:
    theta = np.linspace(0.0, 2.0 * np.pi, 32)
    z_values = np.linspace(config.receiver_z_min, config.receiver_z_max, 8)
    theta_grid, z_grid = np.meshgrid(theta, z_values)
    x_grid = config.tower_x + config.receiver_radius * np.cos(theta_grid)
    y_grid = config.tower_y + config.receiver_radius * np.sin(theta_grid)
    ax.plot_surface(
        x_grid,
        y_grid,
        z_grid,
        color=RECEIVER_COLOR,
        alpha=0.9,
        linewidth=0,
        shade=True,
    )
    ax.plot(
        [config.tower_x, config.tower_x],
        [config.tower_y, config.tower_y],
        [0.0, config.receiver_z_min],
        color="#667085",
        linewidth=4.5,
        solid_capstyle="round",
    )


def plot_3d_comparison(
    *,
    figure_dir: Path,
    partitioned_parameters: PartitionedRingParameters,
    partitioned_coordinates: np.ndarray,
    partitioned_powers: np.ndarray,
    campo_parameters: CampoParameters,
    campo_coordinates: np.ndarray,
    campo_powers: np.ndarray,
) -> Path:
    all_powers = np.concatenate((partitioned_powers, campo_powers))
    norm = colors.Normalize(
        vmin=float(np.percentile(all_powers, 1.0)),
        vmax=float(np.percentile(all_powers, 99.0)),
    )
    figure = plt.figure(figsize=(16.0, 8.0))
    axes = (
        figure.add_subplot(1, 2, 1, projection="3d"),
        figure.add_subplot(1, 2, 2, projection="3d"),
    )
    layouts = (
        (
            axes[0],
            "方案A：分区交错同心圆",
            partitioned_parameters,
            partitioned_coordinates,
            partitioned_powers,
        ),
        (
            axes[1],
            "方案B：改进 Campo",
            campo_parameters,
            campo_coordinates,
            campo_powers,
        ),
    )
    scatter = None
    for ax, title, parameters, coordinates, powers in layouts:
        config = replace(
            FieldConfig(),
            tower_x=parameters.tower_x,
            tower_y=parameters.tower_y,
            mirror_width=parameters.mirror_width,
            mirror_height=parameters.mirror_height,
            mirror_center_z=parameters.installation_height,
        )
        z = np.full(coordinates.shape[0], config.mirror_center_z)
        scatter = ax.scatter(
            coordinates[:, 0],
            coordinates[:, 1],
            z,
            c=powers,
            cmap=POWER_CMAP,
            norm=norm,
            s=7,
            linewidths=0,
            alpha=0.95,
            rasterized=True,
        )
        draw_receiver(ax, config)
        selected = representative_indices(
            coordinates,
            config.tower_x,
            config.tower_y,
        )
        for index in selected:
            ax.plot(
                [coordinates[index, 0], config.tower_x],
                [coordinates[index, 1], config.tower_y],
                [config.mirror_center_z, config.receiver_center_z],
                color=RAY_COLOR,
                alpha=0.32,
                linewidth=0.7,
            )
        ax.set_title(title, fontsize=13, fontweight="bold", pad=12)
        ax.set_xlabel("x / East (m)", labelpad=8)
        ax.set_ylabel("y / North (m)", labelpad=8)
        ax.set_zlabel("z (m)", labelpad=6)
        ax.set_xlim(-360, 360)
        ax.set_ylim(-360, 360)
        ax.set_zlim(0, 105)
        ax.view_init(elev=27, azim=-61)
        ax.set_box_aspect((1.0, 1.0, 0.33))
        ax.grid(True)
    if scatter is not None:
        colorbar_axis = figure.add_axes((0.925, 0.20, 0.014, 0.60))
        colorbar = figure.colorbar(
            scatter,
            cax=colorbar_axis,
        )
        colorbar.set_label("单镜年平均输出热功率 (kW)")
    figure.suptitle(
        "图2-4  两种候选布局的三维镜场与代表性中心光路",
        fontsize=16,
        fontweight="bold",
        y=0.97,
    )
    figure.text(
        0.5,
        0.925,
        "光路线仅用于展示镜位至吸收器中心的空间关系",
        ha="center",
        fontsize=10,
        color="#526075",
    )
    figure.subplots_adjust(left=0.02, right=0.87, bottom=0.04, top=0.89, wspace=0.08)
    path = figure_dir / "14_图2-4_两种候选布局三维镜场与代表性中心光路.png"
    figure.savefig(path, dpi=220)
    plt.close(figure)
    return path


def build_question2_figures(
    *,
    output_dir: str | Path,
    comparison: dict,
    parameters: dict[str, LayoutParameters],
    evaluations: dict[str, FieldEvaluation],
) -> tuple[Path, ...]:
    """由两种候选布局的正式复算对象生成四张论文图。"""

    configure_matplotlib()
    figure_dir = Path(output_dir)
    figure_dir.mkdir(parents=True, exist_ok=True)
    write_monthly_comparison_data(figure_dir, evaluations)
    partitioned_parameters = parameters["partitioned"]
    campo_parameters = parameters["campo"]
    partitioned_evaluation = evaluations["partitioned"]
    campo_evaluation = evaluations["campo"]
    partitioned_coordinates = partitioned_evaluation.coordinates
    campo_coordinates = campo_evaluation.coordinates
    partitioned_powers = np.asarray(
        [
            row.average_output_power_kw
            for row in partitioned_evaluation.solution.mirror_annual_results
        ],
        dtype=float,
    )
    campo_powers = np.asarray(
        [
            row.average_output_power_kw
            for row in campo_evaluation.solution.mirror_annual_results
        ],
        dtype=float,
    )

    return (
        plot_layout_comparison(
            figure_dir=figure_dir,
            comparison=comparison,
            partitioned_parameters=partitioned_parameters,
            partitioned_coordinates=partitioned_coordinates,
            partitioned_powers=partitioned_powers,
            campo_parameters=campo_parameters,
            campo_coordinates=campo_coordinates,
            campo_powers=campo_powers,
        ),
        plot_metric_comparison(
            figure_dir=figure_dir,
            comparison=comparison,
            partitioned_optical=(
                partitioned_evaluation.solution.annual_result.average_optical_efficiency
            ),
            campo_optical=(
                campo_evaluation.solution.annual_result.average_optical_efficiency
            ),
        ),
        plot_monthly_comparison(
            figure_dir=figure_dir,
            partitioned_evaluation=partitioned_evaluation,
            campo_evaluation=campo_evaluation,
            comparison=comparison,
        ),
        plot_3d_comparison(
            figure_dir=figure_dir,
            partitioned_parameters=partitioned_parameters,
            partitioned_coordinates=partitioned_coordinates,
            partitioned_powers=partitioned_powers,
            campo_parameters=campo_parameters,
            campo_coordinates=campo_coordinates,
            campo_powers=campo_powers,
        ),
    )


def write_monthly_comparison_data(
    output_dir: str | Path,
    evaluations: dict[str, FieldEvaluation],
) -> Path:
    """落盘图2-3使用的两种布局月度数据，便于论文逐项引用。"""

    destination = Path(output_dir) / "15_双布局月平均对比数据.csv"
    rows: list[dict[str, object]] = []
    for kind, label in (
        ("partitioned", "分区交错同心圆"),
        ("campo", "改进 Campo"),
    ):
        for record in evaluations[kind].solution.monthly_results:
            rows.append(
                {
                    "layout": kind,
                    "layout_label": label,
                    "month": record.month,
                    "average_optical_efficiency": (record.average_optical_efficiency),
                    "average_cosine_efficiency": (record.average_cosine_efficiency),
                    "average_shadow_blocking_efficiency": (
                        record.average_shadow_blocking_efficiency
                    ),
                    "average_atmospheric_efficiency": (
                        record.average_atmospheric_efficiency
                    ),
                    "average_truncation_efficiency": (
                        record.average_truncation_efficiency
                    ),
                    "field_output_mw": record.field_output_mw,
                    "unit_area_output_kw_m2": (record.unit_area_output_kw_m2),
                }
            )
    with destination.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)
    return destination


def build_question2_figures_from_output(
    output_dir: str | Path,
) -> tuple[Path, ...]:
    """读取扁平交付目录，统一正式精度复算后重新生成四张图。"""

    destination = Path(output_dir)
    comparison = load_json(destination / "02_双布局比较.json")
    (
        partitioned_parameters,
        _,
        partitioned_evaluation,
    ) = build_partitioned_result(comparison)
    campo_parameters, _, campo_evaluation = build_campo_result(
        comparison,
        destination,
    )
    return build_question2_figures(
        output_dir=destination,
        comparison=comparison,
        parameters={
            "partitioned": partitioned_parameters,
            "campo": campo_parameters,
        },
        evaluations={
            "partitioned": partitioned_evaluation,
            "campo": campo_evaluation,
        },
    )


# ========================================================================
# 第二问命令行流程
# ========================================================================

import argparse
import json
from dataclasses import asdict, replace
from pathlib import Path
from typing import Sequence


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_OUTPUT = PROJECT_ROOT / "outputs" / "q2"
DEFAULT_TEMPLATE = PROJECT_ROOT / "task" / "A" / "result2.xlsx"


def _smoke_profile() -> EvaluationProfile:
    return EvaluationProfile(
        name="smoke",
        solver=SolverConfig(
            shadow_grid_size=3,
            truncation_rays=8,
            neighbor_radius_m=60.0,
            truncation_chunk_size=64,
            sobol_seed=2023,
        ),
        months=(6,),
        solar_times=(12.0,),
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="独立优化分区圆环和改进 Campo 两种问题二镜场"
    )
    parser.add_argument(
        "--layout",
        choices=("both", "partitioned", "campo"),
        default="both",
    )
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument("--initial-samples", type=int, default=16)
    parser.add_argument("--retained-starts", type=int, default=3)
    parser.add_argument("--max-cycles", type=int, default=4)
    parser.add_argument("--coarse-stride", type=int, default=4)
    parser.add_argument("--extent-window", type=int, default=2)
    parser.add_argument("--seed", type=int, default=2023)
    parser.add_argument(
        "--resume-comparison",
        type=Path,
        default=None,
        help="从上一阶段的 02_双布局比较.json 参数继续局部搜索",
    )
    parser.add_argument(
        "--search-profile",
        choices=("exploration", "refinement"),
        default="exploration",
        help="非烟雾搜索使用的数值离散精度",
    )
    parser.add_argument(
        "--step-level-count",
        type=int,
        choices=(1, 2, 3),
        default=3,
        help="本阶段连续使用几档步长",
    )
    parser.add_argument(
        "--step-level-start",
        type=int,
        choices=(1, 2, 3),
        default=1,
        help="本阶段从第几档步长开始（1 为粗、3 为细）",
    )
    parser.add_argument(
        "--prune-rounds",
        type=int,
        default=10,
        help="胜出布局结构化删镜的最大轮数；0 表示跳过",
    )
    parser.add_argument(
        "--prune-pairs-per-round",
        type=int,
        default=None,
        help="每轮最多复算的外层对称镜位对；默认全部",
    )
    parser.add_argument(
        "--smoke",
        action="store_true",
        help="仅用 6 月正午、3×3 阴影网格和 8 条截断光线验证流程",
    )
    parser.add_argument(
        "--skip-x-check",
        action="store_true",
        help="跳过塔东西坐标 {-10,-5,0,5,10} m 的少量复核",
    )
    parser.add_argument(
        "--skip-figures",
        action="store_true",
        help="不生成四张正式结果图",
    )
    parser.add_argument(
        "--figures-only",
        action="store_true",
        help="读取输出目录中的正式结果并重新生成四张图",
    )
    parser.add_argument(
        "--run-validation",
        action="store_true",
        help="额外运行 20×20 阴影网格、512 条截断光线的加密复算",
    )
    return parser


def _validate_args(args: argparse.Namespace) -> None:
    for name in (
        "initial_samples",
        "retained_starts",
        "coarse_stride",
    ):
        if getattr(args, name) < 1:
            raise SystemExit(f"--{name.replace('_', '-')} 必须大于等于 1。")
    if args.extent_window < 0:
        raise SystemExit("--extent-window 不能小于 0。")
    if args.max_cycles < 0:
        raise SystemExit("--max-cycles 不能小于 0。")
    if args.step_level_start + args.step_level_count - 1 > 3:
        raise SystemExit("--step-level-start 与 --step-level-count 超出三档步长。")
    if args.prune_rounds < 0:
        raise SystemExit("--prune-rounds 不能小于 0。")
    if args.prune_pairs_per_round is not None and args.prune_pairs_per_round < 1:
        raise SystemExit("--prune-pairs-per-round 必须大于等于 1。")


def run(argv: Sequence[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    _validate_args(args)
    if args.figures_only:
        for path in build_question2_figures_from_output(args.output):
            print(f"输出：{path}")
        return 0
    if args.smoke:
        search_profile = _smoke_profile()
    elif args.search_profile == "refinement":
        search_profile = refinement_profile()
    else:
        search_profile = exploration_profile()
    verification_profile = _smoke_profile() if args.smoke else final_profile()
    cache = EvaluationCache()
    optimized: dict[str, object] = {}
    step_start = args.step_level_start - 1
    step_stop = step_start + args.step_level_count
    partitioned_steps = PARTITIONED_STEP_LEVELS[step_start:step_stop]
    campo_steps = CAMPO_STEP_LEVELS[step_start:step_stop]
    resumed = None
    if args.resume_comparison is not None:
        if not args.resume_comparison.exists():
            raise SystemExit(f"找不到恢复文件：{args.resume_comparison}")
        resumed = json.loads(args.resume_comparison.read_text(encoding="utf-8"))

    if args.layout in ("both", "partitioned"):
        print("开始独立优化方案 A：分区交错同心圆")
        if resumed is not None:
            optimized["partitioned"] = refine_partitioned(
                PartitionedRingParameters(**resumed["partitioned"]["parameters"]),
                profile=search_profile,
                step_levels=partitioned_steps,
                maximum_cycles_per_level=args.max_cycles,
                coarse_stride=args.coarse_stride,
                window=args.extent_window,
                cache=cache,
                progress=print,
            )
        else:
            optimized["partitioned"] = optimize_partitioned(
                profile=search_profile,
                initial_sample_count=args.initial_samples,
                retained_starts=args.retained_starts,
                seed=args.seed,
                step_levels=partitioned_steps,
                maximum_cycles_per_level=args.max_cycles,
                coarse_stride=args.coarse_stride,
                window=args.extent_window,
                cache=cache,
                progress=print,
            )
    if args.layout in ("both", "campo"):
        print("开始独立优化方案 B：改进 Campo 径向交错")
        if resumed is not None:
            optimized["campo"] = refine_campo(
                CampoParameters(**resumed["campo"]["parameters"]),
                profile=search_profile,
                step_levels=campo_steps,
                maximum_cycles_per_level=args.max_cycles,
                coarse_stride=args.coarse_stride,
                window=args.extent_window,
                cache=cache,
                progress=print,
            )
        else:
            optimized["campo"] = optimize_campo(
                profile=search_profile,
                initial_sample_count=args.initial_samples,
                retained_starts=args.retained_starts,
                seed=args.seed + 1,
                step_levels=campo_steps,
                maximum_cycles_per_level=args.max_cycles,
                coarse_stride=args.coarse_stride,
                window=args.extent_window,
                cache=cache,
                progress=print,
            )

    verified: dict[str, tuple[object, object, object, object]] = {}
    for kind, result in optimized.items():
        parameters = result.best.parameters
        x_check_scan = None
        if not args.smoke and not args.skip_x_check:
            center_layout = (
                generate_partitioned_layout(parameters)
                if kind == "partitioned"
                else generate_campo_layout(parameters)
            )
            center_scan = scan_layout_extents(
                center_layout,
                parameters,
                verification_profile,
                coarse_stride=args.coarse_stride,
                window=args.extent_window,
                cache=cache,
            )
            selected_parameters = parameters
            selected_scan = center_scan
            for tower_x in (-10.0, -5.0, 5.0, 10.0):
                candidate_parameters = replace(
                    parameters,
                    tower_x=tower_x,
                )
                candidate_layout = (
                    generate_partitioned_layout(candidate_parameters)
                    if kind == "partitioned"
                    else generate_campo_layout(candidate_parameters)
                )
                candidate_scan = scan_layout_extents(
                    candidate_layout,
                    candidate_parameters,
                    verification_profile,
                    coarse_stride=args.coarse_stride,
                    window=args.extent_window,
                    cache=cache,
                )
                selected = better_evaluation(
                    selected_scan.best,
                    candidate_scan.best,
                )
                selected_feasible = selected_scan.best.is_feasible()
                candidate_feasible = candidate_scan.best.is_feasible()
                if candidate_feasible != selected_feasible:
                    stable_gain = candidate_feasible
                elif candidate_feasible:
                    stable_gain = (
                        candidate_scan.best.unit_area_power_kw_m2
                        - selected_scan.best.unit_area_power_kw_m2
                        > 1e-4
                    )
                else:
                    stable_gain = (
                        candidate_scan.best.annual_power_mw
                        - selected_scan.best.annual_power_mw
                        > 1e-3
                    )
                if selected is candidate_scan.best and stable_gain:
                    selected_parameters = candidate_parameters
                    selected_scan = candidate_scan
            parameters = selected_parameters
            x_check_scan = selected_scan
        layout = (
            generate_partitioned_layout(parameters)
            if kind == "partitioned"
            else generate_campo_layout(parameters)
        )
        precision_label = "烟雾测试精度" if args.smoke else "问题一最终精度"
        print(f"使用统一{precision_label}复算 {kind}")
        if x_check_scan is None:
            scan = scan_layout_extents(
                layout,
                parameters,
                verification_profile,
                coarse_stride=args.coarse_stride,
                window=args.extent_window,
                cache=cache,
            )
        else:
            scan = x_check_scan
        verified[kind] = (parameters, result, layout, scan)

    verified_values = list(verified.items())
    winner_kind, winner_bundle = verified_values[0]
    for kind, bundle in verified_values[1:]:
        if (
            better_evaluation(
                winner_bundle[3].best,
                bundle[3].best,
            )
            is bundle[3].best
        ):
            winner_kind, winner_bundle = kind, bundle

    winner_parameters, _, winner_layout, winner_scan = winner_bundle
    winner_evaluation = winner_scan.best

    if args.prune_rounds and abs(winner_parameters.tower_x) <= 1e-9:
        print("对胜出布局执行外层东西对称镜位修剪")
        prune = prune_outer_symmetric_pairs(
            layout=winner_layout,
            parameters=winner_parameters,
            initial=winner_evaluation,
            profile=verification_profile,
            maximum_rounds=args.prune_rounds,
            maximum_pairs_per_round=args.prune_pairs_per_round,
            cache=cache,
        )
        winner_evaluation = prune.best
    elif args.prune_rounds:
        print("塔东西坐标不为 0，跳过要求南北轴对称的外层镜位修剪")

    args.output.mkdir(parents=True, exist_ok=True)
    comparison = {
        kind: {
            "parameters": asdict(bundle[0]),
            "ring_count": bundle[3].best.ring_count,
            "mirror_count": bundle[3].best.mirror_count,
            "total_area_m2": bundle[3].best.total_area_m2,
            "annual_power_mw": bundle[3].best.annual_power_mw,
            "unit_area_power_kw_m2": (bundle[3].best.unit_area_power_kw_m2),
        }
        for kind, bundle in verified.items()
    }
    comparison[winner_kind].update(
        {
            "ring_count": winner_evaluation.ring_count,
            "mirror_count": winner_evaluation.mirror_count,
            "total_area_m2": winner_evaluation.total_area_m2,
            "annual_power_mw": winner_evaluation.annual_power_mw,
            "unit_area_power_kw_m2": (winner_evaluation.unit_area_power_kw_m2),
        }
    )
    comparison["winner"] = winner_kind
    comparison_path = args.output / "02_双布局比较.json"
    comparison_path.write_text(
        json.dumps(comparison, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    written = write_question2_results(
        output_dir=args.output,
        layout_name=winner_kind,
        parameters=winner_parameters,
        evaluation=winner_evaluation,
        result2_template=args.template,
        comparison=comparison,
    )
    if args.run_validation and not args.smoke:
        dense_profile = EvaluationProfile(
            name="dense-validation",
            solver=SolverConfig(
                shadow_grid_size=20,
                truncation_rays=512,
                neighbor_radius_m=80.0,
                truncation_chunk_size=128,
                sobol_seed=2023,
            ),
        )
        dense_evaluation = evaluate_coordinates(
            layout_kind=winner_kind,
            ring_count=winner_evaluation.ring_count,
            coordinates=winner_evaluation.coordinates,
            parameters=winner_parameters,
            profile=dense_profile,
        )
        written["dense_validation"] = write_high_precision_validation(
            output_dir=args.output,
            evaluation=dense_evaluation,
            profile=dense_profile,
        )
    if not args.skip_figures and len(verified) == 2:
        figure_evaluations = {kind: bundle[3].best for kind, bundle in verified.items()}
        figure_evaluations[winner_kind] = winner_evaluation
        figure_parameters = {kind: bundle[0] for kind, bundle in verified.items()}
        for path in build_question2_figures(
            output_dir=args.output,
            comparison=comparison,
            parameters=figure_parameters,
            evaluations=figure_evaluations,
        ):
            written[path.stem] = path
    elif not args.skip_figures:
        print("仅优化一种布局，跳过双布局对比图。")

    print(
        "\n第二问烟雾测试结果（不可作为正式年平均结论）"
        if args.smoke
        else "\n第二问结果"
    )
    print(f"胜出布局：{winner_kind}")
    print(f"镜子数：{winner_evaluation.mirror_count}")
    print(f"总镜面面积：{winner_evaluation.total_area_m2:.3f} m²")
    target_power_mw = 42.0
    print(f"年平均输出热功率约束下限：{target_power_mw:.6f} MW")
    print(f"最终年平均输出热功率：{winner_evaluation.annual_power_mw:.6f} MW")
    print(
        "相对约束下限的功率余量："
        f"{winner_evaluation.annual_power_mw - target_power_mw:.6f} MW"
    )
    print(
        f"单位面积年平均输出热功率：{winner_evaluation.unit_area_power_kw_m2:.6f} kW/m²"
    )
    print(f"双布局比较：{comparison_path}")
    for path in written.values():
        print(f"输出：{path}")
    return 0


def main() -> None:
    raise SystemExit(run())


if __name__ == "__main__":
    main()
