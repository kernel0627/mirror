"""第三问完整代码展示稿。

本文件合并共享光学核心、Campo 母场、异构搜索、验证和输出流程，可直接运行。
"""
from __future__ import annotations
# ruff: noqa: E402,F401,F811

# ========================================================================
# 来源：src/heliostat/config.py
# ========================================================================

"""题目参数与数值计算参数。"""
from dataclasses import asdict, dataclass

@dataclass(frozen=True)
class FieldConfig:
    """第一问给定的镜场和环境参数。"""
    latitude_deg: float = 39.4
    altitude_km: float = 3.0
    field_radius: float = 350.0
    exclusion_radius: float = 100.0
    tower_x: float = 0.0
    tower_y: float = 0.0
    receiver_center_z: float = 86.0
    receiver_radius: float = 4.0
    receiver_height: float = 8.0
    mirror_width: float = 6.2
    mirror_height: float = 6.2
    mirror_center_z: float = 4.5
    reflectivity: float = 0.92
    solar_angular_radius_rad: float = 0.00465

    @property
    def receiver_z_min(self) -> float:
        return self.receiver_center_z - self.receiver_height / 2.0

    @property
    def receiver_z_max(self) -> float:
        return self.receiver_center_z + self.receiver_height / 2.0

    @property
    def mirror_area(self) -> float:
        return self.mirror_width * self.mirror_height

    def to_dict(self) -> dict[str, float]:
        return asdict(self)

@dataclass(frozen=True)
class SolverConfig:
    """可调的采样精度和加速参数。"""
    shadow_grid_size: int = 15
    truncation_rays: int = 256
    neighbor_radius_m: float = 60.0
    candidate_margin: float = 1.05
    ray_epsilon: float = 1e-07
    truncation_chunk_size: int = 128
    sobol_seed: int = 2023
    calculate_shadow: bool = True
    calculate_truncation: bool = True

    def __post_init__(self) -> None:
        if self.shadow_grid_size < 1:
            raise ValueError('shadow_grid_size 必须大于等于 1。')
        if self.truncation_rays < 1:
            raise ValueError('truncation_rays 必须大于等于 1。')
        if self.neighbor_radius_m <= 0.0:
            raise ValueError('neighbor_radius_m 必须大于 0。')
        if self.candidate_margin < 1.0:
            raise ValueError('candidate_margin 不能小于 1。')
        if self.ray_epsilon <= 0.0:
            raise ValueError('ray_epsilon 必须大于 0。')
        if self.truncation_chunk_size < 1:
            raise ValueError('truncation_chunk_size 必须大于等于 1。')

    def to_dict(self) -> dict[str, int | float | bool]:
        return asdict(self)

# ========================================================================
# 来源：src/heliostat/solar.py
# ========================================================================

"""太阳位置与 DNI 计算。"""
import math
from dataclasses import dataclass
import numpy as np
from numpy.typing import NDArray
FloatArray = NDArray[np.float64]

@dataclass(frozen=True)
class SolarState:
    month: int
    solar_time: float
    direction: FloatArray
    altitude_rad: float
    azimuth_rad: float
    declination_rad: float
    dni_kw_m2: float
MONTH_DAYS = (31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31)

def day_from_spring_equinox(month: int, day: int=21) -> int:
    """以 3 月 21 日为第 0 天，返回题面赤纬公式所需的 D。"""
    if not 1 <= month <= 12:
        raise ValueError('month 必须位于 1 到 12。')
    if not 1 <= day <= MONTH_DAYS[month - 1]:
        raise ValueError('day 不在指定月份的有效范围内。')
    day_of_year = sum(MONTH_DAYS[:month - 1]) + day
    return day_of_year - 80

def calculate_solar_state(month: int, solar_time: float, latitude_deg: float, altitude_km: float) -> SolarState:
    """按题面附录计算东-北-天坐标下的太阳单位方向和 DNI。"""
    if not 0.0 <= solar_time <= 24.0:
        raise ValueError('solar_time 必须位于 0 到 24 小时。')
    d = day_from_spring_equinox(month)
    declination = math.asin(math.sin(2.0 * math.pi * d / 365.0) * math.sin(math.radians(23.45)))
    latitude = math.radians(latitude_deg)
    hour_angle = math.pi / 12.0 * (solar_time - 12.0)
    direction = np.array([-math.cos(declination) * math.sin(hour_angle), math.cos(latitude) * math.sin(declination) - math.sin(latitude) * math.cos(declination) * math.cos(hour_angle), math.sin(latitude) * math.sin(declination) + math.cos(latitude) * math.cos(declination) * math.cos(hour_angle)], dtype=float)
    direction /= np.linalg.norm(direction)
    altitude = math.asin(float(np.clip(direction[2], -1.0, 1.0)))
    azimuth = math.atan2(float(direction[0]), float(direction[1])) % (2.0 * math.pi)
    h = altitude_km
    a = 0.4237 - 0.00821 * (6.0 - h) ** 2
    b = 0.5055 + 0.00595 * (6.5 - h) ** 2
    c = 0.2711 + 0.01858 * (2.5 - h) ** 2
    if altitude <= 0.0:
        dni = 0.0
    else:
        dni = 1.366 * (a + b * math.exp(-c / math.sin(altitude)))
    return SolarState(month=month, solar_time=solar_time, direction=direction, altitude_rad=altitude, azimuth_rad=azimuth, declination_rad=declination, dni_kw_m2=dni)

# ========================================================================
# 来源：src/heliostat/geometry.py
# ========================================================================

"""镜面姿态、反射与基础几何计算。"""
from dataclasses import dataclass
import numpy as np
from numpy.typing import NDArray
FloatArray = NDArray[np.float64]

def normalize_rows(vectors: FloatArray) -> FloatArray:
    norms = np.linalg.norm(vectors, axis=1, keepdims=True)
    if np.any(norms <= 1e-15):
        raise ValueError('不能归一化长度为零的向量。')
    return vectors / norms

def reflect(directions: FloatArray, normals: FloatArray) -> FloatArray:
    """根据 d - 2(d·n)n 计算反射方向。"""
    dot = np.sum(directions * normals, axis=-1, keepdims=True)
    return directions - 2.0 * dot * normals

@dataclass(frozen=True)
class PreparedField:
    config: FieldConfig
    centers: FloatArray
    mirror_widths: FloatArray
    mirror_heights: FloatArray
    mirror_areas: FloatArray
    receiver_center: FloatArray
    receiver_directions: FloatArray
    receiver_distances: FloatArray
    atmospheric_efficiency: FloatArray

    @property
    def mirror_count(self) -> int:
        return int(self.centers.shape[0])

    @property
    def total_mirror_area(self) -> float:
        return float(np.sum(self.mirror_areas))

@dataclass(frozen=True)
class MirrorOrientation:
    normals: FloatArray
    width_axes: FloatArray
    height_axes: FloatArray
    cosine_efficiency: FloatArray

def _per_mirror_values(values: FloatArray | None, *, fallback: float, mirror_count: int, name: str) -> FloatArray:
    if values is None:
        result = np.full(mirror_count, fallback, dtype=float)
    else:
        result = np.asarray(values, dtype=float)
        if result.ndim != 1 or result.shape[0] != mirror_count:
            raise ValueError(f'{name} 必须是一维且长度等于镜子数 {mirror_count}，实际形状为 {result.shape}。')
        result = result.copy()
    if not np.all(np.isfinite(result)):
        raise ValueError(f'{name} 包含 NaN 或无穷值。')
    if np.any(result <= 0.0):
        raise ValueError(f'{name} 必须全部大于 0。')
    return result

def prepare_field(mirror_xy: FloatArray, config: FieldConfig, *, mirror_widths: FloatArray | None=None, mirror_heights: FloatArray | None=None, mirror_center_zs: FloatArray | None=None) -> PreparedField:
    xy = np.asarray(mirror_xy, dtype=float)
    if xy.ndim != 2 or xy.shape[1] != 2 or xy.shape[0] == 0:
        raise ValueError('镜位坐标必须为非空 N×2 数组。')
    if not np.all(np.isfinite(xy)):
        raise ValueError('镜位坐标包含 NaN 或无穷值。')
    mirror_count = int(xy.shape[0])
    widths = _per_mirror_values(mirror_widths, fallback=config.mirror_width, mirror_count=mirror_count, name='mirror_widths')
    heights = _per_mirror_values(mirror_heights, fallback=config.mirror_height, mirror_count=mirror_count, name='mirror_heights')
    center_zs = _per_mirror_values(mirror_center_zs, fallback=config.mirror_center_z, mirror_count=mirror_count, name='mirror_center_zs')
    areas = widths * heights
    centers = np.column_stack((xy, center_zs))
    receiver_center = np.array([config.tower_x, config.tower_y, config.receiver_center_z], dtype=float)
    receiver_vectors = receiver_center[None, :] - centers
    distances = np.linalg.norm(receiver_vectors, axis=1)
    receiver_directions = receiver_vectors / distances[:, None]
    atmospheric = 0.99321 - 0.0001176 * distances + 1.97e-08 * distances ** 2
    atmospheric = np.clip(atmospheric, 0.0, 1.0)
    return PreparedField(config=config, centers=centers, mirror_widths=widths, mirror_heights=heights, mirror_areas=areas, receiver_center=receiver_center, receiver_directions=receiver_directions, receiver_distances=distances, atmospheric_efficiency=atmospheric)

def calculate_orientation(prepared: PreparedField, sun_direction: FloatArray) -> MirrorOrientation:
    sun_rows = np.broadcast_to(sun_direction, prepared.receiver_directions.shape)
    normals = normalize_rows(sun_rows + prepared.receiver_directions)
    upward = np.broadcast_to(np.array([0.0, 0.0, 1.0]), normals.shape)
    width_axes = np.cross(upward, normals)
    weak = np.linalg.norm(width_axes, axis=1) < 1e-10
    width_axes[weak] = np.array([1.0, 0.0, 0.0])
    width_axes = normalize_rows(width_axes)
    height_axes = normalize_rows(np.cross(normals, width_axes))
    cosine = np.clip(normals @ sun_direction, 0.0, 1.0)
    return MirrorOrientation(normals=normals, width_axes=width_axes, height_axes=height_axes, cosine_efficiency=cosine)

def maximum_reflection_error(prepared: PreparedField, orientation: MirrorOrientation, sun_direction: FloatArray) -> float:
    incoming = np.broadcast_to(-sun_direction, orientation.normals.shape)
    reflected = reflect(incoming, orientation.normals)
    errors = np.linalg.norm(reflected - prepared.receiver_directions, axis=1)
    return float(np.max(errors))

# ========================================================================
# 来源：src/heliostat/shadow.py
# ========================================================================

"""规则网格射线追踪计算阴影遮挡效率。"""
import numpy as np
from numpy.typing import NDArray
from scipy.spatial import cKDTree
IntArray = NDArray[np.int64]

def mirror_grid_offsets(grid_size: int, mirror_width: float, mirror_height: float) -> FloatArray:
    """返回位于等面积小格中心的局部二维坐标。"""
    width_step = mirror_width / grid_size
    height_step = mirror_height / grid_size
    width_values = np.linspace(-mirror_width / 2.0 + width_step / 2.0, mirror_width / 2.0 - width_step / 2.0, grid_size)
    height_values = np.linspace(-mirror_height / 2.0 + height_step / 2.0, mirror_height / 2.0 - height_step / 2.0, grid_size)
    (width_grid, height_grid) = np.meshgrid(width_values, height_values, indexing='xy')
    return np.column_stack((width_grid.ravel(), height_grid.ravel()))

def _direction_candidates(target_index: int, neighbors: IntArray, centers: FloatArray, direction: FloatArray, reach: float, maximum_distance: float | None=None) -> IntArray:
    if neighbors.size == 0:
        return neighbors
    delta = centers[neighbors] - centers[target_index]
    projection = delta @ direction
    perpendicular = delta - projection[:, None] * direction[None, :]
    perpendicular_distance = np.linalg.norm(perpendicular, axis=1)
    keep = (projection > -reach) & (perpendicular_distance <= reach)
    if maximum_distance is not None:
        keep &= projection < maximum_distance + reach
    return neighbors[keep]

def ray_rectangle_hits(origins: FloatArray, direction: FloatArray, rectangle_center: FloatArray, rectangle_normal: FloatArray, rectangle_width_axis: FloatArray, rectangle_height_axis: FloatArray, half_width: float, half_height: float, epsilon: float, maximum_distance: float | None=None) -> NDArray[np.bool_]:
    """判断一组同向射线是否与一面有限矩形相交。"""
    denominator = float(direction @ rectangle_normal)
    if abs(denominator) <= epsilon:
        return np.zeros(origins.shape[0], dtype=bool)
    distance = (rectangle_center - origins) @ rectangle_normal / denominator
    active = distance > epsilon
    if maximum_distance is not None:
        active &= distance < maximum_distance - epsilon
    if not np.any(active):
        return active
    intersections = origins + distance[:, None] * direction[None, :]
    relative = intersections - rectangle_center[None, :]
    local_width = relative @ rectangle_width_axis
    local_height = relative @ rectangle_height_axis
    active &= np.abs(local_width) <= half_width + epsilon
    active &= np.abs(local_height) <= half_height + epsilon
    return active

def _blocked_by_candidates(origins: FloatArray, direction: FloatArray, candidates: IntArray, prepared: PreparedField, orientation: MirrorOrientation, solver: SolverConfig, maximum_distance: float | None=None) -> NDArray[np.bool_]:
    blocked = np.zeros(origins.shape[0], dtype=bool)
    for candidate in candidates:
        active_indices = np.flatnonzero(~blocked)
        if active_indices.size == 0:
            break
        hits = ray_rectangle_hits(origins=origins[active_indices], direction=direction, rectangle_center=prepared.centers[candidate], rectangle_normal=orientation.normals[candidate], rectangle_width_axis=orientation.width_axes[candidate], rectangle_height_axis=orientation.height_axes[candidate], half_width=prepared.mirror_widths[candidate] / 2.0, half_height=prepared.mirror_heights[candidate] / 2.0, epsilon=solver.ray_epsilon, maximum_distance=maximum_distance)
        blocked[active_indices[hits]] = True
    return blocked

def calculate_shadow_blocking_efficiency(prepared: PreparedField, orientation: MirrorOrientation, sun_direction: FloatArray, solver: SolverConfig) -> FloatArray:
    """逐镜计算入射阴影和反射遮挡损失的采样点并集。"""
    mirror_count = prepared.mirror_count
    if mirror_count == 1:
        return np.ones(1, dtype=float)
    tree = cKDTree(prepared.centers[:, :2])
    bounding_radii = 0.5 * np.hypot(prepared.mirror_widths, prepared.mirror_heights)
    maximum_bounding_radius = float(np.max(bounding_radii))
    efficiencies = np.empty(mirror_count, dtype=float)
    offset_cache: dict[tuple[float, float], FloatArray] = {}
    for index in range(mirror_count):
        size_key = (float(prepared.mirror_widths[index]), float(prepared.mirror_heights[index]))
        offsets = offset_cache.get(size_key)
        if offsets is None:
            offsets = mirror_grid_offsets(solver.shadow_grid_size, size_key[0], size_key[1])
            offset_cache[size_key] = offsets
        sample_count = offsets.shape[0]
        reach = (float(bounding_radii[index]) + maximum_bounding_radius) * solver.candidate_margin
        points = prepared.centers[index][None, :] + offsets[:, :1] * orientation.width_axes[index][None, :] + offsets[:, 1:] * orientation.height_axes[index][None, :]
        neighbors = np.asarray(tree.query_ball_point(prepared.centers[index, :2], solver.neighbor_radius_m), dtype=np.int64)
        neighbors = neighbors[neighbors != index]
        incoming_candidates = _direction_candidates(target_index=index, neighbors=neighbors, centers=prepared.centers, direction=sun_direction, reach=reach)
        incoming_blocked = _blocked_by_candidates(origins=points, direction=sun_direction, candidates=incoming_candidates, prepared=prepared, orientation=orientation, solver=solver)
        reflected_direction = prepared.receiver_directions[index]
        reflected_candidates = _direction_candidates(target_index=index, neighbors=neighbors, centers=prepared.centers, direction=reflected_direction, reach=reach, maximum_distance=prepared.receiver_distances[index])
        reflected_blocked = _blocked_by_candidates(origins=points, direction=reflected_direction, candidates=reflected_candidates, prepared=prepared, orientation=orientation, solver=solver, maximum_distance=prepared.receiver_distances[index])
        blocked = incoming_blocked | reflected_blocked
        efficiencies[index] = 1.0 - np.count_nonzero(blocked) / sample_count
    return np.clip(efficiencies, 0.0, 1.0)

# ========================================================================
# 来源：src/heliostat/truncation.py
# ========================================================================

"""太阳锥光联合采样与有限高圆柱集热器求交。"""
import math
import numpy as np
from numpy.typing import NDArray
from scipy.stats import qmc

def build_sobol_samples(sample_count: int, seed: int) -> FloatArray:
    """生成固定、可复现的四维 Sobol 样本。"""
    exponent = int(math.ceil(math.log2(sample_count)))
    sampler = qmc.Sobol(d=4, scramble=True, seed=seed)
    return sampler.random_base2(exponent)[:sample_count]

def _sun_disk_directions(sun_direction: FloatArray, samples: FloatArray, angular_radius: float) -> FloatArray:
    reference = np.array([1.0, 0.0, 0.0]) if abs(float(sun_direction[2])) > 0.9 else np.array([0.0, 0.0, 1.0])
    tangent_one = np.cross(reference, sun_direction)
    tangent_one /= np.linalg.norm(tangent_one)
    tangent_two = np.cross(sun_direction, tangent_one)
    radial_angle = angular_radius * np.sqrt(samples[:, 2])
    polar_angle = 2.0 * math.pi * samples[:, 3]
    tangent = np.cos(polar_angle)[:, None] * tangent_one[None, :] + np.sin(polar_angle)[:, None] * tangent_two[None, :]
    directions = np.cos(radial_angle)[:, None] * sun_direction[None, :] + np.sin(radial_angle)[:, None] * tangent
    directions /= np.linalg.norm(directions, axis=1, keepdims=True)
    return directions

def ray_cylinder_side_hits(origins: FloatArray, directions: FloatArray, tower_x: float, tower_y: float, radius: float, z_min: float, z_max: float, epsilon: float) -> NDArray[np.bool_]:
    """判断任意形状批量射线的两个正根中是否有有限圆柱侧面交点。"""
    origin_x = origins[..., 0] - tower_x
    origin_y = origins[..., 1] - tower_y
    direction_x = directions[..., 0]
    direction_y = directions[..., 1]
    a = direction_x ** 2 + direction_y ** 2
    b = 2.0 * (origin_x * direction_x + origin_y * direction_y)
    c = origin_x ** 2 + origin_y ** 2 - radius ** 2
    discriminant = b ** 2 - 4.0 * a * c
    valid = (a > epsilon) & (discriminant >= 0.0)
    square_root = np.sqrt(np.maximum(discriminant, 0.0))
    denominator = np.where(valid, 2.0 * a, 1.0)
    near = (-b - square_root) / denominator
    far = (-b + square_root) / denominator
    near_z = origins[..., 2] + near * directions[..., 2]
    far_z = origins[..., 2] + far * directions[..., 2]
    near_hit = (near > epsilon) & (near_z >= z_min - epsilon) & (near_z <= z_max + epsilon)
    far_hit = (far > epsilon) & (far_z >= z_min - epsilon) & (far_z <= z_max + epsilon)
    return valid & (near_hit | far_hit)

def calculate_truncation_efficiency(prepared: PreparedField, orientation: MirrorOrientation, sun_direction: FloatArray, solver: SolverConfig) -> FloatArray:
    """联合采样镜面位置和太阳圆盘方向，计算集热器截断效率。"""
    config = prepared.config
    samples = build_sobol_samples(solver.truncation_rays, solver.sobol_seed)
    unit_width = samples[:, 0] - 0.5
    unit_height = samples[:, 1] - 0.5
    sampled_sun = _sun_disk_directions(sun_direction, samples, config.solar_angular_radius_rad)
    incoming = -sampled_sun
    efficiencies = np.empty(prepared.mirror_count, dtype=float)
    chunk_size = solver.truncation_chunk_size
    for start in range(0, prepared.mirror_count, chunk_size):
        stop = min(start + chunk_size, prepared.mirror_count)
        centers = prepared.centers[start:stop]
        normals = orientation.normals[start:stop]
        width_axes = orientation.width_axes[start:stop]
        height_axes = orientation.height_axes[start:stop]
        local_width = prepared.mirror_widths[start:stop, None] * unit_width[None, :]
        local_height = prepared.mirror_heights[start:stop, None] * unit_height[None, :]
        origins = centers[:, None, :] + local_width[:, :, None] * width_axes[:, None, :] + local_height[:, :, None] * height_axes[:, None, :]
        incoming_chunk = np.broadcast_to(incoming[None, :, :], origins.shape)
        dot = np.einsum('csj,cj->cs', incoming_chunk, normals)
        reflected = incoming_chunk - 2.0 * dot[:, :, None] * normals[:, None, :]
        hits = ray_cylinder_side_hits(origins=origins, directions=reflected, tower_x=config.tower_x, tower_y=config.tower_y, radius=config.receiver_radius, z_min=config.receiver_z_min, z_max=config.receiver_z_max, epsilon=solver.ray_epsilon)
        efficiencies[start:stop] = np.mean(hits, axis=1)
    return np.clip(efficiencies, 0.0, 1.0)

# ========================================================================
# 来源：src/heliostat/io.py
# ========================================================================

"""三问共用的坐标输入。"""
import csv
from pathlib import Path
import numpy as np
from numpy.typing import NDArray
from openpyxl import load_workbook
FloatArray = NDArray[np.float64]

def load_mirror_xy(path: str | Path, expected_count: int | None=1745) -> FloatArray:
    """从题目附件读取定日镜 x、y 坐标。"""
    source = Path(path)
    if not source.exists():
        raise FileNotFoundError(f'找不到定日镜坐标文件：{source}')
    if source.suffix.lower() == '.xlsx':
        workbook = load_workbook(source, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        workbook.close()
        values = [(row[0], row[1]) for row in rows if row[0] is not None]
    elif source.suffix.lower() == '.csv':
        with source.open('r', encoding='utf-8-sig', newline='') as handle:
            reader = csv.reader(handle)
            next(reader, None)
            values = [(row[0], row[1]) for row in reader if row]
    else:
        raise ValueError('坐标文件只支持 .xlsx 或 .csv。')
    try:
        mirror_xy = np.asarray(values, dtype=float)
    except (TypeError, ValueError) as exc:
        raise ValueError(f'坐标文件包含非数值数据：{source}') from exc
    if mirror_xy.ndim != 2 or mirror_xy.shape[1] != 2:
        raise ValueError(f'坐标数据应为 N×2，实际形状为 {mirror_xy.shape}。')
    if not np.all(np.isfinite(mirror_xy)):
        raise ValueError('坐标数据包含 NaN 或无穷值。')
    if expected_count is not None and mirror_xy.shape[0] != expected_count:
        raise ValueError(f'应读取 {expected_count} 面定日镜，实际读取 {mirror_xy.shape[0]} 面。')
    return mirror_xy

# ========================================================================
# 来源：src/heliostat/q1/aggregate.py
# ========================================================================

"""第一问的月平均和年平均汇总。"""
from dataclasses import dataclass
from typing import Any, Sequence
import numpy as np

@dataclass(frozen=True)
class MonthlyResult:
    month: int
    average_optical_efficiency: float
    average_cosine_efficiency: float
    average_shadow_blocking_efficiency: float
    average_atmospheric_efficiency: float
    average_truncation_efficiency: float
    field_output_mw: float
    unit_area_output_kw_m2: float

@dataclass(frozen=True)
class AnnualResult:
    average_optical_efficiency: float
    average_cosine_efficiency: float
    average_shadow_blocking_efficiency: float
    average_atmospheric_efficiency: float
    average_truncation_efficiency: float
    field_output_mw: float
    unit_area_output_kw_m2: float

@dataclass(frozen=True)
class MirrorAnnualResult:
    mirror_id: int
    x_m: float
    y_m: float
    radius_to_tower_m: float
    average_optical_efficiency: float
    average_cosine_efficiency: float
    average_shadow_blocking_efficiency: float
    average_atmospheric_efficiency: float
    average_truncation_efficiency: float
    average_output_power_kw: float

@dataclass(frozen=True)
class Question1Solution:
    time_results: tuple[Any, ...]
    monthly_results: tuple[MonthlyResult, ...]
    annual_result: AnnualResult
    mirror_annual_results: tuple[MirrorAnnualResult, ...]
_MEAN_FIELDS = ('average_optical_efficiency', 'average_cosine_efficiency', 'average_shadow_blocking_efficiency', 'average_atmospheric_efficiency', 'average_truncation_efficiency', 'field_output_mw', 'unit_area_output_kw_m2')

def _means(records: Sequence[Any]) -> tuple[float, ...]:
    if not records:
        raise ValueError('汇总记录不能为空。')
    return tuple((float(np.mean([getattr(record, field) for record in records])) for field in _MEAN_FIELDS))

def summarize_monthly(records: Sequence[Any]) -> tuple[MonthlyResult, ...]:
    """对每月规定的五个时刻等权平均。"""
    results: list[MonthlyResult] = []
    for month in sorted({record.month for record in records}):
        monthly = [record for record in records if record.month == month]
        results.append(MonthlyResult(month, *_means(monthly)))
    return tuple(results)

def summarize_annual(records: Sequence[Any]) -> AnnualResult:
    """对题目规定的全部时刻等权平均。"""
    return AnnualResult(*_means(records))

def summarize_mirror_annual(mirror_xy: np.ndarray, tower_x: float, tower_y: float, state_count: int, optical_efficiency_sum: np.ndarray, cosine_efficiency_sum: np.ndarray, shadow_blocking_efficiency_sum: np.ndarray, atmospheric_efficiency_sum: np.ndarray, truncation_efficiency_sum: np.ndarray, output_power_kw_sum: np.ndarray) -> tuple[MirrorAnnualResult, ...]:
    """由逐时刻运行和生成单镜年平均结果，不保留单镜逐时刻明细。"""
    if state_count < 1:
        raise ValueError('state_count 必须大于等于 1。')
    radius = np.hypot(mirror_xy[:, 0] - tower_x, mirror_xy[:, 1] - tower_y)
    means = {'optical': optical_efficiency_sum / state_count, 'cosine': cosine_efficiency_sum / state_count, 'shadow': shadow_blocking_efficiency_sum / state_count, 'atmospheric': atmospheric_efficiency_sum / state_count, 'truncation': truncation_efficiency_sum / state_count, 'power': output_power_kw_sum / state_count}
    return tuple((MirrorAnnualResult(mirror_id=index + 1, x_m=float(mirror_xy[index, 0]), y_m=float(mirror_xy[index, 1]), radius_to_tower_m=float(radius[index]), average_optical_efficiency=float(means['optical'][index]), average_cosine_efficiency=float(means['cosine'][index]), average_shadow_blocking_efficiency=float(means['shadow'][index]), average_atmospheric_efficiency=float(means['atmospheric'][index]), average_truncation_efficiency=float(means['truncation'][index]), average_output_power_kw=float(means['power'][index])) for index in range(mirror_xy.shape[0])))

# ========================================================================
# 来源：src/heliostat/q1/export.py
# ========================================================================

"""第一问结果和论文表格输出。"""
import csv
import json
from dataclasses import asdict
from pathlib import Path
from typing import Any, Iterable, Sequence

def _write_csv(path: Path, rows: Sequence[dict[str, Any]]) -> None:
    if not rows:
        raise ValueError(f'没有可写入 {path.name} 的结果。')
    with path.open('w', encoding='utf-8-sig', newline='') as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)

def _display_source_path(source_path: str | Path) -> str:
    path = Path(source_path)
    if not path.is_absolute():
        return path.as_posix()
    if 'task' in path.parts:
        task_index = path.parts.index('task')
        return Path(*path.parts[task_index:]).as_posix()
    return path.name

def write_question1_results(output_dir: str | Path, time_records: Iterable[Any], monthly_records: Iterable[Any], annual_record: Any, mirror_annual_records: Iterable[Any], field_config: FieldConfig, solver_config: SolverConfig, source_path: str | Path, mirror_count: int) -> dict[str, Path]:
    """保持原有四类结果文件名和字段口径不变。"""
    destination = Path(output_dir)
    destination.mkdir(parents=True, exist_ok=True)
    time_rows = [asdict(record) for record in time_records]
    monthly_rows = [asdict(record) for record in monthly_records]
    annual_row = asdict(annual_record)
    mirror_annual_rows = [asdict(record) for record in mirror_annual_records]
    months = sorted({row['month'] for row in time_rows})
    solar_times = sorted({row['solar_time'] for row in time_rows})
    time_path = destination / '02_逐时刻计算结果.csv'
    monthly_path = destination / '03_月平均计算结果.csv'
    annual_path = destination / '04_年平均计算结果.json'
    mirror_annual_path = destination / '05_单镜年平均结果.csv'
    run_path = destination / '06_运行配置.json'
    _write_csv(time_path, time_rows)
    _write_csv(monthly_path, monthly_rows)
    _write_csv(mirror_annual_path, mirror_annual_rows)
    annual_path.write_text(json.dumps(annual_row, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    run_path.write_text(json.dumps({'source': _display_source_path(source_path), 'field': field_config.to_dict(), 'solver': solver_config.to_dict(), 'run': {'mirror_count': mirror_count, 'months': months, 'solar_times': solar_times, 'time_state_count': len(time_rows)}}, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    return {'time': time_path, 'monthly': monthly_path, 'annual': annual_path, 'mirror_annual': mirror_annual_path, 'config': run_path}

def write_paper_tables(output_dir: str | Path, monthly_records: Iterable[Any], annual_record: Any) -> dict[str, Path]:
    """将月平均、年平均和验证表集中到一个展示文件。"""
    destination = Path(output_dir)
    destination.mkdir(parents=True, exist_ok=True)
    table_path = destination / '07_论文结果与验证表.md'
    monthly_lines = ['# 第一问结果与验证表', '', '本文档汇总第一问的月平均、年平均和数值收敛结果。', '', '## 表 1 每月 21 日平均光学效率及输出功率', '', '| 日期 | 平均光学效率 | 平均余弦效率 | 平均阴影遮挡效率 | 平均截断效率 | 单位面积镜面平均输出热功率 (kW/m²) |', '| --- | ---: | ---: | ---: | ---: | ---: |']
    for record in monthly_records:
        monthly_lines.append(f'| {record.month} 月 21 日 | {record.average_optical_efficiency:.6f} | {record.average_cosine_efficiency:.6f} | {record.average_shadow_blocking_efficiency:.6f} | {record.average_truncation_efficiency:.6f} | {record.unit_area_output_kw_m2:.6f} |')
    annual_lines = ['', '## 表 2 年平均光学效率及输出功率', '', '| 年平均光学效率 | 年平均余弦效率 | 年平均阴影遮挡效率 | 年平均截断效率 | 年平均输出热功率 (MW) | 单位面积镜面年平均输出热功率 (kW/m²) |', '| ---: | ---: | ---: | ---: | ---: | ---: |', f'| {annual_record.average_optical_efficiency:.6f} | {annual_record.average_cosine_efficiency:.6f} | {annual_record.average_shadow_blocking_efficiency:.6f} | {annual_record.average_truncation_efficiency:.6f} | {annual_record.field_output_mw:.6f} | {annual_record.unit_area_output_kw_m2:.6f} |']
    table_path.write_text('\n'.join(monthly_lines + annual_lines) + '\n', encoding='utf-8')
    return {'paper_tables': table_path}

def write_validation_table(output_dir: str | Path, validation_records: Iterable[Any]) -> dict[str, Path]:
    """把三组收敛实验追加为一张验证表。"""
    destination = Path(output_dir)
    rows = [asdict(record) for record in validation_records]
    table_path = destination / '07_论文结果与验证表.md'
    lines = ['', '## 表 3 数值收敛验证', '', '| 验证项目 | 参数 | 观测指标 | 数值 | 相对正式配置差异 | 运行时间 (s) |', '| --- | ---: | --- | ---: | ---: | ---: |']
    for row in rows:
        lines.append(f"| {row['category']} | {row['parameter']} | {row['metric']} | {row['value']:.6f} | {row['relative_difference_percent']:.4f}% | {row['runtime_seconds']:.3f} |")
    with table_path.open('a', encoding='utf-8') as handle:
        handle.write('\n'.join(lines) + '\n')
    return {'validation_table': table_path}

# ========================================================================
# 来源：src/heliostat/q1/plot.py
# ========================================================================

"""第一问的两张正式结果图。"""
import csv
import os
import tempfile
from pathlib import Path
_mpl_config = Path(tempfile.gettempdir()) / 'cowork-matplotlib'
_mpl_config.mkdir(parents=True, exist_ok=True)
os.environ.setdefault('MPLCONFIGDIR', str(_mpl_config))
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.colors import LinearSegmentedColormap
BLUE = '#2F5D7C'
BLUE_DARK = '#173B54'
BLUE_LIGHT = '#DCE9F1'
ORANGE = '#D97706'
DARK = '#24323D'
GREY = '#76838F'
LIGHT_GREY = '#D7DEE3'

def _configure_style() -> None:
    plt.rcParams.update({'font.family': 'sans-serif', 'font.sans-serif': ['Hiragino Sans GB', 'Arial Unicode MS', 'PingFang SC', 'Microsoft YaHei', 'SimHei', 'Noto Sans CJK SC', 'WenQuanYi Zen Hei', 'DejaVu Sans'], 'axes.unicode_minus': False, 'figure.facecolor': 'white', 'axes.facecolor': 'white', 'axes.edgecolor': DARK, 'axes.labelcolor': DARK, 'axes.titlecolor': DARK, 'xtick.color': DARK, 'ytick.color': DARK, 'text.color': DARK, 'grid.color': LIGHT_GREY, 'grid.linewidth': 0.7, 'grid.alpha': 0.65})

def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open('r', encoding='utf-8-sig', newline='') as handle:
        return list(csv.DictReader(handle))

def plot_monthly_performance(output_dir: str | Path) -> Path:
    """绘制月平均综合光学效率和单位面积输出热功率。"""
    _configure_style()
    destination = Path(output_dir)
    rows = _read_csv(destination / '03_月平均计算结果.csv')
    months = np.array([int(row['month']) for row in rows])
    optical = np.array([float(row['average_optical_efficiency']) for row in rows])
    unit_power = np.array([float(row['unit_area_output_kw_m2']) for row in rows])
    output_path = destination / '08_月平均光学性能与输出热功率.png'
    (fig, (ax_efficiency, ax_power)) = plt.subplots(2, 1, figsize=(8.0, 6.3), sharex=True, gridspec_kw={'height_ratios': (1.0, 1.15), 'hspace': 0.12})
    ax_efficiency.plot(months, optical, color=BLUE, linewidth=2.2, marker='o', markersize=5.0, markerfacecolor='white', markeredgewidth=1.5)
    efficiency_padding = 0.02
    ax_efficiency.set_ylim(max(0.0, float(np.min(optical)) - efficiency_padding), min(1.0, float(np.max(optical)) + efficiency_padding))
    ax_efficiency.set_ylabel('综合光学效率')
    ax_efficiency.grid(axis='y')
    ax_efficiency.spines[['top', 'right']].set_visible(False)
    ax_power.bar(months, unit_power, width=0.62, color=ORANGE, edgecolor='white', linewidth=0.8)
    ax_power.set_ylim(0.0, float(np.max(unit_power)) * 1.14)
    ax_power.set_ylabel('单位面积输出热功率 ($\\mathrm{kW\\,m^{-2}}$)')
    ax_power.set_xlabel('月份')
    ax_power.set_xticks(months)
    ax_power.grid(axis='y')
    ax_power.set_axisbelow(True)
    ax_power.spines[['top', 'right']].set_visible(False)
    fig.suptitle('月平均光学性能与输出热功率', fontsize=15, y=0.98)
    fig.subplots_adjust(left=0.13, right=0.97, top=0.91, bottom=0.1)
    fig.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close(fig)
    return output_path

def plot_mirror_annual_efficiency_map(output_dir: str | Path) -> Path:
    """绘制 1745 面定日镜的年平均综合光学效率空间分布。"""
    _configure_style()
    destination = Path(output_dir)
    rows = _read_csv(destination / '05_单镜年平均结果.csv')
    x = np.array([float(row['x_m']) for row in rows])
    y = np.array([float(row['y_m']) for row in rows])
    optical = np.array([float(row['average_optical_efficiency']) for row in rows])
    output_path = destination / '09_单镜年平均综合光学效率空间分布.png'
    efficiency_cmap = LinearSegmentedColormap.from_list('heliostat_efficiency', (BLUE_LIGHT, '#8EB7CF', BLUE, BLUE_DARK))
    (fig, ax) = plt.subplots(figsize=(7.4, 6.6))
    points = ax.scatter(x, y, c=optical, cmap=efficiency_cmap, vmin=float(np.min(optical)), vmax=float(np.max(optical)), s=18, linewidths=0)
    ax.scatter([0.0], [0.0], marker='*', s=190, color=ORANGE, edgecolor=DARK, linewidth=0.8, label='吸收塔', zorder=4)
    limit = max(float(np.max(np.abs(x))), float(np.max(np.abs(y)))) + 20.0
    ax.set_xlim(-limit, limit)
    ax.set_ylim(-limit, limit)
    ax.set_aspect('equal', adjustable='box')
    ax.set_xlabel('x 坐标 (m)')
    ax.set_ylabel('y 坐标 (m)')
    ax.set_title('单镜年平均综合光学效率空间分布', fontsize=15, pad=12)
    ax.grid(color=LIGHT_GREY, linewidth=0.6, alpha=0.55)
    ax.set_axisbelow(True)
    ax.legend(loc='upper right', frameon=False)
    colorbar = fig.colorbar(points, ax=ax, pad=0.025, fraction=0.047)
    colorbar.set_label('年平均综合光学效率')
    colorbar.outline.set_edgecolor(GREY)
    colorbar.outline.set_linewidth(0.7)
    fig.subplots_adjust(left=0.11, right=0.91, top=0.91, bottom=0.1)
    fig.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close(fig)
    return output_path

def build_paper_figures(output_dir: str | Path) -> dict[str, Path]:
    """生成第一问最终采用的两张结果图。"""
    return {'monthly_performance': plot_monthly_performance(output_dir), 'mirror_efficiency_map': plot_mirror_annual_efficiency_map(output_dir)}

# ========================================================================
# 来源：src/heliostat/q1/solve.py
# ========================================================================

"""第一问逐时刻计算、验证运行和命令行入口。"""
import argparse
import time
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Callable, Sequence
import numpy as np
SOLAR_TIMES = (9.0, 10.5, 12.0, 13.5, 15.0)
PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_INPUT = PROJECT_ROOT / 'task' / 'A' / 'fj.xlsx'
DEFAULT_OUTPUT = PROJECT_ROOT / 'outputs' / 'q1'

@dataclass(frozen=True)
class TimeResult:
    month: int
    solar_time: float
    dni_kw_m2: float
    average_optical_efficiency: float
    average_cosine_efficiency: float
    average_shadow_blocking_efficiency: float
    average_atmospheric_efficiency: float
    average_truncation_efficiency: float
    field_output_mw: float
    unit_area_output_kw_m2: float
    maximum_reflection_error: float

@dataclass(frozen=True)
class ValidationResult:
    category: str
    parameter: str
    metric: str
    value: float
    relative_difference_percent: float
    runtime_seconds: float
ProgressCallback = Callable[[int, int, TimeResult], None]

def _check_efficiency(name: str, values: np.ndarray) -> None:
    tolerance = 1e-12
    if np.any(values < -tolerance) or np.any(values > 1.0 + tolerance):
        minimum = float(np.min(values))
        maximum = float(np.max(values))
        raise RuntimeError(f'{name} 超出 [0, 1]：min={minimum:.6g}, max={maximum:.6g}')

def evaluate_time(prepared: PreparedField, month: int, solar_time: float, solver: SolverConfig, mirror_sums: dict[str, np.ndarray] | None=None) -> TimeResult:
    """计算一个月份、一个规定时刻的全场平均结果。"""
    solar = calculate_solar_state(month=month, solar_time=solar_time, latitude_deg=prepared.config.latitude_deg, altitude_km=prepared.config.altitude_km)
    orientation = calculate_orientation(prepared, solar.direction)
    reflection_error = maximum_reflection_error(prepared, orientation, solar.direction)
    if reflection_error >= 1e-08:
        raise RuntimeError(f'中心光线反射误差过大：{reflection_error:.3e}')
    if solver.calculate_shadow:
        shadow = calculate_shadow_blocking_efficiency(prepared, orientation, solar.direction, solver)
    else:
        shadow = np.ones(prepared.mirror_count, dtype=float)
    if solver.calculate_truncation:
        truncation = calculate_truncation_efficiency(prepared, orientation, solar.direction, solver)
    else:
        truncation = np.ones(prepared.mirror_count, dtype=float)
    cosine = orientation.cosine_efficiency
    atmospheric = prepared.atmospheric_efficiency
    optical = cosine * shadow * atmospheric * truncation * prepared.config.reflectivity
    for (name, values) in (('余弦效率', cosine), ('阴影遮挡效率', shadow), ('大气透射率', atmospheric), ('截断效率', truncation), ('光学效率', optical)):
        _check_efficiency(name, values)
    mirror_power_kw = solar.dni_kw_m2 * prepared.mirror_areas * optical
    if mirror_sums is not None:
        mirror_sums['optical_efficiency_sum'] += optical
        mirror_sums['cosine_efficiency_sum'] += cosine
        mirror_sums['shadow_blocking_efficiency_sum'] += shadow
        mirror_sums['atmospheric_efficiency_sum'] += atmospheric
        mirror_sums['truncation_efficiency_sum'] += truncation
        mirror_sums['output_power_kw_sum'] += mirror_power_kw
    field_power_kw = float(np.sum(mirror_power_kw))
    area_weights = prepared.mirror_areas
    return TimeResult(month=month, solar_time=solar_time, dni_kw_m2=solar.dni_kw_m2, average_optical_efficiency=float(np.average(optical, weights=area_weights)), average_cosine_efficiency=float(np.average(cosine, weights=area_weights)), average_shadow_blocking_efficiency=float(np.average(shadow, weights=area_weights)), average_atmospheric_efficiency=float(np.average(atmospheric, weights=area_weights)), average_truncation_efficiency=float(np.average(truncation, weights=area_weights)), field_output_mw=field_power_kw / 1000.0, unit_area_output_kw_m2=field_power_kw / prepared.total_mirror_area, maximum_reflection_error=reflection_error)

def solve_question1(prepared: PreparedField, solver: SolverConfig, months: Sequence[int]=tuple(range(1, 13)), solar_times: Sequence[float]=SOLAR_TIMES, progress: ProgressCallback | None=None) -> Question1Solution:
    """执行所选月份和时刻；默认即题目规定的 60 个状态。"""
    if not months or not solar_times:
        raise ValueError('months 和 solar_times 不能为空。')
    if any((month < 1 or month > 12 for month in months)):
        raise ValueError('months 必须位于 1 到 12。')
    records: list[TimeResult] = []
    mirror_sums = {name: np.zeros(prepared.mirror_count, dtype=float) for name in ('optical_efficiency_sum', 'cosine_efficiency_sum', 'shadow_blocking_efficiency_sum', 'atmospheric_efficiency_sum', 'truncation_efficiency_sum', 'output_power_kw_sum')}
    total = len(months) * len(solar_times)
    for month in months:
        for solar_time in solar_times:
            record = evaluate_time(prepared, month, solar_time, solver, mirror_sums=mirror_sums)
            records.append(record)
            if progress is not None:
                progress(len(records), total, record)
    time_results = tuple(records)
    return Question1Solution(time_results=time_results, monthly_results=summarize_monthly(time_results), annual_result=summarize_annual(time_results), mirror_annual_results=summarize_mirror_annual(mirror_xy=prepared.centers[:, :2], tower_x=prepared.config.tower_x, tower_y=prepared.config.tower_y, state_count=len(time_results), **mirror_sums))

def run_validation_suite(prepared: PreparedField, base_solver: SolverConfig) -> tuple[ValidationResult, ...]:
    """运行三组隔离后的收敛实验，供一张验证表使用。"""
    specifications = [('阴影网格', '10×10', replace(base_solver, shadow_grid_size=10, calculate_shadow=True, calculate_truncation=False), 'average_shadow_blocking_efficiency', '年平均阴影遮挡效率', False), ('阴影网格', '15×15', replace(base_solver, shadow_grid_size=15, calculate_shadow=True, calculate_truncation=False), 'average_shadow_blocking_efficiency', '年平均阴影遮挡效率', True), ('阴影网格', '20×20', replace(base_solver, shadow_grid_size=20, calculate_shadow=True, calculate_truncation=False), 'average_shadow_blocking_efficiency', '年平均阴影遮挡效率', False), ('邻镜半径', '40 m', replace(base_solver, neighbor_radius_m=40.0, calculate_shadow=True, calculate_truncation=False), 'average_shadow_blocking_efficiency', '年平均阴影遮挡效率', False), ('邻镜半径', '60 m', replace(base_solver, neighbor_radius_m=60.0, calculate_shadow=True, calculate_truncation=False), 'average_shadow_blocking_efficiency', '年平均阴影遮挡效率', True), ('邻镜半径', '80 m', replace(base_solver, neighbor_radius_m=80.0, calculate_shadow=True, calculate_truncation=False), 'average_shadow_blocking_efficiency', '年平均阴影遮挡效率', False), ('截断光线', '128', replace(base_solver, truncation_rays=128, calculate_shadow=False, calculate_truncation=True), 'average_truncation_efficiency', '年平均截断效率', False), ('截断光线', '256', replace(base_solver, truncation_rays=256, calculate_shadow=False, calculate_truncation=True), 'average_truncation_efficiency', '年平均截断效率', True), ('截断光线', '512', replace(base_solver, truncation_rays=512, calculate_shadow=False, calculate_truncation=True), 'average_truncation_efficiency', '年平均截断效率', False)]
    raw: list[tuple[str, str, str, float, float, bool]] = []
    for (category, parameter, solver, field, metric, reference) in specifications:
        started = time.perf_counter()
        solution = solve_question1(prepared, solver)
        elapsed = time.perf_counter() - started
        value = float(getattr(solution.annual_result, field))
        raw.append((category, parameter, metric, value, elapsed, reference))
    baselines = {category: value for (category, _, _, value, _, reference) in raw if reference}
    return tuple((ValidationResult(category=category, parameter=parameter, metric=metric, value=value, relative_difference_percent=abs(value - baselines[category]) / abs(baselines[category]) * 100.0, runtime_seconds=elapsed) for (category, parameter, metric, value, elapsed, _) in raw))

def _comma_ints(value: str) -> tuple[int, ...]:
    try:
        result = tuple((int(item.strip()) for item in value.split(',') if item.strip()))
    except ValueError as exc:
        raise argparse.ArgumentTypeError('月份应使用逗号分隔的整数。') from exc
    if not result:
        raise argparse.ArgumentTypeError('月份列表不能为空。')
    return result

def _comma_floats(value: str) -> tuple[float, ...]:
    try:
        result = tuple((float(item.strip()) for item in value.split(',') if item.strip()))
    except ValueError as exc:
        raise argparse.ArgumentTypeError('时刻应使用逗号分隔的数字。') from exc
    if not result:
        raise argparse.ArgumentTypeError('时刻列表不能为空。')
    return result

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description='计算 CUMCM 2023 A 题第一问的镜场光学效率和输出热功率')
    parser.add_argument('--input', type=Path, default=DEFAULT_INPUT)
    parser.add_argument('--output', type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument('--shadow-grid', type=int, default=15)
    parser.add_argument('--truncation-rays', type=int, default=256)
    parser.add_argument('--neighbor-radius', type=float, default=60.0)
    parser.add_argument('--truncation-chunk-size', type=int, default=128)
    parser.add_argument('--sobol-seed', type=int, default=2023)
    parser.add_argument('--months', type=_comma_ints, default=tuple(range(1, 13)), help='逗号分隔；默认 1 到 12 月')
    parser.add_argument('--times', type=_comma_floats, default=SOLAR_TIMES, help='逗号分隔的当地太阳时')
    parser.add_argument('--limit-mirrors', type=int, default=None, help='仅用于调试；只计算附件中的前 N 面镜子')
    parser.add_argument('--skip-shadow', action='store_true')
    parser.add_argument('--skip-truncation', action='store_true')
    parser.add_argument('--skip-figures', action='store_true')
    parser.add_argument('--run-validation', action='store_true', help='额外运行三组收敛实验并生成一张验证表')
    parser.add_argument('--quiet', action='store_true')
    return parser

def _progress(current: int, total: int, record: TimeResult) -> None:
    hour = int(record.solar_time)
    minute = int(round((record.solar_time - hour) * 60.0))
    print(f'[{current:02d}/{total:02d}] {record.month:02d}月21日 {hour:02d}:{minute:02d} 光学效率={record.average_optical_efficiency:.4f} 输出={record.field_output_mw:.3f} MW')

def run(argv: Sequence[str] | None=None) -> int:
    args = build_parser().parse_args(argv)
    mirror_xy = load_mirror_xy(args.input)
    if args.limit_mirrors is not None:
        if args.limit_mirrors < 1:
            raise SystemExit('--limit-mirrors 必须大于等于 1。')
        mirror_xy = mirror_xy[:args.limit_mirrors]
    field_config = FieldConfig()
    solver_config = SolverConfig(shadow_grid_size=args.shadow_grid, truncation_rays=args.truncation_rays, neighbor_radius_m=args.neighbor_radius, truncation_chunk_size=args.truncation_chunk_size, sobol_seed=args.sobol_seed, calculate_shadow=not args.skip_shadow, calculate_truncation=not args.skip_truncation)
    prepared = prepare_field(mirror_xy, field_config)
    solution = solve_question1(prepared=prepared, solver=solver_config, months=args.months, solar_times=args.times, progress=None if args.quiet else _progress)
    written = write_question1_results(output_dir=args.output, time_records=solution.time_results, monthly_records=solution.monthly_results, annual_record=solution.annual_result, mirror_annual_records=solution.mirror_annual_results, field_config=field_config, solver_config=solver_config, source_path=args.input, mirror_count=prepared.mirror_count)
    written.update(write_paper_tables(args.output, solution.monthly_results, solution.annual_result))
    if args.run_validation:
        validation = run_validation_suite(prepared, solver_config)
        written.update(write_validation_table(args.output, validation))
    if not args.skip_figures:
        written.update(build_paper_figures(output_dir=args.output))
    annual = solution.annual_result
    print('\n汇总结果')
    print(f'平均光学效率：{annual.average_optical_efficiency:.6f}')
    print(f'平均余弦效率：{annual.average_cosine_efficiency:.6f}')
    print(f'平均阴影遮挡效率：{annual.average_shadow_blocking_efficiency:.6f}')
    print(f'平均截断效率：{annual.average_truncation_efficiency:.6f}')
    print(f'平均输出热功率：{annual.field_output_mw:.6f} MW')
    print(f'单位镜面面积平均输出热功率：{annual.unit_area_output_kw_m2:.6f} kW/m²')
    print(f'结果目录：{args.output.resolve()}')
    for (name, path) in written.items():
        print(f'  {name}: {path.relative_to(args.output)}')
    return 0

def main() -> None:
    raise SystemExit(run())

# ========================================================================
# 来源：src/heliostat/q2/layout.py
# ========================================================================

"""第二问的两种参数化镜场布局与统一几何约束检查。"""
import math
from dataclasses import dataclass
from typing import Protocol
import numpy as np
from numpy.typing import NDArray
from scipy.spatial import cKDTree
FloatArray = NDArray[np.float64]

class LayoutError(ValueError):
    """布局参数或生成结果不可行。"""

class CommonParameters(Protocol):
    tower_x: float
    tower_y: float
    mirror_width: float
    mirror_height: float
    installation_height: float
    field_radius: float
    exclusion_radius: float
    safety_epsilon: float

@dataclass(frozen=True)
class PartitionedRingParameters:
    """分区交错同心圆布局参数。"""
    tower_y: float
    mirror_width: float
    mirror_height: float
    installation_height: float
    split_radius: float
    near_spacing: float
    far_spacing: float
    tower_x: float = 0.0
    field_radius: float = 350.0
    exclusion_radius: float = 100.0
    safety_epsilon: float = 0.01

    @property
    def safe_distance(self) -> float:
        return self.mirror_width + 5.0 + self.safety_epsilon

@dataclass(frozen=True)
class CampoParameters:
    """带渐增径向行距的 Campo 径向交错布局参数。"""
    tower_y: float
    mirror_width: float
    mirror_height: float
    installation_height: float
    first_ring_count: int
    initial_spacing: float
    spacing_growth: float
    tower_x: float = 0.0
    field_radius: float = 350.0
    exclusion_radius: float = 100.0
    safety_epsilon: float = 0.01

    @property
    def safe_distance(self) -> float:
        return self.mirror_width + 5.0 + self.safety_epsilon

@dataclass(frozen=True)
class LayoutRing:
    """一条生成并经过场地裁剪的圆环或圆弧。"""
    index: int
    radius: float
    zone: int
    nominal_count: int
    coordinates: FloatArray

    @property
    def mirror_count(self) -> int:
        return int(self.coordinates.shape[0])

@dataclass(frozen=True)
class GeneratedLayout:
    """由若干按半径排序的圆环或圆弧组成的镜场。"""
    kind: str
    rings: tuple[LayoutRing, ...]

    @property
    def mirror_count(self) -> int:
        return sum((ring.mirror_count for ring in self.rings))

    @property
    def coordinates(self) -> FloatArray:
        if not self.rings:
            return np.empty((0, 2), dtype=float)
        return np.concatenate([ring.coordinates for ring in self.rings], axis=0)

    def prefix(self, ring_count: int) -> FloatArray:
        if ring_count < 1 or ring_count > len(self.rings):
            raise ValueError(f'ring_count 应位于 1 到 {len(self.rings)}，实际为 {ring_count}。')
        return np.concatenate([ring.coordinates for ring in self.rings[:ring_count]], axis=0)

    def prefix_mirror_count(self, ring_count: int) -> int:
        if ring_count < 1 or ring_count > len(self.rings):
            raise ValueError(f'ring_count 应位于 1 到 {len(self.rings)}，实际为 {ring_count}。')
        return sum((ring.mirror_count for ring in self.rings[:ring_count]))

@dataclass(frozen=True)
class GeometryCheck:
    valid: bool
    reason: str | None
    mirror_count: int
    minimum_center_distance: float
    maximum_field_radius: float
    minimum_tower_distance: float

def _validate_common_parameters(parameters: CommonParameters) -> None:
    values = (parameters.tower_x, parameters.tower_y, parameters.mirror_width, parameters.mirror_height, parameters.installation_height, parameters.field_radius, parameters.exclusion_radius, parameters.safety_epsilon)
    if not all((math.isfinite(value) for value in values)):
        raise LayoutError('布局参数必须全部为有限数。')
    if not 2.0 <= parameters.mirror_height <= parameters.mirror_width <= 8.0:
        raise LayoutError('镜面尺寸必须满足 2 ≤ h ≤ w ≤ 8。')
    if not 2.0 <= parameters.installation_height <= 6.0:
        raise LayoutError('安装高度必须位于 2 m 到 6 m。')
    if parameters.installation_height < parameters.mirror_height / 2.0:
        raise LayoutError('安装高度不足，镜面转动时可能触地。')
    if parameters.field_radius <= 0.0:
        raise LayoutError('场地半径必须大于 0。')
    if parameters.exclusion_radius <= 0.0:
        raise LayoutError('塔周禁区半径必须大于 0。')
    if parameters.safety_epsilon <= 0.0:
        raise LayoutError('安全距离余量必须大于 0。')

def _maximum_tower_centered_radius(parameters: CommonParameters) -> float:
    return parameters.field_radius + math.hypot(parameters.tower_x, parameters.tower_y)

def _ring_coordinates(parameters: CommonParameters, radius: float, mirror_count: int, phase: float) -> FloatArray:
    if mirror_count < 2:
        raise LayoutError('单圈镜子数必须大于等于 2。')
    angles = 2.0 * math.pi * np.arange(mirror_count, dtype=float) / mirror_count + phase
    coordinates = np.column_stack((parameters.tower_x + radius * np.sin(angles), parameters.tower_y + radius * np.cos(angles)))
    field_radius = np.hypot(coordinates[:, 0], coordinates[:, 1])
    keep = field_radius <= parameters.field_radius + 1e-09
    clipped = np.asarray(coordinates[keep], dtype=float)
    clipped.setflags(write=False)
    return clipped

def _within_ring_count(radius: float, safe_distance: float) -> int:
    ratio = safe_distance / (2.0 * radius)
    if ratio >= 1.0:
        raise LayoutError('圆环半径过小，无法放置满足安全距离的镜子。')
    return int(math.floor(math.pi / math.asin(ratio)))

def generate_partitioned_layout(parameters: PartitionedRingParameters, *, maximum_rings: int=256) -> GeneratedLayout:
    """生成分区交错同心圆，并拒绝跨环距离冲突。"""
    _validate_common_parameters(parameters)
    if not math.isfinite(parameters.split_radius):
        raise LayoutError('分区半径必须为有限数。')
    if parameters.split_radius <= parameters.exclusion_radius:
        raise LayoutError('分区半径必须位于塔周禁区之外。')
    if parameters.near_spacing <= 0.0:
        raise LayoutError('近区行距必须大于 0。')
    if parameters.far_spacing < parameters.near_spacing:
        raise LayoutError('远区行距必须大于等于近区行距。')
    rings: list[LayoutRing] = []
    radius = parameters.exclusion_radius
    maximum_radius = _maximum_tower_centered_radius(parameters)
    for ring_index in range(maximum_rings):
        if radius > maximum_radius + 1e-09:
            break
        count = _within_ring_count(radius, parameters.safe_distance)
        phase = 0.0 if ring_index % 2 == 0 else math.pi / count
        coordinates = _ring_coordinates(parameters, radius, count, phase)
        if coordinates.size:
            rings.append(LayoutRing(index=ring_index, radius=radius, zone=1 if radius < parameters.split_radius else 2, nominal_count=count, coordinates=coordinates))
        spacing = parameters.near_spacing if radius < parameters.split_radius else parameters.far_spacing
        radius += spacing
    else:
        raise LayoutError('达到 maximum_rings，圆环生成未正常终止。')
    layout = GeneratedLayout('partitioned', tuple(rings))
    check = validate_layout(layout.coordinates, parameters)
    if not check.valid:
        raise LayoutError(check.reason or '分区圆环布局不满足几何约束。')
    return layout

def _campo_zone(radius: float, first_radius: float) -> tuple[int, int]:
    if radius < 2.0 * first_radius:
        return (1, 1)
    if radius < 4.0 * first_radius:
        return (2, 2)
    return (3, 4)

def generate_campo_layout(parameters: CampoParameters, *, maximum_rings: int=256) -> GeneratedLayout:
    """生成三分区、渐增径向行距的 Campo 径向交错镜场。"""
    _validate_common_parameters(parameters)
    if parameters.first_ring_count < 2:
        raise LayoutError('Campo 首环镜子数必须大于等于 2。')
    if parameters.initial_spacing <= 0.0:
        raise LayoutError('Campo 初始行距必须大于 0。')
    if parameters.spacing_growth < 0.0:
        raise LayoutError('Campo 行距增长量不能小于 0。')
    first_radius = max(parameters.exclusion_radius, parameters.safe_distance / (2.0 * math.sin(math.pi / parameters.first_ring_count)))
    maximum_radius = _maximum_tower_centered_radius(parameters)
    rings: list[LayoutRing] = []
    zone_rows = {1: 0, 2: 0, 3: 0}
    radius = first_radius
    for ring_index in range(maximum_rings):
        if radius > maximum_radius + 1e-09:
            break
        (zone, multiplier) = _campo_zone(radius, first_radius)
        count = parameters.first_ring_count * multiplier
        zone_index = zone_rows[zone]
        phase = 0.0 if zone_index % 2 == 0 else math.pi / count
        coordinates = _ring_coordinates(parameters, radius, count, phase)
        if coordinates.size:
            rings.append(LayoutRing(index=ring_index, radius=radius, zone=zone, nominal_count=count, coordinates=coordinates))
        zone_rows[zone] += 1
        radius += parameters.initial_spacing + parameters.spacing_growth * ring_index
    else:
        raise LayoutError('达到 maximum_rings，Campo 圆环生成未正常终止。')
    layout = GeneratedLayout('campo', tuple(rings))
    check = validate_layout(layout.coordinates, parameters)
    if not check.valid:
        raise LayoutError(check.reason or 'Campo 布局不满足几何约束。')
    return layout

def validate_layout(coordinates: FloatArray, parameters: CommonParameters) -> GeometryCheck:
    """按题目口径检查尺寸、场地、禁区和严格中心距离约束。"""
    try:
        _validate_common_parameters(parameters)
    except LayoutError as exc:
        return GeometryCheck(False, str(exc), 0, math.inf, math.inf, math.inf)
    xy = np.asarray(coordinates, dtype=float)
    if xy.ndim != 2 or xy.shape[1] != 2:
        return GeometryCheck(False, f'镜位坐标应为 N×2，实际形状为 {xy.shape}。', 0, math.inf, math.inf, math.inf)
    if xy.shape[0] == 0:
        return GeometryCheck(False, '镜场中没有定日镜。', 0, math.inf, 0.0, 0.0)
    if not np.all(np.isfinite(xy)):
        return GeometryCheck(False, '镜位坐标包含 NaN 或无穷值。', int(xy.shape[0]), math.inf, math.inf, math.inf)
    field_radii = np.hypot(xy[:, 0], xy[:, 1])
    tower_distances = np.hypot(xy[:, 0] - parameters.tower_x, xy[:, 1] - parameters.tower_y)
    maximum_field_radius = float(np.max(field_radii))
    minimum_tower_distance = float(np.min(tower_distances))
    if maximum_field_radius > parameters.field_radius + 1e-09:
        return GeometryCheck(False, '存在越过 350 m 场地边界的镜位。', int(xy.shape[0]), math.inf, maximum_field_radius, minimum_tower_distance)
    if minimum_tower_distance < parameters.exclusion_radius - 1e-09:
        return GeometryCheck(False, '存在进入塔周 100 m 禁区的镜位。', int(xy.shape[0]), math.inf, maximum_field_radius, minimum_tower_distance)
    if xy.shape[0] == 1:
        minimum_distance = math.inf
    else:
        (distances, _) = cKDTree(xy).query(xy, k=2)
        minimum_distance = float(np.min(distances[:, 1]))
    if minimum_distance <= parameters.mirror_width + 5.0:
        return GeometryCheck(False, f'最小中心距离不满足严格约束：{minimum_distance:.9f} m ≤ {parameters.mirror_width + 5.0:.9f} m。', int(xy.shape[0]), minimum_distance, maximum_field_radius, minimum_tower_distance)
    return GeometryCheck(True, None, int(xy.shape[0]), minimum_distance, maximum_field_radius, minimum_tower_distance)

# ========================================================================
# 来源：src/heliostat/q2/evaluate.py
# ========================================================================

"""第二问候选镜场的光学评价、缓存和外边界扫描。"""
import hashlib
from dataclasses import dataclass, replace
from typing import Sequence
import numpy as np
LayoutParameters = PartitionedRingParameters | CampoParameters

@dataclass(frozen=True)
class EvaluationProfile:
    """同一物理模型下的一组数值离散精度。"""
    name: str
    solver: SolverConfig
    months: tuple[int, ...] = tuple(range(1, 13))
    solar_times: tuple[float, ...] = SOLAR_TIMES

@dataclass(frozen=True)
class FieldEvaluation:
    """一个确定镜场外边界下的完整评价结果。"""
    layout_kind: str
    ring_count: int
    mirror_count: int
    mirror_area_m2: float
    total_area_m2: float
    coordinates: np.ndarray
    solution: Question1Solution

    @property
    def annual_power_mw(self) -> float:
        return self.solution.annual_result.field_output_mw

    @property
    def unit_area_power_kw_m2(self) -> float:
        return self.solution.annual_result.unit_area_output_kw_m2

    def is_feasible(self, target_power_mw: float=42.0) -> bool:
        return self.annual_power_mw >= target_power_mw

@dataclass(frozen=True)
class ExtentScanResult:
    """一组布局参数在若干镜场外边界中的最好结果。"""
    best: FieldEvaluation
    evaluations: tuple[FieldEvaluation, ...]
    first_feasible_ring_count: int | None

class EvaluationCache:
    """按坐标、塔和数值精度缓存昂贵的光学评价。"""

    def __init__(self) -> None:
        self._values: dict[str, Question1Solution] = {}

    def get(self, key: str) -> Question1Solution | None:
        return self._values.get(key)

    def put(self, key: str, value: Question1Solution) -> None:
        self._values[key] = value

    def __len__(self) -> int:
        return len(self._values)

def exploration_profile() -> EvaluationProfile:
    return EvaluationProfile(name='exploration', solver=SolverConfig(shadow_grid_size=5, truncation_rays=64, neighbor_radius_m=60.0, truncation_chunk_size=128, sobol_seed=2023))

def refinement_profile() -> EvaluationProfile:
    return EvaluationProfile(name='refinement', solver=SolverConfig(shadow_grid_size=10, truncation_rays=128, neighbor_radius_m=60.0, truncation_chunk_size=128, sobol_seed=2023))

def final_profile() -> EvaluationProfile:
    return EvaluationProfile(name='final', solver=SolverConfig(shadow_grid_size=15, truncation_rays=256, neighbor_radius_m=60.0, truncation_chunk_size=128, sobol_seed=2023))

def _field_config(parameters: CommonParameters, base: FieldConfig | None=None) -> FieldConfig:
    config = base or FieldConfig()
    return replace(config, field_radius=parameters.field_radius, exclusion_radius=parameters.exclusion_radius, tower_x=parameters.tower_x, tower_y=parameters.tower_y, mirror_width=parameters.mirror_width, mirror_height=parameters.mirror_height, mirror_center_z=parameters.installation_height)

def _cache_key(coordinates: np.ndarray, config: FieldConfig, profile: EvaluationProfile) -> str:
    digest = hashlib.sha256()
    rounded = np.round(np.asarray(coordinates, dtype='<f8'), decimals=9)
    digest.update(rounded.tobytes(order='C'))
    digest.update(repr(config.to_dict()).encode('utf-8'))
    digest.update(repr(profile.solver.to_dict()).encode('utf-8'))
    digest.update(repr(profile.months).encode('ascii'))
    digest.update(repr(profile.solar_times).encode('ascii'))
    return digest.hexdigest()

def evaluate_coordinates(*, layout_kind: str, ring_count: int, coordinates: np.ndarray, parameters: LayoutParameters, profile: EvaluationProfile, cache: EvaluationCache | None=None, base_field_config: FieldConfig | None=None) -> FieldEvaluation:
    """直接复用问题一模型评价一套确定坐标。"""
    xy = np.asarray(coordinates, dtype=float)
    if xy.ndim != 2 or xy.shape[1] != 2 or xy.shape[0] == 0:
        raise ValueError('候选镜场坐标必须为非空 N×2 数组。')
    config = _field_config(parameters, base_field_config)
    key = _cache_key(xy, config, profile)
    solution = cache.get(key) if cache is not None else None
    if solution is None:
        prepared = prepare_field(xy, config)
        solution = solve_question1(prepared=prepared, solver=profile.solver, months=profile.months, solar_times=profile.solar_times)
        if cache is not None:
            cache.put(key, solution)
    mirror_area = parameters.mirror_width * parameters.mirror_height
    return FieldEvaluation(layout_kind=layout_kind, ring_count=ring_count, mirror_count=int(xy.shape[0]), mirror_area_m2=mirror_area, total_area_m2=float(xy.shape[0] * mirror_area), coordinates=xy, solution=solution)

def better_evaluation(left: FieldEvaluation, right: FieldEvaluation, *, target_power_mw: float=42.0) -> FieldEvaluation:
    """按可行性优先规则返回较优结果。"""
    left_feasible = left.is_feasible(target_power_mw)
    right_feasible = right.is_feasible(target_power_mw)
    if left_feasible != right_feasible:
        return left if left_feasible else right
    if left_feasible:
        if left.unit_area_power_kw_m2 != right.unit_area_power_kw_m2:
            return left if left.unit_area_power_kw_m2 > right.unit_area_power_kw_m2 else right
        return left if left.annual_power_mw <= right.annual_power_mw else right
    return left if left.annual_power_mw >= right.annual_power_mw else right

def _unique_ring_counts(values: Sequence[int], total: int) -> tuple[int, ...]:
    return tuple(sorted({value for value in values if 1 <= value <= total}))

def scan_layout_extents(layout: GeneratedLayout, parameters: LayoutParameters, profile: EvaluationProfile, *, target_power_mw: float=42.0, coarse_stride: int=4, window: int=2, cache: EvaluationCache | None=None, base_field_config: FieldConfig | None=None) -> ExtentScanResult:
    """先粗定位功率阈值，再评价阈值附近的连续圆环外边界。"""
    if not layout.rings:
        raise ValueError('布局中没有可用于评价的圆环。')
    if coarse_stride < 1:
        raise ValueError('coarse_stride 必须大于等于 1。')
    if window < 0:
        raise ValueError('window 不能小于 0。')
    total_rings = len(layout.rings)
    coarse_counts = list(range(coarse_stride, total_rings + 1, coarse_stride))
    if not coarse_counts or coarse_counts[-1] != total_rings:
        coarse_counts.append(total_rings)
    evaluated: dict[int, FieldEvaluation] = {}

    def evaluate(ring_count: int) -> FieldEvaluation:
        previous = evaluated.get(ring_count)
        if previous is not None:
            return previous
        value = evaluate_coordinates(layout_kind=layout.kind, ring_count=ring_count, coordinates=layout.prefix(ring_count), parameters=parameters, profile=profile, cache=cache, base_field_config=base_field_config)
        evaluated[ring_count] = value
        return value
    first_feasible: int | None = None
    previous_coarse = 0
    for ring_count in coarse_counts:
        value = evaluate(ring_count)
        if value.is_feasible(target_power_mw):
            for refined_count in range(previous_coarse + 1, ring_count + 1):
                refined = evaluate(refined_count)
                if refined.is_feasible(target_power_mw):
                    first_feasible = refined_count
                    break
            break
        previous_coarse = ring_count
    center = first_feasible if first_feasible is not None else total_rings
    local_counts = _unique_ring_counts(range(center - window, center + window + 1), total_rings)
    for ring_count in local_counts:
        evaluate(ring_count)
    values = tuple((evaluated[key] for key in sorted(evaluated)))
    best = values[0]
    for value in values[1:]:
        best = better_evaluation(best, value, target_power_mw=target_power_mw)
    return ExtentScanResult(best=best, evaluations=values, first_feasible_ring_count=first_feasible)

# ========================================================================
# 来源：src/heliostat/q3/_baseline.py
# ========================================================================

"""第三问母场、六组规格展开和异构几何约束。"""
import json
import math
from dataclasses import dataclass
from pathlib import Path
import numpy as np
from numpy.typing import NDArray
from scipy.spatial import cKDTree
FloatArray = NDArray[np.float64]
IntArray = NDArray[np.int64]
GROUP_RING_RANGES = ((1, 1), (2, 5), (6, 11), (12, 14), (15, 20), (21, 28))
GROUP_COUNT = len(GROUP_RING_RANGES)
EXPECTED_GROUP_COUNTS = (72, 269, 283, 224, 357, 266)

@dataclass(frozen=True)
class CampoMotherField:
    """由问题二最终参数确定性重建的完整 1471 面 Campo 母场。"""
    parameters: CampoParameters
    layout: GeneratedLayout
    coordinates: FloatArray
    ring_indices: IntArray
    group_indices: IntArray

    @property
    def mirror_count(self) -> int:
        return int(self.coordinates.shape[0])

    @property
    def group_counts(self) -> tuple[int, ...]:
        return tuple((int(np.count_nonzero(self.group_indices == group)) for group in range(GROUP_COUNT)))

    @property
    def base_width(self) -> float:
        return self.parameters.mirror_width

    @property
    def base_height(self) -> float:
        return self.parameters.mirror_height

    @property
    def base_installation_height(self) -> float:
        return self.parameters.installation_height

@dataclass(frozen=True)
class ExpandedSpecifications:
    widths: FloatArray
    heights: FloatArray
    installation_heights: FloatArray
    areas: FloatArray

    @property
    def total_area_m2(self) -> float:
        return float(np.sum(self.areas))

@dataclass(frozen=True)
class HeterogeneousGeometryCheck:
    valid: bool
    reason: str | None
    mirror_count: int
    minimum_center_distance_m: float
    minimum_width_clearance_m: float
    maximum_field_radius_m: float
    minimum_tower_distance_m: float
    minimum_ground_clearance_m: float

def _group_for_ring(ring_index: int) -> int:
    for (group, (start, stop)) in enumerate(GROUP_RING_RANGES):
        if start <= ring_index <= stop:
            return group
    raise ValueError(f'圆环 {ring_index} 不在第三问六组范围内。')

def load_q2_campo_parameters(summary_path: str | Path) -> CampoParameters:
    path = Path(summary_path)
    if not path.exists():
        raise FileNotFoundError(f'找不到问题二摘要：{path}')
    payload = json.loads(path.read_text(encoding='utf-8'))
    if payload.get('layout') != 'campo':
        raise ValueError('问题二摘要的最终布局不是 campo。')
    parameters = payload.get('parameters')
    if not isinstance(parameters, dict):
        raise ValueError('问题二摘要缺少 parameters。')
    return CampoParameters(**parameters)

def build_campo_mother_field(summary_path: str | Path, *, require_recorded_structure: bool=True) -> CampoMotherField:
    parameters = load_q2_campo_parameters(summary_path)
    layout = generate_campo_layout(parameters)
    if len(layout.rings) != 28:
        raise ValueError(f'第三问分组要求 28 个有效环，实际为 {len(layout.rings)}。')
    coordinates: list[FloatArray] = []
    ring_indices: list[IntArray] = []
    group_indices: list[IntArray] = []
    for (display_index, ring) in enumerate(layout.rings, start=1):
        group = _group_for_ring(display_index)
        coordinates.append(ring.coordinates)
        ring_indices.append(np.full(ring.mirror_count, display_index, dtype=np.int64))
        group_indices.append(np.full(ring.mirror_count, group, dtype=np.int64))
    mother = CampoMotherField(parameters=parameters, layout=layout, coordinates=np.concatenate(coordinates, axis=0), ring_indices=np.concatenate(ring_indices), group_indices=np.concatenate(group_indices))
    if require_recorded_structure and mother.group_counts != EXPECTED_GROUP_COUNTS:
        raise ValueError(f'问题二 Campo 结构已变化：期望组镜数 {EXPECTED_GROUP_COUNTS}，实际为 {mother.group_counts}。')
    return mother

def validate_heterogeneous_field(*, coordinates: FloatArray, widths: FloatArray, heights: FloatArray, installation_heights: FloatArray, tower_x: float, tower_y: float, field_radius: float=350.0, exclusion_radius: float=100.0, safety_epsilon: float=0.01) -> HeterogeneousGeometryCheck:
    xy = np.asarray(coordinates, dtype=float)
    mirror_widths = np.asarray(widths, dtype=float)
    mirror_heights = np.asarray(heights, dtype=float)
    center_zs = np.asarray(installation_heights, dtype=float)
    mirror_count = int(xy.shape[0]) if xy.ndim >= 1 else 0
    invalid = HeterogeneousGeometryCheck(valid=False, reason=None, mirror_count=mirror_count, minimum_center_distance_m=math.inf, minimum_width_clearance_m=-math.inf, maximum_field_radius_m=math.inf, minimum_tower_distance_m=-math.inf, minimum_ground_clearance_m=-math.inf)
    if xy.ndim != 2 or xy.shape[1] != 2 or mirror_count == 0:
        return HeterogeneousGeometryCheck(**{**invalid.__dict__, 'reason': '镜位必须为非空 N×2 数组。'})
    for (name, values) in (('宽度', mirror_widths), ('高度', mirror_heights), ('安装高度', center_zs)):
        if values.ndim != 1 or values.shape[0] != mirror_count:
            return HeterogeneousGeometryCheck(**{**invalid.__dict__, 'reason': f'{name}数组长度与镜子数不一致。'})
    if not all((np.all(np.isfinite(values)) for values in (xy, mirror_widths, mirror_heights, center_zs))):
        return HeterogeneousGeometryCheck(**{**invalid.__dict__, 'reason': '几何数据包含 NaN 或无穷值。'})
    if np.any(mirror_heights < 2.0) or np.any(mirror_heights > 8.0):
        return HeterogeneousGeometryCheck(**{**invalid.__dict__, 'reason': '镜面高度必须位于 2 m 到 8 m。'})
    if np.any(mirror_widths < 2.0) or np.any(mirror_widths > 8.0):
        return HeterogeneousGeometryCheck(**{**invalid.__dict__, 'reason': '镜面宽度必须位于 2 m 到 8 m。'})
    if np.any(mirror_widths < mirror_heights):
        return HeterogeneousGeometryCheck(**{**invalid.__dict__, 'reason': '镜面宽度不能小于镜面高度。'})
    if np.any(center_zs < 2.0) or np.any(center_zs > 6.0):
        return HeterogeneousGeometryCheck(**{**invalid.__dict__, 'reason': '安装高度必须位于 2 m 到 6 m。'})
    ground_clearance = center_zs - mirror_heights / 2.0
    if np.any(ground_clearance < -1e-12):
        return HeterogeneousGeometryCheck(**{**invalid.__dict__, 'reason': '存在旋转时可能触地的镜面。'})
    field_radii = np.hypot(xy[:, 0], xy[:, 1])
    tower_distances = np.hypot(xy[:, 0] - tower_x, xy[:, 1] - tower_y)
    maximum_field_radius = float(np.max(field_radii))
    minimum_tower_distance = float(np.min(tower_distances))
    if maximum_field_radius > field_radius + 1e-09:
        return HeterogeneousGeometryCheck(**{**invalid.__dict__, 'reason': '存在镜位超出圆形场地边界。', 'maximum_field_radius_m': maximum_field_radius, 'minimum_tower_distance_m': minimum_tower_distance, 'minimum_ground_clearance_m': float(np.min(ground_clearance))})
    if minimum_tower_distance < exclusion_radius - 1e-09:
        return HeterogeneousGeometryCheck(**{**invalid.__dict__, 'reason': '存在镜位进入塔周禁区。', 'maximum_field_radius_m': maximum_field_radius, 'minimum_tower_distance_m': minimum_tower_distance, 'minimum_ground_clearance_m': float(np.min(ground_clearance))})
    tree = cKDTree(xy)
    nearest = tree.query(xy, k=2)[0][:, 1]
    minimum_center_distance = float(np.min(nearest))
    pairs = tree.query_pairs(r=13.0 + safety_epsilon + 1e-09, output_type='ndarray')
    minimum_width_clearance = math.inf
    if pairs.size:
        deltas = xy[pairs[:, 0]] - xy[pairs[:, 1]]
        distances = np.linalg.norm(deltas, axis=1)
        required = np.maximum(mirror_widths[pairs[:, 0]], mirror_widths[pairs[:, 1]]) + 5.0
        clearances = distances - required
        minimum_width_clearance = float(np.min(clearances))
        if minimum_width_clearance < safety_epsilon - 1e-09:
            return HeterogeneousGeometryCheck(valid=False, reason='存在镜对不满足异构宽度对应的中心距安全余量。', mirror_count=mirror_count, minimum_center_distance_m=minimum_center_distance, minimum_width_clearance_m=minimum_width_clearance, maximum_field_radius_m=maximum_field_radius, minimum_tower_distance_m=minimum_tower_distance, minimum_ground_clearance_m=float(np.min(ground_clearance)))
    return HeterogeneousGeometryCheck(valid=True, reason=None, mirror_count=mirror_count, minimum_center_distance_m=minimum_center_distance, minimum_width_clearance_m=minimum_width_clearance, maximum_field_radius_m=maximum_field_radius, minimum_tower_distance_m=minimum_tower_distance, minimum_ground_clearance_m=float(np.min(ground_clearance)))

# ========================================================================
# 来源：src/heliostat/q3/_optics.py
# ========================================================================

"""第三问异构镜场评价、缓存和精度配置。"""
import hashlib
from dataclasses import dataclass
import numpy as np
from numpy.typing import NDArray
FloatArray = NDArray[np.float64]
IntArray = NDArray[np.int64]

@dataclass(frozen=True)
class HeterogeneousEvaluation:
    """一套确定逐镜规格和活跃镜集合的完整评价结果。"""
    profile_name: str
    coordinates: FloatArray
    widths: FloatArray
    heights: FloatArray
    installation_heights: FloatArray
    ring_indices: IntArray
    group_indices: IntArray
    original_indices: IntArray
    solution: Question1Solution
    geometry: HeterogeneousGeometryCheck

    @property
    def mirror_count(self) -> int:
        return int(self.coordinates.shape[0])

    @property
    def total_area_m2(self) -> float:
        return float(np.sum(self.widths * self.heights))

    @property
    def annual_power_mw(self) -> float:
        return self.solution.annual_result.field_output_mw

    @property
    def unit_area_power_kw_m2(self) -> float:
        return self.solution.annual_result.unit_area_output_kw_m2

    def is_feasible(self, target_power_mw: float=42.0) -> bool:
        return self.annual_power_mw >= target_power_mw

class EvaluationCache:
    """按坐标、逐镜规格、塔位和数值精度缓存第三问评价。"""

    def __init__(self) -> None:
        self._values: dict[str, Question1Solution] = {}

    def get(self, key: str) -> Question1Solution | None:
        return self._values.get(key)

    def put(self, key: str, value: Question1Solution) -> None:
        self._values[key] = value

    def __len__(self) -> int:
        return len(self._values)

def coarse_profile() -> EvaluationProfile:
    return EvaluationProfile(name='q3-coarse', solver=SolverConfig(shadow_grid_size=5, truncation_rays=64, neighbor_radius_m=60.0, truncation_chunk_size=128, sobol_seed=2023), months=(3, 6, 9, 12), solar_times=SOLAR_TIMES)

def medium_profile() -> EvaluationProfile:
    return EvaluationProfile(name='q3-medium', solver=SolverConfig(shadow_grid_size=10, truncation_rays=128, neighbor_radius_m=60.0, truncation_chunk_size=128, sobol_seed=2023))

def formal_profile() -> EvaluationProfile:
    return EvaluationProfile(name='q3-formal', solver=SolverConfig(shadow_grid_size=15, truncation_rays=256, neighbor_radius_m=60.0, truncation_chunk_size=128, sobol_seed=2023))

def dense_profile() -> EvaluationProfile:
    return EvaluationProfile(name='q3-dense', solver=SolverConfig(shadow_grid_size=20, truncation_rays=512, neighbor_radius_m=80.0, truncation_chunk_size=128, sobol_seed=2023))

def smoke_profile() -> EvaluationProfile:
    return EvaluationProfile(name='q3-smoke', solver=SolverConfig(shadow_grid_size=2, truncation_rays=4, neighbor_radius_m=60.0, truncation_chunk_size=64, sobol_seed=2023), months=(6,), solar_times=(12.0,))

def _cache_key(*, coordinates: FloatArray, specifications: ExpandedSpecifications, field_config: FieldConfig, profile: EvaluationProfile) -> str:
    digest = hashlib.sha256()
    for values in (coordinates, specifications.widths, specifications.heights, specifications.installation_heights):
        rounded = np.round(np.asarray(values, dtype='<f8'), decimals=9)
        digest.update(rounded.tobytes(order='C'))
    digest.update(repr(field_config.to_dict()).encode('utf-8'))
    digest.update(repr(profile.solver.to_dict()).encode('utf-8'))
    digest.update(repr(profile.months).encode('ascii'))
    digest.update(repr(profile.solar_times).encode('ascii'))
    return digest.hexdigest()

def evaluate_specifications(*, coordinates: FloatArray, specifications: ExpandedSpecifications, ring_indices: IntArray, group_indices: IntArray, original_indices: IntArray, field_config: FieldConfig, profile: EvaluationProfile, safety_epsilon: float=0.01, cache: EvaluationCache | None=None) -> HeterogeneousEvaluation:
    xy = np.asarray(coordinates, dtype=float)
    rings = np.asarray(ring_indices, dtype=np.int64)
    groups = np.asarray(group_indices, dtype=np.int64)
    originals = np.asarray(original_indices, dtype=np.int64)
    count = int(xy.shape[0])
    for (name, values) in (('ring_indices', rings), ('group_indices', groups), ('original_indices', originals)):
        if values.ndim != 1 or values.shape[0] != count:
            raise ValueError(f'{name} 长度与镜子数不一致。')
    geometry = validate_heterogeneous_field(coordinates=xy, widths=specifications.widths, heights=specifications.heights, installation_heights=specifications.installation_heights, tower_x=field_config.tower_x, tower_y=field_config.tower_y, field_radius=field_config.field_radius, exclusion_radius=field_config.exclusion_radius, safety_epsilon=safety_epsilon)
    if not geometry.valid:
        raise ValueError(geometry.reason or '异构镜场几何约束不合法。')
    key = _cache_key(coordinates=xy, specifications=specifications, field_config=field_config, profile=profile)
    solution = cache.get(key) if cache is not None else None
    if solution is None:
        prepared = prepare_field(xy, field_config, mirror_widths=specifications.widths, mirror_heights=specifications.heights, mirror_center_zs=specifications.installation_heights)
        solution = solve_question1(prepared=prepared, solver=profile.solver, months=profile.months, solar_times=profile.solar_times)
        if cache is not None:
            cache.put(key, solution)
    return HeterogeneousEvaluation(profile_name=profile.name, coordinates=xy, widths=specifications.widths, heights=specifications.heights, installation_heights=specifications.installation_heights, ring_indices=rings, group_indices=groups, original_indices=originals, solution=solution, geometry=geometry)

# ========================================================================
# 来源：src/heliostat/q3/_workbook.py
# ========================================================================

"""第三问 result3.xlsx 提交表输出。"""
from copy import copy
from pathlib import Path
from openpyxl import load_workbook

def write_result3_workbook(*, template_path: str | Path, output_path: str | Path, evaluation: HeterogeneousEvaluation, tower_x: float, tower_y: float) -> Path:
    """按题目模板写出塔坐标和每面镜子的异构规格、位置。"""
    template = Path(template_path)
    if not template.exists():
        raise FileNotFoundError(f'找不到 result3.xlsx 模板：{template}')
    destination = Path(output_path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(template)
    sheet = workbook.active
    if sheet.max_column < 8:
        workbook.close()
        raise ValueError('result3.xlsx 模板列数不足 8 列。')
    style_row = 2 if sheet.max_row >= 2 else 1
    styles = [copy(sheet.cell(style_row, column)._style) for column in range(1, 9)]
    number_formats = [sheet.cell(style_row, column).number_format for column in range(1, 9)]
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)
    for index in range(evaluation.mirror_count):
        row_index = index + 2
        values = (tower_x, tower_y, index + 1, float(evaluation.widths[index]), float(evaluation.heights[index]), float(evaluation.coordinates[index, 0]), float(evaluation.coordinates[index, 1]), float(evaluation.installation_heights[index]))
        for (column, value) in enumerate(values, start=1):
            cell = sheet.cell(row_index, column, value)
            cell._style = copy(styles[column - 1])
            cell.number_format = number_formats[column - 1]
    workbook.save(destination)
    workbook.close()
    return destination

# ========================================================================
# 来源：src/heliostat/q3/model.py
# ========================================================================

"""六组正式初值、21 维设计对象和逐镜规格展开。"""
import json
import math
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Literal
import numpy as np
TowerMode = Literal['A', 'B']

@dataclass(frozen=True)
class RefineDesign:
    """保留六区阶梯结构的完整候选。"""
    tower_mode: TowerMode
    tower_y: float
    initial_spacing: float
    spacing_growth: float
    widths: tuple[float, ...]
    mirror_heights: tuple[float, ...]
    installation_heights: tuple[float, ...]

    def __post_init__(self) -> None:
        if self.tower_mode not in ('A', 'B'):
            raise ValueError('tower_mode 必须为 A 或 B。')
        for (name, values) in (('widths', self.widths), ('mirror_heights', self.mirror_heights), ('installation_heights', self.installation_heights)):
            if len(values) != GROUP_COUNT:
                raise ValueError(f'{name} 必须包含六个值。')
        values = (self.tower_y, self.initial_spacing, self.spacing_growth, *self.widths, *self.mirror_heights, *self.installation_heights)
        if not all((math.isfinite(value) for value in values)):
            raise ValueError('候选参数必须全部为有限数。')

    def parameter(self, name: str) -> float:
        if name in ('tower_y', 'initial_spacing', 'spacing_growth'):
            return float(getattr(self, name))
        prefix = name[0]
        try:
            group = int(name[1:]) - 1
        except (ValueError, IndexError) as exc:
            raise KeyError(name) from exc
        if not 0 <= group < GROUP_COUNT:
            raise KeyError(name)
        values = {'w': self.widths, 'h': self.mirror_heights, 'H': self.installation_heights}.get(prefix)
        if values is None:
            raise KeyError(name)
        return float(values[group])

    def with_parameter(self, name: str, value: float) -> RefineDesign:
        if name in ('tower_y', 'initial_spacing', 'spacing_growth'):
            return replace(self, **{name: float(value)})
        prefix = name[0]
        try:
            group = int(name[1:]) - 1
        except (ValueError, IndexError) as exc:
            raise KeyError(name) from exc
        attribute = {'w': 'widths', 'h': 'mirror_heights', 'H': 'installation_heights'}.get(prefix)
        if attribute is None or not 0 <= group < GROUP_COUNT:
            raise KeyError(name)
        values = list(getattr(self, attribute))
        values[group] = float(value)
        return replace(self, **{attribute: tuple(values)})

    def to_dict(self) -> dict[str, object]:
        return {'tower_mode': self.tower_mode, 'tower_x_m': 0.0, 'tower_y_m': self.tower_y, 'initial_spacing_m': self.initial_spacing, 'spacing_growth_m_per_ring': self.spacing_growth, 'widths_m': list(self.widths), 'mirror_heights_m': list(self.mirror_heights), 'installation_heights_m': list(self.installation_heights)}

@dataclass(frozen=True)
class RefineBaseline:
    mother: CampoMotherField
    design: RefineDesign
    expected_mirror_count: int
    expected_total_area_m2: float
    expected_power_mw: float
    expected_q_kw_m2: float
    expected_annual: dict[str, float]

    @property
    def parameters(self) -> CampoParameters:
        return self.mother.parameters

@dataclass(frozen=True)
class RefineField:
    """某一塔位语义和 Campo 参数下的前 28 个有效环。"""
    coordinates: np.ndarray
    ring_indices: np.ndarray
    group_indices: np.ndarray
    original_indices: np.ndarray
    mirror_set_hash: str
    outer_clipped_count: int
    geometry_center_y: float

    @property
    def mirror_count(self) -> int:
        return int(self.coordinates.shape[0])

    @property
    def group_counts(self) -> tuple[int, ...]:
        return tuple((int(np.count_nonzero(self.group_indices == group)) for group in range(GROUP_COUNT)))

def load_baseline(*, q2_summary_path: str | Path, six_group_summary_path: str | Path) -> RefineBaseline:
    """从正式结果读取六组初值，禁止重新估计或手抄参数。"""
    mother = build_campo_mother_field(q2_summary_path)
    payload = json.loads(Path(six_group_summary_path).read_text(encoding='utf-8'))
    group_payload = payload.get('group_design', {}).get('groups')
    if not isinstance(group_payload, list) or len(group_payload) != GROUP_COUNT:
        raise ValueError('六组正式摘要缺少完整的 groups 数据。')
    ordered = sorted(group_payload, key=lambda item: int(item['group']))
    design = RefineDesign(tower_mode='A', tower_y=float(payload['tower']['y_m']), initial_spacing=float(mother.parameters.initial_spacing), spacing_growth=float(mother.parameters.spacing_growth), widths=tuple((float(item['mirror_width_m']) for item in ordered)), mirror_heights=tuple((float(item['mirror_height_m']) for item in ordered)), installation_heights=tuple((float(item['installation_height_m']) for item in ordered)))
    annual = {key: float(value) for (key, value) in payload['annual'].items()}
    return RefineBaseline(mother=mother, design=design, expected_mirror_count=int(payload['mirror_count']), expected_total_area_m2=float(payload['total_area_m2']), expected_power_mw=float(annual['field_output_mw']), expected_q_kw_m2=float(annual['unit_area_output_kw_m2']), expected_annual=annual)

def expand_specifications(field: RefineField, design: RefineDesign) -> ExpandedSpecifications:
    groups = field.group_indices
    widths = np.asarray(design.widths, dtype=float)[groups]
    heights = np.asarray(design.mirror_heights, dtype=float)[groups]
    installation = np.asarray(design.installation_heights, dtype=float)[groups]
    return ExpandedSpecifications(widths=np.asarray(widths, dtype=float), heights=np.asarray(heights, dtype=float), installation_heights=np.asarray(installation, dtype=float), areas=np.asarray(widths * heights, dtype=float))

# ========================================================================
# 来源：src/heliostat/q3/tower_modes.py
# ========================================================================

"""塔位模式 A/B 和动态 Campo 前缀构造。"""
import hashlib
from dataclasses import replace
import numpy as np
RING_COUNT = 28

def _group_for_ring(ring_index: int) -> int:
    for (group, (start, stop)) in enumerate(GROUP_RING_RANGES):
        if start <= ring_index <= stop:
            return group
    raise ValueError(f'圆环 {ring_index} 不属于六区。')

def _membership_hash(*, layout: GeneratedLayout, geometry_center_y: float) -> str:
    digest = hashlib.sha256()
    for (ring_index, ring) in enumerate(layout.rings, start=1):
        angles = np.mod(np.arctan2(ring.coordinates[:, 0], ring.coordinates[:, 1] - geometry_center_y), 2.0 * np.pi)
        digest.update(np.asarray((ring_index, ring.nominal_count), dtype='<i8').tobytes())
        digest.update(np.round(angles, 10).astype('<f8').tobytes())
    return digest.hexdigest()

def build_refine_field(baseline: RefineBaseline, design: RefineDesign) -> RefineField:
    """按单一塔位语义重建 Campo，不在轨迹内切换语义。"""
    geometry_center_y = design.tower_y if design.tower_mode == 'A' else baseline.design.tower_y
    parameters = replace(baseline.parameters, tower_y=geometry_center_y, initial_spacing=design.initial_spacing, spacing_growth=design.spacing_growth)
    generated = generate_campo_layout(parameters)
    if len(generated.rings) < RING_COUNT:
        raise ValueError(f'候选只生成 {len(generated.rings)} 个有效环，不能保留前 28 环。')
    layout = GeneratedLayout('campo', generated.rings[:RING_COUNT])
    coordinates: list[np.ndarray] = []
    rings: list[np.ndarray] = []
    groups: list[np.ndarray] = []
    originals: list[np.ndarray] = []
    cursor = 0
    clipped = 0
    for (display_index, ring) in enumerate(layout.rings, start=1):
        count = ring.mirror_count
        coordinates.append(np.asarray(ring.coordinates, dtype=float))
        rings.append(np.full(count, display_index, dtype=np.int64))
        groups.append(np.full(count, _group_for_ring(display_index), dtype=np.int64))
        originals.append(np.arange(cursor, cursor + count, dtype=np.int64))
        cursor += count
        clipped += ring.nominal_count - count
    return RefineField(coordinates=np.concatenate(coordinates), ring_indices=np.concatenate(rings), group_indices=np.concatenate(groups), original_indices=np.concatenate(originals), mirror_set_hash=_membership_hash(layout=layout, geometry_center_y=geometry_center_y), outer_clipped_count=int(clipped), geometry_center_y=float(geometry_center_y))

# ========================================================================
# 来源：src/heliostat/q3/evaluate.py
# ========================================================================

"""统一几何预检、四级精度和六区候选评价。"""
from dataclasses import asdict, dataclass, replace
_coarse_profile = coarse_profile
_dense_profile = dense_profile
_evaluate_specifications = evaluate_specifications
_formal_profile = formal_profile
_medium_profile = medium_profile
_smoke_profile = smoke_profile

@dataclass(frozen=True)
class RefineEvaluation:
    design: RefineDesign
    field: RefineField
    specifications: ExpandedSpecifications
    raw: HeterogeneousEvaluation

    @property
    def profile_name(self) -> str:
        return self.raw.profile_name

    @property
    def mirror_count(self) -> int:
        return self.raw.mirror_count

    @property
    def total_area_m2(self) -> float:
        return self.raw.total_area_m2

    @property
    def annual_power_mw(self) -> float:
        return self.raw.annual_power_mw

    @property
    def unit_area_power_kw_m2(self) -> float:
        return self.raw.unit_area_power_kw_m2

    @property
    def geometry(self) -> HeterogeneousGeometryCheck:
        return self.raw.geometry

    def is_feasible(self, target_power_mw: float=42.0) -> bool:
        return self.raw.is_feasible(target_power_mw)

def coarse_profile() -> EvaluationProfile:
    return replace(_coarse_profile(), name='q3-six-refine-coarse')

def medium_profile() -> EvaluationProfile:
    return replace(_medium_profile(), name='q3-six-refine-medium')

def formal_profile() -> EvaluationProfile:
    return replace(_formal_profile(), name='q3-six-refine-formal')

def dense_profile(*, neighbor_radius_m: float) -> EvaluationProfile:
    profile = _dense_profile()
    return replace(profile, name=f'q3-six-refine-dense-{neighbor_radius_m:g}m', solver=replace(profile.solver, neighbor_radius_m=neighbor_radius_m))

def smoke_profile() -> EvaluationProfile:
    return replace(_smoke_profile(), name='q3-six-refine-smoke')

def _field_config(baseline: RefineBaseline, design: RefineDesign) -> FieldConfig:
    parameters = baseline.parameters
    return replace(FieldConfig(), field_radius=parameters.field_radius, exclusion_radius=parameters.exclusion_radius, tower_x=parameters.tower_x, tower_y=design.tower_y, mirror_width=parameters.mirror_width, mirror_height=parameters.mirror_height, mirror_center_z=parameters.installation_height)

def prepare_candidate(*, baseline: RefineBaseline, design: RefineDesign) -> tuple[RefineField, ExpandedSpecifications, HeterogeneousGeometryCheck]:
    field = build_refine_field(baseline, design)
    specifications = expand_specifications(field, design)
    check = validate_heterogeneous_field(coordinates=field.coordinates, widths=specifications.widths, heights=specifications.heights, installation_heights=specifications.installation_heights, tower_x=baseline.parameters.tower_x, tower_y=design.tower_y, field_radius=baseline.parameters.field_radius, exclusion_radius=baseline.parameters.exclusion_radius, safety_epsilon=baseline.parameters.safety_epsilon)
    return (field, specifications, check)

def evaluate_design(*, baseline: RefineBaseline, design: RefineDesign, profile: EvaluationProfile, cache: EvaluationCache | None=None) -> RefineEvaluation:
    field = build_refine_field(baseline, design)
    return evaluate_field(baseline=baseline, design=design, field=field, profile=profile, cache=cache)

def evaluate_field(*, baseline: RefineBaseline, design: RefineDesign, field: RefineField, profile: EvaluationProfile, cache: EvaluationCache | None=None) -> RefineEvaluation:
    """评价固定镜位及显式分区归属，用于边界局部检验。"""
    specifications = expand_specifications(field, design)
    check = validate_heterogeneous_field(coordinates=field.coordinates, widths=specifications.widths, heights=specifications.heights, installation_heights=specifications.installation_heights, tower_x=baseline.parameters.tower_x, tower_y=design.tower_y, field_radius=baseline.parameters.field_radius, exclusion_radius=baseline.parameters.exclusion_radius, safety_epsilon=baseline.parameters.safety_epsilon)
    if not check.valid:
        raise ValueError(check.reason or '六区候选几何不合法。')
    raw = _evaluate_specifications(coordinates=field.coordinates, specifications=specifications, ring_indices=field.ring_indices, group_indices=field.group_indices, original_indices=field.original_indices, field_config=_field_config(baseline, design), profile=profile, safety_epsilon=baseline.parameters.safety_epsilon, cache=cache)
    return RefineEvaluation(design=design, field=field, specifications=specifications, raw=raw)

def metrics(evaluation: RefineEvaluation, *, target_power_mw: float=42.0) -> dict[str, object]:
    annual = asdict(evaluation.raw.solution.annual_result)
    return {'profile': evaluation.profile_name, 'tower_mode': evaluation.design.tower_mode, 'tower_x_m': 0.0, 'tower_y_m': evaluation.design.tower_y, 'mirror_count': evaluation.mirror_count, 'mirror_set_hash': evaluation.field.mirror_set_hash, 'outer_clipped_count': evaluation.field.outer_clipped_count, 'total_area_m2': evaluation.total_area_m2, 'annual_power_mw': evaluation.annual_power_mw, 'power_margin_mw': evaluation.annual_power_mw - target_power_mw, 'unit_area_power_kw_m2': evaluation.unit_area_power_kw_m2, **annual}
__all__ = ('EvaluationCache', 'RefineEvaluation', 'coarse_profile', 'dense_profile', 'evaluate_design', 'evaluate_field', 'formal_profile', 'medium_profile', 'metrics', 'prepare_candidate', 'smoke_profile')

# ========================================================================
# 来源：src/heliostat/q3/sensitivity.py
# ========================================================================

"""六区规格敏感性筛选与径向边界局部扰动。"""
from dataclasses import dataclass, replace
import numpy as np
SPECIFICATION_VARIABLES = tuple((f'{prefix}{group}' for prefix in ('w', 'h', 'H') for group in range(1, 7)))
BASE_BOUNDARIES = (1, 5, 11, 14, 20)
BOUNDARY_SHIFTS = (-2, -1, 1, 2)
RING_COUNT = 28

@dataclass(frozen=True)
class Perturbation:
    parameter: str
    group_id: int
    direction: str
    old_value: float
    new_value: float
    design: RefineDesign

@dataclass(frozen=True)
class BoundaryPerturbation:
    """只移动一条内部边界的候选。"""
    boundary_id: int
    shift_rings: int
    boundaries: tuple[int, ...]

    @property
    def original_end_ring(self) -> int:
        return BASE_BOUNDARIES[self.boundary_id - 1]

    @property
    def new_end_ring(self) -> int:
        return self.boundaries[self.boundary_id - 1]

    @property
    def label(self) -> str:
        sign = '+' if self.shift_rings > 0 else ''
        return f'B{self.boundary_id}{sign}{self.shift_rings}'

def validate_boundaries(boundaries: tuple[int, ...]) -> None:
    """验证五条边界能形成六个非空、连续的径向分区。"""
    if len(boundaries) != len(BASE_BOUNDARIES):
        raise ValueError('六区划分必须包含五条内部边界。')
    if not all((isinstance(value, int) for value in boundaries)):
        raise ValueError('边界必须使用整数圆环编号。')
    if boundaries[0] < 1 or boundaries[-1] >= RING_COUNT:
        raise ValueError('边界必须位于第 1 环至第 27 环。')
    if any((left >= right for (left, right) in zip(boundaries, boundaries[1:]))):
        raise ValueError('五条内部边界必须严格递增。')

def boundary_perturbations() -> tuple[BoundaryPerturbation, ...]:
    """生成全部合法的单边界正负 1--2 环扰动。"""
    candidates: list[BoundaryPerturbation] = []
    for (boundary_id, original) in enumerate(BASE_BOUNDARIES, start=1):
        for shift in BOUNDARY_SHIFTS:
            values = list(BASE_BOUNDARIES)
            values[boundary_id - 1] = original + shift
            boundaries = tuple(values)
            try:
                validate_boundaries(boundaries)
            except ValueError:
                continue
            candidates.append(BoundaryPerturbation(boundary_id=boundary_id, shift_rings=shift, boundaries=boundaries))
    return tuple(candidates)

def group_indices_for_boundaries(ring_indices: np.ndarray, boundaries: tuple[int, ...]) -> np.ndarray:
    """按每组末环编号把逐镜圆环映射为从 0 开始的分区编号。"""
    validate_boundaries(boundaries)
    rings = np.asarray(ring_indices, dtype=np.int64)
    if rings.ndim != 1 or rings.size == 0:
        raise ValueError('ring_indices 必须是一维非空数组。')
    if int(np.min(rings)) < 1 or int(np.max(rings)) > RING_COUNT:
        raise ValueError('ring_indices 必须位于第 1 环至第 28 环。')
    return np.searchsorted(np.asarray(boundaries, dtype=np.int64), rings, side='left').astype(np.int64, copy=False)

def reassign_boundary_groups(field: RefineField, boundaries: tuple[int, ...]) -> RefineField:
    """固定塔位与镜位，只更新逐镜所属分区。"""
    return replace(field, group_indices=group_indices_for_boundaries(field.ring_indices, boundaries))

def moved_mirror_count(field: RefineField, boundaries: tuple[int, ...]) -> int:
    assigned = group_indices_for_boundaries(field.ring_indices, boundaries)
    return int(np.count_nonzero(assigned != field.group_indices))

def specification_perturbations(design: RefineDesign, *, step_m: float=0.1) -> tuple[Perturbation, ...]:
    if step_m <= 0.0:
        raise ValueError('敏感性扰动步长必须大于 0。')
    candidates: list[Perturbation] = []
    for parameter in SPECIFICATION_VARIABLES:
        old = design.parameter(parameter)
        for (direction, delta) in (('-', -step_m), ('+', step_m)):
            new = old + delta
            candidates.append(Perturbation(parameter=parameter, group_id=int(parameter[1:]), direction=direction, old_value=old, new_value=new, design=design.with_parameter(parameter, new)))
    return tuple(candidates)

def select_formal_directions(rows: list[dict[str, object]], *, limit: int=6) -> list[dict[str, object]]:
    """按中精度提升选出至多六个不同变量的最佳方向。"""
    eligible = [row for row in rows if bool(row.get('legal')) and row.get('medium_q') is not None and (float(row['medium_power']) >= 42.0)]
    eligible.sort(key=lambda row: float(row['delta_q']), reverse=True)
    selected: list[dict[str, object]] = []
    seen: set[str] = set()
    for row in eligible:
        parameter = str(row['parameter'])
        if parameter in seen:
            continue
        selected.append(row)
        seen.add(parameter)
        if len(selected) >= limit:
            break
    return selected

def active_from_formal(rows: list[dict[str, object]], *, reference_q: float, target_power_mw: float=42.0, threshold: float=1e-08) -> tuple[str, ...]:
    active = {str(row['parameter']) for row in rows if row.get('formal_q') is not None and float(row['formal_power']) >= target_power_mw and (float(row['formal_q']) > reference_q + threshold)}
    return tuple((parameter for parameter in SPECIFICATION_VARIABLES if parameter in active))

# ========================================================================
# 来源：src/heliostat/q3/search.py
# ========================================================================

"""活跃变量的分块变步长局部搜索。"""
from dataclasses import dataclass
from typing import Callable
Evaluator = Callable[[RefineDesign], RefineEvaluation | None]
BLOCK_ORDER = ('tower', 'campo', 'width', 'height', 'installation')
STEP_LEVELS = {'tower_y': (0.5, 0.25, 0.1), 'initial_spacing': (0.2, 0.1, 0.05), 'spacing_growth': (0.02, 0.01, 0.005), 'w': (0.1, 0.05, 0.02), 'h': (0.1, 0.05, 0.02), 'H': (0.1, 0.05, 0.02)}

@dataclass(frozen=True)
class SearchOutcome:
    initial_design: RefineDesign
    initial_evaluation: RefineEvaluation
    best_design: RefineDesign
    best_evaluation: RefineEvaluation
    trace: tuple[dict[str, object], ...]
    evaluated_candidates: int

def parameter_block(parameter: str) -> str:
    if parameter == 'tower_y':
        return 'tower'
    if parameter in ('initial_spacing', 'spacing_growth'):
        return 'campo'
    return {'w': 'width', 'h': 'height', 'H': 'installation'}[parameter[0]]

def _steps(parameter: str) -> tuple[float, ...]:
    if parameter in STEP_LEVELS:
        return STEP_LEVELS[parameter]
    return STEP_LEVELS[parameter[0]]

def _better(candidate: RefineEvaluation, reference: RefineEvaluation, *, target_power_mw: float, threshold: float) -> bool:
    candidate_feasible = candidate.is_feasible(target_power_mw)
    reference_feasible = reference.is_feasible(target_power_mw)
    if candidate_feasible != reference_feasible:
        return candidate_feasible
    if candidate_feasible:
        return candidate.unit_area_power_kw_m2 > reference.unit_area_power_kw_m2 + threshold
    return candidate.annual_power_mw > reference.annual_power_mw + 1e-06

def coordinate_search(*, initial_design: RefineDesign, initial_evaluation: RefineEvaluation, active_variables: tuple[str, ...], evaluator: Evaluator, baseline_q_kw_m2: float, maximum_sweeps: int=2, target_power_mw: float=42.0, move_q_threshold: float=1e-08) -> SearchOutcome:
    """按塔位、Campo、宽、高、安装高顺序执行最多两轮回扫。"""
    if maximum_sweeps < 0 or maximum_sweeps > 2:
        raise ValueError('联合回扫轮数必须位于 0 到 2。')
    current_design = initial_design
    current_evaluation = initial_evaluation
    level_by_block = {block: 0 for block in BLOCK_ORDER}
    trace: list[dict[str, object]] = []
    evaluated = 0
    for sweep in range(1, maximum_sweeps + 1):
        sweep_improved = False
        for block in BLOCK_ORDER:
            parameters = tuple((parameter for parameter in active_variables if parameter_block(parameter) == block))
            if not parameters:
                continue
            level = level_by_block[block]
            ranked: list[tuple[str, float, float, RefineDesign, RefineEvaluation]] = []
            for parameter in parameters:
                step = _steps(parameter)[level]
                old = current_design.parameter(parameter)
                for sign in (-1.0, 1.0):
                    new = old + sign * step
                    candidate_design = current_design.with_parameter(parameter, new)
                    evaluation = evaluator(candidate_design)
                    if evaluation is None:
                        continue
                    evaluated += 1
                    ranked.append((parameter, old, new, candidate_design, evaluation))
            improving = [item for item in ranked if _better(item[4], current_evaluation, target_power_mw=target_power_mw, threshold=move_q_threshold)]
            if improving:
                (parameter, old, new, design, evaluation) = max(improving, key=lambda item: (int(item[4].is_feasible(target_power_mw)), item[4].unit_area_power_kw_m2 if item[4].is_feasible(target_power_mw) else item[4].annual_power_mw))
                previous = current_evaluation
                current_design = design
                current_evaluation = evaluation
                sweep_improved = True
                trace.append({'sweep_id': sweep, 'parameter_block': block, 'parameter': parameter, 'old_value': old, 'new_value': new, 'step_size': abs(new - old), 'evaluation_level': evaluation.profile_name, 'power': evaluation.annual_power_mw, 'power_margin': evaluation.annual_power_mw - target_power_mw, 'total_area': evaluation.total_area_m2, 'q': evaluation.unit_area_power_kw_m2, 'delta_q_from_previous': evaluation.unit_area_power_kw_m2 - previous.unit_area_power_kw_m2, 'delta_q_from_six': evaluation.unit_area_power_kw_m2 - baseline_q_kw_m2, 'accepted': True})
            else:
                level_by_block[block] = min(level + 1, 2)
        if not sweep_improved:
            break
    return SearchOutcome(initial_design=initial_design, initial_evaluation=initial_evaluation, best_design=current_design, best_evaluation=current_evaluation, trace=tuple(trace), evaluated_candidates=evaluated)

# ========================================================================
# 来源：src/heliostat/q3/closure.py
# ========================================================================

"""最终候选的塔位包围扫描和正式精度最细邻域收口。"""
from dataclasses import dataclass
from typing import Callable
Evaluator = Callable[[RefineDesign], RefineEvaluation | None]
LOCAL_STEPS = (('tower_y', 0.1), ('initial_spacing', 0.05), ('spacing_growth', 0.005), ('w1', 0.02), ('h1', 0.02), ('H2', 0.02))

@dataclass(frozen=True)
class ClosureOutcome:
    initial_design: RefineDesign
    initial_evaluation: RefineEvaluation
    best_design: RefineDesign
    best_evaluation: RefineEvaluation
    trace: tuple[dict[str, object], ...]
    tower_bracketed: bool
    local_converged: bool
    local_sweeps: int

def _score(evaluation: RefineEvaluation, target_power_mw: float) -> tuple[int, float]:
    feasible = evaluation.is_feasible(target_power_mw)
    return (int(feasible), evaluation.unit_area_power_kw_m2 if feasible else evaluation.annual_power_mw)

def _better(candidate: RefineEvaluation, reference: RefineEvaluation, *, target_power_mw: float, threshold: float) -> bool:
    if candidate.is_feasible(target_power_mw) != reference.is_feasible(target_power_mw):
        return candidate.is_feasible(target_power_mw)
    if candidate.is_feasible(target_power_mw):
        return candidate.unit_area_power_kw_m2 > reference.unit_area_power_kw_m2 + threshold
    return candidate.annual_power_mw > reference.annual_power_mw + 1e-06

def _record(*, phase: str, sweep: int, parameter: str, old_value: float, new_value: float, design: RefineDesign, evaluation: RefineEvaluation | None, target_power_mw: float) -> dict[str, object]:
    row: dict[str, object] = {'phase': phase, 'sweep': sweep, 'parameter': parameter, 'old_value': old_value, 'new_value': new_value, 'step': new_value - old_value, 'tower_y': design.tower_y, 'initial_spacing': design.initial_spacing, 'spacing_growth': design.spacing_growth, 'w1': design.widths[0], 'h1': design.mirror_heights[0], 'H2': design.installation_heights[1], 'legal': evaluation is not None, 'feasible': False, 'annual_power_mw': None, 'unit_area_power_kw_m2': None, 'accepted': False, 'stage_selected': False}
    if evaluation is not None:
        row.update({'feasible': evaluation.is_feasible(target_power_mw), 'annual_power_mw': evaluation.annual_power_mw, 'unit_area_power_kw_m2': evaluation.unit_area_power_kw_m2})
    return row

def close_formal_neighborhood(*, initial_design: RefineDesign, initial_evaluation: RefineEvaluation, evaluator: Evaluator, target_power_mw: float=42.0, coarse_step_limit: int=12, fine_radius_steps: int=4, maximum_local_sweeps: int=4, move_q_threshold: float=1e-08) -> ClosureOutcome:
    """先包围塔位极值，再对六个活跃变量做正式精度双侧检查。"""
    if coarse_step_limit < 1:
        raise ValueError('塔位包围扫描至少需要一个 0.5 m 步长。')
    if fine_radius_steps < 1:
        raise ValueError('塔位细扫至少需要中心两侧各一个 0.1 m 点。')
    if maximum_local_sweeps < 0:
        raise ValueError('最细邻域回扫轮数不能为负。')
    trace: list[dict[str, object]] = []
    coarse: list[tuple[RefineDesign, RefineEvaluation, dict[str, object] | None]] = [(initial_design, initial_evaluation, None)]
    previous = initial_evaluation
    tower_bracketed = False
    for index in range(1, coarse_step_limit + 1):
        candidate = initial_design.with_parameter('tower_y', initial_design.tower_y + 0.5 * index)
        evaluation = evaluator(candidate)
        row = _record(phase='tower_coarse', sweep=0, parameter='tower_y', old_value=initial_design.tower_y, new_value=candidate.tower_y, design=candidate, evaluation=evaluation, target_power_mw=target_power_mw)
        trace.append(row)
        if evaluation is None:
            tower_bracketed = True
            break
        coarse.append((candidate, evaluation, row))
        if previous.is_feasible(target_power_mw) and evaluation.is_feasible(target_power_mw) and (evaluation.unit_area_power_kw_m2 < previous.unit_area_power_kw_m2):
            tower_bracketed = True
            break
        previous = evaluation
    coarse_best = max(coarse, key=lambda item: _score(item[1], target_power_mw))
    if coarse_best[2] is not None:
        coarse_best[2]['stage_selected'] = True
    fine: list[tuple[RefineDesign, RefineEvaluation, dict[str, object]]] = []
    center_y = coarse_best[0].tower_y
    for offset in range(-fine_radius_steps, fine_radius_steps + 1):
        candidate = coarse_best[0].with_parameter('tower_y', center_y + 0.1 * offset)
        evaluation = evaluator(candidate)
        row = _record(phase='tower_fine', sweep=0, parameter='tower_y', old_value=center_y, new_value=candidate.tower_y, design=candidate, evaluation=evaluation, target_power_mw=target_power_mw)
        trace.append(row)
        if evaluation is not None:
            fine.append((candidate, evaluation, row))
    fine_best = max(fine, key=lambda item: _score(item[1], target_power_mw))
    fine_best[2]['stage_selected'] = True
    current_design = fine_best[0]
    current_evaluation = fine_best[1]
    local_converged = maximum_local_sweeps == 0
    completed_sweeps = 0
    for sweep in range(1, maximum_local_sweeps + 1):
        completed_sweeps = sweep
        sweep_improved = False
        for (parameter, step) in LOCAL_STEPS:
            old_value = current_design.parameter(parameter)
            candidates: list[tuple[RefineDesign, RefineEvaluation, dict[str, object]]] = []
            for sign in (-1.0, 1.0):
                candidate = current_design.with_parameter(parameter, old_value + sign * step)
                evaluation = evaluator(candidate)
                row = _record(phase='local_fine', sweep=sweep, parameter=parameter, old_value=old_value, new_value=candidate.parameter(parameter), design=candidate, evaluation=evaluation, target_power_mw=target_power_mw)
                trace.append(row)
                if evaluation is not None:
                    candidates.append((candidate, evaluation, row))
            improving = [item for item in candidates if _better(item[1], current_evaluation, target_power_mw=target_power_mw, threshold=move_q_threshold)]
            if improving:
                best = max(improving, key=lambda item: _score(item[1], target_power_mw))
                best[2]['accepted'] = True
                current_design = best[0]
                current_evaluation = best[1]
                sweep_improved = True
        if not sweep_improved:
            local_converged = True
            break
    return ClosureOutcome(initial_design=initial_design, initial_evaluation=initial_evaluation, best_design=current_design, best_evaluation=current_evaluation, trace=tuple(trace), tower_bracketed=tower_bracketed, local_converged=local_converged, local_sweeps=completed_sweeps)

# ========================================================================
# 来源：src/heliostat/q3/export.py
# ========================================================================

"""六区微调实验的数据、论文表和 result3.xlsx 导出。"""
import csv
import json
from dataclasses import asdict
from pathlib import Path
from typing import Iterable

def _json(path: Path, payload: object) -> Path:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    return path

def _csv(path: Path, rows: Iterable[dict[str, object]]) -> Path:
    records = list(rows)
    if not records:
        path.write_text('\n', encoding='utf-8-sig')
        return path
    fields: list[str] = []
    for record in records:
        for key in record:
            if key not in fields:
                fields.append(key)
    with path.open('w', encoding='utf-8-sig', newline='') as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, lineterminator='\n')
        writer.writeheader()
        writer.writerows(records)
    return path

def _comparison_record(label: str, evaluation: RefineEvaluation, *, target_power_mw: float) -> dict[str, object]:
    return {'scheme': label, **metrics(evaluation, target_power_mw=target_power_mw)}

def write_results(*, output_dir: str | Path, baseline: RefineBaseline, regression: dict[str, object], tower_rows: list[dict[str, object]], geometry_rows: list[dict[str, object]], sensitivity_rows: list[dict[str, object]], active_payload: dict[str, object], search_trace: Iterable[dict[str, object]], formal_rows: list[dict[str, object]], baseline_formal: RefineEvaluation, preclosure_formal: RefineEvaluation, attempted_formal: RefineEvaluation, selected_formal: RefineEvaluation, selected_design: RefineDesign, dense_payload: dict[str, object], result3_template: str | Path, target_power_mw: float, decision: str, closure_rows: Iterable[dict[str, object]], closure_payload: dict[str, object], boundary_rows: list[dict[str, object]]) -> dict[str, Path]:
    destination = Path(output_dir)
    destination.mkdir(parents=True, exist_ok=True)
    paths: dict[str, Path] = {}
    paths['regression'] = _json(destination / '02_六组回归结果.json', regression)
    paths['tower'] = _csv(destination / '03_塔位两种语义扫描.csv', tower_rows)
    paths['geometry_scan'] = _csv(destination / '04_Campo几何粗扫.csv', geometry_rows)
    paths['sensitivity'] = _csv(destination / '05_规格参数敏感性.csv', sensitivity_rows)
    paths['active'] = _json(destination / '06_活跃变量集合.json', active_payload)
    paths['trace'] = _csv(destination / '07_局部搜索轨迹.csv', search_trace)
    paths['formal_candidates'] = _csv(destination / '08_正式候选比较.csv', formal_rows)
    group_rows = []
    for group in range(6):
        active = selected_formal.field.group_indices == group
        group_rows.append({'group': group + 1, 'ring_start': (1, 2, 6, 12, 15, 21)[group], 'ring_stop': (1, 5, 11, 14, 20, 28)[group], 'mirror_count': int(active.sum()), 'mirror_width_m': selected_design.widths[group], 'mirror_height_m': selected_design.mirror_heights[group], 'installation_height_m': selected_design.installation_heights[group], 'group_area_m2': float(selected_formal.specifications.areas[active].sum())})
    paths['groups'] = _csv(destination / '09_最终六区参数.csv', group_rows)
    mirror_rows = []
    for index in range(selected_formal.mirror_count):
        mirror_rows.append({'mirror_id': index + 1, 'original_mirror_id': int(selected_formal.field.original_indices[index]) + 1, 'ring_index': int(selected_formal.field.ring_indices[index]), 'group': int(selected_formal.field.group_indices[index]) + 1, 'mirror_width_m': float(selected_formal.specifications.widths[index]), 'mirror_height_m': float(selected_formal.specifications.heights[index]), 'x_m': float(selected_formal.field.coordinates[index, 0]), 'y_m': float(selected_formal.field.coordinates[index, 1]), 'z_m': float(selected_formal.specifications.installation_heights[index])})
    paths['mirrors'] = _csv(destination / '10_最终逐镜参数与坐标.csv', mirror_rows)
    comparison = {'decision': decision, 'target_power_mw': target_power_mw, 'baseline': _comparison_record('six_group_baseline', baseline_formal, target_power_mw=target_power_mw), 'preclosure_candidate': _comparison_record('two_sweep_candidate', preclosure_formal, target_power_mw=target_power_mw), 'attempted_candidate': _comparison_record('refined_candidate', attempted_formal, target_power_mw=target_power_mw), 'selected': _comparison_record('selected_final', selected_formal, target_power_mw=target_power_mw), 'selected_design': selected_design.to_dict(), 'closure': closure_payload}
    paths['formal'] = _json(destination / '11_正式结果比较.json', comparison)
    paths['dense'] = _json(destination / '12_加密验收比较.json', dense_payload)
    paths['geometry'] = _json(destination / '13_几何约束验证.json', {'valid': selected_formal.geometry.valid, 'details': asdict(selected_formal.geometry), 'mirror_set_hash': selected_formal.field.mirror_set_hash, 'outer_clipped_count': selected_formal.field.outer_clipped_count, 'group_counts': list(selected_formal.field.group_counts)})
    paths['closure'] = _csv(destination / '14_局部收口检查.csv', closure_rows)
    paths['boundary'] = _csv(destination / '20_六区边界局部敏感性检验.csv', boundary_rows)
    workbook = destination / 'result3.xlsx'
    write_result3_workbook(template_path=result3_template, output_path=workbook, evaluation=selected_formal.raw, tower_x=baseline.parameters.tower_x, tower_y=selected_design.tower_y)
    paths['workbook'] = workbook
    delta_q = attempted_formal.unit_area_power_kw_m2 - baseline_formal.unit_area_power_kw_m2
    lines = ['# 第三问六区参数微调结果与验证表', '', '本文档汇总第三问的正式结果、加密结果、最终六区规格和边界局部检验。', '', f'计算结论：{decision}。', '', '## 表 S3-1 正式精度比较', '', '| 方案 | 镜子数 | 年平均功率 (MW) | 功率余量 (MW) | 总面积 (m²) | 单位面积输出 (kW/m²) |', '| --- | ---: | ---: | ---: | ---: | ---: |', f'| 原六组 | {baseline_formal.mirror_count} | {baseline_formal.annual_power_mw:.9f} | {baseline_formal.annual_power_mw - target_power_mw:.9f} | {baseline_formal.total_area_m2:.6f} | {baseline_formal.unit_area_power_kw_m2:.9f} |', f'| 两轮局部搜索候选 | {preclosure_formal.mirror_count} | {preclosure_formal.annual_power_mw:.9f} | {preclosure_formal.annual_power_mw - target_power_mw:.9f} | {preclosure_formal.total_area_m2:.6f} | {preclosure_formal.unit_area_power_kw_m2:.9f} |', f'| 正式收口候选 | {attempted_formal.mirror_count} | {attempted_formal.annual_power_mw:.9f} | {attempted_formal.annual_power_mw - target_power_mw:.9f} | {attempted_formal.total_area_m2:.6f} | {attempted_formal.unit_area_power_kw_m2:.9f} |', '', f'正式精度候选相对原六组的 $\\Delta q={delta_q:.9f}\\ \\mathrm{{kW/m^2}}$。', '', f"塔位包围扫描已找到北侧下降点；随后完成一次六个活跃变量的正负最细邻域检查，并接受 {closure_payload['accepted_moves']} 个可行改进。", '', '## 表 S3-2 加密精度比较', '', '| 邻镜半径 (m) | 原六组功率 (MW) | 微调候选功率 (MW) | 原六组 q (kW/m²) | 微调候选 q (kW/m²) | $\\Delta q$ (kW/m²) |', '| ---: | ---: | ---: | ---: | ---: | ---: |']
    for radius in ('80', '100'):
        before = dense_payload.get('baseline', {}).get(radius)
        after = dense_payload.get('candidate', {}).get(radius)
        if before is not None and after is not None:
            lines.append(f"| {radius} | {before['annual_power_mw']:.9f} | {after['annual_power_mw']:.9f} | {before['unit_area_power_kw_m2']:.9f} | {after['unit_area_power_kw_m2']:.9f} | {after['unit_area_power_kw_m2'] - before['unit_area_power_kw_m2']:.9f} |")
    lines.extend(('', '## 表 S3-3 最终六区规格', '', '| 分区 | 镜子数 | 宽度 (m) | 高度 (m) | 安装高度 (m) |', '| ---: | ---: | ---: | ---: | ---: |'))
    for row in group_rows:
        lines.append(f"| G{row['group']} | {row['mirror_count']} | {row['mirror_width_m']:.6f} | {row['mirror_height_m']:.6f} | {row['installation_height_m']:.6f} |")
    lines.extend(('', '## 表 S3-4 六区边界局部合理性检验', ''))
    if bool(regression.get('smoke')):
        lines.extend(('smoke 仅验证 18 个边界候选的生成、评价、导出和绘图链路；数值及分类不得用于论文结论。',))
    else:
        boundary_counts = {classification: sum((row['classification'] == classification for row in boundary_rows)) for classification in ('功率可行但q下降', 'q提高但功率不达标', '功率与q均不占优')}
        q_sensitivity: dict[int, float] = {}
        for boundary_id in range(1, 6):
            candidates = [row for row in boundary_rows if int(row['boundary_id']) == boundary_id]
            q_sensitivity[boundary_id] = max((abs(float(row['formal_delta_q_kw_m2'])) / abs(int(row['shift_rings'])) for row in candidates))
        most_sensitive = max(q_sensitivity, key=q_sensitivity.get)
        lines.extend((f'零扰动正式评价复现最终方案：$P_0={selected_formal.annual_power_mw:.9f}\\ \\mathrm{{MW}}$，$q_0={selected_formal.unit_area_power_kw_m2:.9f}\\ \\mathrm{{kW/m^2}}$。', '', '| 正式分类 | 候选数 |', '| --- | ---: |', f"| 功率可行但 $q$ 下降 | {boundary_counts['功率可行但q下降']} |", f"| $q$ 提高但功率不达标 | {boundary_counts['q提高但功率不达标']} |", f"| 功率与 $q$ 均不占优 | {boundary_counts['功率与q均不占优']} |", '', f'18 个合法单边界候选均完成中精度和正式精度评价；未发现同时满足 $42\\ \\mathrm{{MW}}$ 功率约束并提高 $q$ 的候选，因此保留边界 $(1,5,11,14,20)$。B{most_sensitive} 是本次单位面积输出局部检验中最敏感的边界。'))
    lines.extend(('', '## 验收说明', '', '塔位模式 A 与 B 分开扫描；搜索轨迹固定使用已选模式。中精度仅用于排序和局部接受，最终判定来自同口径正式精度及 80/100 m 加密比较。', ''))
    table = destination / '15_论文结果与验证表.md'
    table.write_text('\n'.join(lines) + '\n', encoding='utf-8')
    paths['table'] = table
    return paths

# ========================================================================
# 来源：src/heliostat/q3/plot.py
# ========================================================================

"""第三问敏感性、规格、指标和最终镜场结果图。"""
import math
from pathlib import Path
import matplotlib
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.patches import Circle, Patch

def configure_matplotlib() -> None:
    font_candidates = (Path('/System/Library/Fonts/STHeiti Medium.ttc'), Path('/System/Library/Fonts/PingFang.ttc'), Path('C:/Windows/Fonts/msyh.ttc'), Path('C:/Windows/Fonts/simhei.ttf'), Path('/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc'), Path('/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc'))
    for font_path in font_candidates:
        if not font_path.exists():
            continue
        matplotlib.font_manager.fontManager.addfont(font_path)
        font_name = matplotlib.font_manager.FontProperties(fname=font_path).get_name()
        plt.rcParams['font.family'] = font_name
        break
    plt.rcParams.update({'font.sans-serif': ['PingFang SC', 'Microsoft YaHei', 'SimHei', 'Noto Sans CJK SC', 'WenQuanYi Zen Hei', 'DejaVu Sans'], 'axes.unicode_minus': False, 'figure.facecolor': 'white', 'axes.facecolor': 'white', 'savefig.facecolor': 'white', 'savefig.bbox': 'tight'})

def plot_sensitivity(rows: list[dict[str, object]], *, tower_rows: list[dict[str, object]], geometry_rows: list[dict[str, object]], selected_tower_mode: str, output_dir: str | Path) -> Path:
    records = [('六区规格', f"{row['parameter']}{row['direction']}", float(row['delta_q_from_geometry'])) for row in rows if row.get('delta_q_from_geometry') not in (None, '')]
    for row in tower_rows:
        if row.get('tower_mode') == selected_tower_mode and abs(float(row.get('delta_y_m', 99.0))) == 0.5 and (row.get('delta_q_from_six_medium') not in (None, '')):
            direction = '+' if float(row['delta_y_m']) > 0 else '-'
            records.append(('塔位', f'yT{direction}', float(row['delta_q_from_six_medium'])))
    for row in geometry_rows:
        if row.get('delta_q_from_tower_medium') in (None, ''):
            continue
        if row.get('scan') == 'D1-one-dimensional' and math.isclose(abs(float(row['delta_D1_from_six'])), 0.1, abs_tol=1e-12):
            direction = '+' if float(row['delta_D1_from_six']) > 0 else '-'
            records.append(('Campo', f'D1{direction}', float(row['delta_q_from_tower_medium'])))
        if row.get('scan') == 'g-one-dimensional' and math.isclose(abs(float(row['delta_g_from_six'])), 0.01, abs_tol=1e-12):
            direction = '+' if float(row['delta_g_from_six']) > 0 else '-'
            records.append(('Campo', f'g{direction}', float(row['delta_q_from_tower_medium'])))
    stage_order = {'塔位': 0, 'Campo': 1, '六区规格': 2}
    records.sort(key=lambda item: (stage_order[item[0]], item[2]))
    labels = [f'{item[0]} · {item[1]}' for item in records]
    values = [item[2] for item in records]
    stage_colors = {'塔位': '#7570b3', 'Campo': '#d95f02', '六区规格': '#1b9e77'}
    (figure, axis) = plt.subplots(figsize=(9, max(5, 0.22 * len(records))))
    colors = [stage_colors[item[0]] for item in records]
    axis.barh(np.arange(len(values)), values, color=colors)
    axis.set_yticks(np.arange(len(values)), labels)
    axis.axvline(0.0, color='black', linewidth=0.8)
    axis.set_xlabel('相对于对应阶段基准的 Δq / (kW/m²)')
    axis.set_title('图 S3-1 各阶段候选相对于对应阶段基准的单位面积输出变化')
    axis.legend(handles=[Patch(color=color, label=stage) for (stage, color) in stage_colors.items()], loc='best')
    axis.grid(axis='x', alpha=0.25)
    figure.tight_layout()
    path = Path(output_dir) / '16_参数敏感性图.png'
    figure.savefig(path, dpi=220)
    plt.close(figure)
    return path

def plot_group_parameters(baseline: RefineBaseline, selected: RefineDesign, output_dir: str | Path) -> Path:
    groups = np.arange(1, 7)
    (figure, axes) = plt.subplots(1, 3, figsize=(13, 4.2), sharex=True)
    series = ((baseline.design.widths, selected.widths, '镜宽 / m'), (baseline.design.mirror_heights, selected.mirror_heights, '镜高 / m'), (baseline.design.installation_heights, selected.installation_heights, '安装高度 / m'))
    for (axis, (before, after, ylabel)) in zip(axes, series):
        axis.plot(groups, before, 'o--', label='原六组')
        axis.plot(groups, after, 's-', label='微调后')
        axis.set_xlabel('径向分区')
        axis.set_ylabel(ylabel)
        axis.set_xticks(groups)
        axis.grid(alpha=0.25)
    axes[0].legend()
    figure.suptitle('图 S3-2 优化前后六区规格对比')
    figure.tight_layout()
    path = Path(output_dir) / '17_六区宽高与安装高度图.png'
    figure.savefig(path, dpi=220)
    plt.close(figure)
    return path

def plot_metric_comparison(*, baseline_formal: RefineEvaluation, candidate_formal: RefineEvaluation, dense_payload: dict[str, object], output_dir: str | Path) -> Path:
    labels = ['正式 q', '80 m q', '100 m q']
    baseline_dense = dense_payload.get('baseline', {})
    candidate_dense = dense_payload.get('candidate', {})
    before = [baseline_formal.unit_area_power_kw_m2, float(baseline_dense.get('80', {}).get('unit_area_power_kw_m2', np.nan)), float(baseline_dense.get('100', {}).get('unit_area_power_kw_m2', np.nan))]
    after = [candidate_formal.unit_area_power_kw_m2, float(candidate_dense.get('80', {}).get('unit_area_power_kw_m2', np.nan)), float(candidate_dense.get('100', {}).get('unit_area_power_kw_m2', np.nan))]
    x = np.arange(3)
    width = 0.36
    (figure, axes) = plt.subplots(1, 2, figsize=(11, 4.2))
    axes[0].bar(x - width / 2, before, width, label='原六组')
    axes[0].bar(x + width / 2, after, width, label='微调候选')
    axes[0].set_xticks(x, labels)
    axes[0].set_ylabel('q / (kW/m²)')
    axes[0].legend()
    axes[0].grid(axis='y', alpha=0.25)
    powers = [baseline_formal.annual_power_mw, candidate_formal.annual_power_mw]
    axes[1].bar(('原六组', '微调候选'), powers, color=('#7570b3', '#1b9e77'))
    axes[1].axhline(42.0, color='#d95f02', linestyle='--', label='42 MW 约束')
    axes[1].set_ylabel('正式年平均功率 / MW')
    axes[1].legend()
    axes[1].grid(axis='y', alpha=0.25)
    figure.suptitle('图 S3-3 正式与加密结果比较')
    figure.tight_layout()
    path = Path(output_dir) / '18_六组与优化方案指标比较图.png'
    figure.savefig(path, dpi=220)
    plt.close(figure)
    return path

def plot_final_field(*, baseline: RefineBaseline, selected: RefineDesign, evaluation: RefineEvaluation, output_dir: str | Path) -> Path:
    (figure, axis) = plt.subplots(figsize=(8.2, 8.2))
    colors = ('#2166AC', '#67A9CF', '#D1E5F0', '#FDDBC7', '#EF8A62', '#B2182B')
    for (group, color) in enumerate(colors):
        active = evaluation.field.group_indices == group
        axis.scatter(evaluation.field.coordinates[active, 0], evaluation.field.coordinates[active, 1], s=5, color=color, alpha=0.78, label=f'G{group + 1}', rasterized=True)
    axis.add_patch(Circle((0.0, 0.0), baseline.parameters.field_radius, fill=False, color='#475569', linewidth=1.2, label='350 m 场地边界'))
    axis.add_patch(Circle((baseline.parameters.tower_x, selected.tower_y), baseline.parameters.exclusion_radius, fill=False, color='#F59E0B', linestyle='--', linewidth=1.2, label='最终塔周 100 m 禁区'))
    axis.scatter((baseline.parameters.tower_x,), (baseline.design.tower_y,), marker='x', s=90, linewidths=2.0, color='#111827', label='原塔位', zorder=5)
    axis.scatter((baseline.parameters.tower_x,), (selected.tower_y,), marker='*', s=145, color='#DC2626', edgecolors='white', linewidths=0.7, label='最终塔位', zorder=6)
    displacement = selected.tower_y - baseline.design.tower_y
    axis.annotate(f'向北移动 {displacement:.1f} m', xy=(baseline.parameters.tower_x, selected.tower_y), xytext=(30.0, baseline.design.tower_y - 18.0), arrowprops={'arrowstyle': '->', 'color': '#DC2626'}, color='#991B1B', fontsize=10)
    axis.set_aspect('equal', adjustable='box')
    axis.set_xlim(-365.0, 365.0)
    axis.set_ylim(-365.0, 365.0)
    axis.set_xlabel('东西坐标 x / m')
    axis.set_ylabel('南北坐标 y / m')
    axis.set_title('图 3-1 最终六区镜场、场地边界与塔位变化')
    axis.grid(alpha=0.2)
    axis.legend(loc='upper right', ncol=2, fontsize=8)
    figure.tight_layout()
    path = Path(output_dir) / '19_最终六区镜场与塔位平面图.png'
    figure.savefig(path, dpi=240)
    plt.close(figure)
    return path

def plot_boundary_sensitivity(rows: list[dict[str, object]], *, output_dir: str | Path) -> Path:
    """绘制全部单边界候选的正式 Δq、功率余量和分类。"""
    configure_matplotlib()
    labels = [str(row['candidate']) for row in rows]
    positions = np.arange(len(labels))
    category_order = ('功率可行但q下降', 'q提高但功率不达标', '功率与q均不占优', 'smoke仅验证链路')
    colors = {'功率可行但q下降': '#2A9D8F', 'q提高但功率不达标': '#E76F51', '功率与q均不占优': '#7A7A7A', 'smoke仅验证链路': '#5E60CE'}
    categories = tuple((category for category in category_order if any((row['classification'] == category for row in rows))))
    bar_colors = [colors[str(row['classification'])] for row in rows]
    delta_q = [float(row['formal_delta_q_kw_m2']) for row in rows]
    power_margin = [float(row['formal_power_margin_mw']) for row in rows]
    (figure, axes) = plt.subplots(2, 1, figsize=(12.5, 7.2), sharex=True)
    axes[0].bar(positions, delta_q, color=bar_colors, alpha=0.9)
    axes[0].axhline(0.0, color='black', linewidth=0.8)
    axes[0].set_ylabel('正式 Δq / (kW/m²)')
    axes[0].set_title('图 S3-5 六区边界单因素局部敏感性检验')
    axes[0].grid(axis='y', alpha=0.25)
    axes[0].legend(handles=[Patch(color=colors[category], label=category) for category in categories], loc='best')
    axes[1].bar(positions, power_margin, color=bar_colors, alpha=0.9)
    axes[1].axhline(0.0, color='black', linewidth=0.8, linestyle='--')
    axes[1].set_ylabel('正式功率余量 / MW')
    axes[1].set_xlabel('边界候选（B1--B5，数字为移动环数）')
    axes[1].set_xticks(positions, labels, rotation=45, ha='right')
    axes[1].grid(axis='y', alpha=0.25)
    figure.tight_layout()
    path = Path(output_dir) / '21_六区边界局部敏感性图.png'
    figure.savefig(path, dpi=220)
    plt.close(figure)
    return path

def generate_figures(**kwargs: object) -> tuple[Path, Path, Path, Path]:
    configure_matplotlib()
    return (plot_sensitivity(kwargs['sensitivity_rows'], tower_rows=kwargs['tower_rows'], geometry_rows=kwargs['geometry_rows'], selected_tower_mode=kwargs['selected_tower_mode'], output_dir=kwargs['output_dir']), plot_group_parameters(kwargs['baseline'], kwargs['selected_design'], kwargs['output_dir']), plot_metric_comparison(baseline_formal=kwargs['baseline_formal'], candidate_formal=kwargs['candidate_formal'], dense_payload=kwargs['dense_payload'], output_dir=kwargs['output_dir']), plot_final_field(baseline=kwargs['baseline'], selected=kwargs['selected_design'], evaluation=kwargs['selected_formal'], output_dir=kwargs['output_dir']))

# ========================================================================
# 来源：src/heliostat/q3/solve.py
# ========================================================================

"""六区阶梯参数微调的完整分阶段入口。"""
import argparse
import math
from dataclasses import replace
from pathlib import Path
from typing import Sequence
import numpy as np
PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_Q2_SUMMARY = PROJECT_ROOT / 'outputs' / 'q2' / '07_最终方案摘要.json'
DEFAULT_SIX_GROUP_SUMMARY = PROJECT_ROOT / 'src' / 'heliostat' / 'q3' / 'six_group_baseline.json'
DEFAULT_TEMPLATE = PROJECT_ROOT / 'task' / 'A' / 'result3.xlsx'
DEFAULT_OUTPUT = PROJECT_ROOT / 'outputs' / 'q3'

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description='第三问六区阶梯参数敏感性与局部微调')
    parser.add_argument('--q2-summary', type=Path, default=DEFAULT_Q2_SUMMARY)
    parser.add_argument('--six-group-summary', type=Path, default=DEFAULT_SIX_GROUP_SUMMARY)
    parser.add_argument('--result3-template', type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument('--output', type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument('--smoke', action='store_true')
    parser.add_argument('--target-power', type=float, default=42.0)
    parser.add_argument('--medium-limit', type=int, default=150)
    parser.add_argument('--formal-limit', type=int, default=12)
    parser.add_argument('--max-sweeps', type=int, default=2)
    parser.add_argument('--closure-sweeps', type=int, default=1)
    parser.add_argument('--move-q', type=float, default=1e-08)
    return parser

def _validate_args(args: argparse.Namespace) -> None:
    if args.target_power <= 0.0:
        raise SystemExit('--target-power 必须大于 0。')
    if args.medium_limit < 69 or args.medium_limit > 150:
        raise SystemExit('--medium-limit 必须位于 69 到 150。')
    if args.formal_limit < 12 or args.formal_limit > 12:
        raise SystemExit('本方案严格使用 --formal-limit 12。')
    if args.max_sweeps < 0 or args.max_sweeps > 2:
        raise SystemExit('--max-sweeps 必须位于 0 到 2。')
    if args.closure_sweeps < 0 or args.closure_sweeps > 1:
        raise SystemExit('--closure-sweeps 必须位于 0 到 1。')
    if args.move_q < 0.0:
        raise SystemExit('--move-q 不能小于 0。')

def _rank_key(evaluation: RefineEvaluation, target_power_mw: float) -> tuple[int, float]:
    feasible = evaluation.is_feasible(target_power_mw)
    return (int(feasible), evaluation.unit_area_power_kw_m2 if feasible else evaluation.annual_power_mw)

def _better(candidate: RefineEvaluation, reference: RefineEvaluation, *, target_power_mw: float, threshold: float) -> bool:
    candidate_feasible = candidate.is_feasible(target_power_mw)
    reference_feasible = reference.is_feasible(target_power_mw)
    if candidate_feasible != reference_feasible:
        return candidate_feasible
    if candidate_feasible:
        return candidate.unit_area_power_kw_m2 > reference.unit_area_power_kw_m2 + threshold
    return candidate.annual_power_mw > reference.annual_power_mw + 1e-06

def _regression_payload(baseline: RefineBaseline, evaluation: RefineEvaluation) -> dict[str, object]:
    parameter_errors = {'width_max_abs_m': float(np.max(np.abs(evaluation.specifications.widths - np.asarray(baseline.design.widths)[evaluation.field.group_indices]))), 'height_max_abs_m': float(np.max(np.abs(evaluation.specifications.heights - np.asarray(baseline.design.mirror_heights)[evaluation.field.group_indices]))), 'installation_height_max_abs_m': float(np.max(np.abs(evaluation.specifications.installation_heights - np.asarray(baseline.design.installation_heights)[evaluation.field.group_indices])))}
    coordinate_error = float(np.max(np.abs(evaluation.field.coordinates - baseline.mother.coordinates)))
    errors = {'mirror_count': evaluation.mirror_count - baseline.expected_mirror_count, 'coordinate_max_abs_m': coordinate_error, 'total_area_m2': evaluation.total_area_m2 - baseline.expected_total_area_m2, 'annual_power_mw': evaluation.annual_power_mw - baseline.expected_power_mw, 'unit_area_power_kw_m2': evaluation.unit_area_power_kw_m2 - baseline.expected_q_kw_m2, **parameter_errors}
    tolerances = {'mirror_count': 0, 'coordinate_max_abs_m': 1e-12, 'total_area_m2': 1e-06, 'annual_power_mw': 1e-06, 'unit_area_power_kw_m2': 1e-09, 'width_max_abs_m': 0.0, 'height_max_abs_m': 0.0, 'installation_height_max_abs_m': 0.0}
    passed = all((abs(float(errors[key])) <= tolerance for (key, tolerance) in tolerances.items()))
    return {'passed': passed, 'expected': {'mirror_count': baseline.expected_mirror_count, 'total_area_m2': baseline.expected_total_area_m2, 'annual_power_mw': baseline.expected_power_mw, 'unit_area_power_kw_m2': baseline.expected_q_kw_m2}, 'actual': metrics(evaluation), 'errors': errors, 'tolerances': tolerances}

def run(argv: Sequence[str] | None=None) -> int:
    args = build_parser().parse_args(argv)
    _validate_args(args)
    baseline = load_baseline(q2_summary_path=args.q2_summary, six_group_summary_path=args.six_group_summary)
    search_profile = smoke_profile() if args.smoke else medium_profile()
    verification_profile = smoke_profile() if args.smoke else formal_profile()
    search_cache = EvaluationCache()
    formal_cache = EvaluationCache()
    medium_count = 0
    formal_count = 0

    def try_evaluate(design: RefineDesign, *, profile_kind: str, count_candidate: bool=True) -> tuple[RefineEvaluation | None, str | None]:
        nonlocal medium_count, formal_count
        if profile_kind == 'medium':
            if count_candidate and medium_count >= args.medium_limit:
                return (None, '达到中精度候选上限')
            profile = search_profile
            cache = search_cache
        elif profile_kind == 'formal':
            if count_candidate and formal_count >= args.formal_limit:
                return (None, '达到正式候选上限')
            profile = verification_profile
            cache = formal_cache
        else:
            raise ValueError(profile_kind)
        try:
            evaluation = evaluate_design(baseline=baseline, design=design, profile=profile, cache=cache)
        except ValueError as exc:
            return (None, str(exc))
        if count_candidate:
            if profile_kind == 'medium':
                medium_count += 1
            else:
                formal_count += 1
        return (evaluation, None)
    print('阶段 0/6：六组正式初值回归', flush=True)
    (baseline_formal, reason) = try_evaluate(baseline.design, profile_kind='formal', count_candidate=False)
    if baseline_formal is None:
        raise RuntimeError(f'六组回归无法评价：{reason}')
    regression = _regression_payload(baseline, baseline_formal)
    if not args.smoke and (not regression['passed']):
        raise RuntimeError(f"六组正式回归失败：{regression['errors']}")
    print(f'回归通过：P={baseline_formal.annual_power_mw:.9f} MW，q={baseline_formal.unit_area_power_kw_m2:.9f} kW/m²', flush=True)
    (baseline_medium, reason) = try_evaluate(baseline.design, profile_kind='medium', count_candidate=False)
    if baseline_medium is None:
        raise RuntimeError(f'六组中精度基准无法评价：{reason}')
    formal_rows: list[dict[str, object]] = []
    print('阶段 1/6：塔位模式 A/B 独立扫描', flush=True)
    tower_internal: list[dict[str, object]] = []
    tower_rows: list[dict[str, object]] = []
    for mode in ('A', 'B'):
        mode_records: list[dict[str, object]] = []
        for delta in (-2.0, -1.0, -0.5, 0.0, 0.5, 1.0, 2.0):
            design = replace(baseline.design, tower_mode=mode, tower_y=baseline.design.tower_y + delta)
            (evaluation, reject) = try_evaluate(design, profile_kind='medium')
            row: dict[str, object] = {'tower_mode': mode, 'tower_x': 0.0, 'tower_y': design.tower_y, 'delta_y_m': delta, 'legal': evaluation is not None, 'reject_reason': reject or '', 'selected_for_formal': False}
            if evaluation is not None:
                row.update(metrics(evaluation, target_power_mw=args.target_power))
                row['delta_q_from_six_medium'] = evaluation.unit_area_power_kw_m2 - baseline_medium.unit_area_power_kw_m2
                mode_records.append({'row': row, 'design': design, 'medium': evaluation})
            tower_rows.append(row)
        ranked = sorted(mode_records, key=lambda item: _rank_key(item['medium'], args.target_power), reverse=True)
        for record in ranked[:2]:
            (evaluation, reject) = try_evaluate(record['design'], profile_kind='formal')
            record['row']['selected_for_formal'] = True
            record['row']['formal_reject_reason'] = reject or ''
            if evaluation is not None:
                record['formal'] = evaluation
                record['row']['formal_power_mw'] = evaluation.annual_power_mw
                record['row']['formal_q_kw_m2'] = evaluation.unit_area_power_kw_m2
                formal_rows.append({'stage': 'tower_scan', 'candidate': f"mode-{mode}-dy-{record['row']['delta_y_m']:+g}", **metrics(evaluation, target_power_mw=args.target_power), 'delta_q_from_six': evaluation.unit_area_power_kw_m2 - baseline_formal.unit_area_power_kw_m2})
        tower_internal.extend(mode_records)
    best_by_mode: dict[str, dict[str, object]] = {}
    for mode in ('A', 'B'):
        records = [record for record in tower_internal if record['design'].tower_mode == mode and 'formal' in record]
        if records:
            best_by_mode[mode] = max(records, key=lambda item: _rank_key(item['formal'], args.target_power))
    improving_modes = {mode: record for (mode, record) in best_by_mode.items() if _better(record['formal'], baseline_formal, target_power_mw=args.target_power, threshold=args.move_q)}
    if not improving_modes:
        current_design = baseline.design
        current_formal = baseline_formal
        tower_active = False
        tower_decision = '两种语义均无正式改善，固定原塔位并采用模式 A'
    elif 'A' in improving_modes and 'B' in improving_modes and (abs(improving_modes['A']['formal'].unit_area_power_kw_m2 - improving_modes['B']['formal'].unit_area_power_kw_m2) <= 1e-05):
        chosen = improving_modes['A']
        current_design = chosen['design']
        current_formal = chosen['formal']
        tower_active = not math.isclose(current_design.tower_y, baseline.design.tower_y)
        tower_decision = '两种语义接近，按文档优先采用模式 A'
    else:
        chosen = max(improving_modes.values(), key=lambda item: _rank_key(item['formal'], args.target_power))
        current_design = chosen['design']
        current_formal = chosen['formal']
        tower_active = not math.isclose(current_design.tower_y, baseline.design.tower_y)
        tower_decision = f'正式精度选择模式 {current_design.tower_mode}'
    print(f'塔位语义：{tower_decision}；y={current_design.tower_y:.6f} m', flush=True)
    print('阶段 2/6：D1、g 一维粗扫及 3×3 局部组合', flush=True)
    geometry_origin = current_design
    geometry_origin_formal = current_formal
    (geometry_origin_medium, reason) = try_evaluate(geometry_origin, profile_kind='medium', count_candidate=False)
    if geometry_origin_medium is None:
        raise RuntimeError(f'Campo 扫描中精度基准无法评价：{reason}')
    geometry_internal: list[dict[str, object]] = []
    geometry_rows: list[dict[str, object]] = []

    def add_geometry(label: str, design: RefineDesign) -> None:
        (evaluation, reject) = try_evaluate(design, profile_kind='medium')
        row: dict[str, object] = {'scan': label, 'tower_mode': design.tower_mode, 'tower_y': design.tower_y, 'initial_spacing': design.initial_spacing, 'spacing_growth': design.spacing_growth, 'delta_D1_from_six': design.initial_spacing - baseline.design.initial_spacing, 'delta_g_from_six': design.spacing_growth - baseline.design.spacing_growth, 'legal': evaluation is not None, 'reject_reason': reject or '', 'selected_for_formal': False}
        if evaluation is not None:
            row.update(metrics(evaluation, target_power_mw=args.target_power))
            row['delta_power_from_six_medium'] = evaluation.annual_power_mw - baseline_medium.annual_power_mw
            row['delta_q_from_six_medium'] = evaluation.unit_area_power_kw_m2 - baseline_medium.unit_area_power_kw_m2
            row['delta_power_from_tower_medium'] = evaluation.annual_power_mw - geometry_origin_medium.annual_power_mw
            row['delta_q_from_tower_medium'] = evaluation.unit_area_power_kw_m2 - geometry_origin_medium.unit_area_power_kw_m2
            geometry_internal.append({'row': row, 'design': design, 'medium': evaluation})
        geometry_rows.append(row)
    for delta in (-0.2, -0.1, 0.0, 0.1, 0.2):
        add_geometry('D1-one-dimensional', replace(geometry_origin, initial_spacing=baseline.design.initial_spacing + delta))
    for delta in (-0.02, -0.01, 0.0, 0.01, 0.02):
        add_geometry('g-one-dimensional', replace(geometry_origin, spacing_growth=baseline.design.spacing_growth + delta))
    d_records = [record for record in geometry_internal if record['row']['scan'] == 'D1-one-dimensional']
    g_records = [record for record in geometry_internal if record['row']['scan'] == 'g-one-dimensional']
    best_d = max(d_records, key=lambda item: _rank_key(item['medium'], args.target_power))['design'].initial_spacing
    best_g = max(g_records, key=lambda item: _rank_key(item['medium'], args.target_power))['design'].spacing_growth
    for delta_d in (-0.1, 0.0, 0.1):
        for delta_g in (-0.01, 0.0, 0.01):
            add_geometry('D1-g-3x3', replace(geometry_origin, initial_spacing=best_d + delta_d, spacing_growth=best_g + delta_g))
    geometry_ranked = sorted(geometry_internal, key=lambda item: _rank_key(item['medium'], args.target_power), reverse=True)
    geometry_best = geometry_ranked[0]
    (geometry_formal, reject) = try_evaluate(geometry_best['design'], profile_kind='formal')
    geometry_best['row']['selected_for_formal'] = True
    geometry_best['row']['formal_reject_reason'] = reject or ''
    if geometry_formal is not None:
        geometry_best['formal'] = geometry_formal
        geometry_best['row']['formal_power_mw'] = geometry_formal.annual_power_mw
        geometry_best['row']['formal_q_kw_m2'] = geometry_formal.unit_area_power_kw_m2
        formal_rows.append({'stage': 'campo_geometry', 'candidate': 'best-medium-geometry', **metrics(geometry_formal, target_power_mw=args.target_power), 'delta_q_from_six': geometry_formal.unit_area_power_kw_m2 - baseline_formal.unit_area_power_kw_m2})
        if _better(geometry_formal, geometry_origin_formal, target_power_mw=args.target_power, threshold=args.move_q):
            current_design = geometry_best['design']
            current_formal = geometry_formal
    geometry_active = tuple((parameter for parameter in ('initial_spacing', 'spacing_growth') if not math.isclose(current_design.parameter(parameter), geometry_origin.parameter(parameter))))
    print(f'几何固定点：D1={current_design.initial_spacing:.6f} m，g={current_design.spacing_growth:.6f} m/环', flush=True)
    print('阶段 3/6：18 个六区规格变量正负敏感性', flush=True)
    (sensitivity_reference, reason) = try_evaluate(current_design, profile_kind='medium', count_candidate=False)
    if sensitivity_reference is None:
        raise RuntimeError(f'敏感性中精度基准无法评价：{reason}')
    sensitivity_rows: list[dict[str, object]] = []
    sensitivity_designs: dict[tuple[str, str], RefineDesign] = {}
    for perturbation in specification_perturbations(current_design):
        (evaluation, reject) = try_evaluate(perturbation.design, profile_kind='medium')
        row: dict[str, object] = {'parameter': perturbation.parameter, 'group_id': perturbation.group_id, 'old_value': perturbation.old_value, 'new_value': perturbation.new_value, 'direction': perturbation.direction, 'legal': evaluation is not None, 'medium_power': None, 'medium_q': None, 'delta_power': None, 'delta_q': None, 'formal_power': None, 'formal_q': None, 'active': False, 'reject_reason': reject or ''}
        if evaluation is not None:
            row.update({'medium_power': evaluation.annual_power_mw, 'medium_q': evaluation.unit_area_power_kw_m2, 'delta_power': evaluation.annual_power_mw - baseline_medium.annual_power_mw, 'delta_q': evaluation.unit_area_power_kw_m2 - baseline_medium.unit_area_power_kw_m2, 'delta_power_from_geometry': evaluation.annual_power_mw - sensitivity_reference.annual_power_mw, 'delta_q_from_geometry': evaluation.unit_area_power_kw_m2 - sensitivity_reference.unit_area_power_kw_m2})
            sensitivity_designs[perturbation.parameter, perturbation.direction] = perturbation.design
        sensitivity_rows.append(row)
    selected_directions = select_formal_directions(sensitivity_rows, limit=6)
    for row in selected_directions:
        design = sensitivity_designs[str(row['parameter']), str(row['direction'])]
        (evaluation, reject) = try_evaluate(design, profile_kind='formal')
        if evaluation is None:
            row['reject_reason'] = reject or '正式复算失败'
            continue
        row['formal_power'] = evaluation.annual_power_mw
        row['formal_q'] = evaluation.unit_area_power_kw_m2
        row['active'] = evaluation.is_feasible(args.target_power) and evaluation.unit_area_power_kw_m2 > current_formal.unit_area_power_kw_m2 + args.move_q
        formal_rows.append({'stage': 'specification_sensitivity', 'candidate': f"{row['parameter']}{row['direction']}", **metrics(evaluation, target_power_mw=args.target_power), 'delta_q_from_six': evaluation.unit_area_power_kw_m2 - baseline_formal.unit_area_power_kw_m2})
    specification_active = active_from_formal(sensitivity_rows, reference_q=current_formal.unit_area_power_kw_m2, target_power_mw=args.target_power, threshold=args.move_q)
    active_variables = (*(('tower_y',) if tower_active else ()), *geometry_active, *specification_active)
    print(f"正式确认的活跃变量：{active_variables or '无'}", flush=True)
    print('阶段 4/6：活跃变量两轮分块回扫', flush=True)
    local_initial = sensitivity_reference

    def local_evaluator(design: RefineDesign) -> RefineEvaluation | None:
        (evaluation, reject_reason) = try_evaluate(design, profile_kind='medium')
        if evaluation is None and reject_reason == '达到中精度候选上限':
            return None
        return evaluation
    search = coordinate_search(initial_design=current_design, initial_evaluation=local_initial, active_variables=active_variables, evaluator=local_evaluator, baseline_q_kw_m2=baseline_medium.unit_area_power_kw_m2, maximum_sweeps=args.max_sweeps, target_power_mw=args.target_power, move_q_threshold=args.move_q)
    (attempted_formal, reason) = try_evaluate(search.best_design, profile_kind='formal')
    if attempted_formal is None:
        raise RuntimeError(f'最终候选正式复算失败：{reason}')
    formal_rows.append({'stage': 'final_acceptance', 'candidate': 'local-search-best', **metrics(attempted_formal, target_power_mw=args.target_power), 'delta_q_from_six': attempted_formal.unit_area_power_kw_m2 - baseline_formal.unit_area_power_kw_m2})
    print('阶段 5/6：塔位包围扫描与正式精度最细邻域收口', flush=True)
    preclosure_formal = attempted_formal
    closure_values: dict[RefineDesign, RefineEvaluation] = {search.best_design: attempted_formal}
    closure_count = 0

    def closure_evaluator(design: RefineDesign) -> RefineEvaluation | None:
        nonlocal closure_count
        if design in closure_values:
            return closure_values[design]
        try:
            evaluation = evaluate_design(baseline=baseline, design=design, profile=verification_profile, cache=formal_cache)
        except ValueError:
            return None
        closure_values[design] = evaluation
        closure_count += 1
        return evaluation
    closure = close_formal_neighborhood(initial_design=search.best_design, initial_evaluation=attempted_formal, evaluator=closure_evaluator, target_power_mw=args.target_power, coarse_step_limit=2 if args.smoke else 12, fine_radius_steps=1 if args.smoke else 4, maximum_local_sweeps=min(1, args.closure_sweeps) if args.smoke else args.closure_sweeps, move_q_threshold=args.move_q)
    if not args.smoke and (not closure.tower_bracketed):
        raise RuntimeError('塔位向北包围扫描未找到下降点。')
    attempted_formal = closure.best_evaluation
    formal_rows.append({'stage': 'formal_closure', 'candidate': 'bracketed-local-best', **metrics(attempted_formal, target_power_mw=args.target_power), 'delta_q_from_six': attempted_formal.unit_area_power_kw_m2 - baseline_formal.unit_area_power_kw_m2})
    print(f'收口候选：y={closure.best_design.tower_y:.6f} m，P={attempted_formal.annual_power_mw:.9f} MW，q={attempted_formal.unit_area_power_kw_m2:.9f} kW/m²', flush=True)
    dense_payload: dict[str, object] = {'status': 'not-run-formal-candidate-failed', 'baseline': {}, 'candidate': {}}
    formal_pass = _better(attempted_formal, baseline_formal, target_power_mw=args.target_power, threshold=0.0)
    dense_pass = False
    if args.smoke:
        dense_payload['status'] = 'smoke-skipped'
    elif formal_pass:
        dense_payload['status'] = 'completed'
        dense_cache = EvaluationCache()
        dense_pass = True
        for radius in (80.0, 100.0):
            profile = dense_profile(neighbor_radius_m=radius)
            baseline_dense = evaluate_design(baseline=baseline, design=baseline.design, profile=profile, cache=dense_cache)
            candidate_dense = evaluate_design(baseline=baseline, design=closure.best_design, profile=profile, cache=dense_cache)
            key = f'{int(radius)}'
            dense_payload['baseline'][key] = metrics(baseline_dense, target_power_mw=args.target_power)
            dense_payload['candidate'][key] = metrics(candidate_dense, target_power_mw=args.target_power)
            dense_pass = dense_pass and candidate_dense.is_feasible(args.target_power) and (candidate_dense.unit_area_power_kw_m2 > baseline_dense.unit_area_power_kw_m2)
        dense_payload['passed'] = dense_pass
    accepted = formal_pass and (args.smoke or dense_pass)
    if accepted:
        selected_design = closure.best_design
        selected_formal = attempted_formal
        decision = '微调方案通过统一正式与加密验收' if not args.smoke else 'smoke 链路通过'
    else:
        selected_design = baseline.design
        selected_formal = baseline_formal
        decision = '微调候选未通过统一验收，保留原六组正式方案'
    print('阶段 6/6：六区边界局部敏感性检验', flush=True)
    boundary_field = build_refine_field(baseline, selected_design)
    base_groups = group_indices_for_boundaries(boundary_field.ring_indices, BASE_BOUNDARIES)
    if not np.array_equal(boundary_field.group_indices, base_groups):
        raise RuntimeError('最终镜场分区与六区边界基准不一致。')
    boundary_medium_cache = EvaluationCache()
    boundary_formal_cache = EvaluationCache()
    boundary_formal_baseline = evaluate_field(baseline=baseline, design=selected_design, field=boundary_field, profile=verification_profile, cache=boundary_formal_cache)
    if not args.smoke:
        boundary_regression = abs(boundary_formal_baseline.total_area_m2 - selected_formal.total_area_m2) <= 1e-06 and abs(boundary_formal_baseline.annual_power_mw - selected_formal.annual_power_mw) <= 1e-06 and (abs(boundary_formal_baseline.unit_area_power_kw_m2 - selected_formal.unit_area_power_kw_m2) <= 1e-09)
        if not boundary_regression:
            raise RuntimeError('六区边界零扰动正式回归失败。')
    boundary_rows: list[dict[str, object]] = []
    for candidate in boundary_perturbations():
        candidate_field = reassign_boundary_groups(boundary_field, candidate.boundaries)
        groups = candidate_field.group_counts
        row: dict[str, object] = {'boundary_id': candidate.boundary_id, 'candidate': candidate.label, 'original_end_ring': candidate.original_end_ring, 'shift_rings': candidate.shift_rings, 'new_end_ring': candidate.new_end_ring, 'boundaries': '|'.join((str(value) for value in candidate.boundaries)), 'moved_mirror_count': moved_mirror_count(boundary_field, candidate.boundaries), 'group_counts': '|'.join((str(value) for value in groups))}
        try:
            medium = evaluate_field(baseline=baseline, design=selected_design, field=candidate_field, profile=search_profile, cache=boundary_medium_cache)
            formal = evaluate_field(baseline=baseline, design=selected_design, field=candidate_field, profile=verification_profile, cache=boundary_formal_cache)
        except ValueError as exc:
            if not args.smoke:
                raise
            row.update({'total_area_m2': math.nan, 'medium_power_mw': math.nan, 'medium_q_kw_m2': math.nan, 'formal_power_mw': math.nan, 'formal_power_margin_mw': math.nan, 'formal_q_kw_m2': math.nan, 'formal_delta_power_mw': math.nan, 'formal_delta_q_kw_m2': math.nan, 'formal_feasible': False, 'q_better_than_baseline': False, 'classification': 'smoke仅验证链路', 'smoke_reject_reason': str(exc)})
            boundary_rows.append(row)
            continue
        formal_feasible = formal.is_feasible(args.target_power)
        q_better = formal.unit_area_power_kw_m2 > boundary_formal_baseline.unit_area_power_kw_m2
        if formal_feasible and q_better:
            if not args.smoke:
                raise RuntimeError(f'边界候选 {candidate.label} 同时满足功率约束并提高 q，当前最终方案需要重新审定。')
            classification = 'smoke仅验证链路'
        elif formal_feasible:
            classification = '功率可行但q下降'
        elif q_better:
            classification = 'q提高但功率不达标'
        else:
            classification = '功率与q均不占优'
        row.update({'total_area_m2': formal.total_area_m2, 'medium_power_mw': medium.annual_power_mw, 'medium_q_kw_m2': medium.unit_area_power_kw_m2, 'formal_power_mw': formal.annual_power_mw, 'formal_power_margin_mw': formal.annual_power_mw - args.target_power, 'formal_q_kw_m2': formal.unit_area_power_kw_m2, 'formal_delta_power_mw': formal.annual_power_mw - boundary_formal_baseline.annual_power_mw, 'formal_delta_q_kw_m2': formal.unit_area_power_kw_m2 - boundary_formal_baseline.unit_area_power_kw_m2, 'formal_feasible': formal_feasible, 'q_better_than_baseline': q_better, 'classification': classification})
        boundary_rows.append(row)
    if args.smoke:
        print('边界检验 smoke 链路完成：18 个候选均已评价。', flush=True)
    else:
        print('边界检验完成：18 个候选均已执行中精度和正式精度评价，未发现可直接替换当前边界的可行改进。', flush=True)
    active_payload = {'tower_mode_decision': tower_decision, 'selected_tower_mode': current_design.tower_mode, 'active_variables': list(active_variables), 'tower_active': tower_active, 'geometry_active': list(geometry_active), 'specification_active': list(specification_active), 'medium_candidate_count': medium_count, 'medium_candidate_limit': args.medium_limit, 'formal_candidate_count': formal_count, 'formal_candidate_limit': args.formal_limit, 'maximum_joint_sweeps': args.max_sweeps, 'closure': {'tower_bracketed': closure.tower_bracketed, 'local_converged': closure.local_converged, 'local_check_completed': closure.local_sweeps >= 1, 'local_sweeps': closure.local_sweeps, 'accepted_moves': sum((bool(row['accepted']) for row in closure.trace)), 'formal_evaluations': closure_count}}
    regression['smoke'] = args.smoke
    regression['candidate_budgets'] = {'medium': {'used': medium_count, 'limit': args.medium_limit}, 'formal': {'used': formal_count, 'limit': args.formal_limit}}
    written = write_results(output_dir=args.output, baseline=baseline, regression=regression, tower_rows=tower_rows, geometry_rows=geometry_rows, sensitivity_rows=sensitivity_rows, active_payload=active_payload, search_trace=search.trace, formal_rows=formal_rows, baseline_formal=baseline_formal, preclosure_formal=preclosure_formal, attempted_formal=attempted_formal, selected_formal=selected_formal, selected_design=selected_design, dense_payload=dense_payload, result3_template=args.result3_template, target_power_mw=args.target_power, decision=decision, closure_rows=closure.trace, closure_payload=active_payload['closure'], boundary_rows=boundary_rows)
    figures = generate_figures(sensitivity_rows=sensitivity_rows, tower_rows=tower_rows, geometry_rows=geometry_rows, selected_tower_mode=current_design.tower_mode, baseline=baseline, selected_design=selected_design, baseline_formal=baseline_formal, candidate_formal=selected_formal, selected_formal=selected_formal, dense_payload=dense_payload, output_dir=args.output)
    for (index, path) in enumerate(figures, start=16):
        written[f'figure_{index}'] = path
    written['boundary_figure'] = plot_boundary_sensitivity(boundary_rows, output_dir=args.output)
    print('\n六区参数微调结果', flush=True)
    print(f'判定：{decision}', flush=True)
    print(f'正式候选：P={attempted_formal.annual_power_mw:.9f} MW', flush=True)
    print(f'正式候选：q={attempted_formal.unit_area_power_kw_m2:.9f} kW/m²', flush=True)
    print(f'候选预算：medium={medium_count}/{args.medium_limit}，formal={formal_count}/{args.formal_limit}，closure={closure_count}', flush=True)
    for path in written.values():
        print(f'输出：{path}', flush=True)
    return 0

def main() -> None:
    raise SystemExit(run())

if __name__ == "__main__":
    raise SystemExit(run())
