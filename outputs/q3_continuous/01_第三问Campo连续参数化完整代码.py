"""第三问 Campo 连续参数化完整代码展示稿。

本文件与原六组第三问并列存在，合并共享光学核心、Campo 结构标签、连续规格搜索、验证和输出流程。
"""
from __future__ import annotations
# ruff: noqa: E402,F401,F811

# ========================================================================
# 来源：src/heliostat/config.py
# ========================================================================

"""题目参数与数值计算参数。"""
from dataclasses import asdict, dataclass

@dataclass(frozen=True)
class FieldConfig:
    """第一问给定的镜场和环境参数。"""
    latitude_deg: float = 39.4
    altitude_km: float = 3.0
    field_radius: float = 350.0
    exclusion_radius: float = 100.0
    tower_x: float = 0.0
    tower_y: float = 0.0
    receiver_center_z: float = 86.0
    receiver_radius: float = 4.0
    receiver_height: float = 8.0
    mirror_width: float = 6.2
    mirror_height: float = 6.2
    mirror_center_z: float = 4.5
    reflectivity: float = 0.92
    solar_angular_radius_rad: float = 0.00465

    @property
    def receiver_z_min(self) -> float:
        return self.receiver_center_z - self.receiver_height / 2.0

    @property
    def receiver_z_max(self) -> float:
        return self.receiver_center_z + self.receiver_height / 2.0

    @property
    def mirror_area(self) -> float:
        return self.mirror_width * self.mirror_height

    def to_dict(self) -> dict[str, float]:
        return asdict(self)

@dataclass(frozen=True)
class SolverConfig:
    """可调的采样精度和加速参数。"""
    shadow_grid_size: int = 15
    truncation_rays: int = 256
    neighbor_radius_m: float = 60.0
    candidate_margin: float = 1.05
    ray_epsilon: float = 1e-07
    truncation_chunk_size: int = 128
    sobol_seed: int = 2023
    calculate_shadow: bool = True
    calculate_truncation: bool = True

    def __post_init__(self) -> None:
        if self.shadow_grid_size < 1:
            raise ValueError('shadow_grid_size 必须大于等于 1。')
        if self.truncation_rays < 1:
            raise ValueError('truncation_rays 必须大于等于 1。')
        if self.neighbor_radius_m <= 0.0:
            raise ValueError('neighbor_radius_m 必须大于 0。')
        if self.candidate_margin < 1.0:
            raise ValueError('candidate_margin 不能小于 1。')
        if self.ray_epsilon <= 0.0:
            raise ValueError('ray_epsilon 必须大于 0。')
        if self.truncation_chunk_size < 1:
            raise ValueError('truncation_chunk_size 必须大于等于 1。')

    def to_dict(self) -> dict[str, int | float | bool]:
        return asdict(self)

# ========================================================================
# 来源：src/heliostat/solar.py
# ========================================================================

"""太阳位置与 DNI 计算。"""
import math
from dataclasses import dataclass
import numpy as np
from numpy.typing import NDArray
FloatArray = NDArray[np.float64]

@dataclass(frozen=True)
class SolarState:
    month: int
    solar_time: float
    direction: FloatArray
    altitude_rad: float
    azimuth_rad: float
    declination_rad: float
    dni_kw_m2: float
MONTH_DAYS = (31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31)

def day_from_spring_equinox(month: int, day: int=21) -> int:
    """以 3 月 21 日为第 0 天，返回题面赤纬公式所需的 D。"""
    if not 1 <= month <= 12:
        raise ValueError('month 必须位于 1 到 12。')
    if not 1 <= day <= MONTH_DAYS[month - 1]:
        raise ValueError('day 不在指定月份的有效范围内。')
    day_of_year = sum(MONTH_DAYS[:month - 1]) + day
    return day_of_year - 80

def calculate_solar_state(month: int, solar_time: float, latitude_deg: float, altitude_km: float) -> SolarState:
    """按题面附录计算东-北-天坐标下的太阳单位方向和 DNI。"""
    if not 0.0 <= solar_time <= 24.0:
        raise ValueError('solar_time 必须位于 0 到 24 小时。')
    d = day_from_spring_equinox(month)
    declination = math.asin(math.sin(2.0 * math.pi * d / 365.0) * math.sin(math.radians(23.45)))
    latitude = math.radians(latitude_deg)
    hour_angle = math.pi / 12.0 * (solar_time - 12.0)
    direction = np.array([-math.cos(declination) * math.sin(hour_angle), math.cos(latitude) * math.sin(declination) - math.sin(latitude) * math.cos(declination) * math.cos(hour_angle), math.sin(latitude) * math.sin(declination) + math.cos(latitude) * math.cos(declination) * math.cos(hour_angle)], dtype=float)
    direction /= np.linalg.norm(direction)
    altitude = math.asin(float(np.clip(direction[2], -1.0, 1.0)))
    azimuth = math.atan2(float(direction[0]), float(direction[1])) % (2.0 * math.pi)
    h = altitude_km
    a = 0.4237 - 0.00821 * (6.0 - h) ** 2
    b = 0.5055 + 0.00595 * (6.5 - h) ** 2
    c = 0.2711 + 0.01858 * (2.5 - h) ** 2
    if altitude <= 0.0:
        dni = 0.0
    else:
        dni = 1.366 * (a + b * math.exp(-c / math.sin(altitude)))
    return SolarState(month=month, solar_time=solar_time, direction=direction, altitude_rad=altitude, azimuth_rad=azimuth, declination_rad=declination, dni_kw_m2=dni)

# ========================================================================
# 来源：src/heliostat/geometry.py
# ========================================================================

"""镜面姿态、反射与基础几何计算。"""
from dataclasses import dataclass
import numpy as np
from numpy.typing import NDArray
FloatArray = NDArray[np.float64]

def normalize_rows(vectors: FloatArray) -> FloatArray:
    norms = np.linalg.norm(vectors, axis=1, keepdims=True)
    if np.any(norms <= 1e-15):
        raise ValueError('不能归一化长度为零的向量。')
    return vectors / norms

def reflect(directions: FloatArray, normals: FloatArray) -> FloatArray:
    """根据 d - 2(d·n)n 计算反射方向。"""
    dot = np.sum(directions * normals, axis=-1, keepdims=True)
    return directions - 2.0 * dot * normals

@dataclass(frozen=True)
class PreparedField:
    config: FieldConfig
    centers: FloatArray
    mirror_widths: FloatArray
    mirror_heights: FloatArray
    mirror_areas: FloatArray
    receiver_center: FloatArray
    receiver_directions: FloatArray
    receiver_distances: FloatArray
    atmospheric_efficiency: FloatArray

    @property
    def mirror_count(self) -> int:
        return int(self.centers.shape[0])

    @property
    def total_mirror_area(self) -> float:
        return float(np.sum(self.mirror_areas))

@dataclass(frozen=True)
class MirrorOrientation:
    normals: FloatArray
    width_axes: FloatArray
    height_axes: FloatArray
    cosine_efficiency: FloatArray

def _per_mirror_values(values: FloatArray | None, *, fallback: float, mirror_count: int, name: str) -> FloatArray:
    if values is None:
        result = np.full(mirror_count, fallback, dtype=float)
    else:
        result = np.asarray(values, dtype=float)
        if result.ndim != 1 or result.shape[0] != mirror_count:
            raise ValueError(f'{name} 必须是一维且长度等于镜子数 {mirror_count}，实际形状为 {result.shape}。')
        result = result.copy()
    if not np.all(np.isfinite(result)):
        raise ValueError(f'{name} 包含 NaN 或无穷值。')
    if np.any(result <= 0.0):
        raise ValueError(f'{name} 必须全部大于 0。')
    return result

def prepare_field(mirror_xy: FloatArray, config: FieldConfig, *, mirror_widths: FloatArray | None=None, mirror_heights: FloatArray | None=None, mirror_center_zs: FloatArray | None=None) -> PreparedField:
    xy = np.asarray(mirror_xy, dtype=float)
    if xy.ndim != 2 or xy.shape[1] != 2 or xy.shape[0] == 0:
        raise ValueError('镜位坐标必须为非空 N×2 数组。')
    if not np.all(np.isfinite(xy)):
        raise ValueError('镜位坐标包含 NaN 或无穷值。')
    mirror_count = int(xy.shape[0])
    widths = _per_mirror_values(mirror_widths, fallback=config.mirror_width, mirror_count=mirror_count, name='mirror_widths')
    heights = _per_mirror_values(mirror_heights, fallback=config.mirror_height, mirror_count=mirror_count, name='mirror_heights')
    center_zs = _per_mirror_values(mirror_center_zs, fallback=config.mirror_center_z, mirror_count=mirror_count, name='mirror_center_zs')
    areas = widths * heights
    centers = np.column_stack((xy, center_zs))
    receiver_center = np.array([config.tower_x, config.tower_y, config.receiver_center_z], dtype=float)
    receiver_vectors = receiver_center[None, :] - centers
    distances = np.linalg.norm(receiver_vectors, axis=1)
    receiver_directions = receiver_vectors / distances[:, None]
    atmospheric = 0.99321 - 0.0001176 * distances + 1.97e-08 * distances ** 2
    atmospheric = np.clip(atmospheric, 0.0, 1.0)
    return PreparedField(config=config, centers=centers, mirror_widths=widths, mirror_heights=heights, mirror_areas=areas, receiver_center=receiver_center, receiver_directions=receiver_directions, receiver_distances=distances, atmospheric_efficiency=atmospheric)

def calculate_orientation(prepared: PreparedField, sun_direction: FloatArray) -> MirrorOrientation:
    sun_rows = np.broadcast_to(sun_direction, prepared.receiver_directions.shape)
    normals = normalize_rows(sun_rows + prepared.receiver_directions)
    upward = np.broadcast_to(np.array([0.0, 0.0, 1.0]), normals.shape)
    width_axes = np.cross(upward, normals)
    weak = np.linalg.norm(width_axes, axis=1) < 1e-10
    width_axes[weak] = np.array([1.0, 0.0, 0.0])
    width_axes = normalize_rows(width_axes)
    height_axes = normalize_rows(np.cross(normals, width_axes))
    cosine = np.clip(normals @ sun_direction, 0.0, 1.0)
    return MirrorOrientation(normals=normals, width_axes=width_axes, height_axes=height_axes, cosine_efficiency=cosine)

def maximum_reflection_error(prepared: PreparedField, orientation: MirrorOrientation, sun_direction: FloatArray) -> float:
    incoming = np.broadcast_to(-sun_direction, orientation.normals.shape)
    reflected = reflect(incoming, orientation.normals)
    errors = np.linalg.norm(reflected - prepared.receiver_directions, axis=1)
    return float(np.max(errors))

# ========================================================================
# 来源：src/heliostat/shadow.py
# ========================================================================

"""规则网格射线追踪计算阴影遮挡效率。"""
import numpy as np
from numpy.typing import NDArray
from scipy.spatial import cKDTree
IntArray = NDArray[np.int64]

def mirror_grid_offsets(grid_size: int, mirror_width: float, mirror_height: float) -> FloatArray:
    """返回位于等面积小格中心的局部二维坐标。"""
    width_step = mirror_width / grid_size
    height_step = mirror_height / grid_size
    width_values = np.linspace(-mirror_width / 2.0 + width_step / 2.0, mirror_width / 2.0 - width_step / 2.0, grid_size)
    height_values = np.linspace(-mirror_height / 2.0 + height_step / 2.0, mirror_height / 2.0 - height_step / 2.0, grid_size)
    (width_grid, height_grid) = np.meshgrid(width_values, height_values, indexing='xy')
    return np.column_stack((width_grid.ravel(), height_grid.ravel()))

def _direction_candidates(target_index: int, neighbors: IntArray, centers: FloatArray, direction: FloatArray, reach: float, maximum_distance: float | None=None) -> IntArray:
    if neighbors.size == 0:
        return neighbors
    delta = centers[neighbors] - centers[target_index]
    projection = delta @ direction
    perpendicular = delta - projection[:, None] * direction[None, :]
    perpendicular_distance = np.linalg.norm(perpendicular, axis=1)
    keep = (projection > -reach) & (perpendicular_distance <= reach)
    if maximum_distance is not None:
        keep &= projection < maximum_distance + reach
    return neighbors[keep]

def ray_rectangle_hits(origins: FloatArray, direction: FloatArray, rectangle_center: FloatArray, rectangle_normal: FloatArray, rectangle_width_axis: FloatArray, rectangle_height_axis: FloatArray, half_width: float, half_height: float, epsilon: float, maximum_distance: float | None=None) -> NDArray[np.bool_]:
    """判断一组同向射线是否与一面有限矩形相交。"""
    denominator = float(direction @ rectangle_normal)
    if abs(denominator) <= epsilon:
        return np.zeros(origins.shape[0], dtype=bool)
    distance = (rectangle_center - origins) @ rectangle_normal / denominator
    active = distance > epsilon
    if maximum_distance is not None:
        active &= distance < maximum_distance - epsilon
    if not np.any(active):
        return active
    intersections = origins + distance[:, None] * direction[None, :]
    relative = intersections - rectangle_center[None, :]
    local_width = relative @ rectangle_width_axis
    local_height = relative @ rectangle_height_axis
    active &= np.abs(local_width) <= half_width + epsilon
    active &= np.abs(local_height) <= half_height + epsilon
    return active

def _blocked_by_candidates(origins: FloatArray, direction: FloatArray, candidates: IntArray, prepared: PreparedField, orientation: MirrorOrientation, solver: SolverConfig, maximum_distance: float | None=None) -> NDArray[np.bool_]:
    blocked = np.zeros(origins.shape[0], dtype=bool)
    for candidate in candidates:
        active_indices = np.flatnonzero(~blocked)
        if active_indices.size == 0:
            break
        hits = ray_rectangle_hits(origins=origins[active_indices], direction=direction, rectangle_center=prepared.centers[candidate], rectangle_normal=orientation.normals[candidate], rectangle_width_axis=orientation.width_axes[candidate], rectangle_height_axis=orientation.height_axes[candidate], half_width=prepared.mirror_widths[candidate] / 2.0, half_height=prepared.mirror_heights[candidate] / 2.0, epsilon=solver.ray_epsilon, maximum_distance=maximum_distance)
        blocked[active_indices[hits]] = True
    return blocked

def calculate_shadow_blocking_efficiency(prepared: PreparedField, orientation: MirrorOrientation, sun_direction: FloatArray, solver: SolverConfig) -> FloatArray:
    """逐镜计算入射阴影和反射遮挡损失的采样点并集。"""
    mirror_count = prepared.mirror_count
    if mirror_count == 1:
        return np.ones(1, dtype=float)
    tree = cKDTree(prepared.centers[:, :2])
    bounding_radii = 0.5 * np.hypot(prepared.mirror_widths, prepared.mirror_heights)
    maximum_bounding_radius = float(np.max(bounding_radii))
    efficiencies = np.empty(mirror_count, dtype=float)
    offset_cache: dict[tuple[float, float], FloatArray] = {}
    for index in range(mirror_count):
        size_key = (float(prepared.mirror_widths[index]), float(prepared.mirror_heights[index]))
        offsets = offset_cache.get(size_key)
        if offsets is None:
            offsets = mirror_grid_offsets(solver.shadow_grid_size, size_key[0], size_key[1])
            offset_cache[size_key] = offsets
        sample_count = offsets.shape[0]
        reach = (float(bounding_radii[index]) + maximum_bounding_radius) * solver.candidate_margin
        points = prepared.centers[index][None, :] + offsets[:, :1] * orientation.width_axes[index][None, :] + offsets[:, 1:] * orientation.height_axes[index][None, :]
        neighbors = np.asarray(tree.query_ball_point(prepared.centers[index, :2], solver.neighbor_radius_m), dtype=np.int64)
        neighbors = neighbors[neighbors != index]
        incoming_candidates = _direction_candidates(target_index=index, neighbors=neighbors, centers=prepared.centers, direction=sun_direction, reach=reach)
        incoming_blocked = _blocked_by_candidates(origins=points, direction=sun_direction, candidates=incoming_candidates, prepared=prepared, orientation=orientation, solver=solver)
        reflected_direction = prepared.receiver_directions[index]
        reflected_candidates = _direction_candidates(target_index=index, neighbors=neighbors, centers=prepared.centers, direction=reflected_direction, reach=reach, maximum_distance=prepared.receiver_distances[index])
        reflected_blocked = _blocked_by_candidates(origins=points, direction=reflected_direction, candidates=reflected_candidates, prepared=prepared, orientation=orientation, solver=solver, maximum_distance=prepared.receiver_distances[index])
        blocked = incoming_blocked | reflected_blocked
        efficiencies[index] = 1.0 - np.count_nonzero(blocked) / sample_count
    return np.clip(efficiencies, 0.0, 1.0)

# ========================================================================
# 来源：src/heliostat/truncation.py
# ========================================================================

"""太阳锥光联合采样与有限高圆柱集热器求交。"""
import math
import numpy as np
from numpy.typing import NDArray
from scipy.stats import qmc

def build_sobol_samples(sample_count: int, seed: int) -> FloatArray:
    """生成固定、可复现的四维 Sobol 样本。"""
    exponent = int(math.ceil(math.log2(sample_count)))
    sampler = qmc.Sobol(d=4, scramble=True, seed=seed)
    return sampler.random_base2(exponent)[:sample_count]

def _sun_disk_directions(sun_direction: FloatArray, samples: FloatArray, angular_radius: float) -> FloatArray:
    reference = np.array([1.0, 0.0, 0.0]) if abs(float(sun_direction[2])) > 0.9 else np.array([0.0, 0.0, 1.0])
    tangent_one = np.cross(reference, sun_direction)
    tangent_one /= np.linalg.norm(tangent_one)
    tangent_two = np.cross(sun_direction, tangent_one)
    radial_angle = angular_radius * np.sqrt(samples[:, 2])
    polar_angle = 2.0 * math.pi * samples[:, 3]
    tangent = np.cos(polar_angle)[:, None] * tangent_one[None, :] + np.sin(polar_angle)[:, None] * tangent_two[None, :]
    directions = np.cos(radial_angle)[:, None] * sun_direction[None, :] + np.sin(radial_angle)[:, None] * tangent
    directions /= np.linalg.norm(directions, axis=1, keepdims=True)
    return directions

def ray_cylinder_side_hits(origins: FloatArray, directions: FloatArray, tower_x: float, tower_y: float, radius: float, z_min: float, z_max: float, epsilon: float) -> NDArray[np.bool_]:
    """判断任意形状批量射线的两个正根中是否有有限圆柱侧面交点。"""
    origin_x = origins[..., 0] - tower_x
    origin_y = origins[..., 1] - tower_y
    direction_x = directions[..., 0]
    direction_y = directions[..., 1]
    a = direction_x ** 2 + direction_y ** 2
    b = 2.0 * (origin_x * direction_x + origin_y * direction_y)
    c = origin_x ** 2 + origin_y ** 2 - radius ** 2
    discriminant = b ** 2 - 4.0 * a * c
    valid = (a > epsilon) & (discriminant >= 0.0)
    square_root = np.sqrt(np.maximum(discriminant, 0.0))
    denominator = np.where(valid, 2.0 * a, 1.0)
    near = (-b - square_root) / denominator
    far = (-b + square_root) / denominator
    near_z = origins[..., 2] + near * directions[..., 2]
    far_z = origins[..., 2] + far * directions[..., 2]
    near_hit = (near > epsilon) & (near_z >= z_min - epsilon) & (near_z <= z_max + epsilon)
    far_hit = (far > epsilon) & (far_z >= z_min - epsilon) & (far_z <= z_max + epsilon)
    return valid & (near_hit | far_hit)

def calculate_truncation_efficiency(prepared: PreparedField, orientation: MirrorOrientation, sun_direction: FloatArray, solver: SolverConfig) -> FloatArray:
    """联合采样镜面位置和太阳圆盘方向，计算集热器截断效率。"""
    config = prepared.config
    samples = build_sobol_samples(solver.truncation_rays, solver.sobol_seed)
    unit_width = samples[:, 0] - 0.5
    unit_height = samples[:, 1] - 0.5
    sampled_sun = _sun_disk_directions(sun_direction, samples, config.solar_angular_radius_rad)
    incoming = -sampled_sun
    efficiencies = np.empty(prepared.mirror_count, dtype=float)
    chunk_size = solver.truncation_chunk_size
    for start in range(0, prepared.mirror_count, chunk_size):
        stop = min(start + chunk_size, prepared.mirror_count)
        centers = prepared.centers[start:stop]
        normals = orientation.normals[start:stop]
        width_axes = orientation.width_axes[start:stop]
        height_axes = orientation.height_axes[start:stop]
        local_width = prepared.mirror_widths[start:stop, None] * unit_width[None, :]
        local_height = prepared.mirror_heights[start:stop, None] * unit_height[None, :]
        origins = centers[:, None, :] + local_width[:, :, None] * width_axes[:, None, :] + local_height[:, :, None] * height_axes[:, None, :]
        incoming_chunk = np.broadcast_to(incoming[None, :, :], origins.shape)
        dot = np.einsum('csj,cj->cs', incoming_chunk, normals)
        reflected = incoming_chunk - 2.0 * dot[:, :, None] * normals[:, None, :]
        hits = ray_cylinder_side_hits(origins=origins, directions=reflected, tower_x=config.tower_x, tower_y=config.tower_y, radius=config.receiver_radius, z_min=config.receiver_z_min, z_max=config.receiver_z_max, epsilon=solver.ray_epsilon)
        efficiencies[start:stop] = np.mean(hits, axis=1)
    return np.clip(efficiencies, 0.0, 1.0)

# ========================================================================
# 来源：src/heliostat/io.py
# ========================================================================

"""三问共用的坐标输入。"""
import csv
from pathlib import Path
import numpy as np
from numpy.typing import NDArray
from openpyxl import load_workbook
FloatArray = NDArray[np.float64]

def load_mirror_xy(path: str | Path, expected_count: int | None=1745) -> FloatArray:
    """从题目附件读取定日镜 x、y 坐标。"""
    source = Path(path)
    if not source.exists():
        raise FileNotFoundError(f'找不到定日镜坐标文件：{source}')
    if source.suffix.lower() == '.xlsx':
        workbook = load_workbook(source, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        workbook.close()
        values = [(row[0], row[1]) for row in rows if row[0] is not None]
    elif source.suffix.lower() == '.csv':
        with source.open('r', encoding='utf-8-sig', newline='') as handle:
            reader = csv.reader(handle)
            next(reader, None)
            values = [(row[0], row[1]) for row in reader if row]
    else:
        raise ValueError('坐标文件只支持 .xlsx 或 .csv。')
    try:
        mirror_xy = np.asarray(values, dtype=float)
    except (TypeError, ValueError) as exc:
        raise ValueError(f'坐标文件包含非数值数据：{source}') from exc
    if mirror_xy.ndim != 2 or mirror_xy.shape[1] != 2:
        raise ValueError(f'坐标数据应为 N×2，实际形状为 {mirror_xy.shape}。')
    if not np.all(np.isfinite(mirror_xy)):
        raise ValueError('坐标数据包含 NaN 或无穷值。')
    if expected_count is not None and mirror_xy.shape[0] != expected_count:
        raise ValueError(f'应读取 {expected_count} 面定日镜，实际读取 {mirror_xy.shape[0]} 面。')
    return mirror_xy

# ========================================================================
# 来源：src/heliostat/q1/aggregate.py
# ========================================================================

"""第一问的月平均和年平均汇总。"""
from dataclasses import dataclass
from typing import Any, Sequence
import numpy as np

@dataclass(frozen=True)
class MonthlyResult:
    month: int
    average_optical_efficiency: float
    average_cosine_efficiency: float
    average_shadow_blocking_efficiency: float
    average_atmospheric_efficiency: float
    average_truncation_efficiency: float
    field_output_mw: float
    unit_area_output_kw_m2: float

@dataclass(frozen=True)
class AnnualResult:
    average_optical_efficiency: float
    average_cosine_efficiency: float
    average_shadow_blocking_efficiency: float
    average_atmospheric_efficiency: float
    average_truncation_efficiency: float
    field_output_mw: float
    unit_area_output_kw_m2: float

@dataclass(frozen=True)
class MirrorAnnualResult:
    mirror_id: int
    x_m: float
    y_m: float
    radius_to_tower_m: float
    average_optical_efficiency: float
    average_cosine_efficiency: float
    average_shadow_blocking_efficiency: float
    average_atmospheric_efficiency: float
    average_truncation_efficiency: float
    average_output_power_kw: float

@dataclass(frozen=True)
class Question1Solution:
    time_results: tuple[Any, ...]
    monthly_results: tuple[MonthlyResult, ...]
    annual_result: AnnualResult
    mirror_annual_results: tuple[MirrorAnnualResult, ...]
_MEAN_FIELDS = ('average_optical_efficiency', 'average_cosine_efficiency', 'average_shadow_blocking_efficiency', 'average_atmospheric_efficiency', 'average_truncation_efficiency', 'field_output_mw', 'unit_area_output_kw_m2')

def _means(records: Sequence[Any]) -> tuple[float, ...]:
    if not records:
        raise ValueError('汇总记录不能为空。')
    return tuple((float(np.mean([getattr(record, field) for record in records])) for field in _MEAN_FIELDS))

def summarize_monthly(records: Sequence[Any]) -> tuple[MonthlyResult, ...]:
    """对每月规定的五个时刻等权平均。"""
    results: list[MonthlyResult] = []
    for month in sorted({record.month for record in records}):
        monthly = [record for record in records if record.month == month]
        results.append(MonthlyResult(month, *_means(monthly)))
    return tuple(results)

def summarize_annual(records: Sequence[Any]) -> AnnualResult:
    """对题目规定的全部时刻等权平均。"""
    return AnnualResult(*_means(records))

def summarize_mirror_annual(mirror_xy: np.ndarray, tower_x: float, tower_y: float, state_count: int, optical_efficiency_sum: np.ndarray, cosine_efficiency_sum: np.ndarray, shadow_blocking_efficiency_sum: np.ndarray, atmospheric_efficiency_sum: np.ndarray, truncation_efficiency_sum: np.ndarray, output_power_kw_sum: np.ndarray) -> tuple[MirrorAnnualResult, ...]:
    """由逐时刻运行和生成单镜年平均结果，不保留单镜逐时刻明细。"""
    if state_count < 1:
        raise ValueError('state_count 必须大于等于 1。')
    radius = np.hypot(mirror_xy[:, 0] - tower_x, mirror_xy[:, 1] - tower_y)
    means = {'optical': optical_efficiency_sum / state_count, 'cosine': cosine_efficiency_sum / state_count, 'shadow': shadow_blocking_efficiency_sum / state_count, 'atmospheric': atmospheric_efficiency_sum / state_count, 'truncation': truncation_efficiency_sum / state_count, 'power': output_power_kw_sum / state_count}
    return tuple((MirrorAnnualResult(mirror_id=index + 1, x_m=float(mirror_xy[index, 0]), y_m=float(mirror_xy[index, 1]), radius_to_tower_m=float(radius[index]), average_optical_efficiency=float(means['optical'][index]), average_cosine_efficiency=float(means['cosine'][index]), average_shadow_blocking_efficiency=float(means['shadow'][index]), average_atmospheric_efficiency=float(means['atmospheric'][index]), average_truncation_efficiency=float(means['truncation'][index]), average_output_power_kw=float(means['power'][index])) for index in range(mirror_xy.shape[0])))

# ========================================================================
# 来源：src/heliostat/q1/export.py
# ========================================================================

"""第一问结果和论文表格输出。"""
import csv
import json
from dataclasses import asdict
from pathlib import Path
from typing import Any, Iterable, Sequence

def _write_csv(path: Path, rows: Sequence[dict[str, Any]]) -> None:
    if not rows:
        raise ValueError(f'没有可写入 {path.name} 的结果。')
    with path.open('w', encoding='utf-8-sig', newline='') as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)

def _display_source_path(source_path: str | Path) -> str:
    path = Path(source_path)
    if not path.is_absolute():
        return path.as_posix()
    if 'task' in path.parts:
        task_index = path.parts.index('task')
        return Path(*path.parts[task_index:]).as_posix()
    return path.name

def write_question1_results(output_dir: str | Path, time_records: Iterable[Any], monthly_records: Iterable[Any], annual_record: Any, mirror_annual_records: Iterable[Any], field_config: FieldConfig, solver_config: SolverConfig, source_path: str | Path, mirror_count: int) -> dict[str, Path]:
    """保持原有四类结果文件名和字段口径不变。"""
    destination = Path(output_dir)
    destination.mkdir(parents=True, exist_ok=True)
    time_rows = [asdict(record) for record in time_records]
    monthly_rows = [asdict(record) for record in monthly_records]
    annual_row = asdict(annual_record)
    mirror_annual_rows = [asdict(record) for record in mirror_annual_records]
    months = sorted({row['month'] for row in time_rows})
    solar_times = sorted({row['solar_time'] for row in time_rows})
    time_path = destination / '02_逐时刻计算结果.csv'
    monthly_path = destination / '03_月平均计算结果.csv'
    annual_path = destination / '04_年平均计算结果.json'
    mirror_annual_path = destination / '05_单镜年平均结果.csv'
    run_path = destination / '06_运行配置.json'
    _write_csv(time_path, time_rows)
    _write_csv(monthly_path, monthly_rows)
    _write_csv(mirror_annual_path, mirror_annual_rows)
    annual_path.write_text(json.dumps(annual_row, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    run_path.write_text(json.dumps({'source': _display_source_path(source_path), 'field': field_config.to_dict(), 'solver': solver_config.to_dict(), 'run': {'mirror_count': mirror_count, 'months': months, 'solar_times': solar_times, 'time_state_count': len(time_rows)}}, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    return {'time': time_path, 'monthly': monthly_path, 'annual': annual_path, 'mirror_annual': mirror_annual_path, 'config': run_path}

def write_paper_tables(output_dir: str | Path, monthly_records: Iterable[Any], annual_record: Any) -> dict[str, Path]:
    """将月平均、年平均和验证表集中到一个展示文件。"""
    destination = Path(output_dir)
    destination.mkdir(parents=True, exist_ok=True)
    table_path = destination / '07_论文结果与验证表.md'
    monthly_lines = ['# 第一问结果与验证表', '', '## 表 1 每月 21 日平均光学效率及输出功率', '', '| 日期 | 平均光学效率 | 平均余弦效率 | 平均阴影遮挡效率 | 平均截断效率 | 单位面积镜面平均输出热功率 (kW/m²) |', '| --- | ---: | ---: | ---: | ---: | ---: |']
    for record in monthly_records:
        monthly_lines.append(f'| {record.month} 月 21 日 | {record.average_optical_efficiency:.6f} | {record.average_cosine_efficiency:.6f} | {record.average_shadow_blocking_efficiency:.6f} | {record.average_truncation_efficiency:.6f} | {record.unit_area_output_kw_m2:.6f} |')
    annual_lines = ['', '## 表 2 年平均光学效率及输出功率', '', '| 年平均光学效率 | 年平均余弦效率 | 年平均阴影遮挡效率 | 年平均截断效率 | 年平均输出热功率 (MW) | 单位面积镜面年平均输出热功率 (kW/m²) |', '| ---: | ---: | ---: | ---: | ---: | ---: |', f'| {annual_record.average_optical_efficiency:.6f} | {annual_record.average_cosine_efficiency:.6f} | {annual_record.average_shadow_blocking_efficiency:.6f} | {annual_record.average_truncation_efficiency:.6f} | {annual_record.field_output_mw:.6f} | {annual_record.unit_area_output_kw_m2:.6f} |']
    table_path.write_text('\n'.join(monthly_lines + annual_lines) + '\n', encoding='utf-8')
    return {'paper_tables': table_path}

def write_validation_table(output_dir: str | Path, validation_records: Iterable[Any]) -> dict[str, Path]:
    """把三组收敛实验追加为一张验证表。"""
    destination = Path(output_dir)
    rows = [asdict(record) for record in validation_records]
    table_path = destination / '07_论文结果与验证表.md'
    lines = ['', '## 表 3 数值收敛验证', '', '| 验证项目 | 参数 | 观测指标 | 数值 | 相对正式配置差异 | 运行时间 (s) |', '| --- | ---: | --- | ---: | ---: | ---: |']
    for row in rows:
        lines.append(f"| {row['category']} | {row['parameter']} | {row['metric']} | {row['value']:.6f} | {row['relative_difference_percent']:.4f}% | {row['runtime_seconds']:.3f} |")
    with table_path.open('a', encoding='utf-8') as handle:
        handle.write('\n'.join(lines) + '\n')
    return {'validation_table': table_path}

# ========================================================================
# 来源：src/heliostat/q1/solve.py
# ========================================================================

"""第一问逐时刻计算、验证运行和命令行入口。"""
import argparse
import time
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Callable, Sequence
import numpy as np
SOLAR_TIMES = (9.0, 10.5, 12.0, 13.5, 15.0)
PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_INPUT = PROJECT_ROOT / 'task' / 'A' / 'fj.xlsx'
DEFAULT_OUTPUT = PROJECT_ROOT / 'outputs' / 'q1'

@dataclass(frozen=True)
class TimeResult:
    month: int
    solar_time: float
    dni_kw_m2: float
    average_optical_efficiency: float
    average_cosine_efficiency: float
    average_shadow_blocking_efficiency: float
    average_atmospheric_efficiency: float
    average_truncation_efficiency: float
    field_output_mw: float
    unit_area_output_kw_m2: float
    maximum_reflection_error: float

@dataclass(frozen=True)
class ValidationResult:
    category: str
    parameter: str
    metric: str
    value: float
    relative_difference_percent: float
    runtime_seconds: float
ProgressCallback = Callable[[int, int, TimeResult], None]

def _check_efficiency(name: str, values: np.ndarray) -> None:
    tolerance = 1e-12
    if np.any(values < -tolerance) or np.any(values > 1.0 + tolerance):
        minimum = float(np.min(values))
        maximum = float(np.max(values))
        raise RuntimeError(f'{name} 超出 [0, 1]：min={minimum:.6g}, max={maximum:.6g}')

def evaluate_time(prepared: PreparedField, month: int, solar_time: float, solver: SolverConfig, mirror_sums: dict[str, np.ndarray] | None=None) -> TimeResult:
    """计算一个月份、一个规定时刻的全场平均结果。"""
    solar = calculate_solar_state(month=month, solar_time=solar_time, latitude_deg=prepared.config.latitude_deg, altitude_km=prepared.config.altitude_km)
    orientation = calculate_orientation(prepared, solar.direction)
    reflection_error = maximum_reflection_error(prepared, orientation, solar.direction)
    if reflection_error >= 1e-08:
        raise RuntimeError(f'中心光线反射误差过大：{reflection_error:.3e}')
    if solver.calculate_shadow:
        shadow = calculate_shadow_blocking_efficiency(prepared, orientation, solar.direction, solver)
    else:
        shadow = np.ones(prepared.mirror_count, dtype=float)
    if solver.calculate_truncation:
        truncation = calculate_truncation_efficiency(prepared, orientation, solar.direction, solver)
    else:
        truncation = np.ones(prepared.mirror_count, dtype=float)
    cosine = orientation.cosine_efficiency
    atmospheric = prepared.atmospheric_efficiency
    optical = cosine * shadow * atmospheric * truncation * prepared.config.reflectivity
    for (name, values) in (('余弦效率', cosine), ('阴影遮挡效率', shadow), ('大气透射率', atmospheric), ('截断效率', truncation), ('光学效率', optical)):
        _check_efficiency(name, values)
    mirror_power_kw = solar.dni_kw_m2 * prepared.mirror_areas * optical
    if mirror_sums is not None:
        mirror_sums['optical_efficiency_sum'] += optical
        mirror_sums['cosine_efficiency_sum'] += cosine
        mirror_sums['shadow_blocking_efficiency_sum'] += shadow
        mirror_sums['atmospheric_efficiency_sum'] += atmospheric
        mirror_sums['truncation_efficiency_sum'] += truncation
        mirror_sums['output_power_kw_sum'] += mirror_power_kw
    field_power_kw = float(np.sum(mirror_power_kw))
    area_weights = prepared.mirror_areas
    return TimeResult(month=month, solar_time=solar_time, dni_kw_m2=solar.dni_kw_m2, average_optical_efficiency=float(np.average(optical, weights=area_weights)), average_cosine_efficiency=float(np.average(cosine, weights=area_weights)), average_shadow_blocking_efficiency=float(np.average(shadow, weights=area_weights)), average_atmospheric_efficiency=float(np.average(atmospheric, weights=area_weights)), average_truncation_efficiency=float(np.average(truncation, weights=area_weights)), field_output_mw=field_power_kw / 1000.0, unit_area_output_kw_m2=field_power_kw / prepared.total_mirror_area, maximum_reflection_error=reflection_error)

def solve_question1(prepared: PreparedField, solver: SolverConfig, months: Sequence[int]=tuple(range(1, 13)), solar_times: Sequence[float]=SOLAR_TIMES, progress: ProgressCallback | None=None) -> Question1Solution:
    """执行所选月份和时刻；默认即题目规定的 60 个状态。"""
    if not months or not solar_times:
        raise ValueError('months 和 solar_times 不能为空。')
    if any((month < 1 or month > 12 for month in months)):
        raise ValueError('months 必须位于 1 到 12。')
    records: list[TimeResult] = []
    mirror_sums = {name: np.zeros(prepared.mirror_count, dtype=float) for name in ('optical_efficiency_sum', 'cosine_efficiency_sum', 'shadow_blocking_efficiency_sum', 'atmospheric_efficiency_sum', 'truncation_efficiency_sum', 'output_power_kw_sum')}
    total = len(months) * len(solar_times)
    for month in months:
        for solar_time in solar_times:
            record = evaluate_time(prepared, month, solar_time, solver, mirror_sums=mirror_sums)
            records.append(record)
            if progress is not None:
                progress(len(records), total, record)
    time_results = tuple(records)
    return Question1Solution(time_results=time_results, monthly_results=summarize_monthly(time_results), annual_result=summarize_annual(time_results), mirror_annual_results=summarize_mirror_annual(mirror_xy=prepared.centers[:, :2], tower_x=prepared.config.tower_x, tower_y=prepared.config.tower_y, state_count=len(time_results), **mirror_sums))

def run_validation_suite(prepared: PreparedField, base_solver: SolverConfig) -> tuple[ValidationResult, ...]:
    """运行三组隔离后的收敛实验，供一张验证表使用。"""
    specifications = [('阴影网格', '10×10', replace(base_solver, shadow_grid_size=10, calculate_shadow=True, calculate_truncation=False), 'average_shadow_blocking_efficiency', '年平均阴影遮挡效率', False), ('阴影网格', '15×15', replace(base_solver, shadow_grid_size=15, calculate_shadow=True, calculate_truncation=False), 'average_shadow_blocking_efficiency', '年平均阴影遮挡效率', True), ('阴影网格', '20×20', replace(base_solver, shadow_grid_size=20, calculate_shadow=True, calculate_truncation=False), 'average_shadow_blocking_efficiency', '年平均阴影遮挡效率', False), ('邻镜半径', '40 m', replace(base_solver, neighbor_radius_m=40.0, calculate_shadow=True, calculate_truncation=False), 'average_shadow_blocking_efficiency', '年平均阴影遮挡效率', False), ('邻镜半径', '60 m', replace(base_solver, neighbor_radius_m=60.0, calculate_shadow=True, calculate_truncation=False), 'average_shadow_blocking_efficiency', '年平均阴影遮挡效率', True), ('邻镜半径', '80 m', replace(base_solver, neighbor_radius_m=80.0, calculate_shadow=True, calculate_truncation=False), 'average_shadow_blocking_efficiency', '年平均阴影遮挡效率', False), ('截断光线', '128', replace(base_solver, truncation_rays=128, calculate_shadow=False, calculate_truncation=True), 'average_truncation_efficiency', '年平均截断效率', False), ('截断光线', '256', replace(base_solver, truncation_rays=256, calculate_shadow=False, calculate_truncation=True), 'average_truncation_efficiency', '年平均截断效率', True), ('截断光线', '512', replace(base_solver, truncation_rays=512, calculate_shadow=False, calculate_truncation=True), 'average_truncation_efficiency', '年平均截断效率', False)]
    raw: list[tuple[str, str, str, float, float, bool]] = []
    for (category, parameter, solver, field, metric, reference) in specifications:
        started = time.perf_counter()
        solution = solve_question1(prepared, solver)
        elapsed = time.perf_counter() - started
        value = float(getattr(solution.annual_result, field))
        raw.append((category, parameter, metric, value, elapsed, reference))
    baselines = {category: value for (category, _, _, value, _, reference) in raw if reference}
    return tuple((ValidationResult(category=category, parameter=parameter, metric=metric, value=value, relative_difference_percent=abs(value - baselines[category]) / abs(baselines[category]) * 100.0, runtime_seconds=elapsed) for (category, parameter, metric, value, elapsed, _) in raw))

def _comma_ints(value: str) -> tuple[int, ...]:
    try:
        result = tuple((int(item.strip()) for item in value.split(',') if item.strip()))
    except ValueError as exc:
        raise argparse.ArgumentTypeError('月份应使用逗号分隔的整数。') from exc
    if not result:
        raise argparse.ArgumentTypeError('月份列表不能为空。')
    return result

def _comma_floats(value: str) -> tuple[float, ...]:
    try:
        result = tuple((float(item.strip()) for item in value.split(',') if item.strip()))
    except ValueError as exc:
        raise argparse.ArgumentTypeError('时刻应使用逗号分隔的数字。') from exc
    if not result:
        raise argparse.ArgumentTypeError('时刻列表不能为空。')
    return result

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description='计算 CUMCM 2023 A 题第一问的镜场光学效率和输出热功率')
    parser.add_argument('--input', type=Path, default=DEFAULT_INPUT)
    parser.add_argument('--output', type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument('--shadow-grid', type=int, default=15)
    parser.add_argument('--truncation-rays', type=int, default=256)
    parser.add_argument('--neighbor-radius', type=float, default=60.0)
    parser.add_argument('--truncation-chunk-size', type=int, default=128)
    parser.add_argument('--sobol-seed', type=int, default=2023)
    parser.add_argument('--months', type=_comma_ints, default=tuple(range(1, 13)), help='逗号分隔；默认 1 到 12 月')
    parser.add_argument('--times', type=_comma_floats, default=SOLAR_TIMES, help='逗号分隔的当地太阳时')
    parser.add_argument('--limit-mirrors', type=int, default=None, help='仅用于调试；只计算附件中的前 N 面镜子')
    parser.add_argument('--skip-shadow', action='store_true')
    parser.add_argument('--skip-truncation', action='store_true')
    parser.add_argument('--skip-figures', action='store_true')
    parser.add_argument('--run-validation', action='store_true', help='额外运行三组收敛实验并生成一张验证表')
    parser.add_argument('--quiet', action='store_true')
    return parser

def _progress(current: int, total: int, record: TimeResult) -> None:
    hour = int(record.solar_time)
    minute = int(round((record.solar_time - hour) * 60.0))
    print(f'[{current:02d}/{total:02d}] {record.month:02d}月21日 {hour:02d}:{minute:02d} 光学效率={record.average_optical_efficiency:.4f} 输出={record.field_output_mw:.3f} MW')

def run(argv: Sequence[str] | None=None) -> int:
    args = build_parser().parse_args(argv)
    mirror_xy = load_mirror_xy(args.input)
    if args.limit_mirrors is not None:
        if args.limit_mirrors < 1:
            raise SystemExit('--limit-mirrors 必须大于等于 1。')
        mirror_xy = mirror_xy[:args.limit_mirrors]
    field_config = FieldConfig()
    solver_config = SolverConfig(shadow_grid_size=args.shadow_grid, truncation_rays=args.truncation_rays, neighbor_radius_m=args.neighbor_radius, truncation_chunk_size=args.truncation_chunk_size, sobol_seed=args.sobol_seed, calculate_shadow=not args.skip_shadow, calculate_truncation=not args.skip_truncation)
    prepared = prepare_field(mirror_xy, field_config)
    solution = solve_question1(prepared=prepared, solver=solver_config, months=args.months, solar_times=args.times, progress=None if args.quiet else _progress)
    written = write_question1_results(output_dir=args.output, time_records=solution.time_results, monthly_records=solution.monthly_results, annual_record=solution.annual_result, mirror_annual_records=solution.mirror_annual_results, field_config=field_config, solver_config=solver_config, source_path=args.input, mirror_count=prepared.mirror_count)
    written.update(write_paper_tables(args.output, solution.monthly_results, solution.annual_result))
    if args.run_validation:
        validation = run_validation_suite(prepared, solver_config)
        written.update(write_validation_table(args.output, validation))
    if not args.skip_figures:
        written.update(build_paper_figures(output_dir=args.output))
    annual = solution.annual_result
    print('\n汇总结果')
    print(f'平均光学效率：{annual.average_optical_efficiency:.6f}')
    print(f'平均余弦效率：{annual.average_cosine_efficiency:.6f}')
    print(f'平均阴影遮挡效率：{annual.average_shadow_blocking_efficiency:.6f}')
    print(f'平均截断效率：{annual.average_truncation_efficiency:.6f}')
    print(f'平均输出热功率：{annual.field_output_mw:.6f} MW')
    print(f'单位镜面面积平均输出热功率：{annual.unit_area_output_kw_m2:.6f} kW/m²')
    print(f'结果目录：{args.output.resolve()}')
    for (name, path) in written.items():
        print(f'  {name}: {path.relative_to(args.output)}')
    return 0

def main() -> None:
    raise SystemExit(run())

# ========================================================================
# 来源：src/heliostat/q2/layout.py
# ========================================================================

"""第二问的两种参数化镜场布局与统一几何约束检查。"""
import math
from dataclasses import dataclass
from typing import Protocol
import numpy as np
from numpy.typing import NDArray
from scipy.spatial import cKDTree
FloatArray = NDArray[np.float64]

class LayoutError(ValueError):
    """布局参数或生成结果不可行。"""

class CommonParameters(Protocol):
    tower_x: float
    tower_y: float
    mirror_width: float
    mirror_height: float
    installation_height: float
    field_radius: float
    exclusion_radius: float
    safety_epsilon: float

@dataclass(frozen=True)
class PartitionedRingParameters:
    """分区交错同心圆布局参数。"""
    tower_y: float
    mirror_width: float
    mirror_height: float
    installation_height: float
    split_radius: float
    near_spacing: float
    far_spacing: float
    tower_x: float = 0.0
    field_radius: float = 350.0
    exclusion_radius: float = 100.0
    safety_epsilon: float = 0.01

    @property
    def safe_distance(self) -> float:
        return self.mirror_width + 5.0 + self.safety_epsilon

@dataclass(frozen=True)
class CampoParameters:
    """带渐增径向行距的 Campo 径向交错布局参数。"""
    tower_y: float
    mirror_width: float
    mirror_height: float
    installation_height: float
    first_ring_count: int
    initial_spacing: float
    spacing_growth: float
    tower_x: float = 0.0
    field_radius: float = 350.0
    exclusion_radius: float = 100.0
    safety_epsilon: float = 0.01

    @property
    def safe_distance(self) -> float:
        return self.mirror_width + 5.0 + self.safety_epsilon

@dataclass(frozen=True)
class LayoutRing:
    """一条生成并经过场地裁剪的圆环或圆弧。"""
    index: int
    radius: float
    zone: int
    nominal_count: int
    coordinates: FloatArray

    @property
    def mirror_count(self) -> int:
        return int(self.coordinates.shape[0])

@dataclass(frozen=True)
class GeneratedLayout:
    """由若干按半径排序的圆环或圆弧组成的镜场。"""
    kind: str
    rings: tuple[LayoutRing, ...]

    @property
    def mirror_count(self) -> int:
        return sum((ring.mirror_count for ring in self.rings))

    @property
    def coordinates(self) -> FloatArray:
        if not self.rings:
            return np.empty((0, 2), dtype=float)
        return np.concatenate([ring.coordinates for ring in self.rings], axis=0)

    def prefix(self, ring_count: int) -> FloatArray:
        if ring_count < 1 or ring_count > len(self.rings):
            raise ValueError(f'ring_count 应位于 1 到 {len(self.rings)}，实际为 {ring_count}。')
        return np.concatenate([ring.coordinates for ring in self.rings[:ring_count]], axis=0)

    def prefix_mirror_count(self, ring_count: int) -> int:
        if ring_count < 1 or ring_count > len(self.rings):
            raise ValueError(f'ring_count 应位于 1 到 {len(self.rings)}，实际为 {ring_count}。')
        return sum((ring.mirror_count for ring in self.rings[:ring_count]))

@dataclass(frozen=True)
class GeometryCheck:
    valid: bool
    reason: str | None
    mirror_count: int
    minimum_center_distance: float
    maximum_field_radius: float
    minimum_tower_distance: float

def _validate_common_parameters(parameters: CommonParameters) -> None:
    values = (parameters.tower_x, parameters.tower_y, parameters.mirror_width, parameters.mirror_height, parameters.installation_height, parameters.field_radius, parameters.exclusion_radius, parameters.safety_epsilon)
    if not all((math.isfinite(value) for value in values)):
        raise LayoutError('布局参数必须全部为有限数。')
    if not 2.0 <= parameters.mirror_height <= parameters.mirror_width <= 8.0:
        raise LayoutError('镜面尺寸必须满足 2 ≤ h ≤ w ≤ 8。')
    if not 2.0 <= parameters.installation_height <= 6.0:
        raise LayoutError('安装高度必须位于 2 m 到 6 m。')
    if parameters.installation_height < parameters.mirror_height / 2.0:
        raise LayoutError('安装高度不足，镜面转动时可能触地。')
    if parameters.field_radius <= 0.0:
        raise LayoutError('场地半径必须大于 0。')
    if parameters.exclusion_radius <= 0.0:
        raise LayoutError('塔周禁区半径必须大于 0。')
    if parameters.safety_epsilon <= 0.0:
        raise LayoutError('安全距离余量必须大于 0。')

def _maximum_tower_centered_radius(parameters: CommonParameters) -> float:
    return parameters.field_radius + math.hypot(parameters.tower_x, parameters.tower_y)

def _ring_coordinates(parameters: CommonParameters, radius: float, mirror_count: int, phase: float) -> FloatArray:
    if mirror_count < 2:
        raise LayoutError('单圈镜子数必须大于等于 2。')
    angles = 2.0 * math.pi * np.arange(mirror_count, dtype=float) / mirror_count + phase
    coordinates = np.column_stack((parameters.tower_x + radius * np.sin(angles), parameters.tower_y + radius * np.cos(angles)))
    field_radius = np.hypot(coordinates[:, 0], coordinates[:, 1])
    keep = field_radius <= parameters.field_radius + 1e-09
    clipped = np.asarray(coordinates[keep], dtype=float)
    clipped.setflags(write=False)
    return clipped

def _within_ring_count(radius: float, safe_distance: float) -> int:
    ratio = safe_distance / (2.0 * radius)
    if ratio >= 1.0:
        raise LayoutError('圆环半径过小，无法放置满足安全距离的镜子。')
    return int(math.floor(math.pi / math.asin(ratio)))

def generate_partitioned_layout(parameters: PartitionedRingParameters, *, maximum_rings: int=256) -> GeneratedLayout:
    """生成分区交错同心圆，并拒绝跨环距离冲突。"""
    _validate_common_parameters(parameters)
    if not math.isfinite(parameters.split_radius):
        raise LayoutError('分区半径必须为有限数。')
    if parameters.split_radius <= parameters.exclusion_radius:
        raise LayoutError('分区半径必须位于塔周禁区之外。')
    if parameters.near_spacing <= 0.0:
        raise LayoutError('近区行距必须大于 0。')
    if parameters.far_spacing < parameters.near_spacing:
        raise LayoutError('远区行距必须大于等于近区行距。')
    rings: list[LayoutRing] = []
    radius = parameters.exclusion_radius
    maximum_radius = _maximum_tower_centered_radius(parameters)
    for ring_index in range(maximum_rings):
        if radius > maximum_radius + 1e-09:
            break
        count = _within_ring_count(radius, parameters.safe_distance)
        phase = 0.0 if ring_index % 2 == 0 else math.pi / count
        coordinates = _ring_coordinates(parameters, radius, count, phase)
        if coordinates.size:
            rings.append(LayoutRing(index=ring_index, radius=radius, zone=1 if radius < parameters.split_radius else 2, nominal_count=count, coordinates=coordinates))
        spacing = parameters.near_spacing if radius < parameters.split_radius else parameters.far_spacing
        radius += spacing
    else:
        raise LayoutError('达到 maximum_rings，圆环生成未正常终止。')
    layout = GeneratedLayout('partitioned', tuple(rings))
    check = validate_layout(layout.coordinates, parameters)
    if not check.valid:
        raise LayoutError(check.reason or '分区圆环布局不满足几何约束。')
    return layout

def _campo_zone(radius: float, first_radius: float) -> tuple[int, int]:
    if radius < 2.0 * first_radius:
        return (1, 1)
    if radius < 4.0 * first_radius:
        return (2, 2)
    return (3, 4)

def generate_campo_layout(parameters: CampoParameters, *, maximum_rings: int=256) -> GeneratedLayout:
    """生成三分区、渐增径向行距的 Campo 径向交错镜场。"""
    _validate_common_parameters(parameters)
    if parameters.first_ring_count < 2:
        raise LayoutError('Campo 首环镜子数必须大于等于 2。')
    if parameters.initial_spacing <= 0.0:
        raise LayoutError('Campo 初始行距必须大于 0。')
    if parameters.spacing_growth < 0.0:
        raise LayoutError('Campo 行距增长量不能小于 0。')
    first_radius = max(parameters.exclusion_radius, parameters.safe_distance / (2.0 * math.sin(math.pi / parameters.first_ring_count)))
    maximum_radius = _maximum_tower_centered_radius(parameters)
    rings: list[LayoutRing] = []
    zone_rows = {1: 0, 2: 0, 3: 0}
    radius = first_radius
    for ring_index in range(maximum_rings):
        if radius > maximum_radius + 1e-09:
            break
        (zone, multiplier) = _campo_zone(radius, first_radius)
        count = parameters.first_ring_count * multiplier
        zone_index = zone_rows[zone]
        phase = 0.0 if zone_index % 2 == 0 else math.pi / count
        coordinates = _ring_coordinates(parameters, radius, count, phase)
        if coordinates.size:
            rings.append(LayoutRing(index=ring_index, radius=radius, zone=zone, nominal_count=count, coordinates=coordinates))
        zone_rows[zone] += 1
        radius += parameters.initial_spacing + parameters.spacing_growth * ring_index
    else:
        raise LayoutError('达到 maximum_rings，Campo 圆环生成未正常终止。')
    layout = GeneratedLayout('campo', tuple(rings))
    check = validate_layout(layout.coordinates, parameters)
    if not check.valid:
        raise LayoutError(check.reason or 'Campo 布局不满足几何约束。')
    return layout

def validate_layout(coordinates: FloatArray, parameters: CommonParameters) -> GeometryCheck:
    """按题目口径检查尺寸、场地、禁区和严格中心距离约束。"""
    try:
        _validate_common_parameters(parameters)
    except LayoutError as exc:
        return GeometryCheck(False, str(exc), 0, math.inf, math.inf, math.inf)
    xy = np.asarray(coordinates, dtype=float)
    if xy.ndim != 2 or xy.shape[1] != 2:
        return GeometryCheck(False, f'镜位坐标应为 N×2，实际形状为 {xy.shape}。', 0, math.inf, math.inf, math.inf)
    if xy.shape[0] == 0:
        return GeometryCheck(False, '镜场中没有定日镜。', 0, math.inf, 0.0, 0.0)
    if not np.all(np.isfinite(xy)):
        return GeometryCheck(False, '镜位坐标包含 NaN 或无穷值。', int(xy.shape[0]), math.inf, math.inf, math.inf)
    field_radii = np.hypot(xy[:, 0], xy[:, 1])
    tower_distances = np.hypot(xy[:, 0] - parameters.tower_x, xy[:, 1] - parameters.tower_y)
    maximum_field_radius = float(np.max(field_radii))
    minimum_tower_distance = float(np.min(tower_distances))
    if maximum_field_radius > parameters.field_radius + 1e-09:
        return GeometryCheck(False, '存在越过 350 m 场地边界的镜位。', int(xy.shape[0]), math.inf, maximum_field_radius, minimum_tower_distance)
    if minimum_tower_distance < parameters.exclusion_radius - 1e-09:
        return GeometryCheck(False, '存在进入塔周 100 m 禁区的镜位。', int(xy.shape[0]), math.inf, maximum_field_radius, minimum_tower_distance)
    if xy.shape[0] == 1:
        minimum_distance = math.inf
    else:
        (distances, _) = cKDTree(xy).query(xy, k=2)
        minimum_distance = float(np.min(distances[:, 1]))
    if minimum_distance <= parameters.mirror_width + 5.0:
        return GeometryCheck(False, f'最小中心距离不满足严格约束：{minimum_distance:.9f} m ≤ {parameters.mirror_width + 5.0:.9f} m。', int(xy.shape[0]), minimum_distance, maximum_field_radius, minimum_tower_distance)
    return GeometryCheck(True, None, int(xy.shape[0]), minimum_distance, maximum_field_radius, minimum_tower_distance)

# ========================================================================
# 来源：src/heliostat/q2/evaluate.py
# ========================================================================

"""第二问候选镜场的光学评价、缓存和外边界扫描。"""
import hashlib
from dataclasses import dataclass, replace
from typing import Sequence
import numpy as np
LayoutParameters = PartitionedRingParameters | CampoParameters

@dataclass(frozen=True)
class EvaluationProfile:
    """同一物理模型下的一组数值离散精度。"""
    name: str
    solver: SolverConfig
    months: tuple[int, ...] = tuple(range(1, 13))
    solar_times: tuple[float, ...] = SOLAR_TIMES

@dataclass(frozen=True)
class FieldEvaluation:
    """一个确定镜场外边界下的完整评价结果。"""
    layout_kind: str
    ring_count: int
    mirror_count: int
    mirror_area_m2: float
    total_area_m2: float
    coordinates: np.ndarray
    solution: Question1Solution

    @property
    def annual_power_mw(self) -> float:
        return self.solution.annual_result.field_output_mw

    @property
    def unit_area_power_kw_m2(self) -> float:
        return self.solution.annual_result.unit_area_output_kw_m2

    def is_feasible(self, target_power_mw: float=42.0) -> bool:
        return self.annual_power_mw >= target_power_mw

@dataclass(frozen=True)
class ExtentScanResult:
    """一组布局参数在若干镜场外边界中的最好结果。"""
    best: FieldEvaluation
    evaluations: tuple[FieldEvaluation, ...]
    first_feasible_ring_count: int | None

class EvaluationCache:
    """按坐标、塔和数值精度缓存昂贵的光学评价。"""

    def __init__(self) -> None:
        self._values: dict[str, Question1Solution] = {}

    def get(self, key: str) -> Question1Solution | None:
        return self._values.get(key)

    def put(self, key: str, value: Question1Solution) -> None:
        self._values[key] = value

    def __len__(self) -> int:
        return len(self._values)

def exploration_profile() -> EvaluationProfile:
    return EvaluationProfile(name='exploration', solver=SolverConfig(shadow_grid_size=5, truncation_rays=64, neighbor_radius_m=60.0, truncation_chunk_size=128, sobol_seed=2023))

def refinement_profile() -> EvaluationProfile:
    return EvaluationProfile(name='refinement', solver=SolverConfig(shadow_grid_size=10, truncation_rays=128, neighbor_radius_m=60.0, truncation_chunk_size=128, sobol_seed=2023))

def final_profile() -> EvaluationProfile:
    return EvaluationProfile(name='final', solver=SolverConfig(shadow_grid_size=15, truncation_rays=256, neighbor_radius_m=60.0, truncation_chunk_size=128, sobol_seed=2023))

def _field_config(parameters: CommonParameters, base: FieldConfig | None=None) -> FieldConfig:
    config = base or FieldConfig()
    return replace(config, field_radius=parameters.field_radius, exclusion_radius=parameters.exclusion_radius, tower_x=parameters.tower_x, tower_y=parameters.tower_y, mirror_width=parameters.mirror_width, mirror_height=parameters.mirror_height, mirror_center_z=parameters.installation_height)

def _cache_key(coordinates: np.ndarray, config: FieldConfig, profile: EvaluationProfile) -> str:
    digest = hashlib.sha256()
    rounded = np.round(np.asarray(coordinates, dtype='<f8'), decimals=9)
    digest.update(rounded.tobytes(order='C'))
    digest.update(repr(config.to_dict()).encode('utf-8'))
    digest.update(repr(profile.solver.to_dict()).encode('utf-8'))
    digest.update(repr(profile.months).encode('ascii'))
    digest.update(repr(profile.solar_times).encode('ascii'))
    return digest.hexdigest()

def evaluate_coordinates(*, layout_kind: str, ring_count: int, coordinates: np.ndarray, parameters: LayoutParameters, profile: EvaluationProfile, cache: EvaluationCache | None=None, base_field_config: FieldConfig | None=None) -> FieldEvaluation:
    """直接复用问题一模型评价一套确定坐标。"""
    xy = np.asarray(coordinates, dtype=float)
    if xy.ndim != 2 or xy.shape[1] != 2 or xy.shape[0] == 0:
        raise ValueError('候选镜场坐标必须为非空 N×2 数组。')
    config = _field_config(parameters, base_field_config)
    key = _cache_key(xy, config, profile)
    solution = cache.get(key) if cache is not None else None
    if solution is None:
        prepared = prepare_field(xy, config)
        solution = solve_question1(prepared=prepared, solver=profile.solver, months=profile.months, solar_times=profile.solar_times)
        if cache is not None:
            cache.put(key, solution)
    mirror_area = parameters.mirror_width * parameters.mirror_height
    return FieldEvaluation(layout_kind=layout_kind, ring_count=ring_count, mirror_count=int(xy.shape[0]), mirror_area_m2=mirror_area, total_area_m2=float(xy.shape[0] * mirror_area), coordinates=xy, solution=solution)

def better_evaluation(left: FieldEvaluation, right: FieldEvaluation, *, target_power_mw: float=42.0) -> FieldEvaluation:
    """按可行性优先规则返回较优结果。"""
    left_feasible = left.is_feasible(target_power_mw)
    right_feasible = right.is_feasible(target_power_mw)
    if left_feasible != right_feasible:
        return left if left_feasible else right
    if left_feasible:
        if left.unit_area_power_kw_m2 != right.unit_area_power_kw_m2:
            return left if left.unit_area_power_kw_m2 > right.unit_area_power_kw_m2 else right
        return left if left.annual_power_mw <= right.annual_power_mw else right
    return left if left.annual_power_mw >= right.annual_power_mw else right

def _unique_ring_counts(values: Sequence[int], total: int) -> tuple[int, ...]:
    return tuple(sorted({value for value in values if 1 <= value <= total}))

def scan_layout_extents(layout: GeneratedLayout, parameters: LayoutParameters, profile: EvaluationProfile, *, target_power_mw: float=42.0, coarse_stride: int=4, window: int=2, cache: EvaluationCache | None=None, base_field_config: FieldConfig | None=None) -> ExtentScanResult:
    """先粗定位功率阈值，再评价阈值附近的连续圆环外边界。"""
    if not layout.rings:
        raise ValueError('布局中没有可用于评价的圆环。')
    if coarse_stride < 1:
        raise ValueError('coarse_stride 必须大于等于 1。')
    if window < 0:
        raise ValueError('window 不能小于 0。')
    total_rings = len(layout.rings)
    coarse_counts = list(range(coarse_stride, total_rings + 1, coarse_stride))
    if not coarse_counts or coarse_counts[-1] != total_rings:
        coarse_counts.append(total_rings)
    evaluated: dict[int, FieldEvaluation] = {}

    def evaluate(ring_count: int) -> FieldEvaluation:
        previous = evaluated.get(ring_count)
        if previous is not None:
            return previous
        value = evaluate_coordinates(layout_kind=layout.kind, ring_count=ring_count, coordinates=layout.prefix(ring_count), parameters=parameters, profile=profile, cache=cache, base_field_config=base_field_config)
        evaluated[ring_count] = value
        return value
    first_feasible: int | None = None
    previous_coarse = 0
    for ring_count in coarse_counts:
        value = evaluate(ring_count)
        if value.is_feasible(target_power_mw):
            for refined_count in range(previous_coarse + 1, ring_count + 1):
                refined = evaluate(refined_count)
                if refined.is_feasible(target_power_mw):
                    first_feasible = refined_count
                    break
            break
        previous_coarse = ring_count
    center = first_feasible if first_feasible is not None else total_rings
    local_counts = _unique_ring_counts(range(center - window, center + window + 1), total_rings)
    for ring_count in local_counts:
        evaluate(ring_count)
    values = tuple((evaluated[key] for key in sorted(evaluated)))
    best = values[0]
    for value in values[1:]:
        best = better_evaluation(best, value, target_power_mw=target_power_mw)
    return ExtentScanResult(best=best, evaluations=values, first_feasible_ring_count=first_feasible)

# ========================================================================
# 来源：src/heliostat/q3_continuous/model.py
# ========================================================================

"""独立第三问 Campo 结构特征、连续规格函数和异构几何约束。"""
import csv
import json
import math
from dataclasses import dataclass
from pathlib import Path
import numpy as np
from numpy.typing import NDArray
from scipy.spatial import cKDTree
FloatArray = NDArray[np.float64]
IntArray = NDArray[np.int64]
SUPPORTED_ZONES = (1, 2)
EXPECTED_RING_COUNT = 28
EXPECTED_FULL_MIRROR_COUNT = 1471
EXPECTED_Q2_MIRROR_COUNT = 1469

@dataclass(frozen=True)
class CampoMotherField:
    """由问题二 Campo 参数重建并附带逐镜结构标签的镜场。"""
    parameters: CampoParameters
    layout: GeneratedLayout
    coordinates: FloatArray
    ring_indices: IntArray
    zone_indices: IntArray
    zone_row_indices: IntArray
    normalized_rows: FloatArray
    ring_radii: FloatArray
    azimuth_angles: FloatArray
    azimuth_features: FloatArray
    nominal_ring_counts: IntArray
    actual_ring_counts: IntArray
    original_indices: IntArray

    @property
    def mirror_count(self) -> int:
        return int(self.coordinates.shape[0])

    @property
    def zone_counts(self) -> tuple[int, ...]:
        return tuple((int(np.count_nonzero(self.zone_indices == zone)) for zone in SUPPORTED_ZONES))

    @property
    def zone_ring_counts(self) -> tuple[int, ...]:
        return tuple((int(np.unique(self.ring_indices[self.zone_indices == zone]).size) for zone in SUPPORTED_ZONES))

    @property
    def base_width(self) -> float:
        return self.parameters.mirror_width

    @property
    def base_height(self) -> float:
        return self.parameters.mirror_height

    @property
    def base_installation_height(self) -> float:
        return self.parameters.installation_height

    @property
    def base_total_area_m2(self) -> float:
        return self.mirror_count * self.parameters.mirror_width * self.parameters.mirror_height

@dataclass(frozen=True)
class ContinuousDesign:
    """Campo 区域—行号连续尺寸与安装高度参数。"""
    size_zone1_slope: float = 0.0
    size_zone2_slope: float = 0.0
    size_zone2_offset: float = 0.0
    height_offset: float = 0.0
    height_zone1_slope: float = 0.0
    height_zone2_slope: float = 0.0
    height_zone2_offset: float = 0.0
    size_azimuth: float = 0.0
    height_azimuth: float = 0.0
    area_ratio: float = 1.0

    def __post_init__(self) -> None:
        values = tuple(self.__dict__.values())
        if not all((math.isfinite(value) for value in values)):
            raise ValueError('连续规格参数必须全部为有限数。')
        if self.area_ratio <= 0.0:
            raise ValueError('总面积比例必须大于 0。')

    @classmethod
    def uniform(cls) -> ContinuousDesign:
        return cls()

    @property
    def uses_azimuth(self) -> bool:
        return abs(self.size_azimuth) > 1e-15 or abs(self.height_azimuth) > 1e-15

@dataclass(frozen=True)
class ExpandedSpecifications:
    widths: FloatArray
    heights: FloatArray
    installation_heights: FloatArray
    areas: FloatArray
    scales: FloatArray
    size_shape: FloatArray
    area_normalizer: float

    @property
    def total_area_m2(self) -> float:
        return float(np.sum(self.areas))

@dataclass(frozen=True)
class HeterogeneousGeometryCheck:
    valid: bool
    reason: str | None
    mirror_count: int
    minimum_center_distance_m: float
    minimum_width_clearance_m: float
    maximum_field_radius_m: float
    minimum_tower_distance_m: float
    minimum_ground_clearance_m: float

def load_q2_campo_parameters(summary_path: str | Path) -> CampoParameters:
    path = Path(summary_path)
    if not path.exists():
        raise FileNotFoundError(f'找不到问题二摘要：{path}')
    payload = json.loads(path.read_text(encoding='utf-8'))
    if payload.get('layout') != 'campo':
        raise ValueError('问题二摘要的最终布局不是 campo。')
    parameters = payload.get('parameters')
    if not isinstance(parameters, dict):
        raise ValueError('问题二摘要缺少 parameters。')
    return CampoParameters(**parameters)

def _read_selected_coordinates(path: str | Path) -> FloatArray:
    source = Path(path)
    if not source.exists():
        raise FileNotFoundError(f'找不到问题二最终镜位：{source}')
    rows: list[tuple[float, float]] = []
    with source.open('r', encoding='utf-8-sig', newline='') as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None or not {'x_m', 'y_m'} <= set(reader.fieldnames):
            raise ValueError('问题二镜位文件必须包含 x_m 和 y_m 列。')
        for row in reader:
            rows.append((float(row['x_m']), float(row['y_m'])))
    if not rows:
        raise ValueError('问题二镜位文件为空。')
    return np.asarray(rows, dtype=float)

def _selected_full_indices(full_coordinates: FloatArray, selected_coordinates: FloatArray, *, tolerance: float=1e-07) -> IntArray:
    (distances, indices) = cKDTree(full_coordinates).query(selected_coordinates, k=1)
    if np.any(distances > tolerance):
        maximum = float(np.max(distances))
        raise ValueError(f'问题二最终镜位无法映射回 Campo 生成结构，最大坐标差为 {maximum:.3e} m。')
    if np.unique(indices).size != indices.size:
        raise ValueError('问题二最终镜位映射出现重复 Campo 镜位。')
    return np.asarray(indices, dtype=np.int64)

def _recenter_ring_features(ring_indices: IntArray, azimuth_angles: FloatArray) -> tuple[FloatArray, IntArray]:
    features = np.empty_like(azimuth_angles)
    actual_counts = np.empty(ring_indices.shape[0], dtype=np.int64)
    cosines = np.cos(azimuth_angles)
    for ring in np.unique(ring_indices):
        active = ring_indices == ring
        count = int(np.count_nonzero(active))
        features[active] = cosines[active] - float(np.mean(cosines[active]))
        actual_counts[active] = count
    return (features, actual_counts)

def build_campo_mother_field(summary_path: str | Path, *, selected_coordinates_path: str | Path | None=None, require_recorded_structure: bool=True) -> CampoMotherField:
    """重建完整 Campo 结构，并可筛成问题二正式保留的 1469 面镜场。"""
    parameters = load_q2_campo_parameters(summary_path)
    layout = generate_campo_layout(parameters)
    if len(layout.rings) != EXPECTED_RING_COUNT:
        raise ValueError(f'期望 {EXPECTED_RING_COUNT} 个有效环，实际为 {len(layout.rings)}。')
    zone_rings: dict[int, list[int]] = {}
    for (display_index, ring) in enumerate(layout.rings, start=1):
        zone_rings.setdefault(ring.zone, []).append(display_index)
    unsupported = set(zone_rings) - set(SUPPORTED_ZONES)
    if unsupported:
        raise ValueError(f'当前镜场出现未建模 Campo 区域：{unsupported}。')
    coordinates: list[FloatArray] = []
    ring_indices: list[IntArray] = []
    zone_indices: list[IntArray] = []
    zone_row_indices: list[IntArray] = []
    normalized_rows: list[FloatArray] = []
    ring_radii: list[FloatArray] = []
    nominal_counts: list[IntArray] = []
    for (display_index, ring) in enumerate(layout.rings, start=1):
        rows = zone_rings[ring.zone]
        row_index = rows.index(display_index)
        denominator = max(len(rows) - 1, 1)
        count = ring.mirror_count
        coordinates.append(ring.coordinates)
        ring_indices.append(np.full(count, display_index, dtype=np.int64))
        zone_indices.append(np.full(count, ring.zone, dtype=np.int64))
        zone_row_indices.append(np.full(count, row_index, dtype=np.int64))
        normalized_rows.append(np.full(count, row_index / denominator, dtype=float))
        ring_radii.append(np.full(count, ring.radius, dtype=float))
        nominal_counts.append(np.full(count, ring.nominal_count, dtype=np.int64))
    full_coordinates = np.concatenate(coordinates, axis=0)
    arrays: list[NDArray] = [full_coordinates, np.concatenate(ring_indices), np.concatenate(zone_indices), np.concatenate(zone_row_indices), np.concatenate(normalized_rows), np.concatenate(ring_radii), np.concatenate(nominal_counts), np.arange(full_coordinates.shape[0], dtype=np.int64)]
    if selected_coordinates_path is not None:
        selected = _read_selected_coordinates(selected_coordinates_path)
        selected_indices = _selected_full_indices(full_coordinates, selected)
        arrays = [values[selected_indices] for values in arrays]
    (selected_coordinates, selected_rings, selected_zones, selected_zone_rows, selected_normalized_rows, selected_radii, selected_nominal_counts, selected_originals) = arrays
    dx = selected_coordinates[:, 0] - parameters.tower_x
    dy = selected_coordinates[:, 1] - parameters.tower_y
    azimuth_angles = np.arctan2(dx, dy)
    (azimuth_features, actual_counts) = _recenter_ring_features(np.asarray(selected_rings, dtype=np.int64), np.asarray(azimuth_angles, dtype=float))
    mother = CampoMotherField(parameters=parameters, layout=layout, coordinates=np.asarray(selected_coordinates, dtype=float), ring_indices=np.asarray(selected_rings, dtype=np.int64), zone_indices=np.asarray(selected_zones, dtype=np.int64), zone_row_indices=np.asarray(selected_zone_rows, dtype=np.int64), normalized_rows=np.asarray(selected_normalized_rows, dtype=float), ring_radii=np.asarray(selected_radii, dtype=float), azimuth_angles=np.asarray(azimuth_angles, dtype=float), azimuth_features=np.asarray(azimuth_features, dtype=float), nominal_ring_counts=np.asarray(selected_nominal_counts, dtype=np.int64), actual_ring_counts=np.asarray(actual_counts, dtype=np.int64), original_indices=np.asarray(selected_originals, dtype=np.int64))
    if require_recorded_structure:
        expected = EXPECTED_Q2_MIRROR_COUNT if selected_coordinates_path is not None else EXPECTED_FULL_MIRROR_COUNT
        if mother.mirror_count != expected:
            raise ValueError(f'问题二 Campo 结构已变化：期望 {expected} 面，实际为 {mother.mirror_count} 面。')
        if mother.zone_ring_counts != (11, 17):
            raise ValueError(f'问题二 Campo 区域结构已变化：期望区域环数 (11, 17)，实际为 {mother.zone_ring_counts}。')
    return mother

def size_shape(mother: CampoMotherField, design: ContinuousDesign) -> FloatArray:
    zone1 = mother.zone_indices == 1
    values = np.empty(mother.mirror_count, dtype=float)
    values[zone1] = design.size_zone1_slope * mother.normalized_rows[zone1] + design.size_azimuth * mother.azimuth_features[zone1]
    zone2 = ~zone1
    values[zone2] = design.size_zone2_offset + design.size_zone2_slope * mother.normalized_rows[zone2] + design.size_azimuth * mother.azimuth_features[zone2]
    return values

def installation_height_values(mother: CampoMotherField, design: ContinuousDesign) -> FloatArray:
    zone1 = mother.zone_indices == 1
    values = np.empty(mother.mirror_count, dtype=float)
    common = mother.base_installation_height + design.height_offset
    values[zone1] = common + design.height_zone1_slope * mother.normalized_rows[zone1] + design.height_azimuth * mother.azimuth_features[zone1]
    zone2 = ~zone1
    values[zone2] = common + design.height_zone2_offset + design.height_zone2_slope * mother.normalized_rows[zone2] + design.height_azimuth * mother.azimuth_features[zone2]
    return values

def expand_continuous_design(mother: CampoMotherField, design: ContinuousDesign) -> ExpandedSpecifications:
    shape = size_shape(mother, design)
    denominator = float(np.sum(np.exp(2.0 * shape)))
    target_area = mother.base_total_area_m2 * design.area_ratio
    normalizer = math.sqrt(target_area / (mother.base_width * mother.base_height * denominator))
    scales = normalizer * np.exp(shape)
    widths = mother.base_width * scales
    heights = mother.base_height * scales
    installation_heights = installation_height_values(mother, design)
    return ExpandedSpecifications(widths=widths, heights=heights, installation_heights=installation_heights, areas=widths * heights, scales=scales, size_shape=shape, area_normalizer=normalizer)

def individual_width_caps(coordinates: FloatArray, *, safety_epsilon: float=0.01) -> FloatArray:
    xy = np.asarray(coordinates, dtype=float)
    if xy.ndim != 2 or xy.shape[1] != 2 or xy.shape[0] < 2:
        raise ValueError('至少需要两个 N×2 镜位计算宽度上限。')
    distances = cKDTree(xy).query(xy, k=2)[0][:, 1]
    return np.minimum(8.0, distances - 5.0 - safety_epsilon)

def validate_heterogeneous_field(*, coordinates: FloatArray, widths: FloatArray, heights: FloatArray, installation_heights: FloatArray, tower_x: float, tower_y: float, field_radius: float=350.0, exclusion_radius: float=100.0, safety_epsilon: float=0.01) -> HeterogeneousGeometryCheck:
    xy = np.asarray(coordinates, dtype=float)
    mirror_widths = np.asarray(widths, dtype=float)
    mirror_heights = np.asarray(heights, dtype=float)
    center_zs = np.asarray(installation_heights, dtype=float)
    mirror_count = int(xy.shape[0]) if xy.ndim >= 1 else 0
    invalid = HeterogeneousGeometryCheck(valid=False, reason=None, mirror_count=mirror_count, minimum_center_distance_m=math.inf, minimum_width_clearance_m=-math.inf, maximum_field_radius_m=math.inf, minimum_tower_distance_m=-math.inf, minimum_ground_clearance_m=-math.inf)
    if xy.ndim != 2 or xy.shape[1] != 2 or mirror_count == 0:
        return HeterogeneousGeometryCheck(**{**invalid.__dict__, 'reason': '镜位必须为非空 N×2 数组。'})
    for (name, values) in (('宽度', mirror_widths), ('高度', mirror_heights), ('安装高度', center_zs)):
        if values.ndim != 1 or values.shape[0] != mirror_count:
            return HeterogeneousGeometryCheck(**{**invalid.__dict__, 'reason': f'{name}数组长度与镜子数不一致。'})
    if not all((np.all(np.isfinite(values)) for values in (xy, mirror_widths, mirror_heights, center_zs))):
        return HeterogeneousGeometryCheck(**{**invalid.__dict__, 'reason': '几何数据包含 NaN 或无穷值。'})
    if np.any(mirror_heights < 2.0) or np.any(mirror_heights > 8.0):
        return HeterogeneousGeometryCheck(**{**invalid.__dict__, 'reason': '镜面高度必须位于 2 m 到 8 m。'})
    if np.any(mirror_widths < 2.0) or np.any(mirror_widths > 8.0):
        return HeterogeneousGeometryCheck(**{**invalid.__dict__, 'reason': '镜面宽度必须位于 2 m 到 8 m。'})
    if np.any(mirror_widths < mirror_heights):
        return HeterogeneousGeometryCheck(**{**invalid.__dict__, 'reason': '镜面宽度不能小于镜面高度。'})
    if np.any(center_zs < 2.0) or np.any(center_zs > 6.0):
        return HeterogeneousGeometryCheck(**{**invalid.__dict__, 'reason': '安装高度必须位于 2 m 到 6 m。'})
    ground_clearance = center_zs - mirror_heights / 2.0
    if np.any(ground_clearance < -1e-12):
        return HeterogeneousGeometryCheck(**{**invalid.__dict__, 'reason': '存在旋转时可能触地的镜面。'})
    field_radii = np.hypot(xy[:, 0], xy[:, 1])
    tower_distances = np.hypot(xy[:, 0] - tower_x, xy[:, 1] - tower_y)
    maximum_field_radius = float(np.max(field_radii))
    minimum_tower_distance = float(np.min(tower_distances))
    if maximum_field_radius > field_radius + 1e-09:
        return HeterogeneousGeometryCheck(**{**invalid.__dict__, 'reason': '存在镜位超出圆形场地边界。', 'maximum_field_radius_m': maximum_field_radius, 'minimum_tower_distance_m': minimum_tower_distance, 'minimum_ground_clearance_m': float(np.min(ground_clearance))})
    if minimum_tower_distance < exclusion_radius - 1e-09:
        return HeterogeneousGeometryCheck(**{**invalid.__dict__, 'reason': '存在镜位进入塔周禁区。', 'maximum_field_radius_m': maximum_field_radius, 'minimum_tower_distance_m': minimum_tower_distance, 'minimum_ground_clearance_m': float(np.min(ground_clearance))})
    tree = cKDTree(xy)
    nearest = tree.query(xy, k=2)[0][:, 1]
    minimum_center_distance = float(np.min(nearest))
    pairs = tree.query_pairs(r=13.0 + safety_epsilon + 1e-09, output_type='ndarray')
    minimum_width_clearance = math.inf
    if pairs.size:
        deltas = xy[pairs[:, 0]] - xy[pairs[:, 1]]
        distances = np.linalg.norm(deltas, axis=1)
        required = np.maximum(mirror_widths[pairs[:, 0]], mirror_widths[pairs[:, 1]]) + 5.0
        clearances = distances - required
        minimum_width_clearance = float(np.min(clearances))
        if minimum_width_clearance < safety_epsilon - 1e-09:
            return HeterogeneousGeometryCheck(valid=False, reason='存在镜对不满足异构宽度对应的中心距安全余量。', mirror_count=mirror_count, minimum_center_distance_m=minimum_center_distance, minimum_width_clearance_m=minimum_width_clearance, maximum_field_radius_m=maximum_field_radius, minimum_tower_distance_m=minimum_tower_distance, minimum_ground_clearance_m=float(np.min(ground_clearance)))
    return HeterogeneousGeometryCheck(valid=True, reason=None, mirror_count=mirror_count, minimum_center_distance_m=minimum_center_distance, minimum_width_clearance_m=minimum_width_clearance, maximum_field_radius_m=maximum_field_radius, minimum_tower_distance_m=minimum_tower_distance, minimum_ground_clearance_m=float(np.min(ground_clearance)))

# ========================================================================
# 来源：src/heliostat/q3_continuous/evaluate.py
# ========================================================================

"""独立第三问连续异构镜场评价、缓存和多级精度配置。"""
import hashlib
from dataclasses import dataclass, replace
import numpy as np
from numpy.typing import NDArray
FloatArray = NDArray[np.float64]
IntArray = NDArray[np.int64]

@dataclass(frozen=True)
class HeterogeneousEvaluation:
    """一套逐镜规格与 Campo 结构标签的完整评价结果。"""
    profile_name: str
    coordinates: FloatArray
    widths: FloatArray
    heights: FloatArray
    installation_heights: FloatArray
    ring_indices: IntArray
    zone_indices: IntArray
    zone_row_indices: IntArray
    normalized_rows: FloatArray
    azimuth_angles: FloatArray
    azimuth_features: FloatArray
    nominal_ring_counts: IntArray
    actual_ring_counts: IntArray
    original_indices: IntArray
    solution: Question1Solution
    geometry: HeterogeneousGeometryCheck

    @property
    def mirror_count(self) -> int:
        return int(self.coordinates.shape[0])

    @property
    def total_area_m2(self) -> float:
        return float(np.sum(self.widths * self.heights))

    @property
    def annual_power_mw(self) -> float:
        return self.solution.annual_result.field_output_mw

    @property
    def unit_area_power_kw_m2(self) -> float:
        return self.solution.annual_result.unit_area_output_kw_m2

    def is_feasible(self, target_power_mw: float=42.0) -> bool:
        return self.annual_power_mw >= target_power_mw

class EvaluationCache:
    """按镜位、逐镜规格、塔位和数值精度缓存评价。"""

    def __init__(self) -> None:
        self._values: dict[str, Question1Solution] = {}

    def get(self, key: str) -> Question1Solution | None:
        return self._values.get(key)

    def put(self, key: str, value: Question1Solution) -> None:
        self._values[key] = value

    def __len__(self) -> int:
        return len(self._values)

def coarse_profile() -> EvaluationProfile:
    return EvaluationProfile(name='q3-coarse', solver=SolverConfig(shadow_grid_size=5, truncation_rays=64, neighbor_radius_m=60.0, truncation_chunk_size=128, sobol_seed=2023), months=(3, 6, 9, 12), solar_times=SOLAR_TIMES)

def medium_profile() -> EvaluationProfile:
    return EvaluationProfile(name='q3-medium', solver=SolverConfig(shadow_grid_size=10, truncation_rays=128, neighbor_radius_m=60.0, truncation_chunk_size=128, sobol_seed=2023))

def formal_profile() -> EvaluationProfile:
    return EvaluationProfile(name='q3-formal', solver=SolverConfig(shadow_grid_size=15, truncation_rays=256, neighbor_radius_m=60.0, truncation_chunk_size=128, sobol_seed=2023))

def dense_profile() -> EvaluationProfile:
    return EvaluationProfile(name='q3-dense', solver=SolverConfig(shadow_grid_size=20, truncation_rays=512, neighbor_radius_m=80.0, truncation_chunk_size=128, sobol_seed=2023))

def smoke_profile() -> EvaluationProfile:
    return EvaluationProfile(name='q3-smoke', solver=SolverConfig(shadow_grid_size=2, truncation_rays=4, neighbor_radius_m=60.0, truncation_chunk_size=64, sobol_seed=2023), months=(6,), solar_times=(12.0,))

def field_config_from_mother(mother: CampoMotherField) -> FieldConfig:
    return replace(FieldConfig(), field_radius=mother.parameters.field_radius, exclusion_radius=mother.parameters.exclusion_radius, tower_x=mother.parameters.tower_x, tower_y=mother.parameters.tower_y, mirror_width=mother.base_width, mirror_height=mother.base_height, mirror_center_z=mother.base_installation_height)

def _cache_key(*, coordinates: FloatArray, specifications: ExpandedSpecifications, field_config: FieldConfig, profile: EvaluationProfile) -> str:
    digest = hashlib.sha256()
    for values in (coordinates, specifications.widths, specifications.heights, specifications.installation_heights):
        rounded = np.round(np.asarray(values, dtype='<f8'), decimals=9)
        digest.update(rounded.tobytes(order='C'))
    digest.update(repr(field_config.to_dict()).encode('utf-8'))
    digest.update(repr(profile.solver.to_dict()).encode('utf-8'))
    digest.update(repr(profile.months).encode('ascii'))
    digest.update(repr(profile.solar_times).encode('ascii'))
    return digest.hexdigest()

def evaluate_specifications(*, coordinates: FloatArray, specifications: ExpandedSpecifications, ring_indices: IntArray, zone_indices: IntArray, zone_row_indices: IntArray, normalized_rows: FloatArray, azimuth_angles: FloatArray, azimuth_features: FloatArray, nominal_ring_counts: IntArray, actual_ring_counts: IntArray, original_indices: IntArray, field_config: FieldConfig, profile: EvaluationProfile, safety_epsilon: float=0.01, cache: EvaluationCache | None=None) -> HeterogeneousEvaluation:
    xy = np.asarray(coordinates, dtype=float)
    count = int(xy.shape[0])
    structural = {'ring_indices': np.asarray(ring_indices, dtype=np.int64), 'zone_indices': np.asarray(zone_indices, dtype=np.int64), 'zone_row_indices': np.asarray(zone_row_indices, dtype=np.int64), 'normalized_rows': np.asarray(normalized_rows, dtype=float), 'azimuth_angles': np.asarray(azimuth_angles, dtype=float), 'azimuth_features': np.asarray(azimuth_features, dtype=float), 'nominal_ring_counts': np.asarray(nominal_ring_counts, dtype=np.int64), 'actual_ring_counts': np.asarray(actual_ring_counts, dtype=np.int64), 'original_indices': np.asarray(original_indices, dtype=np.int64)}
    for (name, values) in structural.items():
        if values.ndim != 1 or values.shape[0] != count:
            raise ValueError(f'{name} 长度与镜子数不一致。')
    geometry = validate_heterogeneous_field(coordinates=xy, widths=specifications.widths, heights=specifications.heights, installation_heights=specifications.installation_heights, tower_x=field_config.tower_x, tower_y=field_config.tower_y, field_radius=field_config.field_radius, exclusion_radius=field_config.exclusion_radius, safety_epsilon=safety_epsilon)
    if not geometry.valid:
        raise ValueError(geometry.reason or '异构镜场几何约束不合法。')
    key = _cache_key(coordinates=xy, specifications=specifications, field_config=field_config, profile=profile)
    solution = cache.get(key) if cache is not None else None
    if solution is None:
        prepared = prepare_field(xy, field_config, mirror_widths=specifications.widths, mirror_heights=specifications.heights, mirror_center_zs=specifications.installation_heights)
        solution = solve_question1(prepared=prepared, solver=profile.solver, months=profile.months, solar_times=profile.solar_times)
        if cache is not None:
            cache.put(key, solution)
    return HeterogeneousEvaluation(profile_name=profile.name, coordinates=xy, widths=specifications.widths, heights=specifications.heights, installation_heights=specifications.installation_heights, solution=solution, geometry=geometry, **structural)

def evaluate_design(*, mother: CampoMotherField, design: ContinuousDesign, profile: EvaluationProfile, cache: EvaluationCache | None=None) -> HeterogeneousEvaluation:
    specifications = expand_continuous_design(mother, design)
    return evaluate_specifications(coordinates=mother.coordinates, specifications=specifications, ring_indices=mother.ring_indices, zone_indices=mother.zone_indices, zone_row_indices=mother.zone_row_indices, normalized_rows=mother.normalized_rows, azimuth_angles=mother.azimuth_angles, azimuth_features=mother.azimuth_features, nominal_ring_counts=mother.nominal_ring_counts, actual_ring_counts=mother.actual_ring_counts, original_indices=mother.original_indices, field_config=field_config_from_mother(mother), profile=profile, safety_epsilon=mother.parameters.safety_epsilon, cache=cache)

# ========================================================================
# 来源：src/heliostat/q3_continuous/search.py
# ========================================================================

"""独立第三问 Campo 连续规格诊断与分阶段坐标搜索。"""
from dataclasses import dataclass, replace
from typing import Callable, Iterable, Sequence
import numpy as np
ProgressCallback = Callable[[str], None]

@dataclass(frozen=True)
class CampoDiagnostics:
    radial_rmse_kw_m2: float
    radial_azimuth_rmse_kw_m2: float
    relative_rmse_reduction: float
    azimuth_coefficient_kw_m2: float
    azimuth_recommended: bool

@dataclass(frozen=True)
class SearchStep:
    stage: str
    action: str
    design: ContinuousDesign
    evaluation: HeterogeneousEvaluation
    estimated_power_mw: float

@dataclass(frozen=True)
class SearchOutcome:
    baseline_design: ContinuousDesign
    baseline_evaluation: HeterogeneousEvaluation
    best_design: ContinuousDesign
    best_evaluation: HeterogeneousEvaluation
    diagnostics: CampoDiagnostics
    trace: tuple[SearchStep, ...]
    stage_evaluations: tuple[tuple[str, HeterogeneousEvaluation], ...]

def _design_matrix(mother: CampoMotherField, *, include_azimuth: bool) -> np.ndarray:
    zone1 = (mother.zone_indices == 1).astype(float)
    zone2 = (mother.zone_indices == 2).astype(float)
    columns = [zone1, zone2, zone1 * mother.normalized_rows, zone2 * mother.normalized_rows]
    if include_azimuth:
        columns.append(mother.azimuth_features)
    return np.column_stack(columns)

def diagnose_campo_structure(mother: CampoMotherField, evaluation: HeterogeneousEvaluation, *, recommendation_threshold: float=0.02) -> CampoDiagnostics:
    """比较区域—行号模型与增加同环方位项后的单镜拟合误差。"""
    if evaluation.mirror_count != mother.mirror_count:
        raise ValueError('诊断评价与 Campo 镜场镜子数不一致。')
    mirror_power = np.asarray([record.average_output_power_kw for record in evaluation.solution.mirror_annual_results], dtype=float)
    mirror_area = evaluation.widths * evaluation.heights
    unit_output = mirror_power / mirror_area
    radial = _design_matrix(mother, include_azimuth=False)
    angular = _design_matrix(mother, include_azimuth=True)
    radial_coefficients = np.linalg.lstsq(radial, unit_output, rcond=None)[0]
    angular_coefficients = np.linalg.lstsq(angular, unit_output, rcond=None)[0]
    radial_residual = unit_output - radial @ radial_coefficients
    angular_residual = unit_output - angular @ angular_coefficients
    radial_rmse = float(np.sqrt(np.mean(radial_residual ** 2)))
    angular_rmse = float(np.sqrt(np.mean(angular_residual ** 2)))
    reduction = 0.0 if radial_rmse <= 0.0 else (radial_rmse - angular_rmse) / radial_rmse
    return CampoDiagnostics(radial_rmse_kw_m2=radial_rmse, radial_azimuth_rmse_kw_m2=angular_rmse, relative_rmse_reduction=float(reduction), azimuth_coefficient_kw_m2=float(angular_coefficients[-1]), azimuth_recommended=bool(reduction >= recommendation_threshold))

def _parameter_candidate(design: ContinuousDesign, name: str, delta: float) -> ContinuousDesign:
    return replace(design, **{name: getattr(design, name) + delta})

def _within_search_bounds(design: ContinuousDesign, *, monotone: bool) -> bool:
    if monotone and (design.size_zone1_slope > 1e-12 or design.size_zone2_slope > 1e-12 or design.height_zone1_slope < -1e-12 or (design.height_zone2_slope < -1e-12)):
        return False
    if any((abs(value) > 0.3 for value in (design.size_zone1_slope, design.size_zone2_slope, design.size_zone2_offset))):
        return False
    if abs(design.size_azimuth) > 0.15:
        return False
    if any((abs(value) > 2.0 for value in (design.height_offset, design.height_zone1_slope, design.height_zone2_slope, design.height_zone2_offset))):
        return False
    if abs(design.height_azimuth) > 1.0:
        return False
    return 0.85 <= design.area_ratio <= 1.05

class _SearchContext:

    def __init__(self, *, mother: CampoMotherField, coarse_profile: EvaluationProfile, reference_profile: EvaluationProfile, current_design: ContinuousDesign, current_coarse: HeterogeneousEvaluation, current_reference: HeterogeneousEvaluation, cache: EvaluationCache, target_power_mw: float, q_improvement_threshold: float, progress: ProgressCallback | None, monotone: bool) -> None:
        self.mother = mother
        self.coarse_profile = coarse_profile
        self.reference_profile = reference_profile
        self.current_design = current_design
        self.current_coarse = current_coarse
        self.current_reference = current_reference
        self.cache = cache
        self.target_power_mw = target_power_mw
        self.q_improvement_threshold = q_improvement_threshold
        self.progress = progress
        self.monotone = monotone
        self.trace: list[SearchStep] = []

    def _evaluate(self, design: ContinuousDesign, profile: EvaluationProfile) -> HeterogeneousEvaluation | None:
        if not _within_search_bounds(design, monotone=self.monotone):
            return None
        try:
            return evaluate_design(mother=self.mother, design=design, profile=profile, cache=self.cache)
        except ValueError:
            return None

    def _estimated_power(self, coarse: HeterogeneousEvaluation) -> float:
        return self.current_reference.annual_power_mw + (coarse.annual_power_mw - self.current_coarse.annual_power_mw)

    def _accepts(self, candidate: HeterogeneousEvaluation) -> bool:
        current_feasible = self.current_reference.is_feasible(self.target_power_mw)
        if not current_feasible:
            return candidate.annual_power_mw > self.current_reference.annual_power_mw + 1e-06
        return candidate.is_feasible(self.target_power_mw) and candidate.unit_area_power_kw_m2 > self.current_reference.unit_area_power_kw_m2 + self.q_improvement_threshold

    def try_candidates(self, *, stage: str, candidates: Iterable[tuple[str, ContinuousDesign]]) -> bool:
        ranked: list[tuple[float, str, ContinuousDesign, HeterogeneousEvaluation, float]] = []
        current_feasible = self.current_reference.is_feasible(self.target_power_mw)
        for (action, design) in candidates:
            coarse = self._evaluate(design, self.coarse_profile)
            if coarse is None:
                continue
            estimated_power = self._estimated_power(coarse)
            if estimated_power < self.target_power_mw - 0.35:
                continue
            estimated_q = 1000.0 * estimated_power / coarse.total_area_m2
            score = estimated_q if current_feasible else estimated_power
            ranked.append((score, action, design, coarse, estimated_power))
        ranked.sort(key=lambda item: item[0], reverse=True)
        for (_, action, design, coarse, estimated_power) in ranked:
            reference = self._evaluate(design, self.reference_profile)
            if reference is None or not self._accepts(reference):
                continue
            self.current_design = design
            self.current_coarse = coarse
            self.current_reference = reference
            self.trace.append(SearchStep(stage=stage, action=action, design=design, evaluation=reference, estimated_power_mw=estimated_power))
            if self.progress is not None:
                self.progress(f'{stage} 接受 {action}：P={reference.annual_power_mw:.6f} MW，q={reference.unit_area_power_kw_m2:.6f} kW/m²')
            return True
        return False

def _coordinate_candidates(design: ContinuousDesign, parameter: str, step: float) -> tuple[tuple[str, ContinuousDesign], ...]:
    return tuple(((f"{parameter}{('+' if direction > 0 else '-')}{step:g}", _parameter_candidate(design, parameter, direction * step)) for direction in (-1.0, 1.0)))

def _scan_level(context: _SearchContext, *, stage: str, parameters: Sequence[str], step: float, maximum_cycles: int) -> None:
    for cycle in range(maximum_cycles):
        improved = False
        order: Iterable[str] = parameters if cycle % 2 == 0 else reversed(parameters)
        for parameter in order:
            improved |= context.try_candidates(stage=stage, candidates=_coordinate_candidates(context.current_design, parameter, step))
        if not improved:
            break

def _build_context(*, mother: CampoMotherField, design: ContinuousDesign, coarse_profile: EvaluationProfile, reference_profile: EvaluationProfile, cache: EvaluationCache, target_power_mw: float, q_improvement_threshold: float, progress: ProgressCallback | None, monotone: bool) -> _SearchContext:
    coarse = evaluate_design(mother=mother, design=design, profile=coarse_profile, cache=cache)
    reference = evaluate_design(mother=mother, design=design, profile=reference_profile, cache=cache)
    return _SearchContext(mother=mother, coarse_profile=coarse_profile, reference_profile=reference_profile, current_design=design, current_coarse=coarse, current_reference=reference, cache=cache, target_power_mw=target_power_mw, q_improvement_threshold=q_improvement_threshold, progress=progress, monotone=monotone)

def optimize_continuous_design(*, mother: CampoMotherField, coarse_profile: EvaluationProfile, reference_profile: EvaluationProfile, target_power_mw: float=42.0, include_azimuth: bool=False, monotone: bool=True, maximum_cycles_per_level: int=2, q_improvement_threshold: float=1e-05, height_steps: tuple[float, ...]=(0.4, 0.2, 0.1), size_steps: tuple[float, ...]=(0.04, 0.02, 0.01), area_steps: tuple[float, ...]=(0.005, 0.002, 0.001), cache: EvaluationCache | None=None, progress: ProgressCallback | None=None) -> SearchOutcome:
    if maximum_cycles_per_level < 0:
        raise ValueError('maximum_cycles_per_level 不能小于 0。')
    working_cache = cache or EvaluationCache()
    baseline_design = ContinuousDesign.uniform()
    context = _build_context(mother=mother, design=baseline_design, coarse_profile=coarse_profile, reference_profile=reference_profile, cache=working_cache, target_power_mw=target_power_mw, q_improvement_threshold=q_improvement_threshold, progress=progress, monotone=monotone)
    baseline_evaluation = context.current_reference
    diagnostics = diagnose_campo_structure(mother, baseline_evaluation)
    stages: list[tuple[str, HeterogeneousEvaluation]] = [('q2-uniform', baseline_evaluation)]
    height_parameters = ['height_offset', 'height_zone1_slope', 'height_zone2_slope', 'height_zone2_offset']
    if include_azimuth:
        height_parameters.append('height_azimuth')
    for (level, step) in enumerate(height_steps, start=1):
        _scan_level(context, stage=f'height-L{level}', parameters=height_parameters, step=step, maximum_cycles=maximum_cycles_per_level)
    stages.append(('height-only', context.current_reference))
    size_parameters = ['size_zone1_slope', 'size_zone2_slope', 'size_zone2_offset']
    if include_azimuth:
        size_parameters.append('size_azimuth')
    for (level, step) in enumerate(size_steps, start=1):
        _scan_level(context, stage=f'fixed-area-size-L{level}', parameters=size_parameters, step=step, maximum_cycles=maximum_cycles_per_level)
        _scan_level(context, stage=f'fixed-area-height-rescan-L{level}', parameters=height_parameters, step=min(step * 5.0, height_steps[-1]), maximum_cycles=min(1, maximum_cycles_per_level))
    stages.append(('fixed-area-reallocation', context.current_reference))
    for (level, step) in enumerate(area_steps, start=1):
        for _ in range(maximum_cycles_per_level):
            improved = context.try_candidates(stage=f'area-compression-L{level}', candidates=((f'area_ratio-{step:g}', replace(context.current_design, area_ratio=context.current_design.area_ratio - step)),))
            if not improved:
                break
        _scan_level(context, stage=f'area-size-rescan-L{level}', parameters=size_parameters, step=size_steps[min(level - 1, len(size_steps) - 1)], maximum_cycles=min(1, maximum_cycles_per_level))
        _scan_level(context, stage=f'area-height-rescan-L{level}', parameters=height_parameters, step=height_steps[-1], maximum_cycles=min(1, maximum_cycles_per_level))
    stages.append(('area-compression', context.current_reference))
    return SearchOutcome(baseline_design=baseline_design, baseline_evaluation=baseline_evaluation, best_design=context.current_design, best_evaluation=context.current_reference, diagnostics=diagnostics, trace=tuple(context.trace), stage_evaluations=tuple(stages))

def refine_design_parameters(*, mother: CampoMotherField, initial_design: ContinuousDesign, coarse_profile: EvaluationProfile, reference_profile: EvaluationProfile, parameters: Sequence[str], steps: Sequence[float], stage: str, target_power_mw: float=42.0, monotone: bool=True, maximum_cycles_per_level: int=2, q_improvement_threshold: float=1e-05, cache: EvaluationCache | None=None, progress: ProgressCallback | None=None) -> tuple[ContinuousDesign, HeterogeneousEvaluation, tuple[SearchStep, ...]]:
    working_cache = cache or EvaluationCache()
    context = _build_context(mother=mother, design=initial_design, coarse_profile=coarse_profile, reference_profile=reference_profile, cache=working_cache, target_power_mw=target_power_mw, q_improvement_threshold=q_improvement_threshold, progress=progress, monotone=monotone)
    for (level, step) in enumerate(steps, start=1):
        _scan_level(context, stage=f'{stage}-L{level}', parameters=parameters, step=step, maximum_cycles=maximum_cycles_per_level)
    return (context.current_design, context.current_reference, tuple(context.trace))

# ========================================================================
# 来源：src/heliostat/q3_continuous/prune.py
# ========================================================================

"""独立第三问 Campo 外边界低贡献东西对称镜位复算。"""
from dataclasses import dataclass
import numpy as np

@dataclass(frozen=True)
class PruneStep:
    removed_original_indices: tuple[int, ...]
    evaluation: HeterogeneousEvaluation

@dataclass(frozen=True)
class PruneOutcome:
    initial: HeterogeneousEvaluation
    best: HeterogeneousEvaluation
    steps: tuple[PruneStep, ...]

def symmetric_pairs(evaluation: HeterogeneousEvaluation, *, tolerance: float=1e-07) -> tuple[tuple[int, int], ...]:
    coordinates = evaluation.coordinates
    unused = set(range(evaluation.mirror_count))
    pairs: list[tuple[int, int]] = []
    while unused:
        index = min(unused)
        unused.remove(index)
        (x_m, y_m) = coordinates[index]
        if abs(float(x_m)) <= tolerance:
            continue
        partners = [other for other in unused if abs(float(coordinates[other, 0] + x_m)) <= tolerance and abs(float(coordinates[other, 1] - y_m)) <= tolerance]
        if not partners:
            continue
        partner = min(partners)
        unused.remove(partner)
        pairs.append((index, partner))
    return tuple(pairs)

def _rank_pairs(evaluation: HeterogeneousEvaluation, pairs: tuple[tuple[int, int], ...]) -> list[tuple[int, int]]:
    mirror_power = np.array([record.average_output_power_kw for record in evaluation.solution.mirror_annual_results], dtype=float)

    def key(pair: tuple[int, int]) -> tuple[int, float, int]:
        zones = evaluation.zone_indices[list(pair)]
        rings = evaluation.ring_indices[list(pair)]
        preferred = bool(np.any(zones == 2) and np.any(rings >= 24))
        return (0 if preferred else 1, float(np.sum(mirror_power[list(pair)])), -int(np.max(rings)))
    return sorted(pairs, key=key)

def _remove_pair(*, current: HeterogeneousEvaluation, pair: tuple[int, int], mother: CampoMotherField, profile: EvaluationProfile, cache: EvaluationCache | None) -> HeterogeneousEvaluation:
    active = np.ones(current.mirror_count, dtype=bool)
    active[list(pair)] = False
    specifications = ExpandedSpecifications(widths=current.widths[active], heights=current.heights[active], installation_heights=current.installation_heights[active], areas=current.widths[active] * current.heights[active], scales=current.widths[active] / mother.base_width, size_shape=np.log(current.widths[active] / mother.base_width), area_normalizer=1.0)
    active_rings = current.ring_indices[active]
    active_angles = current.azimuth_angles[active]
    active_features = np.empty(active_angles.shape[0], dtype=float)
    active_counts = np.empty(active_angles.shape[0], dtype=np.int64)
    cosines = np.cos(active_angles)
    for ring in np.unique(active_rings):
        in_ring = active_rings == ring
        count = int(np.count_nonzero(in_ring))
        active_features[in_ring] = cosines[in_ring] - float(np.mean(cosines[in_ring]))
        active_counts[in_ring] = count
    return evaluate_specifications(coordinates=current.coordinates[active], specifications=specifications, ring_indices=active_rings, zone_indices=current.zone_indices[active], zone_row_indices=current.zone_row_indices[active], normalized_rows=current.normalized_rows[active], azimuth_angles=active_angles, azimuth_features=active_features, nominal_ring_counts=current.nominal_ring_counts[active], actual_ring_counts=active_counts, original_indices=current.original_indices[active], field_config=field_config_from_mother(mother), profile=profile, safety_epsilon=mother.parameters.safety_epsilon, cache=cache)

def prune_symmetric_pairs(*, mother: CampoMotherField, initial: HeterogeneousEvaluation, profile: EvaluationProfile, target_power_mw: float=42.0, maximum_rounds: int=4, maximum_pairs_per_round: int=12, q_improvement_threshold: float=1e-05, cache: EvaluationCache | None=None) -> PruneOutcome:
    if maximum_rounds < 0:
        raise ValueError('maximum_rounds 不能小于 0。')
    if maximum_pairs_per_round < 1:
        raise ValueError('maximum_pairs_per_round 必须大于等于 1。')
    if not initial.is_feasible(target_power_mw):
        raise ValueError('结构化删镜要求初始方案满足功率约束。')
    current = initial
    steps: list[PruneStep] = []
    for _ in range(maximum_rounds):
        pairs = _rank_pairs(current, symmetric_pairs(current))
        pairs = pairs[:maximum_pairs_per_round]
        best_pair: tuple[int, int] | None = None
        best_candidate: HeterogeneousEvaluation | None = None
        for pair in pairs:
            candidate = _remove_pair(current=current, pair=pair, mother=mother, profile=profile, cache=cache)
            if not candidate.is_feasible(target_power_mw):
                continue
            if candidate.unit_area_power_kw_m2 <= current.unit_area_power_kw_m2 + q_improvement_threshold:
                continue
            if best_candidate is None or candidate.unit_area_power_kw_m2 > best_candidate.unit_area_power_kw_m2:
                best_pair = pair
                best_candidate = candidate
        if best_pair is None or best_candidate is None:
            break
        removed = tuple((int(current.original_indices[index]) for index in best_pair))
        current = best_candidate
        steps.append(PruneStep(removed, current))
    return PruneOutcome(initial=initial, best=current, steps=tuple(steps))

# ========================================================================
# 来源：src/heliostat/q3_continuous/export.py
# ========================================================================

"""独立第三问连续 Campo 参数、逐镜结果和提交表输出。"""
import csv
import json
from copy import copy
from dataclasses import asdict
from pathlib import Path
from typing import Any, Iterable, Sequence
import numpy as np
from openpyxl import load_workbook
TARGET_ANNUAL_POWER_MW = 42.0

def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        raise ValueError(f'没有可写入 {path.name} 的结果。')
    with path.open('w', encoding='utf-8-sig', newline='') as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)

def write_result3_workbook(*, template_path: str | Path, output_path: str | Path, evaluation: HeterogeneousEvaluation, tower_x: float, tower_y: float) -> Path:
    """按题目模板写出塔坐标和每面镜子的异构规格、位置。"""
    template = Path(template_path)
    if not template.exists():
        raise FileNotFoundError(f'找不到 result3.xlsx 模板：{template}')
    destination = Path(output_path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(template)
    sheet = workbook.active
    if sheet.max_column < 8:
        workbook.close()
        raise ValueError('result3.xlsx 模板列数不足 8 列。')
    style_row = 2 if sheet.max_row >= 2 else 1
    styles = [copy(sheet.cell(style_row, column)._style) for column in range(1, 9)]
    number_formats = [sheet.cell(style_row, column).number_format for column in range(1, 9)]
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)
    for index in range(evaluation.mirror_count):
        row_index = index + 2
        values = (tower_x, tower_y, index + 1, float(evaluation.widths[index]), float(evaluation.heights[index]), float(evaluation.coordinates[index, 0]), float(evaluation.coordinates[index, 1]), float(evaluation.installation_heights[index]))
        for (column, value) in enumerate(values, start=1):
            cell = sheet.cell(row_index, column, value)
            cell._style = copy(styles[column - 1])
            cell.number_format = number_formats[column - 1]
    workbook.save(destination)
    workbook.close()
    return destination

def _zone_rows(evaluation: HeterogeneousEvaluation) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for zone in SUPPORTED_ZONES:
        active = evaluation.zone_indices == zone
        rows.append({'zone': zone, 'ring_count': int(np.unique(evaluation.ring_indices[active]).size), 'mirror_count': int(np.count_nonzero(active)), 'minimum_width_m': float(np.min(evaluation.widths[active])), 'maximum_width_m': float(np.max(evaluation.widths[active])), 'minimum_height_m': float(np.min(evaluation.heights[active])), 'maximum_height_m': float(np.max(evaluation.heights[active])), 'minimum_installation_height_m': float(np.min(evaluation.installation_heights[active])), 'maximum_installation_height_m': float(np.max(evaluation.installation_heights[active])), 'total_area_m2': float(np.sum(evaluation.widths[active] * evaluation.heights[active]))})
    return rows

def _stage_rows(stages: Iterable[tuple[str, HeterogeneousEvaluation]]) -> list[dict[str, Any]]:
    return [{'stage': name, 'profile': evaluation.profile_name, 'mirror_count': evaluation.mirror_count, 'total_area_m2': evaluation.total_area_m2, 'annual_power_mw': evaluation.annual_power_mw, 'unit_area_power_kw_m2': evaluation.unit_area_power_kw_m2} for (name, evaluation) in stages]

def _parameter_rows(design: ContinuousDesign) -> tuple[tuple[str, str, float], ...]:
    return (('size_zone1_slope', '区域 1 尺寸行号趋势', design.size_zone1_slope), ('size_zone2_slope', '区域 2 尺寸行号趋势', design.size_zone2_slope), ('size_zone2_offset', '进入区域 2 的尺寸修正', design.size_zone2_offset), ('size_azimuth', '同环南北尺寸修正', design.size_azimuth), ('height_offset', '全场安装高度平移', design.height_offset), ('height_zone1_slope', '区域 1 高度行号趋势', design.height_zone1_slope), ('height_zone2_slope', '区域 2 高度行号趋势', design.height_zone2_slope), ('height_zone2_offset', '进入区域 2 的高度修正', design.height_zone2_offset), ('height_azimuth', '同环南北高度修正', design.height_azimuth), ('area_ratio', '相对起始镜场总面积比例', design.area_ratio))

def write_question3_results(*, output_dir: str | Path, mother: CampoMotherField, design: ContinuousDesign, evaluation: HeterogeneousEvaluation, result3_template: str | Path, stages: Iterable[tuple[str, HeterogeneousEvaluation]]=(), diagnostics: CampoDiagnostics | None=None, model_name: str='campo-monotone-radial', legacy_comparison: dict[str, Any] | None=None) -> dict[str, Path]:
    destination = Path(output_dir)
    destination.mkdir(parents=True, exist_ok=True)
    coordinate_rows = [{'mirror_id': index + 1, 'original_campo_mirror_id': int(evaluation.original_indices[index]) + 1, 'ring_index': int(evaluation.ring_indices[index]), 'campo_zone': int(evaluation.zone_indices[index]), 'zone_row_index': int(evaluation.zone_row_indices[index]), 'normalized_zone_row': float(evaluation.normalized_rows[index]), 'azimuth_rad': float(evaluation.azimuth_angles[index]), 'centered_azimuth_feature': float(evaluation.azimuth_features[index]), 'nominal_ring_count': int(evaluation.nominal_ring_counts[index]), 'actual_ring_count': int(evaluation.actual_ring_counts[index]), 'mirror_width_m': float(evaluation.widths[index]), 'mirror_height_m': float(evaluation.heights[index]), 'x_m': float(evaluation.coordinates[index, 0]), 'y_m': float(evaluation.coordinates[index, 1]), 'z_m': float(evaluation.installation_heights[index])} for index in range(evaluation.mirror_count)]
    monthly_rows = [asdict(record) for record in evaluation.solution.monthly_results]
    mirror_rows = [{**asdict(record), 'original_campo_mirror_id': int(evaluation.original_indices[index]) + 1, 'ring_index': int(evaluation.ring_indices[index]), 'campo_zone': int(evaluation.zone_indices[index]), 'zone_row_index': int(evaluation.zone_row_indices[index]), 'normalized_zone_row': float(evaluation.normalized_rows[index]), 'centered_azimuth_feature': float(evaluation.azimuth_features[index]), 'mirror_width_m': float(evaluation.widths[index]), 'mirror_height_m': float(evaluation.heights[index]), 'installation_height_m': float(evaluation.installation_heights[index]), 'mirror_area_m2': float(evaluation.widths[index] * evaluation.heights[index])} for (index, record) in enumerate(evaluation.solution.mirror_annual_results)]
    annual = asdict(evaluation.solution.annual_result)
    zones = _zone_rows(evaluation)
    stage_data = _stage_rows(stages)
    stages_path = destination / '02_分阶段方案比较.json'
    coordinates_path = destination / '03_最终逐镜参数与坐标.csv'
    monthly_path = destination / '04_月平均计算结果.csv'
    annual_path = destination / '05_年平均计算结果.json'
    mirror_path = destination / '06_单镜年平均结果.csv'
    summary_path = destination / '07_最终方案摘要.json'
    table_path = destination / '08_论文结果与验证表.md'
    workbook_path = destination / '10_第三问提交结果.xlsx'
    stages_path.write_text(json.dumps(stage_data, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    _write_csv(coordinates_path, coordinate_rows)
    _write_csv(monthly_path, monthly_rows)
    _write_csv(mirror_path, mirror_rows)
    annual_path.write_text(json.dumps(annual, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    summary = {'layout': 'q2-campo-continuous-heterogeneous', 'model': model_name, 'annual_power_constraint_mw': TARGET_ANNUAL_POWER_MW, 'annual_power_margin_mw': evaluation.annual_power_mw - TARGET_ANNUAL_POWER_MW, 'constraint_satisfied': evaluation.is_feasible(TARGET_ANNUAL_POWER_MW), 'tower': {'x_m': mother.parameters.tower_x, 'y_m': mother.parameters.tower_y}, 'mirror_count': evaluation.mirror_count, 'ring_count': int(np.unique(evaluation.ring_indices).size), 'total_area_m2': evaluation.total_area_m2, 'continuous_design': asdict(design), 'zone_summaries': zones, 'diagnostics': asdict(diagnostics) if diagnostics is not None else None, 'geometry': asdict(evaluation.geometry), 'annual': annual}
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    lines = ['# 第三问结果与验证表', '', '## 表 1 功率约束与优化目标', '', '| 年平均功率下限 (MW) | 年平均功率 (MW) | 功率余量 (MW) | 总镜面面积 (m²) | 单位面积年平均输出 (kW/m²) | 是否满足约束 |', '| ---: | ---: | ---: | ---: | ---: | :---: |', f"| {TARGET_ANNUAL_POWER_MW:.6f} | {evaluation.annual_power_mw:.6f} | {evaluation.annual_power_mw - TARGET_ANNUAL_POWER_MW:.6f} | {evaluation.total_area_m2:.3f} | {evaluation.unit_area_power_kw_m2:.6f} | {('是' if evaluation.is_feasible() else '否')} |", '', '## 表 2 连续规格参数', '', '| 参数 | 含义 | 结果 |', '| --- | --- | ---: |']
    for (name, meaning, value) in _parameter_rows(design):
        lines.append(f'| `{name}` | {meaning} | {value:.9f} |')
    lines.extend(['', '## 表 3 Campo 区域规格范围', '', '| 区域 | 圆环数 | 镜子数 | 宽度范围 (m) | 高度范围 (m) | 安装高度范围 (m) | 区域总面积 (m²) |', '| ---: | ---: | ---: | ---: | ---: | ---: | ---: |'])
    for row in zones:
        lines.append(f"| {row['zone']} | {row['ring_count']} | {row['mirror_count']} | {row['minimum_width_m']:.6f}–{row['maximum_width_m']:.6f} | {row['minimum_height_m']:.6f}–{row['maximum_height_m']:.6f} | {row['minimum_installation_height_m']:.6f}–{row['maximum_installation_height_m']:.6f} | {row['total_area_m2']:.3f} |")
    lines.extend(['', '## 表 4 分阶段与模型消融', '', '| 阶段 | 评价精度 | 镜子数 | 总面积 (m²) | 年平均功率 (MW) | 单位面积输出 (kW/m²) |', '| --- | --- | ---: | ---: | ---: | ---: |'])
    for row in stage_data:
        lines.append(f"| {row['stage']} | {row['profile']} | {row['mirror_count']} | {row['total_area_m2']:.3f} | {row['annual_power_mw']:.6f} | {row['unit_area_power_kw_m2']:.6f} |")
    geometry = evaluation.geometry
    lines.extend(['', '## 表 5 异构几何约束复核', '', '| 检查项 | 实际值 | 约束 | 结果 |', '| --- | ---: | ---: | :---: |', f"| 最小镜心距离 (m) | {geometry.minimum_center_distance_m:.9f} | - | {('通过' if geometry.valid else '未通过')} |", f"| 最小异构宽度安全余量 (m) | {geometry.minimum_width_clearance_m:.9f} | ≥ 0.010000000 | {('通过' if geometry.minimum_width_clearance_m >= 0.01 - 1e-09 else '未通过')} |", f"| 最大场地半径 (m) | {geometry.maximum_field_radius_m:.6f} | ≤ 350 | {('通过' if geometry.maximum_field_radius_m <= 350.0 + 1e-09 else '未通过')} |", f"| 最小塔距 (m) | {geometry.minimum_tower_distance_m:.6f} | ≥ 100 | {('通过' if geometry.minimum_tower_distance_m >= 100.0 - 1e-09 else '未通过')} |", f"| 最小不触地余量 (m) | {geometry.minimum_ground_clearance_m:.6f} | ≥ 0 | {('通过' if geometry.minimum_ground_clearance_m >= -1e-09 else '未通过')} |", '', '## 表 6 每月 21 日平均光学效率及输出功率', '', '| 月份 | 光学效率 | 余弦效率 | 阴影遮挡效率 | 截断效率 | 输出热功率 (MW) | 单位面积输出 (kW/m²) |', '| ---: | ---: | ---: | ---: | ---: | ---: | ---: |'])
    for record in evaluation.solution.monthly_results:
        lines.append(f'| {record.month} | {record.average_optical_efficiency:.6f} | {record.average_cosine_efficiency:.6f} | {record.average_shadow_blocking_efficiency:.6f} | {record.average_truncation_efficiency:.6f} | {record.field_output_mw:.6f} | {record.unit_area_output_kw_m2:.6f} |')
    table_path.write_text('\n'.join(lines) + '\n', encoding='utf-8')
    write_result3_workbook(template_path=result3_template, output_path=workbook_path, evaluation=evaluation, tower_x=mother.parameters.tower_x, tower_y=mother.parameters.tower_y)
    written = {'stages': stages_path, 'coordinates': coordinates_path, 'monthly': monthly_path, 'annual': annual_path, 'mirror_annual': mirror_path, 'summary': summary_path, 'paper_table': table_path, 'result3': workbook_path}
    if legacy_comparison is not None:
        legacy_path = destination / '11_原六组对照结果.json'
        legacy_path.write_text(json.dumps(legacy_comparison, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
        written['legacy_six_group_comparison'] = legacy_path
    return written

def write_dense_validation(*, output_dir: str | Path, evaluation: HeterogeneousEvaluation, profile: EvaluationProfile, sensitivity_evaluations: Sequence[tuple[EvaluationProfile, HeterogeneousEvaluation]]=()) -> Path:
    destination = Path(output_dir)
    path = destination / '09_高精度加密验证.json'
    evaluations = ((profile, evaluation), *sensitivity_evaluations)

    def validation_record(item_profile: EvaluationProfile, item_evaluation: HeterogeneousEvaluation) -> dict[str, Any]:
        return {'profile': {'months': len(item_profile.months), 'solar_times_per_month': len(item_profile.solar_times), 'shadow_grid_size': item_profile.solver.shadow_grid_size, 'truncation_rays': item_profile.solver.truncation_rays, 'neighbor_radius_m': item_profile.solver.neighbor_radius_m}, 'annual_power_mw': item_evaluation.annual_power_mw, 'annual_power_margin_mw': item_evaluation.annual_power_mw - TARGET_ANNUAL_POWER_MW, 'unit_area_power_kw_m2': item_evaluation.unit_area_power_kw_m2, 'constraint_satisfied': item_evaluation.is_feasible()}
    payload = {'profile': {'months': len(profile.months), 'solar_times_per_month': len(profile.solar_times), 'shadow_grid_size': profile.solver.shadow_grid_size, 'truncation_rays': profile.solver.truncation_rays, 'neighbor_radius_m': profile.solver.neighbor_radius_m}, 'mirror_count': evaluation.mirror_count, 'total_area_m2': evaluation.total_area_m2, 'annual_power_constraint_mw': TARGET_ANNUAL_POWER_MW, 'annual_power_mw': evaluation.annual_power_mw, 'annual_power_margin_mw': evaluation.annual_power_mw - TARGET_ANNUAL_POWER_MW, 'unit_area_power_kw_m2': evaluation.unit_area_power_kw_m2, 'constraint_satisfied': evaluation.is_feasible(), 'neighbor_radius_sensitivity': [validation_record(item_profile, item_evaluation) for (item_profile, item_evaluation) in evaluations]}
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    table_path = destination / '08_论文结果与验证表.md'
    lines = ['', '## 表 7 高精度加密与邻域敏感性验证', '', '| 阴影网格 | 截断光线 | 邻镜半径 (m) | 年平均功率 (MW) | 功率余量 (MW) | 单位面积输出 (kW/m²) | 是否满足约束 |', '| ---: | ---: | ---: | ---: | ---: | ---: | :---: |']
    for (item_profile, item_evaluation) in evaluations:
        lines.append(f"| {item_profile.solver.shadow_grid_size}×{item_profile.solver.shadow_grid_size} | {item_profile.solver.truncation_rays} | {item_profile.solver.neighbor_radius_m:.0f} | {item_evaluation.annual_power_mw:.6f} | {item_evaluation.annual_power_mw - TARGET_ANNUAL_POWER_MW:.6f} | {item_evaluation.unit_area_power_kw_m2:.6f} | {('是' if item_evaluation.is_feasible() else '否')} |")
    table_content = table_path.read_text(encoding='utf-8')
    marker = '\n## 表 7 '
    if marker in table_content:
        table_content = table_content.split(marker, maxsplit=1)[0].rstrip()
    table_path.write_text(table_content + '\n' + '\n'.join(lines) + '\n', encoding='utf-8')
    return path

# ========================================================================
# 来源：src/heliostat/q3_continuous/solve.py
# ========================================================================

"""独立第三问命令行：Campo 连续规格搜索、消融、复算和输出。"""
import argparse
import json
from dataclasses import replace
from pathlib import Path
from typing import Sequence
import numpy as np
PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_Q2_SUMMARY = PROJECT_ROOT / 'outputs' / 'q2' / '07_最终方案摘要.json'
DEFAULT_Q2_COORDINATES = PROJECT_ROOT / 'outputs' / 'q2' / '03_最终镜位坐标.csv'
DEFAULT_TEMPLATE = PROJECT_ROOT / 'task' / 'A' / 'result3.xlsx'
DEFAULT_OUTPUT = PROJECT_ROOT / 'outputs' / 'q3_continuous'
LEGACY_Q3_OUTPUT = PROJECT_ROOT / 'outputs' / 'q3'

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description='求解 CUMCM 2023 A 题第三问的 Campo 连续异构镜场')
    parser.add_argument('--q2-summary', type=Path, default=DEFAULT_Q2_SUMMARY)
    parser.add_argument('--q2-coordinates', type=Path, default=DEFAULT_Q2_COORDINATES)
    parser.add_argument('--full-campo-prefix', action='store_true', help='使用修剪前 1471 面 Campo 前缀，而不是问题二正式 1469 面镜场。')
    parser.add_argument('--result3-template', type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument('--output', type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument('--smoke', action='store_true')
    parser.add_argument('--target-power', type=float, default=42.0)
    parser.add_argument('--max-cycles', type=int, default=2)
    parser.add_argument('--azimuth-mode', choices=('auto', 'on', 'off'), default='auto')
    parser.add_argument('--skip-relaxed', action='store_true')
    parser.add_argument('--prune-rounds', type=int, default=1)
    parser.add_argument('--prune-pairs-per-round', type=int, default=8)
    parser.add_argument('--run-validation', action='store_true')
    return parser

def _validate_args(args: argparse.Namespace) -> None:
    if args.target_power <= 0.0:
        raise SystemExit('--target-power 必须大于 0。')
    if args.max_cycles < 0:
        raise SystemExit('--max-cycles 不能小于 0。')
    if args.prune_rounds < 0:
        raise SystemExit('--prune-rounds 不能小于 0。')
    if args.prune_pairs_per_round < 1:
        raise SystemExit('--prune-pairs-per-round 必须大于等于 1。')

def _specifications_from_evaluation(source: HeterogeneousEvaluation, mother: CampoMotherField) -> ExpandedSpecifications:
    scales = source.widths / mother.base_width
    return ExpandedSpecifications(widths=source.widths, heights=source.heights, installation_heights=source.installation_heights, areas=source.widths * source.heights, scales=scales, size_shape=np.log(scales), area_normalizer=1.0)

def _reevaluate(*, source: HeterogeneousEvaluation, profile: EvaluationProfile, mother: CampoMotherField, cache: EvaluationCache) -> HeterogeneousEvaluation:
    return evaluate_specifications(coordinates=source.coordinates, specifications=_specifications_from_evaluation(source, mother), ring_indices=source.ring_indices, zone_indices=source.zone_indices, zone_row_indices=source.zone_row_indices, normalized_rows=source.normalized_rows, azimuth_angles=source.azimuth_angles, azimuth_features=source.azimuth_features, nominal_ring_counts=source.nominal_ring_counts, actual_ring_counts=source.actual_ring_counts, original_indices=source.original_indices, field_config=field_config_from_mother(mother), profile=profile, safety_epsilon=mother.parameters.safety_epsilon, cache=cache)

def _load_legacy_comparison(output: Path) -> dict | None:
    path = output / '07_最终方案摘要.json'
    if not path.exists():
        return None
    try:
        payload = json.loads(path.read_text(encoding='utf-8'))
    except (json.JSONDecodeError, OSError):
        return None
    if payload.get('layout') != 'fixed-q2-campo-heterogeneous':
        legacy_path = output / '11_原六组对照结果.json'
        if legacy_path.exists():
            try:
                return json.loads(legacy_path.read_text(encoding='utf-8'))
            except (json.JSONDecodeError, OSError):
                return None
        return None
    return {'model': 'legacy-six-ring-groups', 'role': 'comparison-only', 'mirror_count': payload.get('mirror_count'), 'total_area_m2': payload.get('total_area_m2'), 'group_design': payload.get('group_design'), 'annual': payload.get('annual'), 'geometry': payload.get('geometry')}

def _formal_candidates(*, candidates: Sequence[tuple[str, ContinuousDesign, HeterogeneousEvaluation]], pruned: HeterogeneousEvaluation, pruned_design: ContinuousDesign, mother: CampoMotherField, profile: EvaluationProfile, target_power_mw: float, cache: EvaluationCache) -> tuple[str, ContinuousDesign, HeterogeneousEvaluation, tuple[tuple[str, HeterogeneousEvaluation], ...]]:
    formal: list[tuple[str, ContinuousDesign, HeterogeneousEvaluation]] = []
    for (name, design, evaluation) in candidates:
        formal.append((name, design, _reevaluate(source=evaluation, profile=profile, mother=mother, cache=cache)))
    if not any((np.array_equal(pruned.original_indices, evaluation.original_indices) for (_, _, evaluation) in candidates)):
        formal.append(('outer-boundary-pruned', pruned_design, _reevaluate(source=pruned, profile=profile, mother=mother, cache=cache)))
    feasible = [item for item in formal if item[2].is_feasible(target_power_mw)]
    if not feasible:
        powers = ', '.join((f'{name}={evaluation.annual_power_mw:.6f}' for (name, _, evaluation) in formal))
        raise RuntimeError(f'正式精度下没有满足功率约束的候选：{powers} MW。')
    best = max(feasible, key=lambda item: item[2].unit_area_power_kw_m2)
    return (best[0], best[1], best[2], tuple(((f'formal-{name}', evaluation) for (name, _, evaluation) in formal)))

def run(argv: Sequence[str] | None=None) -> int:
    args = build_parser().parse_args(argv)
    _validate_args(args)
    legacy_comparison = _load_legacy_comparison(LEGACY_Q3_OUTPUT)
    selected_coordinates = None if args.full_campo_prefix else args.q2_coordinates
    mother = build_campo_mother_field(args.q2_summary, selected_coordinates_path=selected_coordinates)
    cache = EvaluationCache()
    if args.smoke:
        coarse = smoke_profile()
        reference = smoke_profile()
        final = smoke_profile()
        maximum_cycles = min(args.max_cycles, 1)
        prune_rounds = min(args.prune_rounds, 1)
        prune_pairs = min(args.prune_pairs_per_round, 2)
    else:
        coarse = coarse_profile()
        reference = medium_profile()
        final = formal_profile()
        maximum_cycles = args.max_cycles
        prune_rounds = args.prune_rounds
        prune_pairs = args.prune_pairs_per_round
    print(f'读取 Campo 镜场：{mother.mirror_count} 面，区域镜数={mother.zone_counts}，区域环数={mother.zone_ring_counts}')
    radial: SearchOutcome = optimize_continuous_design(mother=mother, coarse_profile=coarse, reference_profile=reference, target_power_mw=args.target_power, include_azimuth=False, monotone=True, maximum_cycles_per_level=maximum_cycles, cache=cache, progress=print)
    model_candidates: list[tuple[str, ContinuousDesign, HeterogeneousEvaluation]] = [('q2-uniform', radial.baseline_design, radial.baseline_evaluation), ('campo-monotone-radial', radial.best_design, radial.best_evaluation)]
    stages = list(radial.stage_evaluations)
    current_design = radial.best_design
    current_evaluation = radial.best_evaluation
    current_name = 'campo-monotone-radial'
    use_azimuth = args.azimuth_mode == 'on' or (args.azimuth_mode == 'auto' and radial.diagnostics.azimuth_recommended)
    if use_azimuth:
        (current_design, current_evaluation, _) = refine_design_parameters(mother=mother, initial_design=current_design, coarse_profile=coarse, reference_profile=reference, parameters=('size_azimuth',), steps=(0.04, 0.02, 0.01), stage='azimuth-size', target_power_mw=args.target_power, monotone=True, maximum_cycles_per_level=maximum_cycles, cache=cache, progress=print)
        (current_design, current_evaluation, _) = refine_design_parameters(mother=mother, initial_design=current_design, coarse_profile=coarse, reference_profile=reference, parameters=('height_azimuth',), steps=(0.4, 0.2, 0.1), stage='azimuth-height', target_power_mw=args.target_power, monotone=True, maximum_cycles_per_level=maximum_cycles, cache=cache, progress=print)
        current_name = 'campo-monotone-azimuth'
        model_candidates.append((current_name, current_design, current_evaluation))
        stages.append((current_name, current_evaluation))
    if not args.skip_relaxed:
        (relaxed_design, relaxed_evaluation, _) = refine_design_parameters(mother=mother, initial_design=current_design, coarse_profile=coarse, reference_profile=reference, parameters=('size_zone1_slope', 'size_zone2_slope'), steps=(0.02, 0.01), stage='relaxed-size', target_power_mw=args.target_power, monotone=False, maximum_cycles_per_level=maximum_cycles, cache=cache, progress=print)
        (relaxed_design, relaxed_evaluation, _) = refine_design_parameters(mother=mother, initial_design=relaxed_design, coarse_profile=coarse, reference_profile=reference, parameters=('height_zone1_slope', 'height_zone2_slope'), steps=(0.2, 0.1), stage='relaxed-height', target_power_mw=args.target_power, monotone=False, maximum_cycles_per_level=maximum_cycles, cache=cache, progress=print)
        model_candidates.append(('campo-relaxed', relaxed_design, relaxed_evaluation))
        stages.append(('campo-relaxed', relaxed_evaluation))
        if relaxed_evaluation.is_feasible(args.target_power) and relaxed_evaluation.unit_area_power_kw_m2 > current_evaluation.unit_area_power_kw_m2:
            current_name = 'campo-relaxed'
            current_design = relaxed_design
            current_evaluation = relaxed_evaluation
    pruned = current_evaluation
    if prune_rounds and pruned.is_feasible(args.target_power):
        pruning = prune_symmetric_pairs(mother=mother, initial=pruned, profile=reference, target_power_mw=args.target_power, maximum_rounds=prune_rounds, maximum_pairs_per_round=prune_pairs, cache=cache)
        pruned = pruning.best
        print(f'外边界对称删镜接受 {len(pruning.steps)} 轮，保留 {pruned.mirror_count} 面')
        if pruning.steps:
            stages.append(('outer-boundary-pruned', pruned))
    (selected_name, selected_design, selected, formal_stages) = _formal_candidates(candidates=model_candidates, pruned=pruned, pruned_design=current_design, mother=mother, profile=final, target_power_mw=args.target_power, cache=cache)
    stages.extend(formal_stages)
    stages.append(('formal-final', selected))
    written = write_question3_results(output_dir=args.output, mother=mother, design=selected_design, evaluation=selected, result3_template=args.result3_template, stages=stages, diagnostics=radial.diagnostics, model_name=selected_name, legacy_comparison=legacy_comparison)
    if args.run_validation and (not args.smoke):
        dense_settings = dense_profile()
        dense = _reevaluate(source=selected, profile=dense_settings, mother=mother, cache=cache)
        sensitivity_settings = replace(dense_settings, name='q3-dense-100m', solver=replace(dense_settings.solver, neighbor_radius_m=100.0))
        sensitivity = _reevaluate(source=selected, profile=sensitivity_settings, mother=mother, cache=cache)
        written['dense_validation'] = write_dense_validation(output_dir=args.output, evaluation=dense, profile=dense_settings, sensitivity_evaluations=((sensitivity_settings, sensitivity),))
    print('\n第三问结果' if not args.smoke else '\n第三问烟雾测试结果')
    print(f'最终模型：{selected_name}')
    print(f'镜子数：{selected.mirror_count}')
    print(f'总镜面面积：{selected.total_area_m2:.3f} m²')
    print(f'年平均输出热功率：{selected.annual_power_mw:.6f} MW')
    print(f'单位镜面面积年平均输出：{selected.unit_area_power_kw_m2:.6f} kW/m²')
    print(f'方位项诊断 RMSE 降幅：{100.0 * radial.diagnostics.relative_rmse_reduction:.3f}%')
    for path in written.values():
        print(f'输出：{path}')
    return 0

def main() -> None:
    raise SystemExit(run())

if __name__ == "__main__":
    raise SystemExit(run())
